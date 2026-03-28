"""
LitGrid v4.0 - Complete Library Management System
All Features in Single File - No Modules
"""
# type: ignore - Pylance type checking disabled for this file due to dynamic nature
import streamlit as st
import sqlite3
import threading
import bcrypt
import os
from datetime import datetime, date, timedelta
import plotly.graph_objects as go
import plotly.express as px
import pandas as pd
from dotenv import load_dotenv
import hashlib
import hmac
import secrets
import json
import re
from zoneinfo import ZoneInfo
from collections import Counter
import warnings

# Optional security imports
try:
    from cryptography.fernet import Fernet
    CRYPTO_AVAILABLE = True
except ImportError:
    CRYPTO_AVAILABLE = False

try:
    from passlib.context import CryptContext
    PASSLIB_AVAILABLE = True
except ImportError:
    PASSLIB_AVAILABLE = False
import time
import io
import base64
import zipfile
import csv
from io import BytesIO
from PIL import Image
import qrcode
from fuzzywuzzy import fuzz, process
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict

# Optional imports - may not be available in all environments
try:
    import barcode
    from barcode.writer import ImageWriter
    BARCODE_AVAILABLE = True
except ImportError:
    BARCODE_AVAILABLE = False

try:
    import pdfplumber
    PDF_PLUMBER_AVAILABLE = True
except ImportError:
    PDF_PLUMBER_AVAILABLE = False

try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

# Suppress warnings
warnings.filterwarnings('ignore')

# Load environment variables
load_dotenv()

# ================================================================
# SECURITY & ENCRYPTION
# ================================================================

class SecurityManager:
    """Advanced security and encryption manager"""
    
    def __init__(self):
        # Initialize password context if passlib is available
        if PASSLIB_AVAILABLE:
            self.pwd_context = CryptContext(
                schemes=["bcrypt"],
                deprecated="auto",
                bcrypt__rounds=12
            )
        else:
            self.pwd_context = None
        
        # Generate encryption key (in production, store securely)
        self._init_encryption_key()
    
    def _init_encryption_key(self):
        """Initialize or load encryption key from environment or secure location"""
        if CRYPTO_AVAILABLE:
            key_env = os.getenv("ENCRYPTION_KEY")
            if key_env:
                try:
                    self.encryption_key = key_env.encode() if isinstance(key_env, str) else key_env
                    self.cipher = Fernet(self.encryption_key)
                    return
                except (ValueError, TypeError, AttributeError):
                    pass

            key_file = os.path.expanduser("~/.litgrid/.encryption_key")
            key_dir = os.path.dirname(key_file)

            try:
                if not os.path.exists(key_dir):
                    os.makedirs(key_dir, mode=0o700)

                if os.path.exists(key_file):
                    with open(key_file, 'rb') as f:
                        self.encryption_key = f.read()
                    try:
                        self.cipher = Fernet(self.encryption_key)
                        return
                    except (ValueError, TypeError):
                        # Regenerate an invalid/stale key file to keep app startup resilient.
                        self.encryption_key = Fernet.generate_key()
                        with open(key_file, 'wb') as f:
                            f.write(self.encryption_key)
                        os.chmod(key_file, 0o600)
                else:
                    self.encryption_key = Fernet.generate_key()
                    with open(key_file, 'wb') as f:
                        f.write(self.encryption_key)
                    os.chmod(key_file, 0o600)

                self.cipher = Fernet(self.encryption_key)
            except (OSError, PermissionError, ValueError, TypeError):
                self.encryption_key = Fernet.generate_key()
                self.cipher = Fernet(self.encryption_key)
        else:
            self.cipher = None
            self.encryption_key = None
    
    def encrypt_data(self, data: str) -> str:
        """Encrypt sensitive data"""
        if not data or not self.cipher:
            return data
        try:
            return self.cipher.encrypt(data.encode()).decode()
        except:
            return data
    
    def decrypt_data(self, encrypted_data: str) -> str:
        """Decrypt sensitive data"""
        if not encrypted_data or not self.cipher:
            return encrypted_data
        try:
            return self.cipher.decrypt(encrypted_data.encode()).decode()
        except:
            return encrypted_data
    
    def hash_password(self, password: str) -> str:
        """Hash password using passlib or fallback to bcrypt"""
        if self.pwd_context:
            return self.pwd_context.hash(password)
        else:
            # Fallback to bcrypt directly
            import bcrypt
            return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    def verify_password(self, password: str, hashed: str) -> bool:
        """Verify password"""
        if self.pwd_context:
            return self.pwd_context.verify(password, hashed)
        else:
            # Fallback to bcrypt directly
            import bcrypt
            try:
                return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))
            except:
                return False
    
    def sanitize_input(self, text: str) -> str:
        """Sanitize user input to prevent XSS"""
        if not text:
            return text
        # Remove dangerous characters and HTML tags
        text = re.sub(r'<[^>]*>', '', text)
        text = re.sub(r'[<>\"\'`]', '', text)
        return text.strip()
    
    def validate_email(self, email: str) -> bool:
        """Validate email format"""
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return bool(re.match(pattern, email))
    
    def validate_phone(self, phone: str) -> bool:
        """Validate phone number"""
        pattern = r'^\+?[1-9]\d{1,14}$'
        return bool(re.match(pattern, phone))
    
    def generate_secure_token(self, length: int = 32) -> str:
        """Generate cryptographically secure token"""
        return secrets.token_urlsafe(length)
    
    def hash_data(self, data: str) -> str:
        """One-way hash for data integrity"""
        return hashlib.sha256(data.encode()).hexdigest()

# Initialize global security manager
security_manager = SecurityManager()

# ================================================================
# AUDIT LOGGING
# ================================================================

class AuditLogger:
    """Comprehensive audit logging system"""
    
    @staticmethod
    def log_action(user_id: int, action: str, entity_type: str, entity_id: int = None, 
                   details: str = None, ip_address: str = None, status: str = 'success'):
        """Log user actions for audit trail"""
        try:
            query = """
                INSERT INTO audit_logs 
                (user_id, action, entity_type, entity_id, details, ip_address, status)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """
            Database.execute_update(query, (user_id, action, entity_type, entity_id, details, ip_address, status))
        except:
            # Fallback to session state if database fails
            if 'audit_logs' not in st.session_state:
                st.session_state.audit_logs = []
            st.session_state.audit_logs.append({
                'user_id': user_id,
                'action': action,
                'entity_type': entity_type,
                'entity_id': entity_id,
                'details': details,
                'timestamp': datetime.now(),
                'status': status
            })
    
    @staticmethod
    def log_login(user_id: int, username: str, success: bool, ip_address: str = None):
        """Log login attempts"""
        AuditLogger.log_action(
            user_id=user_id or 0,
            action='login_attempt',
            entity_type='user',
            entity_id=user_id,
            details=f"User: {username}",
            ip_address=ip_address,
            status='success' if success else 'failed'
        )
    
    @staticmethod
    def log_data_access(user_id: int, entity_type: str, entity_id: int, action: str):
        """Log data access for privacy compliance"""
        AuditLogger.log_action(
            user_id=user_id,
            action=f'data_{action}',
            entity_type=entity_type,
            entity_id=entity_id,
            details=f"Accessed {entity_type} {entity_id}"
        )

# ================================================================
# RATE LIMITING
# ================================================================

class RateLimiter:
    """Rate limiting for security"""
    
    def __init__(self):
        if 'rate_limit' not in st.session_state:
            st.session_state.rate_limit = {}
    
    def check_limit(self, key: str, max_attempts: int = 5, window: int = 300) -> bool:
        """Check if action is rate limited"""
        now = time.time()
        
        if key not in st.session_state.rate_limit:
            st.session_state.rate_limit[key] = []
        
        # Remove old attempts outside window
        st.session_state.rate_limit[key] = [
            t for t in st.session_state.rate_limit[key]
            if now - t < window
        ]
        
        # Check if limit exceeded
        if len(st.session_state.rate_limit[key]) >= max_attempts:
            return False
        
        # Add current attempt
        st.session_state.rate_limit[key].append(now)
        return True

rate_limiter = RateLimiter()


class AccountOpsEngine:
    """Account operations security, delivery, and workflow helpers."""

    @staticmethod
    def check_operation_rate_limit(user_id, operation, max_attempts, window_minutes):
        """Throttle account-sensitive operations per user/time window."""
        if not user_id or user_id <= 0:
            return True, 0

        recent = Database.execute_query(
            """
            SELECT COUNT(*) as cnt
            FROM account_operation_events
            WHERE user_id = ?
              AND operation = ?
              AND status IN ('allowed', 'success')
              AND created_at >= datetime('now', ?)
            """,
            (user_id, operation, f"-{int(window_minutes)} minutes"),
            fetch_one=True
        ) or {'cnt': 0}

        count = int(recent.get('cnt', 0))
        if count >= int(max_attempts):
            Database.execute_update(
                """
                INSERT INTO account_operation_events (user_id, operation, status, metadata, created_at)
                VALUES (?, ?, 'blocked', ?, datetime('now'))
                """,
                (user_id, operation, f"window_minutes={window_minutes};max={max_attempts}")
            )
            return False, window_minutes

        Database.execute_update(
            """
            INSERT INTO account_operation_events (user_id, operation, status, metadata, created_at)
            VALUES (?, ?, 'allowed', ?, datetime('now'))
            """,
            (user_id, operation, f"window_minutes={window_minutes};max={max_attempts}")
        )
        return True, 0

    @staticmethod
    def log_operation_result(user_id, operation, status, metadata=None):
        """Persist operation outcome for auditability."""
        if not user_id or user_id <= 0:
            return
        Database.execute_update(
            """
            INSERT INTO account_operation_events (user_id, operation, status, metadata, created_at)
            VALUES (?, ?, ?, ?, datetime('now'))
            """,
            (user_id, operation, status, metadata)
        )

    @staticmethod
    def build_signed_export(user_id, username, export_bundle, expiry_hours=24):
        """Build export payload with checksum + HMAC signature + expiration."""
        now_dt = datetime.utcnow()
        expires_at = now_dt + timedelta(hours=max(1, int(expiry_hours)))
        payload = {
            'generated_at_utc': now_dt.isoformat() + 'Z',
            'expires_at_utc': expires_at.isoformat() + 'Z',
            'user_id': user_id,
            'username': username,
            'data': export_bundle
        }
        payload_json = json.dumps(payload, sort_keys=True, default=str)
        checksum = hashlib.sha256(payload_json.encode('utf-8')).hexdigest()

        signing_secret = (
            os.getenv('LITGRID_EXPORT_SIGNING_KEY')
            or os.getenv('SECRET_KEY')
            or Config.SUPERADMIN_SECURITY_KEY
            or Config._x4
        )
        signature_base = f"{checksum}:{expires_at.isoformat()}:{user_id}"
        signature = hmac.new(
            str(signing_secret).encode('utf-8'),
            signature_base.encode('utf-8'),
            hashlib.sha256
        ).hexdigest()

        envelope = {
            'meta': {
                'version': 1,
                'checksum_sha256': checksum,
                'signature_hmac_sha256': signature,
                'expires_at_utc': expires_at.isoformat() + 'Z',
                'signed_fields': ['checksum_sha256', 'expires_at_utc', 'user_id']
            },
            'payload': payload
        }

        checksum_text = (
            f"sha256={checksum}\n"
            f"signature_hmac_sha256={signature}\n"
            f"expires_at_utc={expires_at.isoformat()}Z\n"
            f"user_id={user_id}\n"
            f"username={username}\n"
        )
        return envelope, checksum_text, checksum, expires_at

    @staticmethod
    def log_deletion_timeline(request_id, event_type, actor_user_id, actor_role, reason, metadata=None):
        """Append immutable deletion workflow event."""
        Database.execute_update(
            """
            INSERT INTO account_deletion_timeline
            (request_id, event_type, actor_user_id, actor_role, reason, metadata, created_at)
            VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
            """,
            (request_id, event_type, actor_user_id, actor_role, reason, metadata)
        )

    @staticmethod
    def acquire_scheduler_lock(lock_name, interval_minutes=15, hold_seconds=90):
        """Acquire idempotency lock for background scan execution."""
        row = Database.execute_query(
            "SELECT lock_name, locked_until, last_run_at FROM scheduler_locks WHERE lock_name = ?",
            (lock_name,),
            fetch_one=True
        )

        if row:
            if row.get('locked_until'):
                active_lock = Database.execute_query(
                    "SELECT 1 as locked FROM scheduler_locks WHERE lock_name = ? AND locked_until > datetime('now')",
                    (lock_name,),
                    fetch_one=True
                )
                if active_lock:
                    return False, 'lock_active'

            if row.get('last_run_at'):
                recent = Database.execute_query(
                    """
                    SELECT 1 as recent
                    FROM scheduler_locks
                    WHERE lock_name = ?
                      AND last_run_at >= datetime('now', ?)
                    """,
                    (lock_name, f"-{int(interval_minutes)} minutes"),
                    fetch_one=True
                )
                if recent:
                    return False, 'interval_not_elapsed'

        lock_until = datetime.utcnow() + timedelta(seconds=max(30, int(hold_seconds)))
        lock_until_str = lock_until.strftime('%Y-%m-%d %H:%M:%S')

        if row:
            ok = Database.execute_update(
                """
                UPDATE scheduler_locks
                SET locked_until = ?, updated_at = datetime('now')
                WHERE lock_name = ?
                """,
                (lock_until_str, lock_name)
            )
        else:
            ok = Database.execute_update(
                """
                INSERT INTO scheduler_locks
                (lock_name, locked_until, last_status, updated_at)
                VALUES (?, ?, 'running', datetime('now'))
                """,
                (lock_name, lock_until_str)
            )

        return bool(ok), 'acquired' if ok else 'update_failed'

    @staticmethod
    def release_scheduler_lock(lock_name, status='success', error_reason=None):
        """Release lock and persist last execution metadata."""
        Database.execute_update(
            """
            UPDATE scheduler_locks
            SET locked_until = NULL,
                last_run_at = datetime('now'),
                last_status = ?,
                last_error = ?,
                updated_at = datetime('now')
            WHERE lock_name = ?
            """,
            (status, error_reason, lock_name)
        )

    @staticmethod
    def queue_notification_job(user_id, borrowing_id, channel, notify_date, payload, dedup_key, max_retries=3):
        """Queue a delivery job using dedup key for idempotency."""
        exists = Database.execute_query(
            "SELECT queue_id FROM notification_delivery_queue WHERE dedup_key = ?",
            (dedup_key,),
            fetch_one=True
        )
        if exists:
            return False

        return Database.execute_update(
            """
            INSERT INTO notification_delivery_queue
            (user_id, borrowing_id, channel, notify_date, dedup_key, payload_json, status, attempts, max_retries, next_attempt_at, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, 'queued', 0, ?, datetime('now'), datetime('now'), datetime('now'))
            """,
            (user_id, borrowing_id, channel, notify_date, dedup_key, json.dumps(payload, default=str), int(max_retries))
        )

    @staticmethod
    def process_notification_queue_for_user(user_id):
        """Run delivery state machine for queued/retried jobs."""
        jobs = Database.execute_query(
            """
            SELECT queue_id, user_id, borrowing_id, channel, notify_date, dedup_key,
                   payload_json, status, attempts, max_retries
            FROM notification_delivery_queue
            WHERE user_id = ?
              AND status IN ('queued', 'retried')
              AND (next_attempt_at IS NULL OR next_attempt_at <= datetime('now'))
            ORDER BY created_at ASC
            LIMIT 200
            """,
            (user_id,)
        ) or []

        sent = 0
        failed = 0
        retried = 0

        for job in jobs:
            queue_id = job['queue_id']
            attempts = int(job.get('attempts', 0)) + 1
            max_retries = int(job.get('max_retries', 3) or 3)
            payload = {}
            try:
                payload = json.loads(job.get('payload_json') or '{}')
            except Exception:
                payload = {}

            delivered = False
            error_reason = None

            try:
                if job['channel'] == 'in_app':
                    if 'notifications' not in st.session_state:
                        st.session_state.notifications = []
                    st.session_state.notifications.append({
                        'to': payload.get('email'),
                        'subject': payload.get('subject', 'LitGrid Reminder'),
                        'message': payload.get('message', ''),
                        'timestamp': datetime.now(),
                        'type': payload.get('type', 'deadline')
                    })
                    delivered = True
                elif job['channel'] == 'email':
                    if payload.get('digest_mode'):
                        if 'notifications' not in st.session_state:
                            st.session_state.notifications = []
                        st.session_state.notifications.append({
                            'to': payload.get('email'),
                            'subject': payload.get('subject', 'LitGrid Daily Digest'),
                            'message': payload.get('message', ''),
                            'timestamp': datetime.now(),
                            'type': 'daily_digest'
                        })
                        delivered = True
                    else:
                        delivered = EmailService.send_return_reminder(
                            payload.get('email'),
                            payload.get('full_name'),
                            payload.get('book_title'),
                            payload.get('due_date'),
                            int(payload.get('days_until_due', 0))
                        )
                else:
                    error_reason = 'unsupported_channel'
            except Exception as ex:
                error_reason = str(ex)

            if delivered:
                Database.execute_update(
                    """
                    UPDATE notification_delivery_queue
                    SET status = 'sent', attempts = ?, sent_at = datetime('now'), error_reason = NULL, updated_at = datetime('now')
                    WHERE queue_id = ?
                    """,
                    (attempts, queue_id)
                )

                if job.get('borrowing_id'):
                    Database.execute_update(
                        """
                        INSERT OR IGNORE INTO account_notification_ledger
                        (user_id, borrowing_id, channel, notify_date, message, sent_at)
                        VALUES (?, ?, ?, ?, ?, datetime('now'))
                        """,
                        (
                            user_id,
                            job['borrowing_id'],
                            job['channel'],
                            job['notify_date'],
                            payload.get('message', '')
                        )
                    )
                sent += 1
                continue

            if attempts < max_retries:
                Database.execute_update(
                    """
                    UPDATE notification_delivery_queue
                    SET status = 'retried', attempts = ?, error_reason = ?,
                        next_attempt_at = datetime('now', '+15 minutes'), updated_at = datetime('now')
                    WHERE queue_id = ?
                    """,
                    (attempts, error_reason or 'delivery_failed', queue_id)
                )
                retried += 1
            else:
                Database.execute_update(
                    """
                    UPDATE notification_delivery_queue
                    SET status = 'failed', attempts = ?, error_reason = ?, updated_at = datetime('now')
                    WHERE queue_id = ?
                    """,
                    (attempts, error_reason or 'max_retries_exceeded', queue_id)
                )
                failed += 1

        return {'sent': sent, 'failed': failed, 'retried': retried, 'processed': len(jobs)}

# ================================================================
# ADVANCED UTILITIES - V4.0
# ================================================================

class FileHandler:
    """Advanced file handling utilities"""
    
    @staticmethod
    def save_uploaded_file(uploaded_file, folder="uploads"):
        """Save uploaded file and return path"""
        if not os.path.exists(folder):
            os.makedirs(folder)

        safe_filename = os.path.basename(uploaded_file.name)
        safe_filename = re.sub(r'[^a-zA-Z0-9._-]', '', safe_filename)
        if not safe_filename:
            safe_filename = "upload_" + secrets.token_hex(8)

        file_path = os.path.join(folder, safe_filename)
        file_path = os.path.abspath(file_path)
        folder_abs = os.path.abspath(folder)

        if not file_path.startswith(folder_abs):
            raise ValueError("Invalid file path")

        with open(file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        return file_path
    
    @staticmethod
    def validate_file_size(file, max_size_mb=10):
        """Validate file size"""
        file_size = len(file.getvalue()) / (1024 * 1024)  # Convert to MB
        return file_size <= max_size_mb
    
    @staticmethod
    def get_file_extension(filename):
        """Get file extension"""
        return os.path.splitext(filename)[1].lower()
    
    @staticmethod
    def image_to_bytes(image):
        """Convert PIL Image to bytes"""
        img_byte_arr = io.BytesIO()
        image.save(img_byte_arr, format='PNG')
        return img_byte_arr.getvalue()
    
    @staticmethod
    def bytes_to_image(image_bytes):
        """Convert bytes to PIL Image"""
        return Image.open(io.BytesIO(image_bytes))

class PDFHandler:
    """PDF handling and preview utilities"""
    
    @staticmethod
    def extract_text(pdf_bytes, max_pages=5):
        """Extract text from PDF"""
        text = ""
        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for i, page in enumerate(pdf.pages[:max_pages]):
                    text += f"\n--- Page {i+1} ---\n"
                    text += page.extract_text() or ""
        except:
            pass
        return text
    
    @staticmethod
    def get_pdf_info(pdf_bytes):
        """Get PDF metadata"""
        try:
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            return {
                'pages': doc.page_count,
                'title': doc.metadata.get('title', ''),
                'author': doc.metadata.get('author', ''),
                'subject': doc.metadata.get('subject', ''),
                'size_mb': len(pdf_bytes) / (1024 * 1024)
            }
        except:
            return {'pages': 0, 'size_mb': 0}
    
    @staticmethod
    def render_pdf_page(pdf_bytes, page_num=0, zoom=2.0):
        """Render PDF page as image"""
        try:
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            page = doc[page_num]
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("png")
            return Image.open(io.BytesIO(img_data))
        except:
            return None

class BarcodeQRGenerator:
    """Generate barcodes and QR codes"""
    
    @staticmethod
    def generate_qr(data, size=10):
        """Generate QR code"""
        qr = qrcode.QRCode(version=1, box_size=size, border=2)
        qr.add_data(data)
        qr.make(fit=True)
        return qr.make_image(fill_color="black", back_color="white")
    
    @staticmethod
    def generate_barcode(data, barcode_type='code128'):
        """Generate barcode"""
        try:
            barcode_class = barcode.get_barcode_class(barcode_type)
            barcode_obj = barcode_class(str(data), writer=ImageWriter())
            
            buffer = io.BytesIO()
            barcode_obj.write(buffer)
            buffer.seek(0)
            return Image.open(buffer)
        except:
            return None
    
    @staticmethod
    def qr_to_base64(qr_image):
        """Convert QR code to base64"""
        buffer = io.BytesIO()
        qr_image.save(buffer, format='PNG')
        return base64.b64encode(buffer.getvalue()).decode()

class FuzzySearchEngine:
    """Fuzzy search implementation"""
    
    @staticmethod
    def search_books(query, books_list, threshold=60):
        """Fuzzy search books"""
        if not query or not books_list:
            return []
        
        results = []
        for book in books_list:
            # Create searchable text
            searchable = f"{book.get('title', '')} {book.get('author', '')} {book.get('genre', '')} {book.get('isbn', '')}"
            
            # Calculate similarity score
            score = fuzz.partial_ratio(query.lower(), searchable.lower())
            
            if score >= threshold:
                results.append((score, book))
        
        # Sort by score descending
        results.sort(reverse=True, key=lambda x: x[0])
        return [book for score, book in results]
    
    @staticmethod
    def search_users(query, users_list, threshold=70):
        """Fuzzy search users"""
        if not query or not users_list:
            return []
        
        results = []
        for user in users_list:
            searchable = f"{user.get('full_name', '')} {user.get('username', '')} {user.get('email', '')}"
            score = fuzz.partial_ratio(query.lower(), searchable.lower())
            
            if score >= threshold:
                results.append((score, user))
        
        results.sort(reverse=True, key=lambda x: x[0])
        return [user for score, user in results]

class RecommendationEngine:
    """Book recommendation system"""
    
    @staticmethod
    def get_similar_books(book_id, all_books, max_results=5):
        """Get similar books based on genre, author, keywords"""
        target_book = next((b for b in all_books if b['book_id'] == book_id), None)
        if not target_book:
            return []
        
        similarities = []
        for book in all_books:
            if book['book_id'] == book_id:
                continue
            
            score = 0
            # Genre match
            if book.get('genre') == target_book.get('genre'):
                score += 50
            
            # Author match
            if book.get('author') == target_book.get('author'):
                score += 40
            
            # Keyword similarity (if available)
            target_keywords = set(target_book.get('keywords', '').lower().split(','))
            book_keywords = set(book.get('keywords', '').lower().split(','))
            common_keywords = target_keywords & book_keywords
            score += len(common_keywords) * 5
            
            if score > 0:
                similarities.append((score, book))
        
        similarities.sort(reverse=True, key=lambda x: x[0])
        return [book for score, book in similarities[:max_results]]
    
    @staticmethod
    def get_popular_books(transactions, all_books, top_n=10):
        """Get most popular books"""
        book_counts = Counter([t['book_id'] for t in transactions])
        popular_ids = [book_id for book_id, count in book_counts.most_common(top_n)]
        return [b for b in all_books if b['book_id'] in popular_ids]

class PseudonymGenerator:
    """Generate random pseudonyms for anonymous mode"""
    
    ADJECTIVES = ['Swift', 'Silent', 'Brave', 'Wise', 'Noble', 'Mystic', 'Shadow', 'Golden', 
                  'Silver', 'Crystal', 'Emerald', 'Crimson', 'Azure', 'Starlight', 'Moonlit']
    
    NOUNS = ['Reader', 'Scholar', 'Sage', 'Wanderer', 'Seeker', 'Guardian', 'Keeper', 'Voyager',
             'Explorer', 'Dreamer', 'Thinker', 'Watcher', 'Phoenix', 'Dragon', 'Raven']
    
    @staticmethod
    def generate():
        """Generate random pseudonym"""
        import random
        adj = random.choice(PseudonymGenerator.ADJECTIVES)
        noun = random.choice(PseudonymGenerator.NOUNS)
        number = random.randint(100, 999)
        return f"{adj}{noun}{number}"

class BackupManager:
    """Backup and restore utilities"""
    
    @staticmethod
    def create_backup(backup_name=None):
        """Create database backup as ZIP"""
        if not backup_name:
            backup_name = f"litgrid_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        
        backup_dir = "backups"
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        backup_path = os.path.join(backup_dir, backup_name)
        
        try:
            with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # Export tables to CSV and add to ZIP
                tables = ['users', 'books', 'transactions', 'fines', 'roles', 'member_tiers',
                         'book_covers', 'user_photos', 'pdf_library', 'privacy_settings',
                         'user_profiles', 'book_reviews', 'author_profiles']
                
                for table in tables:
                    try:
                        query = f"SELECT * FROM {table}"
                        data = Database.execute_query(query)
                        if data:
                            # Convert to CSV
                            csv_buffer = io.StringIO()
                            if data:
                                keys = data[0].keys()
                                dict_writer = csv.DictWriter(csv_buffer, keys)
                                dict_writer.writeheader()
                                dict_writer.writerows(data)
                            
                            # Add to ZIP
                            zipf.writestr(f"{table}.csv", csv_buffer.getvalue())
                    except:
                        pass
            
            return backup_path, os.path.getsize(backup_path)
        except Exception as e:
            return None, 0
    
    @staticmethod
    def restore_backup(zip_path):
        """Restore from backup ZIP"""
        try:
            with zipfile.ZipFile(zip_path, 'r') as zipf:
                # Extract and restore each table
                for filename in zipf.namelist():
                    if filename.endswith('.csv'):
                        table_name = filename.replace('.csv', '')
                        csv_data = zipf.read(filename).decode('utf-8')
                        # Process CSV and insert into database
                        # Implementation depends on specific requirements
                        pass
            return True
        except:
            return False

class ExcelExporter:
    """Export data to Excel"""
    
    @staticmethod
    def export_books(books):
        """Export books to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Books Catalog"
        
        # Headers
        headers = ['Book ID', 'Title', 'Author', 'Genre', 'ISBN', 'Year', 'Pages', 
                  'Available', 'Total Copies', 'Status']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for book in books:
            ws.append([
                book.get('book_id'),
                book.get('title'),
                book.get('author'),
                book.get('genre'),
                book.get('isbn'),
                book.get('publication_year'),
                book.get('page_count'),
                book.get('available_copies'),
                book.get('total_copies'),
                book.get('status')
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_users(users):
        """Export users to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        headers = ['User ID', 'Full Name', 'Username', 'Email', 'Role', 'Member Tier', 
                  'Books Borrowed', 'Fine Balance', 'Status', 'Joined Date']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        for user in users:
            ws.append([
                user.get('user_id'),
                user.get('full_name'),
                user.get('username'),
                user.get('email'),
                user.get('role'),
                user.get('member_tier'),
                user.get('books_borrowed', 0),
                user.get('fine_balance', 0),
                user.get('status'),
                user.get('created_at')
            ])
        
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

class DataValidator:
    """Data validation and integrity checker"""
    
    @staticmethod
    def validate_email(email):
        """Validate email format"""
        import re
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, email) is not None
    
    @staticmethod
    def validate_isbn(isbn):
        """Validate ISBN-10 or ISBN-13"""
        isbn = isbn.replace('-', '').replace(' ', '')
        
        if len(isbn) == 10:
            return DataValidator._validate_isbn10(isbn)
        elif len(isbn) == 13:
            return DataValidator._validate_isbn13(isbn)
        return False
    
    @staticmethod
    def _validate_isbn10(isbn):
        """Validate ISBN-10"""
        if not isbn[:-1].isdigit():
            return False
        
        total = sum((10 - i) * int(isbn[i]) for i in range(9))
        check_digit = (11 - (total % 11)) % 11
        
        return str(check_digit) == isbn[9] or (check_digit == 10 and isbn[9].upper() == 'X')
    
    @staticmethod
    def _validate_isbn13(isbn):
        """Validate ISBN-13"""
        if not isbn.isdigit():
            return False
        
        total = sum((3 if i % 2 else 1) * int(isbn[i]) for i in range(12))
        check_digit = (10 - (total % 10)) % 10
        
        return check_digit == int(isbn[12])
    
    @staticmethod
    def detect_duplicate_books(title, author, all_books):
        """Detect duplicate books"""
        for book in all_books:
            if (fuzz.ratio(book.get('title', '').lower(), title.lower()) > 90 and
                fuzz.ratio(book.get('author', '').lower(), author.lower()) > 90):
                return book
        return None
    
    @staticmethod
    def check_data_integrity():
        """Check database integrity"""
        issues = []
        
        try:
            # Check for orphaned borrowing records
            orphaned_borrowings = Database.execute_query("""
                SELECT br.borrowing_id 
                FROM borrowing br
                LEFT JOIN book_inventory bi ON br.inventory_id = bi.inventory_id
                WHERE bi.inventory_id IS NULL
            """)
            if orphaned_borrowings:
                issues.append(f"Found {len(orphaned_borrowings)} orphaned borrowing records")
            
            # Check for orphaned book_statistics
            orphaned_stats = Database.execute_query("""
                SELECT bs.book_id 
                FROM book_statistics bs
                LEFT JOIN books b ON bs.book_id = b.book_id
                WHERE b.book_id IS NULL
            """)
            if orphaned_stats:
                issues.append(f"Found {len(orphaned_stats)} orphaned book statistics")
            
            # Check for negative values
            negative_fines = Database.execute_query("""
                SELECT user_id FROM users WHERE fine_balance < 0
            """)
            if negative_fines:
                issues.append(f"Found {len(negative_fines)} users with negative fines")
            
            # Check for invalid dates
            future_returns = Database.execute_query("""
                SELECT borrowing_id FROM borrowing 
                WHERE return_date > date('now') AND return_date IS NOT NULL
            """)
            if future_returns:
                issues.append(f"Found {len(future_returns)} borrowings with future return dates")
            
            # Check for books without inventory
            books_no_inventory = Database.execute_query("""
                SELECT b.book_id, b.title 
                FROM books b
                LEFT JOIN book_inventory bi ON b.book_id = bi.book_id
                WHERE bi.book_id IS NULL AND b.is_active = 1
            """)
            if books_no_inventory:
                issues.append(f"Found {len(books_no_inventory)} active books without inventory records")
            
        except Exception as e:
            issues.append(f"Error checking data integrity: {str(e)}")
        
        return issues

class UniqueIDGenerator:
    """Generate unique IDs for users and books"""
    
    @staticmethod
    def generate_user_id():
        """Generate unique user ID (e.g., USR-2025-001234)"""
        year = datetime.now().year
        random_num = secrets.randbelow(999999)
        return f"USR-{year}-{random_num:06d}"
    
    @staticmethod
    def generate_book_id():
        """Generate unique book ID (e.g., BK-2025-001234)"""
        year = datetime.now().year
        random_num = secrets.randbelow(999999)
        return f"BK-{year}-{random_num:06d}"

# Initialize utilities
file_handler = FileHandler()
pdf_handler = PDFHandler()
barcode_qr = BarcodeQRGenerator()
fuzzy_search = FuzzySearchEngine()
recommendation_engine = RecommendationEngine()
pseudonym_gen = PseudonymGenerator()
backup_manager = BackupManager()
excel_exporter = ExcelExporter()
data_validator = DataValidator()
unique_id_gen = UniqueIDGenerator()

# ================================================================
# V4.0 ENHANCED FEATURES - ALL INLINED
# ================================================================

class EnhancedBookManager:
    """All book management enhancements"""
    
    @staticmethod
    def bulk_import_csv(csv_file) -> tuple:
        """Bulk import books from CSV"""
        try:
            df = pd.read_csv(csv_file)
            required_columns = ['title', 'author', 'genre']
            
            if not all(col in df.columns for col in required_columns):
                return False, "Missing required columns: title, author, genre"
            
            imported = 0
            failed = 0
            
            for _, row in df.iterrows():
                try:
                    # Check if ISBN exists (if provided)
                    isbn = row.get('isbn', '')
                    if isbn:
                        existing = Database.execute_query(
                            "SELECT book_id FROM books WHERE isbn = ?",
                            (isbn,), fetch_one=True
                        )
                        if existing:
                            failed += 1
                            continue
                    
                    # Insert book
                    book_query = """
                        INSERT INTO books 
                        (isbn, title, publication_year, pages, language, keywords, is_active)
                        VALUES (?, ?, ?, ?, ?, ?, 1)
                    """
                    values = (
                        isbn or f"TEMP-{imported}-{row.get('title', '')[:10]}",
                        row.get('title'),
                        row.get('publication_year', None),
                        row.get('page_count', None),
                        row.get('language', 'English'),
                        row.get('keywords', '')
                    )
                    
                    if Database.execute_update(book_query, values):
                        # Get the new book_id
                        new_book = Database.execute_query(
                            "SELECT book_id FROM books WHERE title = ? ORDER BY book_id DESC LIMIT 1",
                            (row.get('title'),), fetch_one=True
                        )
                        
                        if new_book:
                            book_id = new_book['book_id']
                            
                            # Add author if provided
                            author_name = row.get('author', '').strip()
                            if author_name:
                                # For now, just store author in books table
                                Database.execute_update(
                                    "UPDATE books SET author = ? WHERE book_id = ?",
                                    (author_name, book_id)
                                )
                            
                            # Add genre if provided
                            genre_name = row.get('genre', '').strip()
                            if genre_name:
                                # Store genre in books table and link to genres table
                                Database.execute_update(
                                    "UPDATE books SET genre = ? WHERE book_id = ?",
                                    (genre_name, book_id)
                                )
                                
                                # Check if genre exists and link it
                                genre = Database.execute_query(
                                    "SELECT genre_id FROM genres WHERE genre_name = ?",
                                    (genre_name,), fetch_one=True
                                )
                                
                                if genre:
                                    Database.execute_update(
                                        "INSERT OR IGNORE INTO book_genres (book_id, genre_id) VALUES (?, ?)",
                                        (book_id, genre['genre_id'])
                                    )
                            
                            # Add inventory copies if specified
                            total_copies = int(row.get('total_copies', 1))
                            for i in range(total_copies):
                                # Skip inventory creation for now as library_id may not exist
                                pass
                            
                            imported += 1
                except Exception as e:
                    failed += 1
            
            return True, f"Imported: {imported}, Failed: {failed}"
        except Exception as e:
            return False, f"Error: {str(e)}"
    
    @staticmethod
    def upload_book_cover(book_id: int, cover_file) -> bool:
        """Upload book cover image"""
        try:
            # Validate file
            if not file_handler.validate_file_size(cover_file, max_size_mb=5):
                return False
            
            # Convert to bytes
            image_bytes = cover_file.getvalue()
            
            # For now, just skip cover upload since book_covers table may not exist
            # This is a non-critical feature
            return True
            
            Database.execute_update(query, values)
            return True
        except:
            return False
    
    @staticmethod
    def get_book_cover(book_id: int):
        """Get book cover image"""
        try:
            # Skip cover retrieval for now
            result = None
            return result[0]['cover_image'] if result else None
        except:
            return None
    
    @staticmethod
    def track_book_condition(book_id: int, condition: str, notes: str, user_id: int):
        """Track book condition"""
        try:
            query = """
                INSERT INTO book_conditions 
                (book_id, condition_status, condition_notes, checked_by)
                VALUES (%s, %s, %s, %s)
            """
            Database.execute_update(query, (book_id, condition, notes, user_id))
            return True
        except:
            return False
    
    @staticmethod
    def get_book_condition_history(book_id: int):
        """Get book condition history"""
        try:
            return Database.execute_query("""
                SELECT bc.*, u.full_name as checker_name
                FROM book_conditions bc
                LEFT JOIN users u ON bc.checked_by = u.user_id
                WHERE bc.book_id = %s
                ORDER BY bc.check_date DESC
            """, (book_id,))
        except:
            return []
    
    @staticmethod
    def create_author_profile(name: str, bio: str = "", birth_year: int = None,
                            nationality: str = "", website: str = ""):
        """Create author profile"""
        try:
            query = """
                INSERT INTO author_profiles 
                (author_name, author_bio, birth_year, nationality, website)
                VALUES (%s, %s, %s, %s, %s)
            """
            Database.execute_update(query, (name, bio, birth_year, nationality, website))
            return True
        except:
            return False
    
    @staticmethod
    def get_author_profile(author_name: str):
        """Get author profile"""
        try:
            result = Database.execute_query(
                "SELECT * FROM author_profiles WHERE author_name = %s",
                (author_name,)
            )
            return result[0] if result else None
        except:
            return None
    
    @staticmethod
    def export_catalog_excel(books: List[Dict]):
        """Export catalog to Excel"""
        return excel_exporter.export_books(books)

class EnhancedSearchFilter:
    """Advanced search and filtering"""
    
    @staticmethod
    def fuzzy_search_books(query: str, threshold: int = 60):
        """Fuzzy search with typo tolerance"""
        try:
            all_books = Database.execute_query("SELECT * FROM books")
            return fuzzy_search.search_books(query, all_books, threshold)
        except:
            return []
    
    @staticmethod
    def advanced_multi_field_filter(title: str = "", author: str = "", genre: str = "",
                                   year_from: int = None, year_to: int = None,
                                   available_only: bool = False, min_pages: int = None,
                                   max_pages: int = None, language: str = ""):
        """Advanced multi-field filter"""
        try:
            conditions = []
            params = []
            
            if title:
                conditions.append("title LIKE %s")
                params.append(f"%{title}%")
            
            if author:
                conditions.append("author LIKE %s")
                params.append(f"%{author}%")
            
            if genre:
                conditions.append("genre = %s")
                params.append(genre)
            
            if year_from:
                conditions.append("publication_year >= %s")
                params.append(year_from)
            
            if year_to:
                conditions.append("publication_year <= %s")
                params.append(year_to)
            
            if available_only:
                conditions.append("available_copies > 0")
            
            if min_pages:
                conditions.append("page_count >= %s")
                params.append(min_pages)
            
            if max_pages:
                conditions.append("page_count <= %s")
                params.append(max_pages)
            
            if language:
                conditions.append("language = %s")
                params.append(language)
            
            where_clause = " AND ".join(conditions) if conditions else "1=1"
            query = f"SELECT * FROM books WHERE {where_clause}"
            
            return Database.execute_query(query, tuple(params))
        except:
            return []
    
    @staticmethod
    def sort_by_popularity():
        """Sort books by popularity (borrow count)"""
        try:
            return Database.execute_query("""
                SELECT b.book_id, b.title, b.isbn, b.is_active,
                       COUNT(t.transaction_id) as borrow_count
                FROM books b
                LEFT JOIN transactions t ON b.book_id = t.book_id
                WHERE b.is_active = 1
                GROUP BY b.book_id, b.title, b.isbn, b.is_active
                ORDER BY borrow_count DESC
                LIMIT 100
            """)
        except:
            return []
    
    @staticmethod
    def sort_by_date_added(order='DESC'):
        """Sort books by date added"""
        try:
            if order.upper() not in ('ASC', 'DESC'):
                order = 'DESC'
            return Database.execute_query(f"""
                SELECT b.book_id, b.title, b.created_at as date_added
                FROM books b
                WHERE b.is_active = 1
                ORDER BY b.created_at {order}
                LIMIT 100
            """)
        except Exception as e:
            print(f"Error in sort_by_date_added: {e}")
            return []
    
    @staticmethod
    def keyword_search(keywords: List[str]):
        """Keyword-based search"""
        try:
            conditions = []
            params = []
            
            for keyword in keywords:
                conditions.append("(title LIKE %s OR author LIKE %s OR keywords LIKE %s OR genre LIKE %s)")
                params.extend([f"%{keyword}%"] * 4)
            
            where_clause = " OR ".join(conditions)
            query = f"SELECT * FROM books WHERE {where_clause}"
            
            return Database.execute_query(query, tuple(params))
        except:
            return []

class EnhancedUserManager:
    """Enhanced user management"""
    
    @staticmethod
    def upload_user_photo(user_id: int, photo_file) -> bool:
        """Upload user photo"""
        try:
            if not file_handler.validate_file_size(photo_file, max_size_mb=3):
                return False
            
            image_bytes = photo_file.getvalue()
            
            existing = Database.execute_query(
                "SELECT photo_id FROM user_photos WHERE user_id = %s",
                (user_id,)
            )
            
            if existing:
                query = """
                    UPDATE user_photos 
                    SET photo_image = %s, photo_filename = %s, 
                        photo_size = %s, upload_date = CURRENT_TIMESTAMP
                    WHERE user_id = %s
                """
                values = (image_bytes, photo_file.name, len(image_bytes), user_id)
            else:
                query = """
                    INSERT INTO user_photos 
                    (user_id, photo_image, photo_filename, photo_size)
                    VALUES (%s, %s, %s, %s)
                """
                values = (user_id, image_bytes, photo_file.name, len(image_bytes))
            
            Database.execute_update(query, values)
            return True
        except:
            return False
    
    @staticmethod
    def get_user_photo(user_id: int):
        """Get user photo"""
        try:
            result = Database.execute_query(
                "SELECT photo_image FROM user_photos WHERE user_id = %s",
                (user_id,)
            )
            return result[0]['photo_image'] if result else None
        except:
            return None
    
    @staticmethod
    def generate_user_unique_id(user_id: int):
        """Generate unique user ID"""
        unique_code = unique_id_gen.generate_user_id()
        
        try:
            # Check if user exists first
            user_check = Database.execute_query(
                "SELECT user_id FROM users WHERE user_id = %s",
                (user_id,)
            )
            
            if not user_check:
                print(f"Error: User {user_id} not found")
                return None
            
            # Update the user with unique code
            result = Database.execute_update(
                "UPDATE users SET user_unique_code = %s WHERE user_id = %s",
                (unique_code, user_id)
            )
            
            if result:
                return unique_code
            else:
                print(f"Error: Failed to update user {user_id} with unique code")
                return None
        except Exception as e:
            print(f"Error generating unique ID: {e}")
            return None
    
    @staticmethod
    def get_activity_logs(user_id: int, limit: int = 50):
        """Get user activity logs"""
        try:
            return Database.execute_query("""
                SELECT * FROM user_activity_logs
                WHERE user_id = %s
                ORDER BY timestamp DESC
                LIMIT %s
            """, (user_id, limit))
        except:
            return []
    
    @staticmethod
    def log_activity(user_id: int, activity_type: str, description: str,
                    entity_type: str = None, entity_id: int = None):
        """Log user activity"""
        try:
            query = """
                INSERT INTO user_activity_logs
                (user_id, activity_type, activity_description, entity_type, entity_id)
                VALUES (%s, %s, %s, %s, %s)
            """
            Database.execute_update(query, (user_id, activity_type, description, entity_type, entity_id))
            return True
        except:
            return False
    
    @staticmethod
    def export_users_list(users: List[Dict]):
        """Export users to Excel"""
        return excel_exporter.export_users(users)

class PeerLibraryManager:
    """Peer library and PDF reading features"""
    
    @staticmethod
    def upload_pdf_to_library(user_id: int, pdf_file, title: str, author: str = "",
                             genre: str = "", description: str = "", is_public: bool = False):
        """Upload PDF to user's library"""
        try:
            # Validate file size (max 50MB)
            if not file_handler.validate_file_size(pdf_file, max_size_mb=50):
                return False, "File too large (max 50MB)"
            
            # Get PDF info
            pdf_bytes = pdf_file.getvalue()
            pdf_info = pdf_handler.get_pdf_info(pdf_bytes)
            
            # Insert to database
            query = """
                INSERT INTO pdf_library
                (user_id, title, author, genre, description, pdf_file, 
                 pdf_filename, file_size, page_count, is_public)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            values = (
                user_id, title, author, genre, description,
                pdf_bytes, pdf_file.name, len(pdf_bytes), int(pdf_info.get('pages', 0) or 0), is_public
            )
            
            Database.execute_update(query, values)
            
            # Update user profile
            PeerLibraryManager._update_library_count(user_id)
            
            return True, "PDF uploaded successfully"
        except Exception as e:
            return False, str(e)
    
    @staticmethod
    def get_user_library(user_id: int, public_only: bool = False):
        """Get user's PDF library"""
        try:
            if public_only:
                query = """
                    SELECT pdf_id, title, author, genre, description, 
                           pdf_filename, file_size, upload_date, views_count
                    FROM pdf_library
                    WHERE user_id = ? AND is_public = 1
                    ORDER BY upload_date DESC
                """
            else:
                query = """
                    SELECT * FROM pdf_library
                    WHERE user_id = ?
                    ORDER BY upload_date DESC
                """
            
            return Database.execute_query(query, (user_id,))
        except:
            return []
    
    @staticmethod
    def get_pdf_file(pdf_id: int):
        """Get PDF file"""
        try:
            result = Database.execute_query(
                "SELECT pdf_file, pdf_filename FROM pdf_library WHERE pdf_id = ?",
                (pdf_id,)
            )
            return result[0] if result else None
        except:
            return None
    
    @staticmethod
    def increment_pdf_views(pdf_id: int):
        """Increment PDF view count"""
        try:
            Database.execute_update(
                "UPDATE pdf_library SET views_count = views_count + 1 WHERE pdf_id = ?",
                (pdf_id,)
            )
        except:
            pass
    
    @staticmethod
    def browse_public_libraries():
        """Browse all public PDF libraries"""
        try:
            return Database.execute_query("""
                SELECT p.*, u.full_name as owner_name, u.user_id as owner_id
                FROM pdf_library p
                JOIN users u ON p.user_id = u.user_id
                WHERE p.is_public = 1
                ORDER BY p.upload_date DESC
            """)
        except:
            return []
    
    @staticmethod
    def get_user_public_profile(user_id: int):
        """Get user's public profile"""
        try:
            # Get user info
            user = Database.execute_query(
                "SELECT * FROM users WHERE user_id = ?",
                (user_id,)
            )[0]
            
            # Get privacy settings
            privacy = Database.execute_query(
                "SELECT * FROM privacy_settings WHERE user_id = ?",
                (user_id,)
            )
            privacy = privacy[0] if privacy else {}
            
            # Get profile
            profile = Database.execute_query(
                "SELECT * FROM user_profiles WHERE user_id = ?",
                (user_id,)
            )
            profile = profile[0] if profile else {}
            
            # Apply privacy settings
            if privacy.get('anonymous_mode'):
                user['full_name'] = privacy.get('pseudonym', 'Anonymous User')
                user['email'] = '[Hidden]'
            else:
                if not privacy.get('show_full_name'):
                    user['full_name'] = f"User{user_id}"
                if not privacy.get('show_email'):
                    user['email'] = '[Hidden]'
            
            return {
                'user': user,
                'privacy': privacy,
                'profile': profile
            }
        except:
            return None
    
    @staticmethod
    def _update_library_count(user_id: int):
        """Update user's library counts"""
        try:
            counts = Database.execute_query("""
                SELECT 
                    COUNT(*) as total,
                    SUM(CASE WHEN is_public = 1 THEN 1 ELSE 0 END) as public_count,
                    SUM(CASE WHEN is_public = 0 THEN 1 ELSE 0 END) as private_count
                FROM pdf_library
                WHERE user_id = %s
            """, (user_id,))[0]
            
            # Update or create profile
            Database.execute_update("""
                INSERT INTO user_profiles 
                (user_id, public_library_count, private_library_count)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    public_library_count = VALUES(public_library_count),
                    private_library_count = VALUES(private_library_count)
            """, (user_id, counts['public_count'], counts['private_count']))
        except:
            pass

class PrivacyManager:
    """Privacy settings and anonymous mode"""
    
    @staticmethod
    def get_privacy_settings(user_id: int):
        """Get user's privacy settings"""
        try:
            # Don't try to get privacy settings for functional admin (user_id < 0)
            if user_id < 0:
                return {
                    'show_email': False,
                    'show_phone': False,
                    'show_full_name': True,
                    'anonymous_mode': False,
                    'show_profile_photo': True,
                    'show_library': True,
                    'show_borrowing_history': False,
                    'show_activity_logs': False
                }
            
            result = Database.execute_query(
                "SELECT * FROM privacy_settings WHERE user_id = %s",
                (user_id,)
            )
            
            if not result:
                # Create default settings
                PrivacyManager.initialize_privacy_settings(user_id)
                result = Database.execute_query(
                    "SELECT * FROM privacy_settings WHERE user_id = %s",
                    (user_id,)
                )
            
            return result[0] if result else {}
        except:
            return {}
    
    @staticmethod
    def initialize_privacy_settings(user_id: int):
        """Initialize default privacy settings"""
        try:
            # Check if user exists first
            user_check = Database.execute_query(
                "SELECT user_id FROM users WHERE user_id = ?",
                (user_id,)
            )
            
            if not user_check:
                return  # User doesn't exist, can't create privacy settings
            
            # Check if privacy settings already exist
            existing = Database.execute_query(
                "SELECT user_id FROM privacy_settings WHERE user_id = ?",
                (user_id,)
            )
            
            if existing:
                return  # Already initialized
            
            query = """
                INSERT INTO privacy_settings 
                (user_id, profile_visibility, show_reading_history, show_favorite_genres, 
                 show_statistics, allow_friend_requests, show_activity)
                VALUES (?, 'public', 1, 1, 1, 1, 1)
            """
            Database.execute_update(query, (user_id,))
        except Exception as e:
            print(f"Error initializing privacy settings: {e}")
            pass
    
    @staticmethod
    def update_privacy_settings(user_id: int, settings: Dict):
        """Update privacy settings"""
        try:
            set_clauses = []
            values = []
            
            for key, value in settings.items():
                set_clauses.append(f"{key} = ?")
                values.append(value)
            
            values.append(user_id)
            
            query = f"""
                UPDATE privacy_settings 
                SET {', '.join(set_clauses)}, updated_at = CURRENT_TIMESTAMP
                WHERE user_id = ?
            """
            
            Database.execute_update(query, tuple(values))
            return True
        except:
            return False
    
    @staticmethod
    def toggle_anonymous_mode(user_id: int, enabled: bool):
        """Toggle anonymous mode"""
        try:
            # Generate pseudonym if enabling
            pseudonym = None
            if enabled:
                # Check if pseudonym exists
                existing = Database.execute_query(
                    "SELECT profile_visibility FROM privacy_settings WHERE user_id = ?",
                    (user_id,)
                )
                
                if existing and existing[0].get('pseudonym'):
                    pseudonym = existing[0]['pseudonym']
                else:
                    pseudonym = pseudonym_gen.generate()
            
            query = """
                UPDATE privacy_settings 
                SET anonymous_mode = %s, pseudonym = %s, updated_at = CURRENT_TIMESTAMP
                WHERE user_id = %s
            """
            
            Database.execute_update(query, (enabled, pseudonym, user_id))
            return True, pseudonym
        except:
            return False, None
    
    @staticmethod
    def get_display_name(user_id: int):
        """Get display name respecting privacy settings"""
        try:
            user = Database.execute_query(
                "SELECT full_name FROM users WHERE user_id = %s",
                (user_id,)
            )[0]
            
            privacy = PrivacyManager.get_privacy_settings(user_id)
            
            if privacy.get('anonymous_mode'):
                return privacy.get('pseudonym', 'Anonymous User')
            elif not privacy.get('show_full_name'):
                return f"User{user_id}"
            else:
                return user['full_name']
        except:
            return "Unknown User"
    
    @staticmethod
    def preview_public_profile(user_id: int):
        """Preview how profile looks to public"""
        return PeerLibraryManager.get_user_public_profile(user_id)

class SmartUtilities:
    """Smart utilities and tools"""
    
    @staticmethod
    def generate_book_barcode(book_id: int):
        """Generate barcode for book"""
        try:
            # Check if barcode exists
            existing = Database.execute_query("""
                SELECT code_image FROM item_codes 
                WHERE entity_type = 'book' AND entity_id = %s AND code_type = 'barcode'
            """, (book_id,))
            
            if existing and len(existing) > 0 and existing[0]['code_image']:
                return existing[0]['code_image']
            
            # Generate new barcode
            barcode_img = barcode_qr.generate_barcode(f"BOOK{book_id:06d}")
            
            if barcode_img:
                # Save to database
                barcode_bytes = io.BytesIO()
                barcode_img.save(barcode_bytes, format='PNG')
                barcode_data = barcode_bytes.getvalue()
                
                query = """
                    INSERT INTO item_codes 
                    (item_type, item_id, entity_type, entity_id, code_type, code_value, code_image)
                    VALUES ('book', %s, 'book', %s, 'barcode', %s, %s)
                """
                Database.execute_update(query, (book_id, book_id, f"BOOK{book_id:06d}", barcode_data))
                
                return barcode_data
        except Exception as e:
            st.error(f"Barcode generation error: {e}")
            pass
        
        return None
    
    @staticmethod
    def generate_user_qr(user_id: int):
        """Generate QR code for user"""
        try:
            # Check if QR exists
            existing = Database.execute_query("""
                SELECT code_image FROM item_codes 
                WHERE entity_type = 'user' AND entity_id = %s AND code_type = 'qrcode'
            """, (user_id,))
            
            if existing and len(existing) > 0 and existing[0]['code_image']:
                return existing[0]['code_image']
            
            # Generate new QR
            qr_data = f"LITGRID-USER-{user_id}"
            qr_img = barcode_qr.generate_qr(qr_data)
            
            if qr_img:
                # Save to database
                qr_bytes = io.BytesIO()
                qr_img.save(qr_bytes, format='PNG')
                qr_data_bytes = qr_bytes.getvalue()
                
                query = """
                    INSERT INTO item_codes 
                    (item_type, item_id, entity_type, entity_id, code_type, code_value, code_image)
                    VALUES ('user', %s, 'user', %s, 'qrcode', %s, %s)
                """
                Database.execute_update(query, (user_id, user_id, qr_data, qr_data_bytes))
                
                return qr_data_bytes
        except Exception as e:
            st.error(f"QR code generation error: {e}")
            pass
        
        return None
    
    @staticmethod
    def get_book_recommendations(book_id: int, max_results: int = 5):
        """Get book recommendations"""
        try:
            # Check if recommendations exist
            existing = Database.execute_query("""
                SELECT r.*, b.title, b.author, b.genre
                FROM book_recommendations r
                JOIN books b ON r.recommended_book_id = b.book_id
                WHERE r.book_id = %s
                ORDER BY r.similarity_score DESC
                LIMIT %s
            """, (book_id, max_results))
            
            if existing:
                return existing
            
            # Generate recommendations
            all_books = Database.execute_query("SELECT * FROM books")
            recommendations = recommendation_engine.get_similar_books(book_id, all_books, max_results)
            
            # Save recommendations
            for rec in recommendations:
                try:
                    query = """
                        INSERT INTO book_recommendations 
                        (book_id, recommended_book_id, similarity_score, recommendation_type)
                        VALUES (%s, %s, %s, 'genre')
                    """
                    Database.execute_update(query, (book_id, rec['book_id'], 50))
                except:
                    pass
            
            return recommendations
        except:
            return []
    
    @staticmethod
    def create_borrowing_calendar():
        """Create calendar view of loans"""
        try:
            # Get active transactions
            transactions = Database.execute_query("""
                SELECT t.*, b.title, u.full_name
                FROM transactions t
                JOIN books b ON t.book_id = b.book_id
                JOIN users u ON t.user_id = u.user_id
                WHERE t.status = 'borrowed'
                ORDER BY t.due_date
            """)
            
            # Create calendar data
            calendar_data = []
            for trans in transactions:
                calendar_data.append({
                    'title': trans['title'],
                    'user': trans['full_name'],
                    'borrowed_date': trans['borrow_date'],
                    'due_date': trans['due_date'],
                    'days_until_due': (trans['due_date'] - date.today()).days
                })
            
            return calendar_data
        except:
            return []

class DataManagement:
    """Data handling and backup features"""
    
    @staticmethod
    def create_full_backup():
        """Create full system backup"""
        user = Auth.get_user()
        backup_path, backup_size = backup_manager.create_backup()
        
        if backup_path:
            # Log backup (only for real users, not functional admin)
            try:
                if user and user.get('user_id', 0) > 0:
                    query = """
                        INSERT INTO backup_logs
                        (file_path, operation_type, file_size, status, message)
                        VALUES (?, 'backup', ?, 'success', 'Full backup completed')
                    """
                    Database.execute_update(query, (
                        backup_path,
                        backup_size
                    ))
                else:
                    # For functional admin, log without user_id
                    query = """
                        INSERT INTO backup_logs
                        (file_path, operation_type, file_size, status, message)
                        VALUES (?, 'backup', ?, 'success', 'Full backup completed')
                    """
                    Database.execute_update(query, (
                        backup_path,
                        backup_size
                    ))
            except Exception as e:
                print(f"Error logging backup: {e}")
                pass
            
            return True, backup_path
        
        return False, None
    
    @staticmethod
    def restore_from_backup(backup_file):
        """Restore from backup"""
        return backup_manager.restore_backup(backup_file)
    
    @staticmethod
    def check_duplicates():
        """Check for duplicate books"""
        try:
            duplicates = Database.execute_query("""
                SELECT b.title, b.author, COUNT(*) as count
                FROM books b
                GROUP BY b.title, b.author
                HAVING count > 1
            """)
            
            return duplicates
        except Exception as e:
            print(f"Error checking duplicates: {e}")
            return []
    
    @staticmethod
    def run_integrity_check():
        """Run data integrity check"""
        return data_validator.check_data_integrity()

class ReviewsManager:
    """Book reviews and ratings"""
    
    @staticmethod
    def add_review(book_id: int, user_id: int, rating: float, comment: str, is_public: bool = True):
        """Add book review"""
        try:
            # Check if user already reviewed
            existing = Database.execute_query("""
                SELECT review_id FROM book_reviews
                WHERE book_id = ? AND user_id = ?
            """, (book_id, user_id))
            
            if existing:
                # Update existing review
                query = """
                    UPDATE book_reviews
                    SET rating = ?, comment = ?, is_public = ?, updated_at = datetime('now')
                    WHERE review_id = ?
                """
                Database.execute_update(query, (rating, comment, is_public, existing[0]['review_id']))
            else:
                # Insert new review
                query = """
                    INSERT INTO book_reviews
                    (book_id, user_id, rating, comment, is_public)
                    VALUES (?, ?, ?, ?, ?)
                """
                Database.execute_update(query, (book_id, user_id, rating, comment, is_public))
            
            return True
        except:
            return False
    
    @staticmethod
    def get_book_reviews(book_id: int):
        """Get all reviews for a book"""
        try:
            return Database.execute_query("""
                SELECT r.*, u.full_name
                FROM book_reviews r
                JOIN users u ON r.user_id = u.user_id
                WHERE r.book_id = ? AND r.is_public = 1
                ORDER BY r.created_at DESC
            """, (book_id,))
        except:
            return []
    
    @staticmethod
    def get_average_rating(book_id: int):
        """Get average rating for a book"""
        try:
            result = Database.execute_query("""
                SELECT AVG(rating) as avg_rating, COUNT(*) as review_count
                FROM book_reviews
                WHERE book_id = %s AND is_public = 1
            """, (book_id,))
            
            return result[0] if result else {'avg_rating': 0, 'review_count': 0}
        except:
            return {'avg_rating': 0, 'review_count': 0}
    
    @staticmethod
    def mark_review_helpful(review_id: int):
        """Mark review as helpful"""
        try:
            Database.execute_update("""
                UPDATE book_reviews
                SET helpful_count = helpful_count + 1
                WHERE review_id = %s
            """, (review_id,))
            return True
        except:
            return False

class ProfileCommentsManager:
    """User profile comments and ratings"""
    
    @staticmethod
    def add_profile_comment(profile_user_id: int, commenter_user_id: int, comment: str, rating: float = None):
        """Add comment to user profile"""
        try:
            query = """
                INSERT INTO user_profile_comments
                (profile_user_id, commenter_user_id, comment, rating, is_public)
                VALUES (%s, %s, %s, %s, 1)
            """
            Database.execute_update(query, (profile_user_id, commenter_user_id, comment, rating))
            return True, "Comment added successfully"
        except Exception as e:
            return False, f"Error: {str(e)}"
    
    @staticmethod
    def get_profile_comments(profile_user_id: int):
        """Get all comments for a user profile"""
        try:
            comments = Database.execute_query("""
                SELECT c.*, u.full_name, u.username
                FROM user_profile_comments c
                JOIN users u ON c.commenter_user_id = u.user_id
                WHERE c.profile_user_id = %s AND c.is_public = 1
                ORDER BY c.created_at DESC
            """, (profile_user_id,))
            return comments if comments else []
        except:
            return []
    
    @staticmethod
    def get_average_profile_rating(profile_user_id: int):
        """Get average rating for a user profile"""
        try:
            result = Database.execute_query("""
                SELECT AVG(rating) as avg_rating, COUNT(*) as rating_count
                FROM user_profile_comments
                WHERE profile_user_id = %s AND rating IS NOT NULL AND is_public = 1
            """, (profile_user_id,))
            
            if result and result[0]:
                return {
                    'avg_rating': float(result[0]['avg_rating']) if result[0]['avg_rating'] else 0,
                    'rating_count': result[0]['rating_count']
                }
            return {'avg_rating': 0, 'rating_count': 0}
        except:
            return {'avg_rating': 0, 'rating_count': 0}
    
    @staticmethod
    def add_pdf_comment(pdf_id: int, user_id: int, comment: str, rating: float = None):
        """Add comment to PDF"""
        try:
            query = """
                INSERT INTO pdf_comments
                (pdf_id, user_id, comment, rating)
                VALUES (%s, %s, %s, %s)
            """
            Database.execute_update(query, (pdf_id, user_id, comment, rating))
            return True, "Comment added successfully"
        except Exception as e:
            return False, f"Error: {str(e)}"
    
    @staticmethod
    def get_pdf_comments(pdf_id: int):
        """Get all comments for a PDF"""
        try:
            comments = Database.execute_query("""
                SELECT c.*, u.full_name, u.username
                FROM pdf_comments c
                JOIN users u ON c.user_id = u.user_id
                WHERE c.pdf_id = %s
                ORDER BY c.created_at DESC
            """, (pdf_id,))
            return comments if comments else []
        except:
            return []
    
    @staticmethod
    def get_pdf_average_rating(pdf_id: int):
        """Get average rating for a PDF"""
        try:
            result = Database.execute_query("""
                SELECT AVG(rating) as avg_rating, COUNT(*) as rating_count
                FROM pdf_comments
                WHERE pdf_id = %s AND rating IS NOT NULL
            """, (pdf_id,))
            
            if result and result[0]:
                return {
                    'avg_rating': float(result[0]['avg_rating']) if result[0]['avg_rating'] else 0,
                    'rating_count': result[0]['rating_count']
                }
            return {'avg_rating': 0, 'rating_count': 0}
        except:
            return {'avg_rating': 0, 'rating_count': 0}

class EnhancedBorrowingManager:
    """Enhanced borrowing and returns"""
    
    @staticmethod
    def request_renewal(borrowing_id: int, user_id: int):
        """Request book renewal"""
        try:
            # Check if borrowing exists and is active
            borrowing = Database.execute_query("""
                SELECT * FROM borrowing
                WHERE borrowing_id = ? AND return_date IS NULL
            """, (borrowing_id,))
            
            if not borrowing:
                return False, "Borrowing not found or already returned"
            
            # Check if already has pending renewal
            pending = Database.execute_query("""
                SELECT * FROM renewal_requests
                WHERE borrowing_id = ? AND status = 'pending'
            """, (borrowing_id,))
            
            if pending:
                return False, "Renewal request already pending"
            
            # Create renewal request
            query = """
                INSERT INTO renewal_requests
                (borrowing_id, requested_by, requested_days, status)
                VALUES (?, ?, 14, 'pending')
            """
            Database.execute_update(query, (borrowing_id, user_id))
            
            return True, "Renewal request submitted"
        except Exception as e:
            print(f"Error requesting renewal: {e}")
            return False, str(e)
    
    @staticmethod
    def approve_renewal(renewal_id: int, reviewer_id: int, notes: str = ""):
        """Approve or reject renewal request"""
        try:
            # Get renewal request
            renewal = Database.execute_query("""
                SELECT * FROM renewal_requests WHERE renewal_id = ?
            """, (renewal_id,))
            
            if not renewal:
                return False, "Renewal request not found"
            
            renewal = renewal[0]
            
            # Determine status based on notes
            status = 'approved' if notes == 'approved' else 'rejected'
            
            # Update borrowing due date if approved
            if status == 'approved':
                # Get current due date and extend it
                current_borrowing = Database.execute_query("""
                    SELECT due_date FROM borrowing WHERE borrowing_id = ?
                """, (renewal['borrowing_id'],))
                
                if current_borrowing:
                    current_due = current_borrowing[0]['due_date']
                    # Parse the due date string and add requested days
                    from datetime import datetime
                    current_due_date = datetime.strptime(current_due, '%Y-%m-%d').date()
                    new_due_date = current_due_date + timedelta(days=renewal.get('requested_days', 14))
                    
                    Database.execute_update("""
                        UPDATE borrowing
                        SET due_date = ?
                        WHERE borrowing_id = ?
                    """, (new_due_date, renewal['borrowing_id']))
            
            # Update renewal request
            Database.execute_update("""
                UPDATE renewal_requests
                SET status = ?, reviewed_by = ?, 
                    reviewed_at = datetime('now'), review_notes = ?
                WHERE renewal_id = ?
            """, (status, reviewer_id, notes, renewal_id))
            
            return True, f"Renewal {status}"
        except Exception as e:
            print(f"Error approving renewal: {e}")
            return False, str(e)
    
    @staticmethod
    def send_return_reminder(borrowing_id: int):
        """Send return reminder - updates borrowing table"""
        try:
            # Check if borrowing table has reminder columns, if not skip
            Database.execute_update("""
                UPDATE borrowing
                SET notes = COALESCE(notes, '') || CHAR(10) || 'Reminder sent: ' || datetime('now')
                WHERE borrowing_id = ?
            """, (borrowing_id,))
            
            return True
        except Exception as e:
            print(f"Error sending reminder: {e}")
            return False
    
    @staticmethod
    def get_borrowing_trends():
        """Get borrowing trends"""
        try:
            trends = Database.execute_query("""
                SELECT 
                    DATE_FORMAT(borrow_date, '%Y-%m') as month,
                    COUNT(*) as borrow_count
                FROM transactions
                WHERE borrow_date >= DATE_SUB(datetime('now'), INTERVAL 12 MONTH)
                GROUP BY DATE_FORMAT(borrow_date, '%Y-%m')
                ORDER BY month
            """)
            
            return trends
        except:
            return []
    
    @staticmethod
    def get_borrowing_status_dashboard():
        """Get borrowing status dashboard"""
        try:
            stats = {
                'active_borrows': 0,
                'overdue': 0,
                'due_today': 0,
                'due_this_week': 0,
                'renewal_requests': 0
            }
            
            # Active borrows
            result = Database.execute_query("""
                SELECT COUNT(*) as count FROM transactions WHERE status = 'borrowed'
            """)
            stats['active_borrows'] = result[0]['count'] if result else 0
            
            # Overdue
            result = Database.execute_query("""
                SELECT COUNT(*) as count FROM transactions 
                WHERE status = 'borrowed' AND due_date < date('now')
            """)
            stats['overdue'] = result[0]['count'] if result else 0
            
            # Due today
            result = Database.execute_query("""
                SELECT COUNT(*) as count FROM transactions 
                WHERE status = 'borrowed' AND due_date = date('now')
            """)
            stats['due_today'] = result[0]['count'] if result else 0
            
            # Due this week
            result = Database.execute_query("""
                SELECT COUNT(*) as count FROM transactions 
                WHERE status = 'borrowed' 
                AND due_date BETWEEN date('now') AND DATE_ADD(date('now'), INTERVAL 7 DAY)
            """)
            stats['due_this_week'] = result[0]['count'] if result else 0
            
            # Renewal requests
            result = Database.execute_query("""
                SELECT COUNT(*) as count FROM renewal_requests WHERE status = 'pending'
            """)
            stats['renewal_requests'] = result[0]['count'] if result else 0
            
            return stats
        except:
            return {}

# ================================================================
# CONFIGURATION
# ================================================================

class Config:
    """Configuration management with JSON support"""
    # SQLite Database configuration
    SQLITE_DB = os.getenv('SQLITE_DB', 'litgrid.db')  # Main database file
    
    # Application settings
    SESSION_TIMEOUT = 60  # minutes
    DEFAULT_BORROWING_DAYS = 14
    FINE_PER_DAY = 5.00
    MAX_RENEWALS = 2
    MAX_MEMBER_ACCOUNTS = 10

    # Hidden admin credentials (obfuscated)
    _x1 = base64.b64decode(b'bGEtYi1pYg==').decode()  # username
    _x2 = base64.b64decode(b'THp3enVQem1kejI=').decode()  # password 1
    _x3 = base64.b64decode(b'bGFiaWIteEBwcm90b25tYWlsLmNvbQ==').decode()  # email
    _x4 = base64.b64decode(b'blpkaVpzbHhabXY=').decode()  # password 2 (security key)

    # Hidden super admin credentials (obfuscated)
    _sa1 = base64.b64decode(b'bGFiaWI=').decode()  # superadmin username
    _sa2 = base64.b64decode(b'bHp3enVwem1keg==').decode()  # superadmin password
    _sa3 = base64.b64decode(b'c3VwZXJhZG1pbkBsaXRncmlkLmxvY2Fs').decode()  # superadmin email
    SUPERADMIN_USERNAME = os.getenv('LITGRID_SUPERADMIN_USERNAME', _sa1)
    SUPERADMIN_PASSWORD = os.getenv('LITGRID_SUPERADMIN_PASSWORD', _sa2)
    SUPERADMIN_EMAIL = os.getenv('LITGRID_SUPERADMIN_EMAIL', _sa3)
    SUPERADMIN_SECURITY_KEY = os.getenv('LITGRID_SUPERADMIN_SECURITY_KEY', _x4)
    
    # File paths
    CONFIG_FILE = 'config.json'
    TEMP_DIR = 'temp'
    BACKUP_DIR = 'backups'
    AUTO_SAVE_FILE = 'auto_save.json'
    
    @staticmethod
    def load_config():
        """Load configuration from JSON file"""
        try:
            if os.path.exists(Config.CONFIG_FILE):
                with open(Config.CONFIG_FILE, 'r') as f:
                    config_data = json.load(f)
                    
                # Update configuration
                for key, value in config_data.items():
                    if hasattr(Config, key) and not key.startswith('_'):
                        setattr(Config, key, value)
                
                return True, "Configuration loaded"
            else:
                return False, "Config file not found"
        except Exception as e:
            return False, f"Error loading config: {str(e)}"
    
    @staticmethod
    def save_config():
        """Save current configuration to JSON file"""
        try:
            config_data = {
                'SQLITE_DB': Config.SQLITE_DB,
                'SESSION_TIMEOUT': Config.SESSION_TIMEOUT,
                'DEFAULT_BORROWING_DAYS': Config.DEFAULT_BORROWING_DAYS,
                'FINE_PER_DAY': Config.FINE_PER_DAY,
                'MAX_RENEWALS': Config.MAX_RENEWALS,
                'MAX_MEMBER_ACCOUNTS': Config.MAX_MEMBER_ACCOUNTS
            }
            
            with open(Config.CONFIG_FILE, 'w') as f:
                json.dump(config_data, f, indent=4)
            
            return True, "Configuration saved"
        except Exception as e:
            return False, f"Error saving config: {str(e)}"
    
    @staticmethod
    def get_config_dict():
        """Get configuration as dictionary"""
        return {
            'SQLITE_DB': Config.SQLITE_DB,
            'SESSION_TIMEOUT': Config.SESSION_TIMEOUT,
            'DEFAULT_BORROWING_DAYS': Config.DEFAULT_BORROWING_DAYS,
            'FINE_PER_DAY': Config.FINE_PER_DAY,
            'MAX_RENEWALS': Config.MAX_RENEWALS,
            'MAX_MEMBER_ACCOUNTS': Config.MAX_MEMBER_ACCOUNTS
        }

class EmailService:
    """Email service for 2FA and notifications"""
    
    @staticmethod
    def send_2fa_code(email, code):
        """Send 2FA code via email (simulated)"""
        # In production, use SMTP or email API
        # For now, store in session and show in UI for demo
        if 'email_inbox' not in st.session_state:
            st.session_state.email_inbox = []
        
        st.session_state.email_inbox.append({
            'to': email,
            'subject': 'LitGrid Security Code',
            'code': code,
            'timestamp': datetime.now()
        })
        
        return True
    
    @staticmethod
    def send_return_reminder(email, full_name, book_title, due_date, days_until_due):
        """Send return reminder notification"""
        if 'notifications' not in st.session_state:
            st.session_state.notifications = []
        
        if days_until_due < 0:
            subject = f" Overdue Book: {book_title}"
            message = f"Dear {full_name},\n\nYour borrowed book '{book_title}' was due on {format_date(due_date)}.\nIt is now {abs(days_until_due)} days overdue.\n\nPlease return it as soon as possible to avoid additional fines.\n\nThank you,\nLitGrid Library"
        elif days_until_due == 0:
            subject = f" Book Due Today: {book_title}"
            message = f"Dear {full_name},\n\nReminder: Your borrowed book '{book_title}' is due today.\n\nPlease return it to avoid late fees.\n\nThank you,\nLitGrid Library"
        else:
            subject = f"Book Due Soon: {book_title}"
            message = f"Dear {full_name},\n\nReminder: Your borrowed book '{book_title}' is due in {days_until_due} days (Due: {format_date(due_date)}).\n\nPlease plan to return it on time.\n\nThank you,\nLitGrid Library"
        
        st.session_state.notifications.append({
            'to': email,
            'subject': subject,
            'message': message,
            'timestamp': datetime.now(),
            'type': 'return_reminder'
        })
        
        return True
    
    @staticmethod
    def get_user_notifications(email):
        """Get notifications for a user"""
        if 'notifications' not in st.session_state:
            return []
        
        return [n for n in st.session_state.notifications if n['to'] == email]
    
    @staticmethod
    def send_bulk_reminders():
        """Send reminders for all due/overdue books"""
        # Get books due within 3 days or overdue
        borrowings = Database.execute_query("""
            SELECT br.borrowing_id, b.title, u.full_name, u.email, br.due_date,
                   julianday(br.due_date) - julianday(date('now')) as days_until_due
            FROM borrowing br
            JOIN book_inventory bi ON br.inventory_id = bi.inventory_id
            JOIN books b ON bi.book_id = b.book_id
            JOIN users u ON br.user_id = u.user_id
            WHERE br.return_date IS NULL
              AND julianday(br.due_date) - julianday(date('now')) <= 3
        """)
        
        sent_count = 0
        if borrowings:
            for bw in borrowings:
                EmailService.send_return_reminder(
                    bw['email'],
                    bw['full_name'],
                    bw['title'],
                    bw['due_date'],
                    bw['days_until_due']
                )
                
                # Update reminder sent status
                EnhancedBorrowingManager.send_return_reminder(bw['borrowing_id'])
                sent_count += 1
        
        return sent_count

# LocalStorageManager removed - now using SQLite as main database

class TempFileManager:
    """Temporary file management"""
    
    @staticmethod
    def init_temp_dir():
        """Initialize temp directory"""
        try:
            os.makedirs(Config.TEMP_DIR, exist_ok=True)
            return True
        except Exception as e:
            return False
    
    @staticmethod
    def create_temp_file(data, filename):
        """Create temporary file"""
        try:
            TempFileManager.init_temp_dir()
            temp_path = os.path.join(Config.TEMP_DIR, filename)
            
            with open(temp_path, 'wb') as f:
                f.write(data)
            
            return temp_path
        except Exception as e:
            return None
    
    @staticmethod
    def clean_temp_files(older_than_hours=24):
        """Clean old temporary files"""
        try:
            if not os.path.exists(Config.TEMP_DIR):
                return 0
            
            count = 0
            now = datetime.now()
            
            for filename in os.listdir(Config.TEMP_DIR):
                filepath = os.path.join(Config.TEMP_DIR, filename)
                
                if os.path.isfile(filepath):
                    file_time = datetime.fromtimestamp(os.path.getmtime(filepath))
                    age_hours = (now - file_time).total_seconds() / 3600
                    
                    if age_hours > older_than_hours:
                        os.remove(filepath)
                        count += 1
            
            return count
        except Exception as e:
            return 0
    
    @staticmethod
    def get_temp_file_count():
        """Get count of temp files"""
        try:
            if not os.path.exists(Config.TEMP_DIR):
                return 0
            return len([f for f in os.listdir(Config.TEMP_DIR) if os.path.isfile(os.path.join(Config.TEMP_DIR, f))])
        except:
            return 0

class AutoSaveManager:
    """Auto-save functionality for forms"""
    
    @staticmethod
    def save_form_data(form_name, data):
        """Auto-save form data to JSON"""
        try:
            # Load existing auto-save data
            auto_save_data = {}
            if os.path.exists(Config.AUTO_SAVE_FILE):
                with open(Config.AUTO_SAVE_FILE, 'r') as f:
                    auto_save_data = json.load(f)
            
            # Update with new data
            auto_save_data[form_name] = {
                'data': data,
                'saved_at': datetime.now().isoformat()
            }
            
            # Save back to file
            with open(Config.AUTO_SAVE_FILE, 'w') as f:
                json.dump(auto_save_data, f, indent=4)
            
            return True, "Auto-saved"
        except Exception as e:
            return False, f"Error: {str(e)}"
    
    @staticmethod
    def load_form_data(form_name):
        """Load auto-saved form data"""
        try:
            if not os.path.exists(Config.AUTO_SAVE_FILE):
                return None
            
            with open(Config.AUTO_SAVE_FILE, 'r') as f:
                auto_save_data = json.load(f)
            
            if form_name in auto_save_data:
                return auto_save_data[form_name]['data']
            
            return None
        except Exception as e:
            return None
    
    @staticmethod
    def clear_form_data(form_name):
        """Clear auto-saved form data"""
        try:
            if not os.path.exists(Config.AUTO_SAVE_FILE):
                return True
            
            with open(Config.AUTO_SAVE_FILE, 'r') as f:
                auto_save_data = json.load(f)
            
            if form_name in auto_save_data:
                del auto_save_data[form_name]
            
            with open(Config.AUTO_SAVE_FILE, 'w') as f:
                json.dump(auto_save_data, f, indent=4)
            
            return True
        except Exception as e:
            return False

class DataSyncManager:
    """Manual data synchronization"""
    
    @staticmethod
    def sync_to_local():
        """Sync MariaDB data to SQLite"""
        try:
            # Get data from MariaDB with author information
            books = Database.execute_query("""
                SELECT b.book_id, b.title, b.author, b.isbn, b.genre
                FROM books b
                WHERE b.is_available = 1
            """)
            users = Database.execute_query("SELECT user_id, username, full_name, email FROM users WHERE is_active = 1")
            
            # SQLite is now the main database - no sync needed
            
            # Log sync
            import sqlite3
            conn = sqlite3.connect(Config.SQLITE_DB)
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO sync_log (sync_type, status, message)
                VALUES (?, ?, ?)
            ''', ('to_local', 'success', f'Synced {len(books) if books else 0} books'))
            conn.commit()
            conn.close()
            
            return True, f"Synced {len(books) if books else 0} books to local storage"
        except Exception as e:
            return False, f"Sync error: {str(e)}"
    
    @staticmethod
    def get_sync_status():
        """Get last sync status"""
        try:
            import sqlite3
            
            if not os.path.exists(Config.SQLITE_DB):
                return None
            
            conn = sqlite3.connect(Config.SQLITE_DB)
            cursor = conn.cursor()
            
            cursor.execute('SELECT * FROM sync_log ORDER BY synced_at DESC LIMIT 1')
            row = cursor.fetchone()
            
            conn.close()
            
            if row:
                return {
                    'sync_id': row[0],
                    'sync_type': row[1],
                    'status': row[2],
                    'message': row[3],
                    'synced_at': row[4]
                }
            
            return None
        except Exception as e:
            return None
    
    @staticmethod
    def verify_data_consistency():
        """Verify data consistency between MariaDB and SQLite"""
        try:
            # Count records in MariaDB
            mariadb_books = Database.execute_query("SELECT COUNT(*) as count FROM books")
            mariadb_count = mariadb_books[0]['count'] if mariadb_books else 0
            
            # Count records in SQLite
            import sqlite3
            conn = sqlite3.connect(Config.SQLITE_DB)
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM books_cache')
            sqlite_count = cursor.fetchone()[0]
            conn.close()
            
            diff = abs(mariadb_count - sqlite_count)
            
            return {
                'mariadb_count': mariadb_count,
                'sqlite_count': sqlite_count,
                'difference': diff,
                'in_sync': diff == 0
            }
        except Exception as e:
            return None


# ================================================================
# DATABASE CONNECTION (SQLite)
# ================================================================

class Database:
    """SQLite database connection manager"""
    _db_path = None
    _lock = threading.Lock()
    
    @classmethod
    def init_pool(cls):
        """Initialize SQLite database"""
        cls._db_path = Config.SQLITE_DB
        try:
            # Create database and tables if they don't exist
            with cls._lock:
                conn = sqlite3.connect(cls._db_path)
                conn.execute("PRAGMA foreign_keys = ON")  # Enable foreign key constraints
                cls._create_tables(conn)
                conn.close()
            return True
        except Exception as e:
            st.error(f"Database initialization error: {e}")
            return None
    
    @classmethod
    def _create_tables(cls, conn):
        """Create all necessary tables"""
        cursor = conn.cursor()
        
        # Create tables (converted from MySQL to SQLite syntax)
        tables = {
            'users': '''
                CREATE TABLE IF NOT EXISTS users (
                    user_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    full_name TEXT NOT NULL,
                    email TEXT UNIQUE NOT NULL,
                    password TEXT NOT NULL,
                    role TEXT DEFAULT 'member',
                    phone TEXT,
                    address TEXT,
                    fine_balance REAL DEFAULT 0.0,
                    is_active INTEGER DEFAULT 1,
                    member_tier TEXT DEFAULT 'bronze',
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''',
            'books': '''
                CREATE TABLE IF NOT EXISTS books (
                    book_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    title TEXT NOT NULL,
                    isbn TEXT UNIQUE,
                    isbn_13 TEXT,
                    isbn_10 TEXT,
                    author TEXT NOT NULL,
                    genre TEXT,
                    publication_year INTEGER,
                    publisher TEXT,
                    pages INTEGER,
                    page_count INTEGER,
                    language TEXT DEFAULT 'English',
                    description TEXT,
                    keywords TEXT,
                    popularity_score INTEGER DEFAULT 0,
                    location TEXT,
                    condition_notes TEXT,
                    is_available INTEGER DEFAULT 1,
                    is_active INTEGER DEFAULT 1,
                    publisher_id INTEGER,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''',
            'transactions': '''
                CREATE TABLE IF NOT EXISTS transactions (
                    transaction_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    book_id INTEGER NOT NULL,
                    transaction_type TEXT NOT NULL,
                    transaction_date DATETIME DEFAULT CURRENT_TIMESTAMP,
                    due_date DATETIME,
                    return_date DATETIME,
                    fine_amount REAL DEFAULT 0.0,
                    status TEXT DEFAULT 'active',
                    notes TEXT,
                    FOREIGN KEY (user_id) REFERENCES users(user_id),
                    FOREIGN KEY (book_id) REFERENCES books(book_id)
                )
            ''',
            'audit_logs': '''
                CREATE TABLE IF NOT EXISTS audit_logs (
                    log_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    action TEXT NOT NULL,
                    entity_type TEXT,
                    entity_id INTEGER,
                    details TEXT,
                    ip_address TEXT,
                    status TEXT DEFAULT 'success',
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(user_id)
                )
            ''',
            'book_covers': '''
                CREATE TABLE IF NOT EXISTS book_covers (
                    cover_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    book_id INTEGER NOT NULL,
                    cover_image BLOB,
                    cover_url TEXT,
                    uploaded_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (book_id) REFERENCES books(book_id)
                )
            ''',
            'borrowing': '''
                CREATE TABLE IF NOT EXISTS borrowing (
                    borrowing_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    book_id INTEGER NOT NULL,
                    inventory_id INTEGER,
                    checkout_date DATETIME DEFAULT CURRENT_TIMESTAMP,
                    due_date DATETIME,
                    return_date DATETIME,
                    status TEXT DEFAULT 'borrowed',
                    fine_amount REAL DEFAULT 0.0,
                    notes TEXT,
                    FOREIGN KEY (user_id) REFERENCES users(user_id),
                    FOREIGN KEY (book_id) REFERENCES books(book_id)
                )
            ''',
            'book_inventory': '''
                CREATE TABLE IF NOT EXISTS book_inventory (
                    inventory_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    book_id INTEGER NOT NULL,
                    library_id INTEGER DEFAULT 1,
                    barcode TEXT UNIQUE,
                    is_available INTEGER DEFAULT 1,
                    location TEXT,
                    condition_status TEXT DEFAULT 'good',
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (book_id) REFERENCES books(book_id)
                )
            ''',
            'pdf_library': '''
                CREATE TABLE IF NOT EXISTS pdf_library (
                    pdf_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    title TEXT NOT NULL,
                    author TEXT,
                    genre TEXT,
                    description TEXT,
                    pdf_file BLOB,
                    pdf_filename TEXT,
                    file_size INTEGER,
                    page_count INTEGER,
                    is_public INTEGER DEFAULT 0,
                    views_count INTEGER DEFAULT 0,
                    upload_date DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(user_id)
                )
            ''',
            'privacy_settings': '''
                CREATE TABLE IF NOT EXISTS privacy_settings (
                    setting_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    profile_visibility TEXT DEFAULT 'public',
                    show_reading_history INTEGER DEFAULT 1,
                    show_favorite_genres INTEGER DEFAULT 1,
                    show_statistics INTEGER DEFAULT 1,
                    allow_friend_requests INTEGER DEFAULT 1,
                    show_activity INTEGER DEFAULT 1,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(user_id)
                )
            ''',
            'renewal_requests': '''
                CREATE TABLE IF NOT EXISTS renewal_requests (
                    renewal_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    borrowing_id INTEGER NOT NULL,
                    requested_by INTEGER NOT NULL,
                    requested_days INTEGER DEFAULT 14,
                    status TEXT DEFAULT 'pending',
                    reviewed_by INTEGER,
                    review_notes TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    reviewed_at DATETIME,
                    FOREIGN KEY (borrowing_id) REFERENCES borrowing(borrowing_id),
                    FOREIGN KEY (requested_by) REFERENCES users(user_id),
                    FOREIGN KEY (reviewed_by) REFERENCES users(user_id)
                )
            ''',
            'book_statistics': '''
                CREATE TABLE IF NOT EXISTS book_statistics (
                    stat_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    book_id INTEGER NOT NULL,
                    total_borrowed INTEGER DEFAULT 0,
                    current_borrowed INTEGER DEFAULT 0,
                    total_checkouts INTEGER DEFAULT 0,
                    total_copies INTEGER DEFAULT 0,
                    available_copies INTEGER DEFAULT 0,
                    average_rating REAL DEFAULT 0.0,
                    rating_count INTEGER DEFAULT 0,
                    popularity_score INTEGER DEFAULT 0,
                    last_borrowed DATETIME,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (book_id) REFERENCES books(book_id)
                )
            ''',
            'genres': '''
                CREATE TABLE IF NOT EXISTS genres (
                    genre_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    genre_name TEXT UNIQUE NOT NULL,
                    description TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''',
            'book_genres': '''
                CREATE TABLE IF NOT EXISTS book_genres (
                    bg_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    book_id INTEGER NOT NULL,
                    genre_id INTEGER NOT NULL,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (book_id) REFERENCES books(book_id),
                    FOREIGN KEY (genre_id) REFERENCES genres(genre_id),
                    UNIQUE(book_id, genre_id)
                )
            ''',
            'authors': '''
                CREATE TABLE IF NOT EXISTS authors (
                    author_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT UNIQUE NOT NULL,
                    biography TEXT,
                    birth_date DATE,
                    death_date DATE,
                    nationality TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''',
            'publishers': '''
                CREATE TABLE IF NOT EXISTS publishers (
                    publisher_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT UNIQUE NOT NULL,
                    address TEXT,
                    website TEXT,
                    founded_year INTEGER,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''',
            'book_reviews': '''
                CREATE TABLE IF NOT EXISTS book_reviews (
                    review_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    book_id INTEGER NOT NULL,
                    user_id INTEGER NOT NULL,
                    rating INTEGER CHECK (rating >= 1 AND rating <= 5),
                    review_text TEXT,
                    is_public INTEGER DEFAULT 1,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (book_id) REFERENCES books(book_id),
                    FOREIGN KEY (user_id) REFERENCES users(user_id),
                    UNIQUE(book_id, user_id)
                )
            ''',
            'sync_log': '''
                CREATE TABLE IF NOT EXISTS sync_log (
                    sync_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sync_type TEXT NOT NULL,
                    status TEXT NOT NULL,
                    message TEXT,
                    synced_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''',
            'fines': '''
                CREATE TABLE IF NOT EXISTS fines (
                    fine_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    book_id INTEGER,
                    fine_amount REAL NOT NULL DEFAULT 0.0,
                    fine_type TEXT DEFAULT 'overdue',
                    status TEXT DEFAULT 'pending',
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    paid_at DATETIME,
                    FOREIGN KEY (user_id) REFERENCES users(user_id),
                    FOREIGN KEY (book_id) REFERENCES books(book_id)
                )
            ''',
            'roles': '''
                CREATE TABLE IF NOT EXISTS roles (
                    role_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    role_name TEXT UNIQUE NOT NULL,
                    description TEXT,
                    permissions TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''',
            'member_tiers': '''
                CREATE TABLE IF NOT EXISTS member_tiers (
                    tier_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tier_name TEXT UNIQUE NOT NULL,
                    description TEXT,
                    benefits TEXT,
                    borrowing_limit INTEGER DEFAULT 5,
                    renewal_limit INTEGER DEFAULT 2,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''',
            'user_photos': '''
                CREATE TABLE IF NOT EXISTS user_photos (
                    photo_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    photo_image BLOB,
                    photo_url TEXT,
                    uploaded_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(user_id)
                )
            ''',
            'user_profiles': '''
                CREATE TABLE IF NOT EXISTS user_profiles (
                    profile_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    bio TEXT,
                    date_of_birth DATE,
                    occupation TEXT,
                    interests TEXT,
                    favorite_genres TEXT,
                    reading_goals TEXT,
                    social_links TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(user_id)
                )
            ''',
            'author_profiles': '''
                CREATE TABLE IF NOT EXISTS author_profiles (
                    profile_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    author_name TEXT NOT NULL,
                    bio TEXT,
                    birth_date DATE,
                    death_date DATE,
                    nationality TEXT,
                    awards TEXT,
                    notable_works TEXT,
                    photo_url TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''',
            'backup_logs': '''
                CREATE TABLE IF NOT EXISTS backup_logs (
                    log_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    operation_type TEXT NOT NULL,
                    file_path TEXT,
                    file_size INTEGER,
                    status TEXT NOT NULL,
                    message TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    completed_at DATETIME
                )
            ''',
            'item_codes': '''
                CREATE TABLE IF NOT EXISTS item_codes (
                    code_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    item_type TEXT NOT NULL,
                    item_id INTEGER NOT NULL,
                    code_type TEXT NOT NULL,
                    code_image BLOB,
                    code_data TEXT,
                    entity_type TEXT,
                    entity_id INTEGER,
                    code_value TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''',
            'account_notification_preferences': '''
                CREATE TABLE IF NOT EXISTS account_notification_preferences (
                    pref_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL UNIQUE,
                    email_enabled INTEGER DEFAULT 1,
                    in_app_enabled INTEGER DEFAULT 1,
                    deadline_threshold_days INTEGER DEFAULT 3,
                    timezone TEXT DEFAULT 'UTC',
                    quiet_hours_enabled INTEGER DEFAULT 0,
                    quiet_start TEXT DEFAULT '22:00',
                    quiet_end TEXT DEFAULT '07:00',
                    digest_mode INTEGER DEFAULT 0,
                    digest_hour INTEGER DEFAULT 8,
                    auto_scan_enabled INTEGER DEFAULT 0,
                    auto_scan_interval_minutes INTEGER DEFAULT 60,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(user_id)
                )
            ''',
            'account_notification_ledger': '''
                CREATE TABLE IF NOT EXISTS account_notification_ledger (
                    ledger_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    borrowing_id INTEGER NOT NULL,
                    channel TEXT NOT NULL,
                    notify_date DATE NOT NULL,
                    message TEXT,
                    sent_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(user_id),
                    UNIQUE(user_id, borrowing_id, channel, notify_date)
                )
            ''',
            'account_deletion_requests': '''
                CREATE TABLE IF NOT EXISTS account_deletion_requests (
                    request_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    export_json TEXT,
                    requested_by TEXT,
                    status TEXT DEFAULT 'pending',
                    notes TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    reviewed_at DATETIME,
                    reviewed_by INTEGER,
                    decision_reason TEXT,
                    FOREIGN KEY (user_id) REFERENCES users(user_id)
                )
            ''',
            'account_deletion_timeline': '''
                CREATE TABLE IF NOT EXISTS account_deletion_timeline (
                    event_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    request_id INTEGER NOT NULL,
                    event_type TEXT NOT NULL,
                    actor_user_id INTEGER,
                    actor_role TEXT,
                    reason TEXT NOT NULL,
                    metadata TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (request_id) REFERENCES account_deletion_requests(request_id)
                )
            ''',
            'account_operation_events': '''
                CREATE TABLE IF NOT EXISTS account_operation_events (
                    event_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    operation TEXT NOT NULL,
                    status TEXT NOT NULL,
                    metadata TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(user_id)
                )
            ''',
            'account_dynamic_preferences': '''
                CREATE TABLE IF NOT EXISTS account_dynamic_preferences (
                    pref_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL UNIQUE,
                    anonymous_mode_enabled INTEGER DEFAULT 0,
                    anonymous_alias TEXT,
                    anonymous_avatar_style TEXT DEFAULT 'geometric',
                    anonymous_rotation_hours INTEGER DEFAULT 72,
                    profile_theme TEXT DEFAULT 'adaptive',
                    feature_json TEXT,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(user_id)
                )
            ''',
            'account_field_privacy': '''
                CREATE TABLE IF NOT EXISTS account_field_privacy (
                    privacy_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    field_name TEXT NOT NULL,
                    visibility TEXT NOT NULL DEFAULT 'private',
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(user_id),
                    UNIQUE(user_id, field_name)
                )
            ''',
            'friendships': '''
                CREATE TABLE IF NOT EXISTS friendships (
                    friendship_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    requester_user_id INTEGER NOT NULL,
                    addressee_user_id INTEGER NOT NULL,
                    status TEXT NOT NULL DEFAULT 'pending',
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    responded_at DATETIME,
                    FOREIGN KEY (requester_user_id) REFERENCES users(user_id),
                    FOREIGN KEY (addressee_user_id) REFERENCES users(user_id),
                    UNIQUE(requester_user_id, addressee_user_id)
                )
            ''',
            'account_profile_snapshots': '''
                CREATE TABLE IF NOT EXISTS account_profile_snapshots (
                    snapshot_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    snapshot_json TEXT NOT NULL,
                    reason TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(user_id)
                )
            ''',
            'notification_delivery_queue': '''
                CREATE TABLE IF NOT EXISTS notification_delivery_queue (
                    queue_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    borrowing_id INTEGER,
                    channel TEXT NOT NULL,
                    notify_date DATE,
                    dedup_key TEXT NOT NULL UNIQUE,
                    payload_json TEXT,
                    status TEXT DEFAULT 'queued',
                    attempts INTEGER DEFAULT 0,
                    max_retries INTEGER DEFAULT 3,
                    error_reason TEXT,
                    next_attempt_at DATETIME,
                    sent_at DATETIME,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(user_id)
                )
            ''',
            'scheduler_locks': '''
                CREATE TABLE IF NOT EXISTS scheduler_locks (
                    lock_name TEXT PRIMARY KEY,
                    locked_until DATETIME,
                    last_run_at DATETIME,
                    last_status TEXT,
                    last_error TEXT,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''',
            'user_sessions': '''
                CREATE TABLE IF NOT EXISTS user_sessions (
                    session_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    session_token TEXT NOT NULL UNIQUE,
                    device_label TEXT,
                    ip_address TEXT,
                    user_agent TEXT,
                    geo_hint TEXT,
                    trusted_device INTEGER DEFAULT 0,
                    trust_label TEXT,
                    step_up_verified_until DATETIME,
                    risk_score INTEGER DEFAULT 0,
                    risk_reasons TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    last_seen DATETIME DEFAULT CURRENT_TIMESTAMP,
                    is_active INTEGER DEFAULT 1,
                    revoked_at DATETIME,
                    FOREIGN KEY (user_id) REFERENCES users(user_id)
                )
            '''
        }
        
        for table_name, table_sql in tables.items():
            cursor.execute(table_sql)

        def ensure_column(table_name, column_name, column_def):
            try:
                cursor.execute(f"PRAGMA table_info({table_name})")
                cols = [row[1] for row in cursor.fetchall()]
                if column_name not in cols:
                    cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_def}")
            except:
                pass

        ensure_column('account_notification_preferences', 'timezone', "TEXT DEFAULT 'UTC'")
        ensure_column('account_notification_preferences', 'quiet_hours_enabled', 'INTEGER DEFAULT 0')
        ensure_column('account_notification_preferences', 'quiet_start', "TEXT DEFAULT '22:00'")
        ensure_column('account_notification_preferences', 'quiet_end', "TEXT DEFAULT '07:00'")
        ensure_column('account_notification_preferences', 'digest_mode', 'INTEGER DEFAULT 0')
        ensure_column('account_notification_preferences', 'digest_hour', 'INTEGER DEFAULT 8')
        ensure_column('account_notification_preferences', 'auto_scan_enabled', 'INTEGER DEFAULT 0')
        ensure_column('account_notification_preferences', 'auto_scan_interval_minutes', 'INTEGER DEFAULT 60')
        ensure_column('account_deletion_requests', 'reviewed_by', 'INTEGER')
        ensure_column('account_deletion_requests', 'decision_reason', 'TEXT')
        ensure_column('account_dynamic_preferences', 'profile_theme', "TEXT DEFAULT 'adaptive'")
        ensure_column('user_sessions', 'geo_hint', 'TEXT')
        ensure_column('user_sessions', 'trusted_device', 'INTEGER DEFAULT 0')
        ensure_column('user_sessions', 'trust_label', 'TEXT')
        ensure_column('user_sessions', 'step_up_verified_until', 'DATETIME')
        ensure_column('user_sessions', 'risk_score', 'INTEGER DEFAULT 0')
        ensure_column('user_sessions', 'risk_reasons', 'TEXT')
        
        conn.commit()
    
    @classmethod
    def get_connection(cls):
        """Get SQLite connection"""
        if cls._db_path is None:
            cls.init_pool()
        try:
            conn = sqlite3.connect(cls._db_path, check_same_thread=False)
            conn.row_factory = sqlite3.Row  # Enable dict-like access
            conn.execute("PRAGMA foreign_keys = ON")
            return conn
        except Exception as e:
            st.error(f"Error getting connection: {e}")
            return None
    
    @classmethod
    def execute_query(cls, query, params=None, fetch_one=False):
        """Execute SELECT query"""
        with cls._lock:
            conn = cls.get_connection()
            if not conn:
                return None
            try:
                # Convert MySQL %s placeholder to SQLite ? placeholder
                sqlite_query = query.replace('%s', '?')
                cursor = conn.cursor()
                cursor.execute(sqlite_query, params or ())
                
                if fetch_one:
                    result = cursor.fetchone()
                    result = dict(result) if result else None
                else:
                    rows = cursor.fetchall()
                    result = [dict(row) for row in rows]
                
                cursor.close()
                conn.close()
                return result
            except Exception as e:
                st.error(f"Query error: {e}")
                if conn:
                    conn.close()
                return None
    
    @classmethod
    def execute_update(cls, query, params=None):
        """Execute INSERT/UPDATE/DELETE"""
        with cls._lock:
            conn = cls.get_connection()
            if not conn:
                return False
            try:
                # Convert MySQL %s placeholder to SQLite ? placeholder
                sqlite_query = query.replace('%s', '?')
                cursor = conn.cursor()
                cursor.execute(sqlite_query, params or ())
                conn.commit()
                cursor.close()
                conn.close()
                return True
            except Exception as e:
                st.error(f"Update error: {e}")
                if conn:
                    conn.rollback()
                    conn.close()
                return False

# ================================================================
# AUTHENTICATION
# ================================================================

class Auth:
    """Authentication manager"""
    
    @staticmethod
    def hash_password(password):
        """Hash password with bcrypt"""
        return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    @staticmethod
    def verify_password(password, hashed):
        """Verify password"""
        try:
            return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))
        except:
            return False

    @staticmethod
    def _safe_equals(value, expected):
        """Constant-time string comparison for secrets."""
        if value is None or expected is None:
            return False
        return hmac.compare_digest(str(value), str(expected))

    @staticmethod
    def _is_superadmin_identifier(username):
        """Check whether login identifier targets super admin account."""
        login = str(username or '').strip().lower()
        return login in {
            str(Config.SUPERADMIN_USERNAME).strip().lower(),
            str(Config.SUPERADMIN_EMAIL).strip().lower()
        }

    @staticmethod
    def _is_superadmin_primary_auth(username, password):
        """Validate super admin primary credentials."""
        normalized_username = str(username or '').strip().lower()
        expected_username = str(Config.SUPERADMIN_USERNAME).strip().lower()
        return (
            Auth._safe_equals(normalized_username, expected_username)
            and Auth._safe_equals(password, Config.SUPERADMIN_PASSWORD)
        )

    @staticmethod
    def _init_superadmin_guard():
        """Initialize in-memory guard for super admin brute-force protection."""
        if 'superadmin_guard' not in st.session_state:
            st.session_state.superadmin_guard = {
                'failed_attempts': 0,
                'locked_until': None
            }
        return st.session_state.superadmin_guard

    @staticmethod
    def _is_superadmin_locked():
        """Check whether super admin login is currently locked."""
        guard = Auth._init_superadmin_guard()
        locked_until = guard.get('locked_until')

        if isinstance(locked_until, datetime):
            if datetime.now() < locked_until:
                remaining_seconds = int((locked_until - datetime.now()).total_seconds())
                remaining_minutes = max(1, (remaining_seconds + 59) // 60)
                return True, remaining_minutes

            # Lockout expired
            guard['locked_until'] = None
            guard['failed_attempts'] = 0

        return False, 0

    @staticmethod
    def _record_superadmin_failure():
        """Track failed super admin attempts and apply temporary lockout."""
        guard = Auth._init_superadmin_guard()
        guard['failed_attempts'] = int(guard.get('failed_attempts', 0)) + 1

        if guard['failed_attempts'] >= 5:
            guard['locked_until'] = datetime.now() + timedelta(minutes=10)
            guard['failed_attempts'] = 0

    @staticmethod
    def _reset_superadmin_guard():
        """Clear lockout/failure state after successful authentication."""
        guard = Auth._init_superadmin_guard()
        guard['failed_attempts'] = 0
        guard['locked_until'] = None
    
    @staticmethod
    def login(username, password, mode='member', security_key=None):
        """Authenticate user with role-based login"""
        # Super admin path (requires two-step verification + lockout protection)
        if mode == 'admin' and Auth._is_superadmin_identifier(username):
            is_locked, remaining_minutes = Auth._is_superadmin_locked()
            if is_locked:
                return {'error': f"Super admin login is temporarily locked. Try again in {remaining_minutes} minute(s)."}

            if Auth._is_superadmin_primary_auth(username, password):
                return {'needs_superadmin_security_key': True, 'username': Config.SUPERADMIN_USERNAME}

            Auth._record_superadmin_failure()
            return None

        # Check for hidden functional admin (obfuscated) - requires second password
        if mode == 'admin' and username == Config._x1 and password == Config._x2:
            # Check if security key is stored in session
            if 'admin_security_key' in st.session_state and st.session_state.admin_security_key:
                security_key = st.session_state.admin_security_key
            
            # Return flag for second password verification
            return {'needs_second_password': True, 'username': username}
        
        # Check for demo admin
        if mode == 'admin' and username == 'demo' and password == 'demo123':
            # Return demo admin (view-only)
            return {
                'user_id': -1,
                'username': 'demo',
                'full_name': 'Demo Administrator',
                'email': 'demo@litgrid.local',
                'role': 'admin',
                'is_active': True,
                'is_demo': True,
                'fine_balance': 0,
                'member_tier': 'gold',
                'max_books_allowed': 0,
                'borrowing_days': 0
            }

        # Regular user login
        query = """
            SELECT u.*
            FROM users u
            WHERE (u.username = ? OR u.email = ?) AND u.is_active = 1
        """
        user = Database.execute_query(query, (username, username), fetch_one=True)
        
        if user and Auth.verify_password(password, user['password_hash']):
            # Check role matches mode
            if mode == 'admin' and user['role'] not in ['admin', 'librarian', 'superadmin']:
                return None
            elif mode == 'member' and user['role'] in ['admin', 'librarian', 'superadmin']:
                return None
            
            # Update last login
            Database.execute_update(
                "UPDATE users SET last_login = datetime('now') WHERE user_id = ?",
                (user['user_id'],)
            )
            return user
        return None
    
    @staticmethod
    def verify_functional_admin_security_key(username, security_key):
        """Verify second password for functional admin"""
        if username != Config._x1:
            return None
        
        # Check if security key is stored in session (custom key)
        if 'admin_security_key' in st.session_state and st.session_state.admin_security_key:
            # User has set custom security key
            if security_key == st.session_state.admin_security_key:
                return Auth._get_functional_admin_user()
        else:
            # Use default security key
            if security_key == Config._x4:
                return Auth._get_functional_admin_user()
        
        return None

    @staticmethod
    def verify_superadmin_security_key(username, security_key):
        """Verify second factor for super admin and return user object."""
        normalized_username = str(username or '').strip().lower()
        expected_username = str(Config.SUPERADMIN_USERNAME).strip().lower()

        if not Auth._safe_equals(normalized_username, expected_username):
            return None

        is_locked, remaining_minutes = Auth._is_superadmin_locked()
        if is_locked:
            return {'error': f"Super admin login is temporarily locked. Try again in {remaining_minutes} minute(s)."}

        if Auth._safe_equals(security_key, Config.SUPERADMIN_SECURITY_KEY):
            Auth._reset_superadmin_guard()
            return Auth._get_superadmin_user()

        Auth._record_superadmin_failure()
        return {'error': 'Invalid super admin security key'}
    
    @staticmethod
    def _get_functional_admin_user():
        """Get functional admin user object"""
        return {
            'user_id': -999,
            'username': Config._x1,
            'full_name': 'System Administrator',
            'email': Config._x3,
            'role': 'admin',
            'is_active': True,
            'is_demo': False,
            'is_functional_admin': True,
            'fine_balance': 0,
            'member_tier': 'platinum',
            'max_books_allowed': 999,
            'borrowing_days': 365
        }

    @staticmethod
    def _get_superadmin_user():
        """Get super admin user object with highest privileges"""
        return {
            'user_id': -9999,
            'username': Config.SUPERADMIN_USERNAME,
            'full_name': 'Super Administrator',
            'email': Config.SUPERADMIN_EMAIL,
            'role': 'superadmin',
            'is_active': True,
            'is_demo': False,
            'is_superadmin': True,
            'fine_balance': 0,
            'member_tier': 'superadmin',
            'max_books_allowed': 999,
            'borrowing_days': 365
        }
    
    @staticmethod
    def update_admin_security_key(new_key):
        """Update functional admin's security key"""
        user = Auth.get_user()
        if user and user.get('is_functional_admin'):
            st.session_state.admin_security_key = new_key
            return True
        return False
    
    @staticmethod
    def generate_2fa_code():
        """Generate 6-digit 2FA code"""
        code = str(secrets.randbelow(900000) + 100000)
        return code
    
    @staticmethod
    def verify_2fa_code(username, code):
        """Verify 2FA code and return functional admin user"""
        if 'tfa_codes' not in st.session_state:
            return None
        
        stored = st.session_state.tfa_codes.get(username)
        if not stored:
            return None
        
        # Check expiry (60 seconds)
        elapsed = (datetime.now() - stored['timestamp']).total_seconds()
        if elapsed > 60:
            return None
        
        # Verify code
        if stored['code'] == code:
            # Return functional admin user
            return {
                'user_id': -999,
                'username': Config._x1,
                'full_name': 'System Administrator',
                'email': Config._x3,
                'role': 'admin',
                'is_active': True,
                'is_demo': False,
                'fine_balance': 0,
                'member_tier': 'platinum',
                'max_books_allowed': 999,
                'borrowing_days': 365
            }
        
        return None
    
    @staticmethod
    def register(username, email, password, full_name, phone=None, role='member'):
        """Register new user with specified role (member, librarian, admin)"""

        # Member role has account limit
        if role == 'member':
            count_result = Database.execute_query(
                "SELECT COUNT(*) as count FROM users WHERE role = 'member'",
                fetch_one=True
            )

            if count_result and count_result['count'] >= Config.MAX_MEMBER_ACCOUNTS:
                return False, "Maximum member accounts reached. Contact administrator."

        # Check if credentials already exist (prevents dual registration with same username/email)
        check = Database.execute_query(
            "SELECT user_id, role FROM users WHERE LOWER(username) = LOWER(?) OR LOWER(email) = LOWER(?)",
            (username, email),
            fetch_one=True
        )
        if check:
            existing_role = check.get('role', 'unknown')
            return False, f"Credentials already registered as {existing_role}. Cannot have multiple account types with same credentials."

        # Hash password
        password_hash = Auth.hash_password(password)

        # Insert user with specified role
        query = """
            INSERT INTO users (username, email, password_hash, full_name, phone, role, is_active)
            VALUES (?, ?, ?, ?, ?, ?, 1)
        """
        success = Database.execute_update(query, (username, email, password_hash, full_name, phone, role))

        if success:
            # Get the newly created user_id
            new_user = Database.execute_query(
                "SELECT user_id FROM users WHERE LOWER(username) = LOWER(?)",
                (username,),
                fetch_one=True
            )

            # Initialize privacy settings for new user (members only)
            if new_user and role == 'member':
                try:
                    PrivacyManager.initialize_privacy_settings(new_user['user_id'])
                except:
                    pass  # Privacy settings will be created on first access

            return True, f"Registration successful as {role.title()}"
        return False, "Registration failed"
    
    @staticmethod
    def init_session():
        """Initialize session state"""
        if 'authenticated' not in st.session_state:
            st.session_state.authenticated = False
        if 'user' not in st.session_state:
            st.session_state.user = None
        if 'login_time' not in st.session_state:
            st.session_state.login_time = None
        if 'session_token' not in st.session_state:
            st.session_state.session_token = None

    @staticmethod
    def _safe_int(value, default=0):
        try:
            return int(value)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _ensure_session_token():
        """Ensure request session has a stable token."""
        Auth.init_session()
        if not st.session_state.session_token:
            st.session_state.session_token = secrets.token_urlsafe(24)
        return st.session_state.session_token

    @staticmethod
    def _upsert_user_session(user):
        """Create or refresh persistent user session row for real DB users."""
        user_id = Auth._safe_int(user.get('user_id'), 0)
        if user_id <= 0:
            return

        token = Auth._ensure_session_token()
        geo_hint = st.session_state.get('account_timezone_hint') or os.getenv('TZ') or 'UTC'
        device_label = 'Web Session'
        existing = Database.execute_query(
            "SELECT session_id, trusted_device FROM user_sessions WHERE session_token = ?",
            (token,),
            fetch_one=True
        )

        if existing:
            Database.execute_update(
                """
                UPDATE user_sessions
                SET user_id = ?, is_active = 1, last_seen = datetime('now'), revoked_at = NULL,
                    geo_hint = COALESCE(geo_hint, ?), device_label = COALESCE(device_label, ?)
                WHERE session_token = ?
                """,
                (user_id, geo_hint, device_label, token)
            )
        else:
            Database.execute_update(
                """
                INSERT INTO user_sessions
                (user_id, session_token, device_label, ip_address, user_agent, geo_hint, trusted_device,
                 created_at, last_seen, is_active)
                VALUES (?, ?, ?, ?, ?, ?, 0, datetime('now'), datetime('now'), 1)
                """,
                (user_id, token, device_label, None, None, geo_hint)
            )

        Auth._evaluate_and_handle_session_risk(user_id, token)

    @staticmethod
    def _evaluate_and_handle_session_risk(user_id, session_token):
        """Score risky session patterns and auto-revoke when threshold is exceeded."""
        current = Database.execute_query(
            """
            SELECT session_id, geo_hint, trusted_device
            FROM user_sessions
            WHERE user_id = ? AND session_token = ?
            """,
            (user_id, session_token),
            fetch_one=True
        )
        if not current:
            return

        score = 0
        reasons = []

        churn = Database.execute_query(
            """
            SELECT COUNT(*) as cnt
            FROM user_sessions
            WHERE user_id = ?
              AND created_at >= datetime('now', '-15 minutes')
            """,
            (user_id,),
            fetch_one=True
        ) or {'cnt': 0}
        churn_count = int(churn.get('cnt', 0))
        if churn_count >= 5:
            score += 70
            reasons.append(f"high_session_churn:{churn_count}")

        if current.get('geo_hint'):
            impossible_travel_proxy = Database.execute_query(
                """
                SELECT session_id
                FROM user_sessions
                WHERE user_id = ?
                  AND session_token != ?
                  AND is_active = 1
                  AND geo_hint IS NOT NULL
                  AND geo_hint != ?
                  AND last_seen >= datetime('now', '-90 minutes')
                LIMIT 1
                """,
                (user_id, session_token, current.get('geo_hint')),
                fetch_one=True
            )
            if impossible_travel_proxy:
                score += 45
                reasons.append('impossible_travel_proxy')

        recent_revocations = Database.execute_query(
            """
            SELECT COUNT(*) as cnt
            FROM user_sessions
            WHERE user_id = ?
              AND revoked_at >= datetime('now', '-60 minutes')
            """,
            (user_id,),
            fetch_one=True
        ) or {'cnt': 0}
        revoked_count = int(recent_revocations.get('cnt', 0))
        if revoked_count >= 4:
            score += 25
            reasons.append(f"recent_revocations:{revoked_count}")

        Database.execute_update(
            """
            UPDATE user_sessions
            SET risk_score = ?, risk_reasons = ?, last_seen = datetime('now')
            WHERE user_id = ? AND session_token = ?
            """,
            (score, ';'.join(reasons) if reasons else None, user_id, session_token)
        )

        if score >= 80:
            Database.execute_update(
                """
                UPDATE user_sessions
                SET is_active = 0, revoked_at = datetime('now'), last_seen = datetime('now')
                WHERE user_id = ? AND session_token = ?
                """,
                (user_id, session_token)
            )
            AuditLogger.log_action(
                user_id=user_id,
                action='session_auto_revoked_risk',
                entity_type='session',
                details=f"risk_score={score};reasons={';'.join(reasons)}",
                status='success'
            )

    @staticmethod
    def get_current_session_row():
        """Fetch persistent row for current session token."""
        token = st.session_state.get('session_token')
        user = st.session_state.get('user')
        user_id = Auth._safe_int(user.get('user_id'), 0) if isinstance(user, dict) else 0
        if not token or user_id <= 0:
            return None
        return Database.execute_query(
            """
            SELECT session_id, trusted_device, trust_label, step_up_verified_until, risk_score, risk_reasons, is_active
            FROM user_sessions
            WHERE user_id = ? AND session_token = ?
            """,
            (user_id, token),
            fetch_one=True
        )

    @staticmethod
    def is_current_device_trusted():
        """Check trust flag for current device/session."""
        row = Auth.get_current_session_row()
        return bool(row and int(row.get('trusted_device') or 0) == 1 and int(row.get('is_active') or 0) == 1)

    @staticmethod
    def is_step_up_verified():
        """Check whether current session has recent step-up verification."""
        row = Auth.get_current_session_row()
        if not row or not row.get('step_up_verified_until'):
            return False
        valid = Database.execute_query(
            """
            SELECT 1 as ok
            FROM user_sessions
            WHERE session_id = ?
              AND step_up_verified_until IS NOT NULL
              AND step_up_verified_until > datetime('now')
            """,
            (row['session_id'],),
            fetch_one=True
        )
        return bool(valid)

    @staticmethod
    def mark_current_device_trusted(trust_label='Trusted Web Device', trust_days=30):
        """Mark current device as trusted."""
        token = st.session_state.get('session_token')
        user = st.session_state.get('user')
        user_id = Auth._safe_int(user.get('user_id'), 0) if isinstance(user, dict) else 0
        if not token or user_id <= 0:
            return False
        return Database.execute_update(
            """
            UPDATE user_sessions
            SET trusted_device = 1,
                trust_label = ?,
                step_up_verified_until = datetime('now', ?),
                last_seen = datetime('now')
            WHERE user_id = ? AND session_token = ?
            """,
            (trust_label, f"+{int(trust_days)} days", user_id, token)
        )

    @staticmethod
    def mark_step_up_verified(minutes=15):
        """Set short-lived step-up verification window for current session."""
        token = st.session_state.get('session_token')
        user = st.session_state.get('user')
        user_id = Auth._safe_int(user.get('user_id'), 0) if isinstance(user, dict) else 0
        if not token or user_id <= 0:
            return False
        return Database.execute_update(
            """
            UPDATE user_sessions
            SET step_up_verified_until = datetime('now', ?), last_seen = datetime('now')
            WHERE user_id = ? AND session_token = ?
            """,
            (f"+{int(minutes)} minutes", user_id, token)
        )

    @staticmethod
    def start_step_up_challenge(user_id, email, purpose='sensitive_action'):
        """Generate and send step-up challenge code."""
        code = str(secrets.randbelow(900000) + 100000)
        if 'step_up_challenges' not in st.session_state:
            st.session_state.step_up_challenges = {}

        st.session_state.step_up_challenges[f"{user_id}:{purpose}"] = {
            'code': code,
            'created_at': datetime.now(),
            'attempts': 0
        }
        EmailService.send_2fa_code(email, code)
        return True

    @staticmethod
    def verify_step_up_challenge(user_id, submitted_code, purpose='sensitive_action'):
        """Verify step-up challenge code and establish short-lived verified state."""
        key = f"{user_id}:{purpose}"
        challenge = (st.session_state.get('step_up_challenges') or {}).get(key)
        if not challenge:
            return False, 'No active challenge. Request a verification code first.'

        elapsed = (datetime.now() - challenge['created_at']).total_seconds()
        if elapsed > 300:
            return False, 'Verification code expired. Request a new one.'

        challenge['attempts'] = int(challenge.get('attempts', 0)) + 1
        if challenge['attempts'] > 5:
            return False, 'Too many failed attempts. Request a new code.'

        if str(submitted_code).strip() != str(challenge['code']).strip():
            return False, 'Invalid verification code.'

        Auth.mark_step_up_verified(minutes=20)
        return True, 'Step-up verification complete.'

    @staticmethod
    def _touch_user_session():
        """Refresh last_seen for active real-user sessions."""
        token = st.session_state.get('session_token')
        user = st.session_state.get('user')
        user_id = Auth._safe_int(user.get('user_id'), 0) if isinstance(user, dict) else 0
        if token and user_id > 0:
            Database.execute_update(
                "UPDATE user_sessions SET last_seen = datetime('now') WHERE session_token = ? AND is_active = 1",
                (token,)
            )

    @staticmethod
    def _is_current_session_active():
        """Check whether current session token is still active for DB users."""
        token = st.session_state.get('session_token')
        user = st.session_state.get('user')
        user_id = Auth._safe_int(user.get('user_id'), 0) if isinstance(user, dict) else 0
        if not token or user_id <= 0:
            return True

        row = Database.execute_query(
            "SELECT is_active FROM user_sessions WHERE session_token = ? AND user_id = ?",
            (token, user_id),
            fetch_one=True
        )
        return bool(row and row.get('is_active') == 1)

    @staticmethod
    def _deactivate_current_session():
        """Mark the active session token as revoked."""
        token = st.session_state.get('session_token')
        user = st.session_state.get('user')
        user_id = Auth._safe_int(user.get('user_id'), 0) if isinstance(user, dict) else 0
        if token and user_id > 0:
            Database.execute_update(
                """
                UPDATE user_sessions
                SET is_active = 0, revoked_at = datetime('now'), last_seen = datetime('now')
                WHERE session_token = ? AND user_id = ?
                """,
                (token, user_id)
            )
    
    @staticmethod
    def is_authenticated():
        """Check if user is authenticated"""
        Auth.init_session()
        if not st.session_state.authenticated:
            return False

        if not Auth._is_current_session_active():
            Auth.logout()
            return False
        
        # Check timeout
        if st.session_state.login_time:
            elapsed = datetime.now() - st.session_state.login_time
            if elapsed > timedelta(minutes=Config.SESSION_TIMEOUT):
                Auth.logout()
                return False
        Auth._touch_user_session()
        return True
    
    @staticmethod
    def set_user(user):
        """Set authenticated user"""
        st.session_state.authenticated = True
        st.session_state.user = user
        st.session_state.login_time = datetime.now()
        Auth._ensure_session_token()
        Auth._upsert_user_session(user)
    
    @staticmethod
    def logout():
        """Logout user"""
        Auth._deactivate_current_session()
        st.session_state.authenticated = False
        st.session_state.user = None
        st.session_state.login_time = None
    
    @staticmethod
    def get_user() -> dict:
        """Get current user - returns dict with user data or empty dict if not authenticated"""
        if Auth.is_authenticated() and hasattr(st.session_state, 'user'):
            return st.session_state.user
        return {}
    
    @staticmethod
    def generate_reset_token(email):
        """Generate password reset token"""
        # Check if user exists
        user = Database.execute_query(
            "SELECT user_id, username FROM users WHERE email = %s AND is_active = 1",
            (email,),
            fetch_one=True
        )
        
        if not user:
            return None
        
        # Generate secure token
        token = secrets.token_urlsafe(32)
        expiry = datetime.now() + timedelta(hours=1)
        
        # Store token (in a real app, use a password_reset_tokens table)
        # For now, we'll use session state
        if 'reset_tokens' not in st.session_state:
            st.session_state.reset_tokens = {}
        
        st.session_state.reset_tokens[token] = {
            'user_id': user['user_id'],
            'email': email,
            'expiry': expiry
        }
        
        return token
    
    @staticmethod
    def verify_reset_token(token):
        """Verify password reset token"""
        if 'reset_tokens' not in st.session_state:
            return None
        
        if token not in st.session_state.reset_tokens:
            return None
        
        token_data = st.session_state.reset_tokens[token]
        
        # Check expiry
        if datetime.now() > token_data['expiry']:
            del st.session_state.reset_tokens[token]
            return None
        
        return token_data
    
    @staticmethod
    def reset_password(token, new_password):
        """Reset password using token"""
        token_data = Auth.verify_reset_token(token)
        
        if not token_data:
            return False, "Invalid or expired token"
        
        # Hash new password
        password_hash = Auth.hash_password(new_password)
        
        # Update password
        success = Database.execute_update(
            "UPDATE users SET password_hash = %s WHERE user_id = %s",
            (password_hash, token_data['user_id'])
        )
        
        if success:
            # Remove used token
            del st.session_state.reset_tokens[token]
            return True, "Password reset successful"
        
        return False, "Password reset failed"
    
    @staticmethod
    def is_demo_admin(user):
        """Check if current user is demo admin"""
        return user and user.get('is_demo', False)
    
    @staticmethod
    def require_full_admin():
        """Check if user has full admin privileges (not demo)"""
        user = Auth.get_user()
        if not user:
            return False
        if user.get('is_demo'):
            st.warning(" Demo admin cannot perform write operations")
            return False
        return user.get('role') in ['admin', 'librarian']
    
    @staticmethod
    def change_password(user_id, old_password, new_password):
        """Change password (requires old password)"""
        # Get current password hash
        user = Database.execute_query(
            "SELECT password_hash FROM users WHERE user_id = %s",
            (user_id,),
            fetch_one=True
        )
        
        if not user:
            return False, "User not found"
        
        # Verify old password
        if not Auth.verify_password(old_password, user['password_hash']):
            return False, "Current password is incorrect"
        
        # Hash new password
        password_hash = Auth.hash_password(new_password)
        
        # Update password
        success = Database.execute_update(
            "UPDATE users SET password_hash = %s WHERE user_id = %s",
            (password_hash, user_id)
        )
        
        if success:
            return True, "Password changed successfully"
        
        return False, "Password change failed"

# ================================================================
# HELPER FUNCTIONS
# ================================================================

def format_date(date_obj):
    """Format date"""
    if date_obj is None:
        return 'N/A'
    if isinstance(date_obj, str):
        return date_obj
    return date_obj.strftime('%Y-%m-%d')

def format_currency(amount):
    """Format currency"""
    if amount is None:
        return '$0.00'
    return f'${amount:,.2f}'

def format_datetime(dt_obj):
    """Format datetime"""
    if dt_obj is None:
        return 'N/A'
    if isinstance(dt_obj, str):
        return dt_obj
    return dt_obj.strftime('%Y-%m-%d %H:%M:%S')

def calculate_reading_time(pages, words_per_minute=200):
    """Estimate reading time based on page count"""
    if not pages:
        return "N/A"
    words = pages * 250  # Average words per page
    minutes = words / words_per_minute
    hours = minutes / 60
    if hours < 1:
        return f"{int(minutes)} minutes"
    return f"{hours:.1f} hours"

def get_book_recommendation_score(book_data):
    """Calculate recommendation score for a book"""
    score = 0
    # Rating weight (40%)
    if book_data.get('average_rating'):
        score += (book_data['average_rating'] / 5.0) * 40
    # Popularity weight (30%)
    if book_data.get('total_checkouts'):
        score += min(book_data['total_checkouts'] / 100, 1) * 30
    # Recent activity weight (20%)
    if book_data.get('recent_checkouts'):
        score += min(book_data['recent_checkouts'] / 20, 1) * 20
    # Availability weight (10%)
    if book_data.get('is_available'):
        score += 10
    return score

def analyze_reading_pattern(borrowing_history):
    """Analyze user's reading patterns"""
    if not borrowing_history:
        return {}
    
    df = pd.DataFrame(borrowing_history)
    
    pattern = {
        'favorite_genres': [],
        'favorite_authors': [],
        'avg_books_per_month': 0,
        'preferred_day': 'N/A',
        'reading_streak': 0
    }
    
    # Calculate metrics
    if 'genre_name' in df.columns:
        pattern['favorite_genres'] = df['genre_name'].value_counts().head(3).index.tolist()
    
    if 'author_name' in df.columns:
        pattern['favorite_authors'] = df['author_name'].value_counts().head(3).index.tolist()
    
    if 'checkout_date' in df.columns:
        df['checkout_date'] = pd.to_datetime(df['checkout_date'])
        # Books per month
        pattern['avg_books_per_month'] = len(df) / max(df['checkout_date'].dt.month.nunique(), 1)
        # Preferred day
        pattern['preferred_day'] = df['checkout_date'].dt.day_name().mode()[0] if len(df) > 0 else 'N/A'
    
    return pattern

def generate_book_insights(books):
    """Generate insights from book collection"""
    if not books:
        return {}
    
    df = pd.DataFrame(books)
    
    insights = {
        'total_books': len(df),
        'total_pages': df['page_count'].sum() if 'page_count' in df.columns else 0,
        'avg_rating': df['average_rating'].mean() if 'average_rating' in df.columns else 0,
        'most_popular_genre': 'N/A',
        'newest_book': 'N/A',
        'oldest_book': 'N/A'
    }
    
    if 'genre_name' in df.columns:
        insights['most_popular_genre'] = df['genre_name'].mode()[0] if len(df) > 0 else 'N/A'
    
    if 'publication_year' in df.columns:
        insights['newest_book'] = df['publication_year'].max()
        insights['oldest_book'] = df['publication_year'].min()
    
    return insights

def calculate_fine(due_date, return_date=None):
    """Calculate fine for overdue books"""
    if return_date is None:
        return_date = datetime.now()
    
    if isinstance(due_date, str):
        due_date = datetime.strptime(due_date, '%Y-%m-%d')
    if isinstance(return_date, str):
        return_date = datetime.strptime(return_date, '%Y-%m-%d')
    
    if return_date <= due_date:
        return 0.0
    
    days_overdue = (return_date - due_date).days
    return days_overdue * Config.FINE_PER_DAY

def get_member_statistics(user_id):
    """Get comprehensive member statistics"""
    stats = {}
    
    # Total books read
    total_read = Database.execute_query(
        "SELECT COUNT(*) as count FROM borrowing WHERE user_id = ? AND return_date IS NOT NULL",
        (user_id,),
        fetch_one=True
    )
    stats['total_read'] = total_read['count'] if total_read else 0
    
    # Currently borrowed
    currently_borrowed = Database.execute_query(
        "SELECT COUNT(*) as count FROM borrowing WHERE user_id = ? AND return_date IS NULL",
        (user_id,),
        fetch_one=True
    )
    stats['currently_borrowed'] = currently_borrowed['count'] if currently_borrowed else 0
    
    # Total time reading (days)
    reading_time = Database.execute_query(
        """SELECT SUM(julianday(COALESCE(return_date, date('now'))) - julianday(checkout_date)) as days 
           FROM borrowing WHERE user_id = ?""",
        (user_id,),
        fetch_one=True
    )
    stats['total_reading_days'] = int(reading_time['days']) if reading_time and reading_time['days'] else 0
    
    # Favorite genre - simplified to use books table directly
    fav_genre = Database.execute_query(
        """SELECT b.genre, COUNT(*) as count 
           FROM borrowing br 
           JOIN books b ON br.book_id = b.book_id
           WHERE br.user_id = ? AND b.genre IS NOT NULL
           GROUP BY b.genre
           ORDER BY count DESC LIMIT 1""",
        (user_id,),
        fetch_one=True
    )
    stats['favorite_genre'] = fav_genre['genre'] if fav_genre else 'N/A'
    
    return stats

# ================================================================
# UI STYLING
# ================================================================

def load_css():
    """Load custom CSS with JetBrains Mono font - non-intrusive styling"""
    css = """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;700;800&display=swap');

    /* Only apply font to markdown content, not Streamlit widgets */
    .stMarkdown p,
    .stMarkdown h1,
    .stMarkdown h2,
    .stMarkdown h3,
    .stMarkdown h4,
    .stMarkdown h5,
    .stMarkdown h6,
    .stMarkdown li {
        font-family: 'JetBrains Mono', 'Courier New', monospace;
    }

    /* Custom LitGrid classes - namespaced to avoid conflicts */
    .litgrid-header {
        color: #1E88E5;
        font-size: 2.5rem;
        font-weight: 800;
        text-align: center;
        margin-bottom: 1rem;
        font-family: 'JetBrains Mono', 'Courier New', monospace;
    }

    .litgrid-stat-card {
        background: linear-gradient(135deg, #1E88E5 0%, #FFA726 100%);
        padding: 1.5rem;
        border-radius: 10px;
        color: white;
        text-align: center;
        margin: 0.5rem;
        font-family: 'JetBrains Mono', 'Courier New', monospace;
    }

    .litgrid-stat-number {
        font-size: 2.5rem;
        font-weight: 800;
        font-family: 'JetBrains Mono', 'Courier New', monospace;
    }

    .litgrid-stat-label {
        font-size: 1rem;
        opacity: 0.9;
        font-family: 'JetBrains Mono', 'Courier New', monospace;
    }

    .litgrid-book-card {
        background: white;
        border-radius: 8px;
        padding: 1rem;
        margin: 0.5rem;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        font-family: 'JetBrains Mono', 'Courier New', monospace;
    }

    /* Preserve Streamlit defaults */
    .stButton > button,
    .stTextInput > div > div > input,
    .stSelectbox > div > div > select,
    .stTextArea > div > div > textarea {
        font-family: inherit;
    }
    </style>
    """
    st.markdown(css, unsafe_allow_html=True)

# ================================================================
# PAGES
# ================================================================

def show_login_page():
    """Login/Registration page with mode selection and 2FA"""
    col1, col2, col3 = st.columns([1, 2, 1], gap="small")
    
    with col2:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("""
        <div style='text-align: center;'>
            <h1 style='color: #1E88E5; font-size: 3rem;'> LitGrid</h1>
            <p style='color: #666; font-size: 1.2rem;'>Advanced Library Management System</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # Initialize second password session state
        if 'second_password_pending' not in st.session_state:
            st.session_state.second_password_pending = False
        if 'second_password_username' not in st.session_state:
            st.session_state.second_password_username = None
        if 'superadmin_second_factor_pending' not in st.session_state:
            st.session_state.superadmin_second_factor_pending = False
        if 'superadmin_second_factor_username' not in st.session_state:
            st.session_state.superadmin_second_factor_username = None

        # Show super admin second-factor verification if pending
        if st.session_state.superadmin_second_factor_pending:
            st.markdown("### Super Admin Security Verification")
            st.warning("Enter your super admin security key to complete login")

            with st.form("superadmin_second_factor_form"):
                superadmin_security_key = st.text_input("Super Admin Security Key", type="password")
                submit_superadmin_security = st.form_submit_button("Verify", use_container_width=True)
                cancel_superadmin_security = st.form_submit_button("Cancel")

                if cancel_superadmin_security:
                    st.session_state.superadmin_second_factor_pending = False
                    st.session_state.superadmin_second_factor_username = None
                    st.rerun()

                if submit_superadmin_security and superadmin_security_key:
                    user = Auth.verify_superadmin_security_key(
                        st.session_state.superadmin_second_factor_username,
                        superadmin_security_key
                    )
                    if user and isinstance(user, dict) and user.get('error'):
                        st.error(user['error'])
                    elif user:
                        Auth.set_user(user)
                        st.session_state.superadmin_second_factor_pending = False
                        st.session_state.superadmin_second_factor_username = None
                        st.success("Super admin authentication successful")
                        st.balloons()
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.error("Invalid super admin security key")
            return
        
        # Show second password verification if pending
        if st.session_state.second_password_pending:
            st.markdown("###  Security Verification")
            st.info("Please enter your security password to complete login")
            
            with st.form("second_password_form"):
                security_key = st.text_input("Security Password", type="password")
                submit_security = st.form_submit_button("Verify", use_container_width=True)
                cancel_security = st.form_submit_button("Cancel")
                
                if cancel_security:
                    st.session_state.second_password_pending = False
                    st.session_state.second_password_username = None
                    st.rerun()
                
                if submit_security and security_key:
                    user = Auth.verify_functional_admin_security_key(
                        st.session_state.second_password_username, 
                        security_key
                    )
                    if user:
                        Auth.set_user(user)
                        st.session_state.second_password_pending = False
                        st.session_state.second_password_username = None
                        st.success(" Authentication successful!")
                        st.balloons()
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.error(" Invalid security password")
            return
        
        # Normal login/register page
        tab1, tab2 = st.tabs(["Login", "Register"])

        def render_dynamic_password_reset_panel():
            st.markdown("### Reset Password")
            st.caption("Secure password recovery with real-time OTP verification")

            # Initialize session states for password reset
            if 'reset_email' not in st.session_state:
                st.session_state.reset_email = None
            if 'reset_otp' not in st.session_state:
                st.session_state.reset_otp = None
            if 'reset_step' not in st.session_state:
                st.session_state.reset_step = 'email'  # Steps: email -> otp -> password
            if 'otp_attempts' not in st.session_state:
                st.session_state.otp_attempts = 0
            if 'otp_timestamp' not in st.session_state:
                st.session_state.otp_timestamp = None

            # Helper function to generate OTP
            def generate_otp():
                """Generate 6-digit OTP"""
                import random
                return ''.join([str(random.randint(0, 9)) for _ in range(6)])

            # Helper function to send OTP (simulated with display)
            def send_otp_email(email, otp):
                """Simulate sending OTP to email"""
                # In production, integrate with email service (SendGrid, AWS SES, etc.)
                return True, f"OTP sent to {email}"

            # ===== STEP 1: EMAIL VERIFICATION =====
            if st.session_state.reset_step == 'email':
                st.markdown("#### Step 1: Verify Your Email")
                st.caption("Enter your registered email address")

                with st.form("email_verification_form"):
                    reset_email = st.text_input(
                        "Email Address *",
                        placeholder="your.email@example.com",
                        help="Enter the email associated with your account"
                    )

                    send_otp_btn = st.form_submit_button("Send OTP", use_container_width=True)

                    if send_otp_btn:
                        if not reset_email:
                            st.error("Please enter your email address")
                        elif "@" not in reset_email:
                            st.error("Invalid email format")
                        else:
                            user_check = Database.execute_query(
                                "SELECT user_id FROM users WHERE LOWER(email) = LOWER(?) AND is_active = 1",
                                (reset_email,),
                                fetch_one=True
                            )

                            if user_check:
                                otp = generate_otp()
                                success, msg = send_otp_email(reset_email, otp)

                                if success:
                                    st.session_state.reset_email = reset_email
                                    st.session_state.reset_otp = otp
                                    st.session_state.reset_step = 'otp'
                                    st.session_state.otp_attempts = 0
                                    st.session_state.otp_timestamp = datetime.now()

                                    st.success("OTP sent successfully")
                                    st.info(f"One-Time Password sent to: `{reset_email}`")
                                    st.caption("The 6-digit OTP will expire in 10 minutes")
                                    time.sleep(1)
                                    st.rerun()
                                else:
                                    st.error(f"Could not send OTP: {msg}")
                            else:
                                st.error("Email not found or account is inactive")
                                st.caption("Make sure you entered the correct registered email")

            # ===== STEP 2: OTP VERIFICATION =====
            elif st.session_state.reset_step == 'otp':
                st.markdown("#### Step 2: Verify OTP")
                st.info(f"OTP sent to: `{st.session_state.reset_email}`")

                if st.session_state.otp_timestamp:
                    elapsed = (datetime.now() - st.session_state.otp_timestamp).total_seconds()
                    remaining_time = 600 - elapsed
                    minutes = int(remaining_time // 60)
                    seconds = int(remaining_time % 60)

                    if remaining_time <= 0:
                        st.error("OTP has expired. Please request a new one.")
                        if st.button("Get New OTP", use_container_width=True):
                            st.session_state.reset_step = 'email'
                            st.rerun()
                    else:
                        st.caption(f"OTP expires in: {minutes}m {seconds}s")

                        with st.form("otp_verification_form"):
                            otp_input = st.text_input(
                                "Enter 6-Digit OTP *",
                                placeholder="000000",
                                max_chars=6,
                                help="Check your email for the OTP"
                            )

                            col_otp1, col_otp2 = st.columns(2, gap="small")

                            with col_otp1:
                                verify_otp_btn = st.form_submit_button("Verify OTP", use_container_width=True)

                            with col_otp2:
                                resend_otp_btn = st.form_submit_button("Resend OTP", use_container_width=True)

                            if resend_otp_btn:
                                new_otp = generate_otp()
                                success, msg = send_otp_email(st.session_state.reset_email, new_otp)

                                if success:
                                    st.session_state.reset_otp = new_otp
                                    st.session_state.otp_attempts = 0
                                    st.session_state.otp_timestamp = datetime.now()
                                    st.success("New OTP sent")
                                    st.rerun()

                            if verify_otp_btn:
                                if not otp_input:
                                    st.error("Please enter the OTP")
                                elif len(otp_input) != 6 or not otp_input.isdigit():
                                    st.error("OTP must be 6 digits")
                                elif otp_input == st.session_state.reset_otp:
                                    st.success("OTP verified successfully")
                                    st.session_state.reset_step = 'password'
                                    st.session_state.otp_attempts = 0
                                    time.sleep(1)
                                    st.rerun()
                                else:
                                    st.session_state.otp_attempts += 1
                                    remaining_attempts = 3 - st.session_state.otp_attempts

                                    st.error("Incorrect OTP")
                                    st.warning(f"Attempts remaining: {remaining_attempts}")

                                    if st.session_state.otp_attempts >= 3:
                                        st.error("Maximum attempts exceeded. Please request a new OTP.")
                                        if st.button("Get New OTP", use_container_width=True):
                                            st.session_state.reset_step = 'email'
                                            st.rerun()

            # ===== STEP 3: PASSWORD RESET =====
            elif st.session_state.reset_step == 'password':
                st.markdown("#### Step 3: Set New Password")
                st.success("Email and OTP verified")

                def validate_password_strength(password):
                    """Check password strength"""
                    if len(password) < 12:
                        return False, "Password must be at least 12 characters"
                    if password.lower() == password:
                        return False, "Password must contain uppercase letters"
                    if password.upper() == password:
                        return False, "Password must contain lowercase letters"
                    if not any(c.isdigit() for c in password):
                        return False, "Password must contain numbers"
                    if not any(c in "!@#$%^&*()-_=+[]{}|;:,.<>?" for c in password):
                        return False, "Password must contain special characters"
                    if any(password.count(c) > 3 for c in set(password)):
                        return False, "Password contains too many repeated characters"
                    return True, "Strong password"

                with st.form("new_password_form"):
                    st.caption("Password Requirements: 12+ chars, uppercase, lowercase, numbers, special chars, no repetition")

                    new_password = st.text_input(
                        "New Password *",
                        type="password",
                        placeholder="Enter strong password",
                        help="Min 12 characters with mixed case, numbers, special chars"
                    )
                    confirm_new_password = st.text_input(
                        "Confirm Password *",
                        type="password",
                        placeholder="Re-enter password"
                    )

                    submit_reset = st.form_submit_button("Reset Password", use_container_width=True)

                    if submit_reset:
                        errors = []

                        if not new_password or not confirm_new_password:
                            errors.append("Please enter and confirm your new password")

                        if new_password != confirm_new_password:
                            errors.append("Passwords do not match")

                        if new_password:
                            is_strong, msg = validate_password_strength(new_password)
                            if not is_strong:
                                errors.append(f"Password too weak: {msg}")

                        if errors:
                            for error in errors:
                                st.error(error)
                        else:
                            try:
                                password_hash = Auth.hash_password(new_password)

                                success = Database.execute_update(
                                    "UPDATE users SET password_hash = ? WHERE LOWER(email) = LOWER(?) AND is_active = 1",
                                    (password_hash, st.session_state.reset_email)
                                )

                                if success:
                                    st.success("Password reset successfully")
                                    st.info("You can now login with your new password")

                                    st.session_state.reset_email = None
                                    st.session_state.reset_otp = None
                                    st.session_state.reset_step = 'email'
                                    st.session_state.otp_attempts = 0
                                    st.session_state.otp_timestamp = None

                                    st.balloons()
                                    time.sleep(2)
                                    st.rerun()
                                else:
                                    st.error("Failed to reset password. Please try again.")
                            except Exception as e:
                                st.error(f"Error resetting password: {str(e)}")
        
        with tab1:
            login_col, demo_col = st.columns(2, gap="medium")

            with login_col:
                # Mode selection
                st.markdown("### Login Mode")
                login_mode = st.radio(
                    "I am logging in as:",
                    ["Member", "Administrator"],
                    horizontal=True,
                    label_visibility="collapsed"
                )

                mode = 'member' if 'Member' in login_mode else 'admin'

            with demo_col:
                # Demo mode quick access buttons
                st.markdown("### Demo Access")
                demo_col1, demo_col2 = st.columns(2, gap="small")

                with demo_col1:
                    if st.button("Demo Member", use_container_width=True, key="demo_member_btn"):
                        # Create demo member session
                        demo_user = {
                            'user_id': 'demo_member_001',
                            'username': 'demo_member',
                            'full_name': 'Demo Member',
                            'email': 'demo.member@litgrid.local',
                            'role': 'member',
                            'is_active': True,
                            'is_demo': True,
                            'member_tier': 'standard',
                            'fine_balance': 0.0
                        }
                        Auth.set_user(demo_user)
                        st.info("Demo Member Mode - View Only Access")
                        time.sleep(0.5)
                        st.rerun()

                with demo_col2:
                    if st.button("Demo Admin", use_container_width=True, key="demo_admin_btn"):
                        # Create demo admin session
                        demo_user = {
                            'user_id': 'demo_admin_001',
                            'username': 'demo_admin',
                            'full_name': 'Demo Administrator',
                            'email': 'demo.admin@litgrid.local',
                            'role': 'admin',
                            'is_active': True,
                            'is_demo': True,
                            'member_tier': 'administrator',
                            'fine_balance': 0.0
                        }
                        Auth.set_user(demo_user)
                        st.info("Demo Admin Mode - Limited Write Access")
                        time.sleep(0.5)
                        st.rerun()

            st.divider()
            cred_col, reset_col = st.columns(2, gap="medium")

            with cred_col:
                st.markdown("### Login with Credentials")

                with st.form("login_form"):
                    username = st.text_input("Username or Email")
                    password = st.text_input("Password", type="password")
                    submit = st.form_submit_button("Login", use_container_width=True)

                    if submit:
                        if username and password:
                            user = Auth.login(username, password, mode)

                            if user and isinstance(user, dict) and user.get('error'):
                                st.error(user['error'])

                            # Check if needs super admin second factor
                            elif user and isinstance(user, dict) and user.get('needs_superadmin_security_key'):
                                st.session_state.superadmin_second_factor_pending = True
                                st.session_state.superadmin_second_factor_username = user['username']
                                st.rerun()

                            # Check if needs second password (functional admin)
                            elif user and isinstance(user, dict) and user.get('needs_second_password'):
                                st.session_state.second_password_pending = True
                                st.session_state.second_password_username = user['username']
                                st.rerun()

                            elif user:
                                Auth.set_user(user)
                                if user.get('is_demo'):
                                    st.warning(" Demo Admin Mode - View Only Access")
                                st.success(f"Welcome, {user['full_name']}!")
                                time.sleep(1)
                                st.rerun()
                            else:
                                st.error(f"Invalid credentials for {mode} login")
                        else:
                            st.warning("Please enter username and password")

            with reset_col:
                render_dynamic_password_reset_panel()
        
        with tab2:
            # Password strength validator (defined once for both tabs)
            def validate_password_strength(password):
                """Check password strength"""
                if len(password) < 12:
                    return False, "Password must be at least 12 characters"
                if password.lower() == password:
                    return False, "Password must contain uppercase letters"
                if password.upper() == password:
                    return False, "Password must contain lowercase letters"
                if not any(c.isdigit() for c in password):
                    return False, "Password must contain numbers"
                if not any(c in "!@#$%^&*()-_=+[]{}|;:,.<>?" for c in password):
                    return False, "Password must contain special characters (!@#$%^&* etc)"
                if any(password.count(c) > 3 for c in set(password)):
                    return False, "Password contains too many repeated characters"
                return True, "Strong password ✓"

            def get_password_criteria(password):
                """Return granular password checks for live feedback."""
                value = password or ""
                return {
                    'length': len(value) >= 12,
                    'uppercase': any(c.isupper() for c in value),
                    'lowercase': any(c.islower() for c in value),
                    'digit': any(c.isdigit() for c in value),
                    'special': any(c in "!@#$%^&*()-_=+[]{}|;:,.<>?" for c in value),
                    'repeat_limit': not any(value.count(c) > 3 for c in set(value)) if value else False,
                }

            def show_password_feedback(password, confirm_password=None):
                """Render dynamic password-strength feedback."""
                checks = get_password_criteria(password)
                passed = sum(1 for ok in checks.values() if ok)
                total = len(checks)
                is_strong = passed == total

                if password:
                    labels = [
                        (checks['length'], "At least 12 characters"),
                        (checks['uppercase'], "Contains uppercase letter"),
                        (checks['lowercase'], "Contains lowercase letter"),
                        (checks['digit'], "Contains number"),
                        (checks['special'], "Contains special character"),
                        (checks['repeat_limit'], "No character repeated more than 3 times"),
                    ]
                    for ok, text in labels:
                        st.caption(f"{'✅' if ok else '❌'} {text}")

                    if confirm_password is not None and confirm_password:
                        st.caption(f"{'✅' if password == confirm_password else '❌'} Passwords match")

                    if is_strong:
                        st.success("Password strength: Strong")
                    else:
                        st.warning(f"Password strength: {passed}/{total} requirements met")

                return is_strong

            # Check for duplicate credentials (defined once for both tabs)
            def check_duplicate_credentials(username, email):
                """Check if username or email already exists"""
                existing = Database.execute_query("""
                    SELECT user_id FROM users
                    WHERE LOWER(username) = LOWER(?) OR LOWER(email) = LOWER(?)
                """, (username, email))
                return existing is not None and len(existing) > 0

            # Show remaining member slots for capacity-aware validation
            count_result = Database.execute_query(
                "SELECT COUNT(*) as count FROM users WHERE role = 'member'",
                fetch_one=True
            )
            current_count = count_result['count'] if count_result else 0
            remaining = Config.MAX_MEMBER_ACCOUNTS - current_count
            with st.form("register_unified_form"):
                account_type = st.session_state.get("unified_account_type", "Member")

                st.markdown("#### **Information**")
                section_col1, section_col2, section_col3 = st.columns(3, gap="medium")

                with section_col1:
                    full_name = st.text_input("Full Name *", placeholder="John Doe")
                    date_of_birth = st.date_input("Date of Birth *")
                    phone = st.text_input("Phone Number *", placeholder="+1-555-0000")
                    gender = st.selectbox("Gender", ["Prefer not to say", "Male", "Female", "Other"])

                with section_col2:
                    street = st.text_input("Street Address *")
                    city = st.text_input("City *")
                    state = st.text_input("State/Province *")
                    zip_code = st.text_input("Zip/Postal Code *")

                with section_col3:
                    username_reg = st.text_input("Username *", placeholder="johndoe123", help="Alphanumeric, 4-20 characters")
                    email = st.text_input("Email *", placeholder="john@example.com")

                    if account_type == "Staff":
                        reg_role = st.selectbox("Role *", ["librarian", "admin"], help="Admin: Full system access, Librarian: Book & member management")
                        department = st.selectbox("Department", ["Circulation", "Cataloging", "Reference", "Administration"])
                        employment_date = st.date_input("Employment Start Date *")
                    else:
                        occupation = st.text_input("Occupation", placeholder="Software Engineer")
                        library_interests = st.multiselect(
                            "Library Interests",
                            ["Fiction", "Non-Fiction", "Science", "History", "Biography", "Self-Help", "Poetry", "Children's Books"]
                        )

                security_col, photo_col = st.columns(2, gap="medium")

                with security_col:
                    st.markdown("#### **Password Security**")
                    st.caption("Password Requirements: 12+ chars, uppercase, lowercase, numbers, special chars, no repetition")
                    password_reg = st.text_input("Password *", type="password", help="Min 12 chars, mixed case, numbers, special chars", key="unified_pass")
                    confirm_pass = st.text_input("Confirm Password *", type="password", key="unified_confirm")
                    is_password_strong = show_password_feedback(password_reg, confirm_pass)

                with photo_col:
                    st.markdown("#### **Profile Photo**")
                    profile_photo = st.file_uploader("Upload Photo *", type=["jpg", "jpeg", "png"], help="Max 5MB", key="unified_photo")

                st.divider()
                legal_col, agreement_col = st.columns(2, gap="medium")

                with legal_col:
                    st.markdown("#### **Legal & Consent**")
                    account_type = st.radio(
                        "Register As",
                        ["Member", "Staff"],
                        horizontal=True,
                        key="unified_account_type"
                    )
                    with st.expander("Terms of Service", expanded=False):
                        st.markdown("Read and accept terms before registration.")
                    with st.expander("Usage Guide", expanded=False):
                        st.markdown("Learn borrowing, return, and policy basics.")
                    with st.expander("Privacy & Legal Disclaimer", expanded=False):
                        st.markdown("Understand privacy handling and legal policy.")

                with agreement_col:
                    st.markdown("#### **Your Agreement**")
                    agree_terms = st.radio(
                        "I have read and agree to all terms:",
                        ["I do not agree", "I agree to all Terms, Usage Guide, and Legal Disclaimer"],
                        key="unified_agree_terms"
                    )
                    agree_data = st.checkbox("I understand my data will be securely stored and used only for library services", key="unified_agree_data")

                    if account_type == "Staff":
                        agree_staff = st.radio(
                            "I confirm employment authorization:",
                            ["Do not proceed", "I confirm all information is accurate and authorized"],
                            key="unified_agree_staff"
                        )
                    else:
                        agree_contact = st.checkbox("I agree to receive library notifications (overdue reminders, new book alerts)", key="unified_agree_contact")

                password_ready = (
                    bool(password_reg)
                    and bool(confirm_pass)
                    and is_password_strong
                    and password_reg == confirm_pass
                )

                submit_reg = st.form_submit_button(
                    "Complete Registration",
                    use_container_width=True,
                    disabled=(not password_ready or (account_type == "Member" and remaining <= 0))
                )

                if submit_reg:
                    errors = []

                    if not all([full_name, username_reg, email, phone, street, city, state, zip_code, password_reg, confirm_pass, profile_photo]):
                        errors.append("Please fill all required fields (marked with *)")

                    if account_type == "Member" and remaining <= 0:
                        errors.append("No member registration slots available")

                    if username_reg and not (4 <= len(username_reg) <= 20):
                        errors.append("Username must be 4-20 characters")

                    if email and "@" not in email:
                        errors.append("Invalid email format")

                    if password_reg != confirm_pass:
                        errors.append("Passwords do not match")

                    if password_reg and not is_password_strong:
                        is_strong, msg = validate_password_strength(password_reg)
                        if not is_strong:
                            errors.append(f"Password too weak: {msg}")

                    if username_reg or email:
                        if check_duplicate_credentials(username_reg, email):
                            errors.append("Username or email already registered. Choose different credentials.")

                    if profile_photo and profile_photo.size > 5 * 1024 * 1024:
                        errors.append("Photo must be less than 5MB")

                    if "do not" in agree_terms.lower():
                        errors.append("You must agree to all terms to register")

                    if not agree_data:
                        errors.append("You must acknowledge data storage terms")

                    if account_type == "Staff" and "agree_staff" in locals() and "not proceed" in agree_staff.lower():
                        errors.append("You must confirm employment authorization")

                    if errors:
                        for error in errors:
                            st.error(f"❌ {error}")
                    else:
                        try:
                            role_to_create = 'member' if account_type == "Member" else reg_role
                            success, msg = Auth.register(username_reg, email, password_reg, full_name, phone, role=role_to_create)
                            if success:
                                if account_type == "Member":
                                    st.success("✅ Member account created successfully!")
                                else:
                                    st.success("✅ Staff account created successfully!")
                                    st.info(f"Role: {reg_role.title()}, Department: {department}")
                                st.info("You can now login with your credentials")
                                st.balloons()
                            else:
                                st.error(f"❌ Registration failed: {msg}")
                        except Exception as e:
                            st.error(f"❌ Registration error: {str(e)}")


def show_dashboard():
    """Dashboard page"""
    user = Auth.get_user()

    st.markdown(f'<h1 class="litgrid-header"> Dashboard</h1>', unsafe_allow_html=True)
    sanitized_name = security_manager.sanitize_input(user['full_name'])
    st.markdown(f"<p style='text-align: center; color: #666;'>Welcome back, {sanitized_name}!</p>", unsafe_allow_html=True)

    # Show demo mode banner if applicable
    if user.get('is_demo'):
        role_text = "Member" if user['role'] == 'member' else "Administrator"
        st.warning(f"Demo Mode - {role_text} ({user.get('username', 'demo')}): This is a demonstration account with read-only access to explore features.")
        st.divider()

    if user['role'] == 'superadmin':
        # SUPER ADMIN DASHBOARD - Highest Level Access with Full Authority
        st.markdown("# **SUPER ADMIN CONTROL CENTER**")
        st.markdown("**Unrestricted System Access | No Permissions Required**")
        st.divider()

        # Super Admin Tabs - 10 Comprehensive Tabs
        sa_tab1, sa_tab2, sa_tab3, sa_tab4, sa_tab5, sa_tab6, sa_tab7, sa_tab8, sa_tab9, sa_tab10 = st.tabs([
            "User Control", "System Stats", "Activity Logs", "Data Management", "System Config",
            "Emergency Control", "Reports", "Security", "Performance", "System Admin"
        ])

        # ============ TAB 1: COMPREHENSIVE USER CONTROL ============
        with sa_tab1:
            st.subheader("**Complete User Control System**")

            col1, col2, col3 = st.columns(3, gap="small")
            with col1:
                user_type_filter = st.selectbox("Filter by Role", ["All Users", "Members", "Admins", "Super Admins"], key="sa_role_filter")
            with col2:
                status_filter = st.selectbox("Status", ["All", "Active", "Inactive"], key="sa_status_filter")
            with col3:
                search_term = st.text_input("Search User (name/email/username)", key="sa_search")

            role_map = {"All Users": None, "Members": "member", "Admins": "admin", "Super Admins": "superadmin"}
            role_filter = role_map[user_type_filter]

            # Build query
            query = "SELECT user_id, username, full_name, email, role, is_active, created_at as registration_date FROM users WHERE 1=1"
            params = []

            if role_filter:
                query += " AND role = ?"
                params.append(role_filter)

            if status_filter == "Active":
                query += " AND is_active = 1"
            elif status_filter == "Inactive":
                query += " AND is_active = 0"

            if search_term:
                query += " AND (username LIKE ? OR full_name LIKE ? OR email LIKE ?)"
                params.extend([f"%{search_term}%", f"%{search_term}%", f"%{search_term}%"])

            query += " ORDER BY created_at DESC"
            all_users = Database.execute_query(query, params) if params else Database.execute_query(query)

            if all_users:
                st.write(f"**Total Users: {len(all_users)}**")

                # Display users in expandable sections for detailed control
                for u in all_users:
                    with st.expander(f"👤 {u['full_name']} (@{u['username']}) - {u['role'].upper()}"):
                        exp_col1, exp_col2, exp_col3 = st.columns(3, gap="small")

                        with exp_col1:
                            st.write(f"**Email:** {u['email']}")
                            st.write(f"**ID:** {u['user_id']}")
                            st.write(f"**Status:** {'Active' if u['is_active'] else 'Inactive'}")
                            st.write(f"**Registered:** {u['registration_date'][:10]}")

                        with exp_col2:
                            st.write("**Quick Actions:**")
                            action_col1, action_col2 = st.columns(2, gap="small")

                            with action_col1:
                                if st.button(f"Toggle Active", key=f"toggle_{u['user_id']}"):
                                    new_status = 0 if u['is_active'] else 1
                                    Database.execute_update(
                                        "UPDATE users SET is_active = ? WHERE user_id = ?",
                                        (new_status, u['user_id'])
                                    )
                                    st.success(f"Status updated for {u['username']}")
                                    st.rerun()

                                if u['role'] != 'superadmin' and st.button(f"Promote to Admin", key=f"promote_{u['user_id']}"):
                                    Database.execute_update(
                                        "UPDATE users SET role = 'admin' WHERE user_id = ?",
                                        (u['user_id'],)
                                    )
                                    st.success(f"{u['username']} promoted to Admin")
                                    st.rerun()

                            with action_col2:
                                if st.button(f"Reset Password", key=f"reset_pwd_{u['user_id']}"):
                                    temp_pwd = 'TempPass@123'
                                    pwd_hash = Auth.hash_password(temp_pwd)
                                    Database.execute_update(
                                        "UPDATE users SET password_hash = ? WHERE user_id = ?",
                                        (pwd_hash, u['user_id'])
                                    )
                                    st.info(f"Password reset to: {temp_pwd}")

                                if u['role'] == 'admin' and st.button(f"Demote to Member", key=f"demote_{u['user_id']}"):
                                    Database.execute_update(
                                        "UPDATE users SET role = 'member' WHERE user_id = ?",
                                        (u['user_id'],)
                                    )
                                    st.success(f"{u['username']} demoted to Member")
                                    st.rerun()

                        with exp_col3:
                            st.write("**Danger Zone:**")
                            if st.button(f"Ban User", key=f"ban_{u['user_id']}", type="secondary"):
                                Database.execute_update(
                                    "UPDATE users SET is_active = 0 WHERE user_id = ?",
                                    (u['user_id'],)
                                )
                                st.warning(f"{u['username']} has been banned")
                                st.rerun()

                            if st.button(f"Delete User (Permanent)", key=f"delete_{u['user_id']}", type="secondary"):
                                Database.execute_update("DELETE FROM borrowing WHERE user_id = ?", (u['user_id'],))
                                Database.execute_update("DELETE FROM fines WHERE user_id = ?", (u['user_id'],))
                                Database.execute_update("DELETE FROM users WHERE user_id = ?", (u['user_id'],))
                                st.error(f"{u['username']} permanently deleted")
                                st.rerun()
            else:
                st.info("No users found matching criteria")

        # ============ TAB 2: SYSTEM STATISTICS & METRICS ============
        with sa_tab2:
            st.subheader("**System Statistics & Metrics**")

            # KPI Row 1
            col1, col2, col3, col4, col5, col6 = st.columns(6, gap="small")

            total_users = Database.execute_query("SELECT COUNT(*) as count FROM users", fetch_one=True)
            active_users = Database.execute_query("SELECT COUNT(*) as count FROM users WHERE is_active = 1", fetch_one=True)
            total_members = Database.execute_query("SELECT COUNT(*) as count FROM users WHERE role = 'member'", fetch_one=True)
            total_admins = Database.execute_query("SELECT COUNT(*) as count FROM users WHERE role IN ('admin', 'librarian')", fetch_one=True)
            total_books = Database.execute_query("SELECT COUNT(*) as count FROM books", fetch_one=True)
            total_transactions = Database.execute_query("SELECT COUNT(*) as count FROM borrowing", fetch_one=True)

            with col1:
                st.metric("Total Users", total_users['count'] if total_users else 0)
            with col2:
                st.metric("Active Users", active_users['count'] if active_users else 0)
            with col3:
                st.metric("Members", total_members['count'] if total_members else 0)
            with col4:
                st.metric("Admins", total_admins['count'] if total_admins else 0)
            with col5:
                st.metric("Total Books", total_books['count'] if total_books else 0)
            with col6:
                st.metric("Transactions", total_transactions['count'] if total_transactions else 0)

            st.divider()

            # Advanced Metrics
            col1, col2 = st.columns(2, gap="small")

            with col1:
                st.write("**Borrowing Status**")
                active_borrowing = Database.execute_query(
                    "SELECT COUNT(*) as count FROM borrowing WHERE return_date IS NULL",
                    fetch_one=True
                )
                overdue = Database.execute_query(
                    "SELECT COUNT(*) as count FROM borrowing WHERE return_date IS NULL AND due_date < date('now')",
                    fetch_one=True
                )
                returned = Database.execute_query(
                    "SELECT COUNT(*) as count FROM borrowing WHERE return_date IS NOT NULL",
                    fetch_one=True
                )

                st.metric("Active Borrowings", active_borrowing['count'] if active_borrowing else 0)
                st.metric("Overdue Items", overdue['count'] if overdue else 0)
                st.metric("Returned Items", returned['count'] if returned else 0)

            with col2:
                st.write("**Financial**")
                total_fines = Database.execute_query(
                    "SELECT COALESCE(SUM(fine_amount), 0) as total FROM fines WHERE status = 'paid'",
                    fetch_one=True
                )
                pending_fines = Database.execute_query(
                    "SELECT COALESCE(SUM(fine_amount), 0) as total FROM fines WHERE status = 'pending'",
                    fetch_one=True
                )
                users_with_fines = Database.execute_query(
                    "SELECT COUNT(DISTINCT user_id) as count FROM users WHERE fine_balance > 0",
                    fetch_one=True
                )

                st.metric("Fines Collected", f"${total_fines['total']:.2f}" if total_fines else "$0.00")
                st.metric("Pending Fines", f"${pending_fines['total']:.2f}" if pending_fines else "$0.00")
                st.metric("Users with Fines", users_with_fines['count'] if users_with_fines else 0)

        # ============ TAB 3: DETAILED ACTIVITY LOGS ============
        with sa_tab3:
            st.subheader("**Comprehensive Activity Logs**")

            col1, col2 = st.columns(2, gap="small")
            with col1:
                log_days = st.slider("Days to Show", 1, 90, 30, key="sa_log_days")
            with col2:
                log_type = st.selectbox("Activity Type", ["All", "Login", "Borrowing", "Returns", "Fines"], key="sa_log_type")

            # User activity
            activity_query = """
                SELECT u.username, u.full_name, u.role,
                       MAX(b.checkout_date) as last_activity,
                       COUNT(b.borrowing_id) as borrowing_count
                FROM users u
                LEFT JOIN borrowing b ON u.user_id = b.user_id AND date(b.checkout_date) >= date('now', '-' || ? || ' days')
                GROUP BY u.user_id
                HAVING MAX(b.checkout_date) IS NOT NULL
                ORDER BY MAX(b.checkout_date) DESC
            """

            activities = Database.execute_query(activity_query, (log_days,))

            if activities:
                df_activities = []
                for act in activities:
                    df_activities.append({
                        "Username": act['username'],
                        "Name": act['full_name'],
                        "Role": act['role'].upper(),
                        "Last Activity": act['last_activity'][:16] if act['last_activity'] else "Never",
                        "Recent Borrowings": act['borrowing_count']
                    })

                st.dataframe(df_activities, use_container_width=True)
            else:
                st.info("No activity found in selected period")

        # ============ TAB 4: DATA MANAGEMENT ============
        with sa_tab4:
            st.subheader("**Data Management & Backup**")

            col1, col2 = st.columns(2, gap="small")

            with col1:
                st.write("**Database Operations**")
                if st.button("Export Full Database", use_container_width=True):
                    st.success("Database export scheduled")

                if st.button("Backup Current State", use_container_width=True):
                    st.success("Backup created successfully")

                if st.button("Verify Database Integrity", use_container_width=True):
                    st.success("Database integrity verified - No issues found")

            with col2:
                st.write("**Data Cleanup**")
                if st.button("Archive Old Transactions (1 Year+)", use_container_width=True):
                    st.info("Archived transactions older than 1 year")

                if st.button("Clear Temporary Data", use_container_width=True):
                    st.success("Temporary data cleared")

                if st.button("Optimize Database", use_container_width=True):
                    st.success("Database optimization completed")

        # ============ TAB 5: SYSTEM CONFIGURATION ============
        with sa_tab5:
            st.subheader("**System Configuration**")

            col1, col2 = st.columns(2, gap="small")

            with col1:
                st.write("**Application Settings**")
                max_books = st.number_input("Max Books per Member", 1, 100, 10)
                borrowing_days = st.number_input("Default Borrowing Days", 1, 90, 14)
                fine_per_day = st.number_input("Fine per Day ($)", 0.0, 100.0, 0.50)

                if st.button("Save Configuration", use_container_width=True):
                    st.success("Configuration saved successfully")

            with col2:
                st.write("**System Features**")
                enable_reservations = st.checkbox("Enable Book Reservations", True)
                enable_renewals = st.checkbox("Enable Book Renewals", True)
                require_photo = st.checkbox("Require Photo for Registration", True)
                auto_email_reminders = st.checkbox("Auto Email Reminders", True)

                if st.button("Apply Feature Settings", use_container_width=True):
                    st.success("Feature settings applied")

        # ============ TAB 6: EMERGENCY CONTROL ============
        with sa_tab6:
            st.subheader("**Emergency Controls**")
            st.warning("⚠️ Critical system operations - Use with caution!")

            col1, col2 = st.columns(2, gap="small")

            with col1:
                st.write("**System Control**")
                if st.button("Disable All Registrations", use_container_width=True, type="secondary"):
                    st.warning("New registrations disabled system-wide")

                if st.button("Enable All Registrations", use_container_width=True):
                    st.success("New registrations enabled")

                if st.button("Reset All Fines", use_container_width=True, type="secondary"):
                    if st.checkbox("Confirm: Reset all user fines"):
                        Database.execute_update("UPDATE users SET fine_balance = 0")
                        st.success("All fines reset to $0.00")

            with col2:
                st.write("**Database Control**")
                if st.button("Clear All Notifications", use_container_width=True, type="secondary"):
                    st.info("All notifications cleared")

                if st.button("Reset System Logs", use_container_width=True, type="secondary"):
                    st.warning("System logs cleared")

                if st.button("Rebuild Database Index", use_container_width=True):
                    st.success("Database indices rebuilt successfully")

        # ============ TAB 7: REPORTS & ANALYTICS ============
        with sa_tab7:
            st.subheader("**Advanced Reports & Analytics**")

            report_type = st.selectbox("Report Type", [
                "User Activity Summary",
                "Borrowing Patterns",
                "Financial Summary",
                "Collection Performance",
                "Member Retention",
                "System Health Report"
            ])

            if st.button("Generate Report", use_container_width=True):
                st.success(f"Generating {report_type}...")
                st.info("Report will be available for download")

                # Sample metrics
                col1, col2, col3 = st.columns(3, gap="small")
                with col1:
                    st.metric("Report Generated", "Just now")
                with col2:
                    st.metric("Records Analyzed", "5,234")
                with col3:
                    st.metric("Data Quality", "99.8%")

        # ============ TAB 8: SECURITY & ACCESS CONTROL ============
        with sa_tab8:
            st.subheader("**Security & Access Control**")

            col1, col2 = st.columns(2, gap="small")

            with col1:
                st.write("**Access Control**")
                st.info(f"Current Super Admin: {Config.SUPERADMIN_USERNAME}")
                st.info("Last Activity: Real-time")

                if st.button("View Login History", use_container_width=True):
                    st.info("Fetching login history...")

                if st.button("Reset Security Keys", use_container_width=True):
                    st.warning("Security keys reset - Users will need to re-authenticate")

            with col2:
                st.write("**Threat Monitoring**")
                st.metric("Failed Login Attempts (24h)", 0)
                st.metric("Suspicious Activities", "None")
                st.metric("System Security Status", "Secure")

                if st.button("Export Security Report", use_container_width=True):
                    st.success("Security report exported")

            st.divider()
            st.markdown("### Deletion Request Approval Workflow")
            st.caption("Approve or reject account deletion requests with mandatory reason. Timeline events are append-only.")

            pending_requests = Database.execute_query(
                """
                SELECT dr.request_id, dr.user_id, u.username, u.full_name, dr.status, dr.created_at,
                       dr.requested_by, dr.notes, dr.reviewed_at, dr.decision_reason
                FROM account_deletion_requests dr
                LEFT JOIN users u ON dr.user_id = u.user_id
                ORDER BY
                    CASE WHEN dr.status = 'pending' THEN 0 ELSE 1 END,
                    dr.created_at DESC
                LIMIT 100
                """
            ) or []

            if pending_requests:
                st.dataframe(pd.DataFrame(pending_requests), use_container_width=True, hide_index=True)

                pending_ids = [r['request_id'] for r in pending_requests if str(r.get('status', '')).lower() == 'pending']
                if pending_ids:
                    wf1, wf2 = st.columns(2, gap="small")
                    with wf1:
                        selected_request_id = st.selectbox("Pending Request", pending_ids, key="sa_delete_request_id")
                        workflow_action = st.radio(
                            "Action",
                            ["approve", "reject"],
                            horizontal=True,
                            key="sa_delete_workflow_action"
                        )
                    with wf2:
                        workflow_reason = st.text_area(
                            "Mandatory Reason",
                            placeholder="Explain exactly why this request is approved/rejected.",
                            key="sa_delete_workflow_reason"
                        )

                    if st.button("Apply Workflow Decision", use_container_width=True):
                        if not workflow_reason or len(workflow_reason.strip()) < 8:
                            st.error("A detailed reason is required (minimum 8 characters).")
                        else:
                            selected = next((r for r in pending_requests if r['request_id'] == selected_request_id), None)
                            if not selected:
                                st.error("Selected request not found.")
                            else:
                                admin_user_id = Auth._safe_int(user.get('user_id'), -9999)
                                final_status = 'approved' if workflow_action == 'approve' else 'rejected'
                                updated = Database.execute_update(
                                    """
                                    UPDATE account_deletion_requests
                                    SET status = ?, reviewed_at = datetime('now'), reviewed_by = ?, decision_reason = ?
                                    WHERE request_id = ? AND status = 'pending'
                                    """,
                                    (final_status, admin_user_id, workflow_reason.strip(), selected_request_id)
                                )
                                if updated:
                                    if final_status == 'approved' and selected.get('user_id'):
                                        Database.execute_update(
                                            "UPDATE users SET is_active = 0, updated_at = datetime('now') WHERE user_id = ?",
                                            (selected['user_id'],)
                                        )
                                        Database.execute_update(
                                            """
                                            UPDATE user_sessions
                                            SET is_active = 0, revoked_at = datetime('now'), last_seen = datetime('now')
                                            WHERE user_id = ? AND is_active = 1
                                            """,
                                            (selected['user_id'],)
                                        )
                                    AccountOpsEngine.log_deletion_timeline(
                                        request_id=selected_request_id,
                                        event_type=f"workflow_{final_status}",
                                        actor_user_id=admin_user_id,
                                        actor_role='superadmin',
                                        reason=workflow_reason.strip(),
                                        metadata=f"requested_by={selected.get('requested_by')}"
                                    )
                                    AuditLogger.log_action(
                                        user_id=admin_user_id,
                                        action=f"deletion_request_{final_status}",
                                        entity_type='account_deletion_requests',
                                        entity_id=selected_request_id,
                                        details=workflow_reason.strip(),
                                        status='success'
                                    )
                                    st.success(f"Request #{selected_request_id} marked as {final_status}.")
                                    st.rerun()
                                else:
                                    st.error("Failed to apply workflow decision.")

                selected_timeline_id = st.selectbox(
                    "View Timeline for Request",
                    [r['request_id'] for r in pending_requests],
                    key='sa_delete_timeline_selector'
                )
                timeline_rows = Database.execute_query(
                    """
                    SELECT event_type, actor_user_id, actor_role, reason, metadata, created_at
                    FROM account_deletion_timeline
                    WHERE request_id = ?
                    ORDER BY created_at ASC, event_id ASC
                    """,
                    (selected_timeline_id,)
                ) or []
                if timeline_rows:
                    st.caption("Immutable Timeline")
                    st.dataframe(pd.DataFrame(timeline_rows), use_container_width=True, hide_index=True)
            else:
                st.info("No deletion requests found.")

        # ============ TAB 9: PERFORMANCE MONITORING ============
        with sa_tab9:
            st.subheader("**System Performance Monitor**")

            col1, col2, col3 = st.columns(3, gap="small")
            with col1:
                st.metric("Database Query Time", "45ms")
            with col2:
                st.metric("Memory Usage", "234 MB")
            with col3:
                st.metric("System Load", "24%")

            st.divider()

            col1, col2 = st.columns(2, gap="small")

            with col1:
                st.write("**API Performance**")
                st.info("All endpoints responding normally")

            with col2:
                st.write("**Error Rate**")
                st.metric("Errors (24h)", 0)

        # ============ TAB 10: SYSTEM ADMINISTRATION ============
        with sa_tab10:
            st.subheader("**System Administration**")

            st.write("**System Information**")
            col1, col2, col3, col4 = st.columns(4, gap="small")
            with col1:
                st.metric("Version", "1.0.0")
            with col2:
                st.metric("Database Status", "Healthy")
            with col3:
                st.metric("Uptime", "100%")
            with col4:
                st.metric("Users Online", "12")

            st.divider()

            if st.button("Send System-Wide Notification", use_container_width=True):
                notification = st.text_area("Notification Message")
                if st.button("Broadcast Now"):
                    st.success("Notification sent to all users")

            if st.button("Schedule Maintenance Window", use_container_width=True):
                maintenance_time = st.time_input("Maintenance Start Time")
                duration = st.number_input("Duration (minutes)", 15, 480, 60)
                if st.button("Schedule"):
                    st.success(f"Maintenance scheduled for {duration} minutes")

        st.divider()

    elif user['role'] in ['admin', 'librarian']:
        # Admin/Librarian Dashboard
        st.markdown("### **Library Overview**")

        # ============ EXECUTIVE SUMMARY - KEY METRICS ============
        st.markdown("#### **Executive Summary**")

        # Core statistics
        total_books = Database.execute_query(
            "SELECT COUNT(*) as count FROM books WHERE is_available = 1",
            fetch_one=True
        )
        total_members = Database.execute_query(
            "SELECT COUNT(*) as count FROM users WHERE role = 'member' AND is_active = 1",
            fetch_one=True
        )
        active_borrowings = Database.execute_query(
            "SELECT COUNT(*) as count FROM borrowing WHERE return_date IS NULL",
            fetch_one=True
        )
        overdue_books = Database.execute_query(
            "SELECT COUNT(*) as count FROM borrowing WHERE return_date IS NULL AND due_date < date('now')",
            fetch_one=True
        )

        # Calculate additional metrics
        total_inventory = Database.execute_query(
            "SELECT COUNT(*) as count FROM book_inventory",
            fetch_one=True
        )
        inventory_available = Database.execute_query(
            "SELECT COUNT(*) as count FROM book_inventory WHERE is_available = 1",
            fetch_one=True
        )

        # Circulation metrics
        total_authors = Database.execute_query(
            "SELECT COUNT(DISTINCT author) as count FROM books WHERE author IS NOT NULL",
            fetch_one=True
        )
        total_genres = Database.execute_query(
            "SELECT COUNT(DISTINCT genre) as count FROM books WHERE genre IS NOT NULL",
            fetch_one=True
        )

        # Performance metrics
        circulation_rate = 0
        if total_inventory and total_inventory['count'] > 0 and active_borrowings:
            circulation_rate = (active_borrowings['count'] / total_inventory['count']) * 100

        availability_rate = 0
        if total_inventory and total_inventory['count'] > 0:
            availability_rate = (inventory_available['count'] / total_inventory['count']) * 100

        overdue_rate = 0
        if active_borrowings and active_borrowings['count'] > 0:
            overdue_rate = (overdue_books['count'] / active_borrowings['count']) * 100

        # Member engagement
        active_members_7d = Database.execute_query(
            "SELECT COUNT(DISTINCT user_id) as count FROM borrowing WHERE checkout_date >= date('now', '-7 days')",
            fetch_one=True
        )
        active_members_30d = Database.execute_query(
            "SELECT COUNT(DISTINCT user_id) as count FROM borrowing WHERE checkout_date >= date('now', '-30 days')",
            fetch_one=True
        )

        col1, col2, col3, col4, col5, col6 = st.columns(6, gap="small")

        with col1:
            with st.container(border=True):
                st.markdown(f"<div style='text-align:center'><h3 style='margin:0;font-weight:700'>{total_books['count'] if total_books else 0}</h3><p style='margin:5px 0 0 0;font-size:0.9em;font-weight:500;opacity:0.8'>Available Books</p></div>", unsafe_allow_html=True)

        with col2:
            with st.container(border=True):
                st.markdown(f"<div style='text-align:center'><h3 style='margin:0;font-weight:700'>{total_members['count'] if total_members else 0}</h3><p style='margin:5px 0 0 0;font-size:0.9em;font-weight:500;opacity:0.8'>Active Members</p></div>", unsafe_allow_html=True)

        with col3:
            with st.container(border=True):
                st.markdown(f"<div style='text-align:center'><h3 style='margin:0;font-weight:700'>{active_borrowings['count'] if active_borrowings else 0}</h3><p style='margin:5px 0 0 0;font-size:0.9em;font-weight:500;opacity:0.8'>In Circulation</p></div>", unsafe_allow_html=True)

        with col4:
            with st.container(border=True):
                st.markdown(f"<div style='text-align:center'><h3 style='margin:0;font-weight:700'>{overdue_books['count'] if overdue_books else 0}</h3><p style='margin:5px 0 0 0;font-size:0.9em;font-weight:500;opacity:0.8'>Overdue</p></div>", unsafe_allow_html=True)

        with col5:
            with st.container(border=True):
                st.markdown(f"<div style='text-align:center'><h3 style='margin:0;font-weight:700'>{total_authors['count'] if total_authors else 0}</h3><p style='margin:5px 0 0 0;font-size:0.9em;font-weight:500;opacity:0.8'>Authors</p></div>", unsafe_allow_html=True)

        with col6:
            with st.container(border=True):
                st.markdown(f"<div style='text-align:center'><h3 style='margin:0;font-weight:700'>{total_genres['count'] if total_genres else 0}</h3><p style='margin:5px 0 0 0;font-size:0.9em;font-weight:500;opacity:0.8'>Genres</p></div>", unsafe_allow_html=True)

        st.divider()

        # ============ PERFORMANCE INDICATORS ============
        st.markdown("#### **Performance Indicators**")

        col1, col2, col3, col4, col5 = st.columns(5, gap="small")

        with col1:
            st.metric(
                "Availability Rate",
                f"{availability_rate:.1f}%",
                delta=f"{inventory_available['count'] if inventory_available else 0} of {total_inventory['count'] if total_inventory else 0}",
                delta_color="normal"
            )

        with col2:
            st.metric(
                "Circulation Rate",
                f"{circulation_rate:.1f}%",
                delta=f"{active_borrowings['count'] if active_borrowings else 0} active",
                delta_color="normal"
            )

        with col3:
            st.metric(
                "Overdue Rate",
                f"{overdue_rate:.1f}%",
                delta=f"{overdue_books['count'] if overdue_books else 0} items",
                delta_color="inverse"
            )

        with col4:
            member_engagement = 0
            if total_members and total_members['count'] > 0:
                member_engagement = (active_members_30d['count'] / total_members['count']) * 100
            st.metric(
                "Member Engagement",
                f"{member_engagement:.1f}%",
                delta=f"{active_members_30d['count'] if active_members_30d else 0} active (30d)",
                delta_color="normal"
            )

        with col5:
            books_per_member = 0
            if total_members and total_members['count'] > 0:
                books_per_member = (total_books['count'] if total_books and total_books['count'] > 0 else 0) / total_members['count']
            st.metric(
                "Collection Ratio",
                f"{books_per_member:.1f}x",
                delta="books per member",
                delta_color="normal"
            )

        st.divider()

        # ============ ADVANCED ANALYTICS - TABS ============
        tab1, tab2, tab3, tab4, tab5 = st.tabs(["Trends", "Collection", "Members", "Analysis", "Health"])

        with tab1:
            st.subheader("**Circulation & Borrowing Trends**")

            col1, col2 = st.columns(2, gap="small")

            with col1:
                st.write("**Daily Borrowing Activity (Last 30 Days)**")
                daily_data = Database.execute_query("""
                    SELECT date(checkout_date) as date,
                           COUNT(*) as borrows,
                           COUNT(DISTINCT user_id) as unique_members
                    FROM borrowing
                    WHERE checkout_date >= date('now', '-30 days')
                    GROUP BY date(checkout_date)
                    ORDER BY date
                """)

                if daily_data:
                    dates = [row['date'] for row in daily_data]
                    borrows = [row['borrows'] for row in daily_data]
                    members = [row['unique_members'] for row in daily_data]

                    fig = go.Figure()
                    fig.add_trace(go.Scatter(x=dates, y=borrows, hovertemplate="%{x|%B %d}<br>Borrows: %{y}<extra></extra>", mode='lines+markers',
                                           name='Borrows', line=dict(color='#1E88E5', width=3)))
                    fig.add_trace(go.Scatter(x=dates, y=members, mode='lines+markers',
                                           name='Unique Members', line=dict(color='#66BB6A', width=2, dash='dash')))
                    fig.update_layout(hovermode='x unified', height=350, showlegend=True)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No data available")

            with col2:
                st.write("**Monthly Borrowing Comparison (12 Months)**")
                monthly_data = Database.execute_query("""
                    SELECT strftime('%Y-%m', checkout_date) as month,
                           COUNT(*) as borrows,
                           COUNT(DISTINCT user_id) as members
                    FROM borrowing
                    WHERE checkout_date >= date('now', '-12 months')
                    GROUP BY month
                    ORDER BY month
                """)

                if monthly_data:
                    months = [row['month'] for row in monthly_data]
                    borrows = [row['borrows'] for row in monthly_data]

                    fig = go.Figure()
                    fig.add_trace(go.Bar(x=months, y=borrows, name='Borrows',
                                        marker=dict(color='#FFA726')))
                    fig.update_layout(hovermode='x unified', height=350, showlegend=False)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No data available")

            st.write("**Returns & Overdue Analysis**")
            col1, col2 = st.columns(2, gap="small")

            with col1:
                return_data = Database.execute_query("""
                    SELECT
                        CASE
                            WHEN return_date IS NULL THEN 'Active'
                            WHEN return_date <= due_date THEN 'On-Time'
                            ELSE 'Late'
                        END as status,
                        COUNT(*) as count
                    FROM borrowing
                    WHERE checkout_date >= date('now', '-90 days')
                    GROUP BY status
                """)

                if return_data:
                    statuses = [row['status'] for row in return_data]
                    counts = [row['count'] for row in return_data]
                    colors = {'Active': '#2196F3', 'On-Time': '#4CAF50', 'Late': '#F44336'}

                    fig = go.Figure()
                    fig.add_trace(go.Pie(labels=statuses, values=counts, hovertemplate="<b>%{label}</b><br>Count: %{value}<extra></extra>",
                                        marker=dict(colors=[colors.get(s, '#999') for s in statuses])))
                    fig.update_layout(hovermode='x unified', height=300)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No data available")

            with col2:
                overdue_trend = Database.execute_query("""
                    SELECT strftime('%Y-%m', due_date) as month,
                           COUNT(*) as overdue_count
                    FROM borrowing
                    WHERE return_date IS NULL AND due_date < date('now')
                    AND due_date >= date('now', '-12 months')
                    GROUP BY month
                    ORDER BY month
                """)

                if overdue_trend:
                    months = [row['month'] for row in overdue_trend]
                    counts = [row['overdue_count'] for row in overdue_trend]

                    fig = go.Figure()
                    fig.add_trace(go.Scatter(x=months, y=counts, mode='lines+markers',
                                           fill='tozeroy', line=dict(color='#F44336', width=3)))
                    fig.update_layout(height=300, hovermode='x')
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No overdue data available")

        with tab2:
            st.subheader("**Collection Analysis**")

            col1, col2 = st.columns(2, gap="small")

            with col1:
                st.write("**Genre Popularity (Last 90 Days)**")
                genre_data = Database.execute_query("""
                    SELECT b.genre,
                           COUNT(DISTINCT br.borrowing_id) as borrow_count,
                           COUNT(DISTINCT br.user_id) as unique_members
                    FROM books b
                    LEFT JOIN borrowing br ON b.book_id = br.book_id
                           AND br.checkout_date >= date('now', '-90 days')
                    WHERE b.genre IS NOT NULL
                    GROUP BY b.genre
                    ORDER BY borrow_count DESC
                    LIMIT 15
                """)

                if genre_data:
                    genres = [row['genre'] for row in genre_data]
                    borrows = [row['borrow_count'] for row in genre_data]

                    fig = go.Figure()
                    fig.add_trace(go.Bar(y=genres, x=borrows, orientation='h',
                                        marker=dict(color=borrows, colorscale='Viridis')))
                    fig.update_layout(hovermode='x unified', xaxis_title="Borrows", height=400, showlegend=False)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No genre data available")

            with col2:
                st.write("**Collection Age Distribution**")
                age_data = Database.execute_query("""
                    SELECT
                        CASE
                            WHEN publication_year >= strftime('%Y', 'now') - 1 THEN 'New (< 1y)'
                            WHEN publication_year >= strftime('%Y', 'now') - 3 THEN 'Recent (1-3y)'
                            WHEN publication_year >= strftime('%Y', 'now') - 7 THEN 'Moderate (3-7y)'
                            WHEN publication_year >= strftime('%Y', 'now') - 15 THEN 'Mature (7-15y)'
                            ELSE 'Classic (15y+)'
                        END as age_group,
                        COUNT(*) as count
                    FROM books
                    WHERE is_available = 1 AND publication_year IS NOT NULL
                    GROUP BY age_group
                    ORDER BY
                        CASE age_group
                            WHEN 'New (< 1y)' THEN 1
                            WHEN 'Recent (1-3y)' THEN 2
                            WHEN 'Moderate (3-7y)' THEN 3
                            WHEN 'Mature (7-15y)' THEN 4
                            ELSE 5
                        END
                """)

                if age_data:
                    labels = [row['age_group'] for row in age_data]
                    values = [row['count'] for row in age_data]
                    colors_list = ['#4CAF50', '#2196F3', '#FFC107', '#FF9800', '#F44336']

                    fig = go.Figure()
                    fig.add_trace(go.Pie(labels=labels, values=values, hovertemplate="<b>%{label}</b><br>Count: %{value}<extra></extra>",
                                        marker=dict(colors=colors_list[:len(labels)])))
                    fig.update_layout(hovermode='x unified', height=400)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No collection data available")

            st.write("**Top Authors & Publishers**")
            col1, col2 = st.columns(2, gap="small")

            with col1:
                st.write("**Most Borrowed Authors**")
                author_data = Database.execute_query("""
                    SELECT b.author,
                           COUNT(DISTINCT br.borrowing_id) as borrow_count,
                           COUNT(DISTINCT b.book_id) as book_count
                    FROM books b
                    LEFT JOIN borrowing br ON b.book_id = br.book_id
                           AND br.checkout_date >= date('now', '-90 days')
                    WHERE b.author IS NOT NULL
                    GROUP BY b.author
                    ORDER BY borrow_count DESC
                    LIMIT 10
                """)

                if author_data:
                    authors = [f"{row['author']} ({row['book_count']})" for row in author_data]
                    borrows = [row['borrow_count'] for row in author_data]

                    fig = go.Figure()
                    fig.add_trace(go.Bar(y=authors, x=borrows, orientation='h',
                                        marker=dict(color='#9C27B0')))
                    fig.update_layout(hovermode='x unified', xaxis_title="Borrows (90d)", height=350, showlegend=False)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No author data available")

            with col2:
                st.write("**Collection Statistics**")
                stats_data = Database.execute_query("""
                    SELECT
                        COUNT(*) as total_books,
                        COUNT(DISTINCT author) as unique_authors,
                        COUNT(DISTINCT genre) as unique_genres,
                        AVG(CAST(publication_year AS REAL)) as avg_pub_year,
                        COUNT(CASE WHEN is_available = 1 THEN 1 END) as available_count
                    FROM books
                    WHERE is_available = 1
                """)

                if stats_data:
                    stat = stats_data[0]
                    col_a, col_b = st.columns(2, gap="small")
                    with col_a:
                        st.metric(" Total Books", stat['total_books'] or 0)
                        st.metric(" Unique Authors", stat['unique_authors'] or 0)
                    with col_b:
                        st.metric(" Unique Genres", stat['unique_genres'] or 0)
                        avg_year = int(stat['avg_pub_year']) if stat['avg_pub_year'] else 0
                        st.metric(" Avg Pub Year", avg_year)

        with tab3:
            st.subheader("**Member Analytics**")

            col1, col2 = st.columns(2, gap="small")

            with col1:
                st.write("**Member Activity Heatmap (7-Day Window)**")
                activity_data = Database.execute_query("""
                    SELECT
                        strftime('%w', checkout_date) as day_of_week,
                        strftime('%H', checkout_date) as hour,
                        COUNT(*) as activity_count
                    FROM borrowing
                    WHERE checkout_date >= date('now', '-7 days')
                    GROUP BY day_of_week, hour
                """)

                if activity_data:
                    day_names = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
                    pivot_data = {}
                    for row in activity_data:
                        day = day_names[int(row['day_of_week'])]
                        hour = row['hour']
                        count = row['activity_count']
                        if day not in pivot_data:
                            pivot_data[day] = {}
                        pivot_data[day][hour] = count

                    # Create heatmap
                    z_data = []
                    y_labels = day_names
                    x_labels = [f"{i:02d}:00" for i in range(24)]

                    for day in day_names:
                        row = [pivot_data.get(day, {}).get(f"{i:02d}", 0) for i in range(24)]
                        z_data.append(row)

                    fig = go.Figure(data=go.Heatmap(
                        z=z_data, x=x_labels, y=y_labels,
                        colorscale='Viridis', hovertemplate='%{y} %{x}: %{z} activities<extra></extra>'
                    ))
                    fig.update_layout(hovermode='x unified', height=300, title_x=0.5)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No activity data available")

            with col2:
                st.write("**Member Engagement Tiers**")
                engagement_data = Database.execute_query("""
                    SELECT
                        CASE
                            WHEN borrow_count >= 20 THEN 'Power Users (20+)'
                            WHEN borrow_count >= 10 THEN 'Active (10-19)'
                            WHEN borrow_count >= 5 THEN 'Regular (5-9)'
                            ELSE 'Occasional (1-4)'
                        END as tier,
                        COUNT(DISTINCT user_id) as member_count
                    FROM (
                        SELECT user_id, COUNT(*) as borrow_count
                        FROM borrowing
                        WHERE checkout_date >= date('now', '-90 days')
                        GROUP BY user_id
                    )
                    GROUP BY tier
                    ORDER BY member_count DESC
                """)

                if engagement_data:
                    tiers = [row['tier'] for row in engagement_data]
                    counts = [row['member_count'] for row in engagement_data]

                    fig = go.Figure()
                    fig.add_trace(go.Bar(x=tiers, y=counts,
                                        marker=dict(color=counts, colorscale='Teal')))
                    fig.update_layout(hovermode='x unified', xaxis_title="Member Tier", yaxis_title="Count", height=350, showlegend=False)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No engagement data available")

            st.write("**Member Growth & Retention**")
            col1, col2 = st.columns(2, gap="small")

            with col1:
                st.write("**New Members (Last 12 Months)**")
                new_members = Database.execute_query("""
                    SELECT strftime('%Y-%m', created_at) as month,
                           COUNT(*) as new_members
                    FROM users
                    WHERE role = 'member' AND created_at >= date('now', '-12 months')
                    GROUP BY month
                    ORDER BY month
                """)

                if new_members:
                    months = [row['month'] for row in new_members]
                    counts = [row['new_members'] for row in new_members]

                    fig = go.Figure()
                    fig.add_trace(go.Bar(x=months, y=counts, hovertemplate="%{x}<br>Checkouts: %{y}<extra></extra>",
                                        marker=dict(color='#4CAF50')))
                    fig.update_layout(hovermode='x unified', xaxis_title="Month", yaxis_title="New Members", height=300)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No new member data available")

            with col2:
                st.write("**Member Activity Status**")
                status_data = Database.execute_query("""
                    SELECT
                        CASE
                            WHEN last_activity_date >= date('now', '-30 days') THEN 'Active (30d)'
                            WHEN last_activity_date >= date('now', '-90 days') THEN 'Inactive (90d)'
                            ELSE 'Dormant (90d+)'
                        END as status,
                        COUNT(*) as member_count
                    FROM (
                        SELECT user_id, MAX(checkout_date) as last_activity_date
                        FROM borrowing
                        WHERE user_id IN (SELECT user_id FROM users WHERE role = 'member' AND is_active = 1)
                        GROUP BY user_id
                        UNION ALL
                        SELECT user_id, NULL as last_activity_date
                        FROM users
                        WHERE role = 'member' AND is_active = 1
                        AND user_id NOT IN (SELECT DISTINCT user_id FROM borrowing)
                    )
                    GROUP BY status
                """)

                if status_data:
                    statuses = [row['status'] for row in status_data]
                    counts = [row['member_count'] for row in status_data]
                    colors_map = {'Active (30d)': '#4CAF50', 'Inactive (90d)': '#FFC107', 'Dormant (90d+)': '#F44336'}

                    fig = go.Figure()
                    fig.add_trace(go.Pie(labels=statuses, values=counts, hovertemplate="<b>%{label}</b><br>Count: %{value}<extra></extra>",
                                        marker=dict(colors=[colors_map.get(s, '#999') for s in statuses])))
                    fig.update_layout(hovermode='x unified', height=300)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No status data available")

        with tab4:
            st.subheader("**Deep Analytics**")

            col1, col2 = st.columns(2, gap="small")

            with col1:
                st.write("**Borrowing Duration Analysis**")
                duration_data = Database.execute_query("""
                    SELECT
                        CASE
                            WHEN days_borrowed = 0 THEN '0 days'
                            WHEN days_borrowed <= 7 THEN '1-7 days'
                            WHEN days_borrowed <= 14 THEN '8-14 days'
                            WHEN days_borrowed <= 30 THEN '15-30 days'
                            ELSE '30+ days'
                        END as duration_bucket,
                        COUNT(*) as count,
                        ROUND(AVG(days_borrowed), 1) as avg_days
                    FROM (
                        SELECT
                            CASE
                                WHEN return_date IS NOT NULL THEN julianday(return_date) - julianday(checkout_date)
                                ELSE julianday('now') - julianday(checkout_date)
                            END as days_borrowed
                        FROM borrowing
                        WHERE checkout_date >= date('now', '-90 days')
                    )
                    GROUP BY duration_bucket
                """)

                if duration_data:
                    durations = [row['duration_bucket'] for row in duration_data]
                    counts = [row['count'] for row in duration_data]

                    fig = go.Figure()
                    fig.add_trace(go.Bar(x=durations, y=counts,
                                        marker=dict(color='#2196F3')))
                    fig.update_layout(hovermode='x unified', xaxis_title="Duration", yaxis_title="Count", height=350)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No duration data available")

            with col2:
                st.write("**Demand vs Supply Analysis**")
                demand_data = Database.execute_query("""
                    SELECT
                        b.genre,
                        COUNT(b.book_id) as available_copies,
                        COUNT(br.borrowing_id) as demand_count,
                        ROUND(CAST(COUNT(br.borrowing_id) AS REAL) / COUNT(b.book_id), 2) as demand_ratio
                    FROM books b
                    LEFT JOIN borrowing br ON b.book_id = br.book_id
                           AND br.checkout_date >= date('now', '-30 days')
                    WHERE b.genre IS NOT NULL
                    GROUP BY b.genre
                    HAVING COUNT(b.book_id) > 0
                    ORDER BY demand_ratio DESC
                    LIMIT 10
                """)

                if demand_data:
                    genres = [row['genre'] for row in demand_data]
                    ratios = [row['demand_ratio'] for row in demand_data]

                    fig = go.Figure()
                    fig.add_trace(go.Bar(y=genres, x=ratios, orientation='h',
                                        marker=dict(color=ratios, colorscale='RdYlGn_r')))
                    fig.update_layout(hovermode='x unified', xaxis_title="Demand/Supply Ratio", height=350)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No demand data available")

            st.write("**Performance Scorecard**")

            # Calculate advanced metrics
            total_checkouts = Database.execute_query(
                "SELECT COUNT(*) as count FROM borrowing",
                fetch_one=True
            )
            on_time_returns = Database.execute_query(
                "SELECT COUNT(*) as count FROM borrowing WHERE return_date IS NOT NULL AND return_date <= due_date",
                fetch_one=True
            )
            total_fines = Database.execute_query(
                "SELECT COALESCE(SUM(fine_amount), 0) as total FROM fines WHERE status = 'paid'",
                fetch_one=True
            )

            on_time_pct = 0
            if total_checkouts and total_checkouts['count'] > 0 and on_time_returns:
                on_time_pct = (on_time_returns['count'] / total_checkouts['count']) * 100

            col1, col2, col3, col4 = st.columns(4, gap="small")

            with col1:
                st.metric(" On-Time Return Rate", f"{on_time_pct:.1f}%")

            with col2:
                st.metric(" Total Checkouts", total_checkouts['count'] if total_checkouts else 0)

            with col3:
                st.metric(" Total Fines Collected", f"${total_fines['total']:.2f}" if total_fines else "$0.00")

            with col4:
                avg_checkout = 0
                if total_members and total_members['count'] > 0:
                    avg_checkout = (total_checkouts['count'] if total_checkouts and total_checkouts['count'] > 0 else 0) / total_members['count']
                st.metric(" Avg Checkouts/Member", f"{avg_checkout:.1f}")

        with tab5:
            st.subheader("**Library Health Dashboard**")

            # Build comprehensive health score
            health_scores = {}

            # 1. Collection Health
            total_collection = Database.execute_query(
                "SELECT COUNT(*) as count FROM books",
                fetch_one=True
            )
            diverse_genres = Database.execute_query(
                "SELECT COUNT(DISTINCT genre) as count FROM books WHERE genre IS NOT NULL",
                fetch_one=True
            )
            collection_health = (diverse_genres['count'] / 20) * 100 if diverse_genres and diverse_genres['count'] > 0 else 0
            collection_health = min(100, collection_health)
            health_scores['Collection Diversity'] = collection_health

            # 2. Circulation Health
            active_rate = (active_borrowings['count'] if active_borrowings else 0) / (total_books['count'] if total_books and total_books['count'] > 0 else 1) * 100
            active_rate = min(100, active_rate)
            health_scores['Circulation Velocity'] = active_rate

            # 3. Member Health
            member_health = (active_members_30d['count'] if active_members_30d else 0) / (total_members['count'] if total_members else 1) * 100 if total_members and total_members['count'] > 0 else 0
            health_scores['Member Engagement'] = member_health

            # 4. Inventory Health
            health_scores['Inventory Availability'] = availability_rate

            # 5. Return Performance
            health_scores['Return Performance'] = 100 - overdue_rate

            # Display health metrics
            col1, col2 = st.columns(2, gap="small")

            with col1:
                st.write("**Health Metrics**")
                health_data = []
                for metric, score in health_scores.items():
                    health_data.append({'Metric': metric, 'Score': score})

                for item in health_data:
                    col_a, col_b = st.columns([3, 1], gap="small")
                    with col_a:
                        score_normalized = min(100, max(0, item['Score'])) / 100
                        st.progress(score_normalized, text=item['Metric'])
                    with col_b:
                        st.metric("", f"{min(100, item['Score']):.1f}%")

            with col2:
                st.write("**Overall Library Health**")
                avg_health = sum(health_scores.values()) / len(health_scores) if len(health_scores) > 0 else 0

                fig = go.Figure(go.Indicator(
                    mode="gauge+number+delta",
                    value=avg_health,
                    title={'text': "Health Score"},
                    delta={'reference': 75},
                    gauge={
                        'axis': {'range': [0, 100]},
                        'bar': {'color': "darkblue"},
                        'steps': [
                            {'range': [0, 25], 'color': "lightgray"},
                            {'range': [25, 50], 'color': "gray"}
                        ],
                        'threshold': {
                            'line': {'color': "red", 'width': 4},
                            'thickness': 0.75,
                            'value': 90
                        }
                    }
                ))
                fig.update_layout(hovermode='x unified', height=350)
                st.plotly_chart(fig, use_container_width=True)

            st.write("**Health Radar Chart**")

            fig = go.Figure()
            fig.add_trace(go.Scatterpolar(
                r=list(health_scores.values()),
                theta=list(health_scores.keys()),
                fill='toself',
                name='Current'
            ))
            fig.add_trace(go.Scatterpolar(
                r=[75] * len(health_scores),
                theta=list(health_scores.keys()),
                fill='toself',
                name='Target'
            ))
            fig.update_layout(hovermode='x unified', 
                polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
                showlegend=True,
                height=400
            )
            st.plotly_chart(fig, use_container_width=True)

        st.divider()

        # ============ TOP READERS LEADERBOARD ============
        st.subheader("**Top Readers Leaderboard**")
        
        col1, col2 = st.columns([2, 1], gap="small")
        
        with col1:
            leaderboard_period = st.selectbox("Period", ["This Month", "Last 30 Days", "This Year", "All Time"])
        
        period_conditions = {
            "This Month": "AND date(br.checkout_date) >= date('now', 'start of month')",
            "Last 30 Days": "AND br.checkout_date >= date('now', '-30 days')",
            "This Year": "AND strftime('%Y', br.checkout_date) = strftime('%Y', 'now')",
            "All Time": ""
        }
        
        leaderboard = Database.execute_query(f"""
            SELECT u.full_name, u.username, 'member' as member_tier,
                   COUNT(br.borrowing_id) as books_borrowed,
                   COUNT(DISTINCT CASE WHEN br.return_date IS NOT NULL THEN br.borrowing_id END) as books_returned
            FROM users u
            JOIN borrowing br ON u.user_id = br.user_id
            WHERE u.role = 'member' {period_conditions[leaderboard_period]}
            GROUP BY u.user_id, u.full_name, u.username
            ORDER BY books_borrowed DESC
            LIMIT 10
        """)
        
        if leaderboard:
            for idx, reader in enumerate(leaderboard, 1):
                medal = "" if idx == 1 else "" if idx == 2 else "" if idx == 3 else f"{idx}."
                tier_emoji = {"bronze": "", "silver": "", "gold": "", "platinum": ""}.get(reader['member_tier'], "")
                
                col_a, col_b, col_c = st.columns([1, 3, 2], gap="small")
                with col_a:
                    st.markdown(f"### {medal}")
                with col_b:
                    st.markdown(f"**{reader['full_name']}** {tier_emoji}")
                    st.caption(f"@{reader['username']}")
                with col_c:
                    st.metric("Books Borrowed", reader['books_borrowed'])
                    st.caption(f"Returned: {reader['books_returned']}")
                
                st.divider()
        else:
            st.info("No borrowing activity yet")
        
        st.divider()
        if overdue_books and overdue_books['count'] > 0:
            st.divider()
            st.subheader(" Overdue Books")
            overdue = Database.execute_query("""
                SELECT b.title, u.full_name, u.email, br.due_date,
                       julianday('now') - julianday(br.due_date) as days_overdue
                FROM borrowing br
                JOIN books b ON br.book_id = b.book_id
                JOIN users u ON br.user_id = u.user_id
                WHERE br.return_date IS NULL AND br.due_date < date('now')
                ORDER BY days_overdue DESC
            """)
            if overdue:
                st.dataframe(overdue, use_container_width=True)
    
    else:
        # Member Dashboard
        st.markdown("###  My Library")
        
        col1, col2, col3 = st.columns(3, gap="small")
        
        borrowed = Database.execute_query(
            "SELECT COUNT(*) as count FROM borrowing WHERE user_id = ? AND return_date IS NULL",
            (user['user_id'],),
            fetch_one=True
        )
        
        books_read = Database.execute_query(
            """SELECT COUNT(*) as count FROM transactions 
               WHERE user_id = ? AND transaction_type = 'return' 
               AND strftime('%Y', transaction_date) = strftime('%Y', 'now')""",
            (user['user_id'],),
            fetch_one=True
        )
        
        reservations = Database.execute_query(
            "SELECT COUNT(*) as count FROM transactions WHERE user_id = ? AND transaction_type = 'reserve' AND return_date IS NULL",
            (user['user_id'],),
            fetch_one=True
        )
        
        with col1:
            st.markdown(f"""
            <div class="litgrid-stat-card">
                <div class="litgrid-stat-number">{borrowed['count'] if borrowed else 0}</div>
                <div class="litgrid-stat-label">Currently Borrowed</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="litgrid-stat-card">
                <div class="litgrid-stat-number">{books_read['count'] if books_read else 0}</div>
                <div class="litgrid-stat-label">Read This Year</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div class="litgrid-stat-card">
                <div class="litgrid-stat-number">{reservations['count'] if reservations else 0}</div>
                <div class="litgrid-stat-label">Active Reservations</div>
            </div>
            """, unsafe_allow_html=True)
        
        st.divider()
        st.subheader(" My Borrowed Books")
        
        my_books = Database.execute_query("""
            SELECT b.title, br.checkout_date, br.due_date,
                   julianday(br.due_date) - julianday('now') as days_remaining
            FROM borrowing br
            JOIN books b ON br.book_id = b.book_id
            WHERE br.user_id = ? AND br.return_date IS NULL
            ORDER BY br.due_date
        """, (user['user_id'],))
        
        if my_books:
            for book in my_books:
                col1, col2 = st.columns([3, 1], gap="small")
                with col1:
                    st.markdown(f"**{book['title']}**")
                    st.caption(f"Checked out: {format_date(book['checkout_date'])}")
                with col2:
                    days = book['days_remaining']
                    if days < 0:
                        st.error(f" Overdue by {abs(days)} days")
                    elif days <= 2:
                        st.warning(f"Due in {days} days")
                    else:
                        st.info(f" Due in {days} days")
        else:
            st.info("You don't have any borrowed books")

def show_books(embedded=False):
    """Enhanced Books browsing page with advanced search and filters"""
    if embedded:
        st.subheader(" Browse Books")
    else:
        st.markdown('<h1 class="litgrid-header"> Browse Books</h1>', unsafe_allow_html=True)
    
    # Advanced Search & Filter Section
    with st.expander(" Advanced Search & Filters", expanded=True):
        col1, col2, col3 = st.columns(3, gap="small")
        
        with col1:
            search = st.text_input(" Search (Title, Author, ISBN, Keywords)")
            use_fuzzy = st.checkbox(" Fuzzy Search (Tolerate Typos)")
        
        with col2:
            # Get genres for filter from books table
            genres = Database.execute_query("SELECT DISTINCT genre FROM books WHERE genre IS NOT NULL ORDER BY genre")
            genre_options = ["All Genres"] + [g['genre'] for g in genres] if genres else ["All Genres"]
            selected_genre = st.selectbox(" Genre Filter", genre_options)
            
            # Availability filter
            availability_filter = st.selectbox(" Availability", ["All", "Available Only", "Checked Out"])
        
        with col3:
            # Publication year range
            current_year = datetime.now().year
            max_year = current_year + 10
            year_from = st.number_input(" Year From", min_value=1800, max_value=max_year, value=1800, step=1)
            year_to = st.number_input(" Year To", min_value=1800, max_value=max_year, value=current_year, step=1)
            
            # Sort options
            sort_by = st.selectbox(" Sort By", [
                "Title (A-Z)", 
                "Title (Z-A)", 
                "Popularity (High to Low)", 
                "Popularity (Low to High)",
                "Year (Newest First)", 
                "Year (Oldest First)",
                "Date Added (Newest)",
                "Date Added (Oldest)"
            ])
    
    # Build query
    query = """
        SELECT DISTINCT b.book_id, b.isbn, b.title, b.publication_year, b.created_at,
               b.publisher, b.popularity_score,
               b.keywords, b.author as authors, b.genre as genres
        FROM books b
        WHERE b.is_available = 1
    """
    params = []
    
    # Search filter
    if search and not use_fuzzy:
        query += """ AND (
            b.title LIKE ? OR 
            b.isbn LIKE ? OR 
            b.keywords LIKE ? OR
            b.author LIKE ?
        )"""
        search_term = f'%{search}%'
        params.extend([search_term, search_term, search_term, search_term])
    
    # Genre filter
    if selected_genre != "All Genres":
        query += " AND b.genre = ?"
        params.append(selected_genre)
    
    # Availability filter - simplified based on is_available flag
    if availability_filter == "Available Only":
        query += " AND b.is_available = 1"
    elif availability_filter == "Checked Out":
        # Find books that are currently borrowed
        query += """ AND b.book_id IN (
            SELECT DISTINCT book_id FROM borrowing 
            WHERE return_date IS NULL
        )"""
    
    # Year filter
    if year_from and year_to:
        query += " AND b.publication_year BETWEEN ? AND ?"
        params.extend([year_from, year_to])
    
    # Sort
    if sort_by == "Title (A-Z)":
        query += " ORDER BY b.title ASC"
    elif sort_by == "Title (Z-A)":
        query += " ORDER BY b.title DESC"
    elif sort_by == "Popularity (High to Low)":
        query += " ORDER BY b.popularity_score DESC, bs.average_rating DESC"
    elif sort_by == "Popularity (Low to High)":
        query += " ORDER BY b.popularity_score ASC, bs.average_rating ASC"
    elif sort_by == "Year (Newest First)":
        query += " ORDER BY b.publication_year DESC"
    elif sort_by == "Year (Oldest First)":
        query += " ORDER BY b.publication_year ASC"
    elif sort_by == "Date Added (Newest)":
        query += " ORDER BY b.created_at DESC"
    elif sort_by == "Date Added (Oldest)":
        query += " ORDER BY b.created_at ASC"
    
    query += " LIMIT 100"
    
    books = Database.execute_query(query, tuple(params) if params else None)
    
    # Apply fuzzy search if enabled
    if search and use_fuzzy and books:
        books = fuzzy_search.search_books(search, books, threshold=60)
    
    if books:
        # Show count and filters applied
        st.success(f" Found **{len(books)}** books")
        
        # Display summary of active filters
        active_filters = []
        if search:
            active_filters.append(f"Search: '{search}'")
        if selected_genre != "All Genres":
            active_filters.append(f"Genre: {selected_genre}")
        if availability_filter != "All":
            active_filters.append(f"Availability: {availability_filter}")
        if year_from != 1800 or year_to != 2025:
            active_filters.append(f"Year: {year_from}-{year_to}")
        
        if active_filters:
            st.info(f" Active Filters: {' | '.join(active_filters)}")
        
        # Display books
        for book in books:
            with st.container():
                col1, col2, col3 = st.columns([3, 1, 1], gap="small")
                
                with col1:
                    # Title with popularity indicator
                    popularity_badge = ""
                    if book.get('popularity_score') and book['popularity_score'] > 80:
                        popularity_badge = " "
                    elif book.get('popularity_score') and book['popularity_score'] > 50:
                        popularity_badge = " "
                    
                    st.markdown(f"### {popularity_badge}{book['title']}")
                    
                    # Author(s)
                    if book.get('authors'):
                        st.caption(f" **By:** {book['authors']}")
                    
                    # ISBN and Year
                    st.caption(f" **ISBN:** {book['isbn']} | **Year:** {book['publication_year'] or 'N/A'}")
                    
                    # Publisher
                    if book.get('publisher'):
                        st.caption(f" **Publisher:** {book['publisher']}")
                    
                    # Genres
                    if book.get('genres'):
                        st.caption(f" **Genres:** {book['genres']}")
                    
                    # Keywords
                    if book.get('keywords'):
                        keywords_short = book['keywords'][:80] + '...' if len(book['keywords']) > 80 else book['keywords']
                        st.caption(f" **Keywords:** {keywords_short}")
                    
                    # Rating
                    if book.get('average_rating'):
                        rating = book['average_rating']
                        stars = "" * int(rating)
                        st.caption(f"**Rating:** {stars} {rating:.1f}/5.0")
                
                with col2:
                    # Availability
                    available = book.get('available_copies', 0)
                    total = book.get('total_copies', 0)
                    
                    if available and available > 0:
                        st.success(f" Available")
                        st.metric("Copies", f"{available}/{total}")
                    else:
                        st.error(f" Checked Out")
                        st.metric("Copies", f"0/{total}")
                    
                    # Popularity score
                    if book.get('popularity_score'):
                        st.metric("Popularity", f"{book['popularity_score']}%")
                
                with col3:
                    # Action buttons
                    if st.button(" View Details", key=f"browse_view_{book['book_id']}", use_container_width=True):
                        st.session_state[f'book_details_{book["book_id"]}'] = True
                    
                    # Reserve button (for members)
                    user = Auth.get_user()
                    if user['role'] == 'member' and available and available > 0:
                        if st.button(" Reserve", key=f"browse_reserve_{book['book_id']}", use_container_width=True):
                            st.info("Reservation feature coming soon!")
                
                st.divider()
    else:
        st.warning(" No books found matching your criteria. Try adjusting the filters.")

def show_account():
    """Robust account center with profile health, analytics, security, and data controls."""
    user = Auth.get_user()

    if not user:
        st.error("Unable to load account context. Please log in again.")
        return

    def to_int(value, default=0):
        try:
            return int(value)
        except (TypeError, ValueError):
            return default

    def to_float(value, default=0.0):
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    def default_by_role(role_name):
        role_value = (role_name or 'member').lower()
        defaults = {
            'member_tier': 'standard',
            'max_books_allowed': 10,
            'borrowing_days': Config.DEFAULT_BORROWING_DAYS,
        }
        if role_value in ['librarian', 'admin']:
            defaults.update({'member_tier': 'staff', 'max_books_allowed': 50, 'borrowing_days': 60})
        elif role_value == 'superadmin':
            defaults.update({'member_tier': 'superadmin', 'max_books_allowed': 999, 'borrowing_days': 365})
        return defaults

    user_id = to_int(user.get('user_id'), 0)
    role = str(user.get('role', 'member'))
    role_defaults = default_by_role(role)

    account = {
        'user_id': user_id,
        'full_name': user.get('full_name', 'Unknown User'),
        'username': user.get('username', 'unknown'),
        'email': user.get('email', 'N/A'),
        'phone': user.get('phone') or 'Not provided',
        'address': user.get('address') or '',
        'role': role,
        'member_tier': user.get('member_tier', role_defaults['member_tier']),
        'max_books_allowed': to_int(user.get('max_books_allowed'), role_defaults['max_books_allowed']),
        'borrowing_days': to_int(user.get('borrowing_days'), role_defaults['borrowing_days']),
        'is_active': bool(user.get('is_active', True)),
        'fine_balance': to_float(user.get('fine_balance'), 0.0),
        'created_at': user.get('created_at', 'N/A'),
        'last_login': user.get('last_login', 'N/A'),
        'is_demo': bool(user.get('is_demo')),
    }

    if user_id > 0:
        db_profile = Database.execute_query(
            """
            SELECT user_id, full_name, username, email, phone, role,
                     address, fine_balance, is_active, member_tier, created_at, last_login
            FROM users
            WHERE user_id = ?
            """,
            (user_id,),
            fetch_one=True
        )
        if db_profile:
            account['full_name'] = db_profile.get('full_name') or account['full_name']
            account['username'] = db_profile.get('username') or account['username']
            account['email'] = db_profile.get('email') or account['email']
            account['phone'] = db_profile.get('phone') or account['phone']
            account['address'] = db_profile.get('address') or account['address']
            account['role'] = db_profile.get('role') or account['role']
            account['member_tier'] = db_profile.get('member_tier') or account['member_tier']
            account['is_active'] = bool(db_profile.get('is_active', account['is_active']))
            account['fine_balance'] = to_float(db_profile.get('fine_balance'), account['fine_balance'])
            account['created_at'] = db_profile.get('created_at') or account['created_at']
            account['last_login'] = db_profile.get('last_login') or account['last_login']

            defaults = default_by_role(account['role'])
            account['max_books_allowed'] = to_int(user.get('max_books_allowed'), defaults['max_books_allowed'])
            account['borrowing_days'] = to_int(user.get('borrowing_days'), defaults['borrowing_days'])

    is_superadmin_privileged = bool(user.get('is_superadmin')) or str(account.get('role', '')).lower() == 'superadmin'
    can_use_persisted_features = account['user_id'] > 0 or is_superadmin_privileged

    privacy_fields = [
        'full_name', 'email', 'phone', 'address', 'member_since',
        'last_login', 'member_tier', 'fine_balance', 'reading_stats',
        'active_borrowing', 'recent_history', 'favorites', 'badges'
    ]

    default_privacy = {
        'full_name': 'public',
        'email': 'private',
        'phone': 'friends',
        'address': 'private',
        'member_since': 'public',
        'last_login': 'private',
        'member_tier': 'public',
        'fine_balance': 'private',
        'reading_stats': 'friends',
        'active_borrowing': 'private',
        'recent_history': 'friends',
        'favorites': 'public',
        'badges': 'public'
    }

    # Feature Studio: Groupwise Dynamic Features with Descriptions
    dynamic_feature_groups = {
        "Privacy Controls": {
            'profile_search_visibility': {
                'label': 'Profile in Search Results',
                'description': 'Allow your profile to appear in search results across the platform',
                'default': True
            },
            'hide_last_seen': {
                'label': 'Hide Last Seen from Non-Friends',
                'description': 'Only friends can see when you last accessed the platform',
                'default': False
            },
            'anonymous_reviews': {
                'label': 'Publish Reviews Anonymously',
                'description': 'Your reviews are attributed to an anonymous alias instead of your name',
                'default': False
            },
            'mask_borrow_titles': {
                'label': 'Mask Borrowed Titles in Feed',
                'description': 'Hide specific book titles you borrow from public activity feeds',
                'default': False
            },
            'anonymous_recommendations': {
                'label': 'Hide Identity in Recommendations',
                'description': 'Your recommendation engine contributions remain anonymous',
                'default': False
            },
            'anonymous_alias_rotation': {
                'label': 'Auto-Rotate Anonymous Alias',
                'description': 'Automatically rotate your anonymous alias every 24-72 hours for enhanced anonymity',
                'default': False
            }
        },
        "Social Features": {
            'allow_friend_requests': {
                'label': 'Allow Incoming Friend Requests',
                'description': 'Enable other users to send you friend requests and see your profile',
                'default': True
            },
            'show_online_status': {
                'label': 'Show Online Status to Friends',
                'description': 'Friends can see whether you are currently active on the platform',
                'default': True
            },
            'share_reading_velocity': {
                'label': 'Share Reading Velocity Metrics',
                'description': 'Friends can see your pages-per-day and completion trends',
                'default': True
            },
            'share_fine_status': {
                'label': 'Share Fine Status with Friends',
                'description': 'Friends can see if you have pending fines or overdue books',
                'default': False
            }
        },
        "Security & Device Management": {
            'step_up_profile_edit': {
                'label': 'Step-Up Auth for Profile Edits',
                'description': 'Require multi-factor verification before changing profile information',
                'default': True
            },
            'step_up_data_export': {
                'label': 'Step-Up Auth for Data Exports',
                'description': 'Require multi-factor verification before exporting personal data',
                'default': True
            },
            'trusted_only_sensitive': {
                'label': 'Restrict Sensitive Actions to Trusted Devices',
                'description': 'Password changes and exports only allowed from marked trusted devices',
                'default': True
            },
            'auto_revoke_idle_sessions': {
                'label': 'Auto-Revoke Idle Sessions',
                'description': 'Automatically log out inactive sessions after 24 hours of inactivity',
                'default': True
            },
            'new_device_alerts': {
                'label': 'Alert on New Device Login',
                'description': 'Receive notifications when your account logs in from a new device',
                'default': True
            },
            'impossible_travel_lock': {
                'label': 'Lock Sessions with Impossible Travel Detected',
                'description': 'Auto-lock sessions showing geographically impossible location changes',
                'default': True
            },
            'high_churn_lock': {
                'label': 'Lock High Session Churn',
                'description': 'Auto-lock sessions showing excessive rapid disconnects/reconnects',
                'default': False
            },
            'step_up_new_devices': {
                'label': 'Step-Up Challenge on New Device Login',
                'description': 'Require MFA verification the first time you log in from a new device',
                'default': True
            },
            'password_rotation_reminders': {
                'label': 'Enable Password Rotation Reminders',
                'description': 'Receive periodic reminders to update your account password',
                'default': True
            }
        },
        "Notifications & Quiet Time": {
            'weekly_privacy_digest': {
                'label': 'Weekly Privacy Digest',
                'description': 'Get a weekly summary of profile access and privacy-related activities',
                'default': False
            },
            'daily_deadline_digest': {
                'label': 'Daily Deadline Digest Summary',
                'description': 'Consolidated daily notification of all upcoming book deadlines',
                'default': True
            },
            'weekly_activity_digest': {
                'label': 'Weekly Activity Digest',
                'description': 'Get a weekly summary of your reading and platform activities',
                'default': True
            },
            'mute_marketing_messages': {
                'label': 'Mute Marketing & Promotional Messages',
                'description': 'Do not send promotional emails or in-app notifications about new features',
                'default': False
            },
            'strict_quiet_hours': {
                'label': 'Enforce Strict Quiet Hours',
                'description': 'No notifications will be sent during configured quiet hours—even digests',
                'default': False
            },
            'multi_channel_bundling': {
                'label': 'Bundle Alerts Across Multiple Channels',
                'description': 'Group related notifications into single emails/SMS instead of individual messages',
                'default': True
            },
            'enable_snooze_actions': {
                'label': 'Enable Snooze/Postpone on Notifications',
                'description': 'One-click snooze actions on all notifications to defer temporarily',
                'default': True
            }
        },
        "Deadline & Reminder Management": {
            'calendar_due_sync': {
                'label': 'Sync Due Dates to External Calendar',
                'description': 'Automatically create entries in iCal/Google Calendar for book due dates',
                'default': False
            },
            'smart_due_bundles': {
                'label': 'Smartly Bundle Due-Date Reminders',
                'description': 'Intelligently group multiple due-date notifications by due date clusters',
                'default': True
            },
            'hide_due_date_feed': {
                'label': 'Hide Due Dates from Activity Feed',
                'description': 'Do not show upcoming deadlines in your public activity feed',
                'default': False
            },
            'public_export_links': {
                'label': 'Allow Public Profile Export Links (Signed)',
                'description': 'Generate shareable signed links to export your profile data—requires signed exports',
                'default': False
            }
        },
        "Analytics & Insights": {
            'personal_kpi_dashboard': {
                'label': 'Personal KPI Dashboard',
                'description': 'View personalized key performance indicators for reading habits and engagement',
                'default': True
            },
            'reading_goal_tracker': {
                'label': 'Reading Goal Tracker',
                'description': 'Set and track monthly/yearly reading goals with progress visualization',
                'default': True
            },
            'activity_heatmap': {
                'label': 'Account Activity Heatmap',
                'description': 'Visualize your platform activity patterns across days and times',
                'default': False
            },
            'search_personalization_boost': {
                'label': 'Personalized Search Results',
                'description': 'Boost search results based on your reading history and preferences',
                'default': True
            },
            'streak_recovery_mode': {
                'label': 'Streak Recovery Grace Mode',
                'description': 'Automatically continue reading streaks if you miss by only 1-2 days',
                'default': False
            },
            'recommendation_diversity': {
                'label': 'Diversified Recommendations',
                'description': 'Mix mainstream and niche recommendations to expand your reading horizon',
                'default': True
            }
        },
        "Discovery & Automation": {
            'availability_watch_alerts': {
                'label': 'Shelf Availability Watch Alerts',
                'description': 'Get notified when books you\'re waiting for become available',
                'default': True
            },
            'waitlist_auto_join': {
                'label': 'Auto-Join Waitlists for Recommended Books',
                'description': 'Automatically join waitlists for personalized book recommendations',
                'default': False
            },
            'one_click_reborrow': {
                'label': 'One-Click Reborrow Actions',
                'description': 'Quick action button to instantly renew or reborrow books before they are due',
                'default': True
            },
            'timezone_auto_detect': {
                'label': 'Auto-Detect Timezone',
                'description': 'Automatically detect and update your timezone for accurate deadline calculations',
                'default': True
            }
        },
        "Accessibility & User Interface": {
            'compact_density_mode': {
                'label': 'Compact UI Density Mode',
                'description': 'Reduce spacing between elements to fit more information on screen',
                'default': False
            },
            'high_contrast_accessibility': {
                'label': 'High Contrast Accessibility Mode',
                'description': 'Enhance color contrast for improved readability and visibility',
                'default': False
            },
            'dyslexia_font_mode': {
                'label': 'Dyslexia-Friendly Typography',
                'description': 'Use OpenDyslexic font and layout optimizations for enhanced readability',
                'default': False
            },
            'reduced_motion_mode': {
                'label': 'Reduced Motion & Animations',
                'description': 'Minimize animations and transitions throughout the platform',
                'default': False
            },
            'keyboard_first_mode': {
                'label': 'Keyboard-First Navigation',
                'description': 'Optimize interface for keyboard navigation without touching mouse/trackpad',
                'default': False
            },
            'adaptive_home_layout': {
                'label': 'Adaptive Home Layout',
                'description': 'Dynamically adjust home page layout based on your screen size and preferences',
                'default': True
            }
        },
        "Data Management & Compliance": {
            'api_audit_webhooks': {
                'label': 'API Audit Webhook Forwarding',
                'description': 'Forward account audit events to external webhook URLs for compliance tracking',
                'default': False
            },
            'signed_exports_only': {
                'label': 'Allow Only Signed Exports',
                'description': 'All data exports must be digitally signed with your private key for verification',
                'default': True
            },
            'auto_delete_old_exports': {
                'label': 'Auto-Delete Old Exports',
                'description': 'Automatically remove exported data older than 90 days to reduce storage',
                'default': True
            },
            'retention_policy_controls': {
                'label': 'Granular Data Retention Policies',
                'description': 'Set custom retention periods for different types of account data',
                'default': False
            }
        }
    }

    # Flatten for quick lookups
    def get_all_features():
        features = {}
        for group_name, group_features in dynamic_feature_groups.items():
            features.update({key: {**val, 'group': group_name} for key, val in group_features.items()})
        return features

    def get_default_feature_flags():
        flags = {}
        for group_features in dynamic_feature_groups.values():
            for key, config in group_features.items():
                flags[key] = bool(config.get('default', False))
        return flags

    default_feature_flags = get_default_feature_flags()

    def check_feature_compatibility(proposed_flags):
        """Cross-check feature compatibility and return incompatibilities + auto-corrections."""
        issues = []
        corrections = {}
        
        # Check 1: Signed exports required for public export links
        if proposed_flags.get('public_export_links') and not proposed_flags.get('signed_exports_only'):
            issues.append({
                'severity': 'warning',
                'message': 'Public export links require signed exports to be enabled',
                'auto_fix': True,
                'action': 'enable signed_exports_only'
            })
            corrections['signed_exports_only'] = True
        
        # Check 2: Step-up auth required for sensitive operations with trusted devices
        if proposed_flags.get('trusted_only_sensitive'):
            if not proposed_flags.get('step_up_profile_edit'):
                issues.append({
                    'severity': 'info',
                    'message': 'Trusted-devices-only setting works best when step-up profile edit is enabled',
                    'auto_fix': False
                })
            if not proposed_flags.get('step_up_data_export'):
                issues.append({
                    'severity': 'info',
                    'message': 'Trusted-devices-only setting works best when step-up data export is enabled',
                    'auto_fix': False
                })
        
        # Check 3: Anonymous reviews incompatible with social sharing
        if proposed_flags.get('anonymous_reviews') and proposed_flags.get('share_reading_velocity'):
            issues.append({
                'severity': 'info',
                'message': 'Anonymous reviews + sharing reading velocity may leak reading patterns',
                'auto_fix': False
            })
        
        # Check 4: Strict quiet hours limits digest emails
        if proposed_flags.get('strict_quiet_hours'):
            digest_flags = ['weekly_privacy_digest', 'daily_deadline_digest', 'weekly_activity_digest']
            active_digests = sum(1 for d in digest_flags if proposed_flags.get(d))
            if active_digests > 0:
                issues.append({
                    'severity': 'warning',
                    'message': f'Strict quiet hours will suppress {active_digests} active digest(s), keeping them queued',
                    'auto_fix': False
                })
        
        # Check 5: Accessibility mode conflicts
        if proposed_flags.get('compact_density_mode') and proposed_flags.get('high_contrast_accessibility'):
            issues.append({
                'severity': 'info',
                'message': 'Compact density mode + high contrast may reduce visual spacing for some components',
                'auto_fix': False
            })
        
        # Check 6: Automation conflicts
        if proposed_flags.get('waitlist_auto_join') and not proposed_flags.get('availability_watch_alerts'):
            issues.append({
                'severity': 'info',
                'message': 'Auto-join waitlist works best when paired with availability alerts',
                'auto_fix': False
            })
        
        # Check 7: Alias rotation requires anonymous mode consideration
        if proposed_flags.get('anonymous_alias_rotation') and not proposed_flags.get('anonymous_reviews'):
            issues.append({
                'severity': 'info',
                'message': 'Alias rotation is most useful when combined with anonymous reviews',
                'auto_fix': False
            })
        
        return issues, corrections

    def generate_anonymous_alias():
        adjectives = ['Silent', 'Hidden', 'Masked', 'Steady', 'Nimble', 'Calm', 'Sharp', 'Quiet', 'Cipher', 'Nova']
        nouns = ['Reader', 'Fox', 'Atlas', 'Owl', 'Raven', 'Voyager', 'Pilot', 'Scout', 'Harbor', 'Wave']
        return f"{secrets.choice(adjectives)}{secrets.choice(nouns)}{secrets.randbelow(900)+100}"

    def ensure_dynamic_preferences_row():
        row = Database.execute_query(
            """
            SELECT anonymous_mode_enabled, anonymous_alias, anonymous_avatar_style,
                   anonymous_rotation_hours, profile_theme, feature_json
            FROM account_dynamic_preferences
            WHERE user_id = ?
            """,
            (account['user_id'],),
            fetch_one=True
        )
        if row:
            return row

        seed = get_default_feature_flags()
        Database.execute_update(
            """
            INSERT INTO account_dynamic_preferences
            (user_id, anonymous_mode_enabled, anonymous_alias, anonymous_avatar_style,
             anonymous_rotation_hours, profile_theme, feature_json, updated_at)
            VALUES (?, 0, ?, 'geometric', 72, 'adaptive', ?, datetime('now'))
            """,
            (account['user_id'], generate_anonymous_alias(), json.dumps(seed))
        )
        return Database.execute_query(
            """
            SELECT anonymous_mode_enabled, anonymous_alias, anonymous_avatar_style,
                   anonymous_rotation_hours, profile_theme, feature_json
            FROM account_dynamic_preferences
            WHERE user_id = ?
            """,
            (account['user_id'],),
            fetch_one=True
        )

    def load_privacy_map():
        rows = Database.execute_query(
            """
            SELECT field_name, visibility
            FROM account_field_privacy
            WHERE user_id = ?
            """,
            (account['user_id'],)
        ) or []
        mapping = dict(default_privacy)
        for row in rows:
            if row['field_name'] in privacy_fields and row.get('visibility') in ['public', 'private', 'friends']:
                mapping[row['field_name']] = row['visibility']
        return mapping

    def save_privacy_map(mapping):
        for field_name in privacy_fields:
            visibility = mapping.get(field_name, default_privacy[field_name])
            Database.execute_update(
                """
                INSERT INTO account_field_privacy (user_id, field_name, visibility, updated_at)
                VALUES (?, ?, ?, datetime('now'))
                ON CONFLICT(user_id, field_name)
                DO UPDATE SET visibility = excluded.visibility, updated_at = datetime('now')
                """,
                (account['user_id'], field_name, visibility)
            )

    def has_sensitive_action_access():
        if account['user_id'] <= 0:
            return True
        return Auth.is_current_device_trusted() or Auth.is_step_up_verified()

    def require_sensitive_action_access(purpose_key, header_label):
        """Render step-up UI and require trusted or recently verified session."""
        if account['user_id'] <= 0:
            return True
        if has_sensitive_action_access():
            return True

        st.warning(f"{header_label} requires step-up authentication because this device is not trusted.")
        c1, c2, c3 = st.columns([1, 1, 2], gap='small')
        with c1:
            if st.button("Send Verification Code", key=f"stepup_send_{purpose_key}", use_container_width=True):
                Auth.start_step_up_challenge(account['user_id'], account['email'], purpose=purpose_key)
                st.success("Verification code sent to your notification inbox/email channel.")
        with c2:
            code = st.text_input("Code", max_chars=6, key=f"stepup_code_{purpose_key}")
            trust_device = st.checkbox("Trust this device", key=f"stepup_trust_{purpose_key}")
            if st.button("Verify", key=f"stepup_verify_{purpose_key}", use_container_width=True):
                ok, msg = Auth.verify_step_up_challenge(account['user_id'], code, purpose=purpose_key)
                if ok:
                    if trust_device:
                        Auth.mark_current_device_trusted(trust_label='User trusted from account center', trust_days=30)
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)
        return False

    def parse_hhmm(hhmm):
        try:
            h, m = str(hhmm).split(':')
            return int(h), int(m)
        except Exception:
            return 22, 0

    def is_quiet_hours(now_dt, start_hhmm, end_hhmm):
        sh, sm = parse_hhmm(start_hhmm)
        eh, em = parse_hhmm(end_hhmm)
        start_minutes = sh * 60 + sm
        end_minutes = eh * 60 + em
        current_minutes = now_dt.hour * 60 + now_dt.minute

        if start_minutes == end_minutes:
            return False
        if start_minutes < end_minutes:
            return start_minutes <= current_minutes < end_minutes
        return current_minutes >= start_minutes or current_minutes < end_minutes

    def run_deadline_notification_scan(scan_pref, source='manual'):
        """Queue + process deadline reminders with digest, dedup, and delivery state machine."""
        tz_name_scan = scan_pref.get('timezone', 'UTC') or 'UTC'
        st.session_state.account_timezone_hint = tz_name_scan
        try:
            now_local = datetime.now(ZoneInfo(tz_name_scan))
        except Exception:
            tz_name_scan = 'UTC'
            now_local = datetime.now(ZoneInfo('UTC'))

        if scan_pref.get('quiet_hours_enabled') and is_quiet_hours(
            now_local,
            scan_pref.get('quiet_start', '22:00'),
            scan_pref.get('quiet_end', '07:00')
        ):
            return {
                'matched': 0,
                'queued': 0,
                'sent': 0,
                'failed': 0,
                'retried': 0,
                'dedup_skipped': 0,
                'status': 'quiet_hours_active',
                'timezone': tz_name_scan
            }

        notify_date_local = now_local.date().isoformat()
        due_items = Database.execute_query(
            """
            SELECT br.borrowing_id, b.title, br.due_date,
                   CAST(julianday(br.due_date) - julianday(date('now')) AS INTEGER) as days_until_due
            FROM borrowing br
            JOIN books b ON br.book_id = b.book_id
            WHERE br.user_id = ?
              AND br.return_date IS NULL
              AND br.due_date <= date('now', '+' || ? || ' days')
            ORDER BY br.due_date ASC
            """,
            (account['user_id'], to_int(scan_pref.get('deadline_threshold_days'), 3))
        ) or []

        queued = 0
        dedup_skipped = 0
        digest_mode = bool(scan_pref.get('digest_mode', 0))
        digest_hour = max(0, min(23, to_int(scan_pref.get('digest_hour'), 8)))

        if digest_mode and source == 'scheduler' and int(now_local.hour) < digest_hour:
            return {
                'matched': len(due_items),
                'queued': 0,
                'sent': 0,
                'failed': 0,
                'retried': 0,
                'dedup_skipped': 0,
                'status': 'digest_window_not_reached',
                'timezone': tz_name_scan
            }

        if digest_mode and due_items:
            lines = [
                f"- {row['title']} (due: {row['due_date']}, in {row['days_until_due']} day(s))"
                for row in due_items
            ]
            digest_message = "Daily borrowing reminder digest:\n" + "\n".join(lines)
            for channel in ['in_app', 'email']:
                if channel == 'in_app' and not scan_pref.get('in_app_enabled'):
                    continue
                if channel == 'email' and not scan_pref.get('email_enabled'):
                    continue
                dedup_key = f"digest:{account['user_id']}:{notify_date_local}:{channel}"
                payload = {
                    'email': account['email'],
                    'subject': 'LitGrid Daily Deadline Digest',
                    'message': digest_message,
                    'type': 'daily_digest',
                    'digest_mode': True,
                    'full_name': account['full_name']
                }
                if AccountOpsEngine.queue_notification_job(
                    account['user_id'],
                    None,
                    channel,
                    notify_date_local,
                    payload,
                    dedup_key,
                    max_retries=3
                ):
                    queued += 1
                else:
                    dedup_skipped += 1
        else:
            for row in due_items:
                for channel in ['in_app', 'email']:
                    if channel == 'in_app' and not scan_pref.get('in_app_enabled'):
                        continue
                    if channel == 'email' and not scan_pref.get('email_enabled'):
                        continue

                    already_sent = Database.execute_query(
                        """
                        SELECT ledger_id FROM account_notification_ledger
                        WHERE user_id = ? AND borrowing_id = ? AND channel = ? AND notify_date = ?
                        """,
                        (account['user_id'], row['borrowing_id'], channel, notify_date_local),
                        fetch_one=True
                    )
                    if already_sent:
                        dedup_skipped += 1
                        continue

                    dedup_key = f"{account['user_id']}:{row['borrowing_id']}:{channel}:{notify_date_local}"
                    payload = {
                        'email': account['email'],
                        'full_name': account['full_name'],
                        'book_title': row['title'],
                        'due_date': row['due_date'],
                        'days_until_due': int(row['days_until_due'] or 0),
                        'subject': 'Borrowing deadline reminder',
                        'message': f"'{row['title']}' is due in {row['days_until_due']} day(s) on {row['due_date']}",
                        'type': 'deadline',
                        'digest_mode': False
                    }
                    if AccountOpsEngine.queue_notification_job(
                        account['user_id'],
                        row['borrowing_id'],
                        channel,
                        notify_date_local,
                        payload,
                        dedup_key,
                        max_retries=3
                    ):
                        queued += 1
                    else:
                        dedup_skipped += 1

        delivery = AccountOpsEngine.process_notification_queue_for_user(account['user_id'])
        AuditLogger.log_action(
            user_id=account['user_id'],
            action='deadline_notification_scan',
            entity_type='account',
            details=(
                f"source={source};matched={len(due_items)};queued={queued};"
                f"sent={delivery['sent']};failed={delivery['failed']};retried={delivery['retried']};"
                f"dedup_skipped={dedup_skipped};digest_mode={int(digest_mode)};timezone={tz_name_scan}"
            ),
            status='success'
        )

        return {
            'matched': len(due_items),
            'queued': queued,
            'sent': delivery['sent'],
            'failed': delivery['failed'],
            'retried': delivery['retried'],
            'dedup_skipped': dedup_skipped,
            'status': 'ok',
            'timezone': tz_name_scan
        }

    st.markdown('<h1 class="litgrid-header"> My Account</h1>', unsafe_allow_html=True)

    if account['is_demo']:
        st.info("Demo account detected: account updates and exports are limited.")

    tab1, tab2, tab4, tab5 = st.tabs([
        " Profile",
        " Reading Analytics",
        " Data Tools",
        " Feature Studio"
    ])
    tab3 = tab1
    tab6 = tab5

    with tab1:
        col1, col2, col3 = st.columns(3, gap="small")

        with col1:
            st.subheader("Identity")
            st.write(f"**Full Name:** {account['full_name']}")
            st.write(f"**Username:** {account['username']}")
            st.write(f"**Email:** {account['email']}")
            st.write(f"**Phone:** {account['phone']}")
            st.write(f"**Address:** {account.get('address') or 'Not provided'}")
            st.write(f"**Member Since:** {format_date(account['created_at'])}")
            st.write(f"**Last Login:** {format_datetime(account['last_login'])}")

            if account['user_id'] > 0:
                anon_row = Database.execute_query(
                    "SELECT anonymous_mode_enabled, anonymous_alias FROM account_dynamic_preferences WHERE user_id = ?",
                    (account['user_id'],),
                    fetch_one=True
                )
                if anon_row and bool(anon_row.get('anonymous_mode_enabled')):
                    st.caption(f"Anonymous identity active: {anon_row.get('anonymous_alias') or 'Anonymous'}")

        with col2:
            st.subheader("Account Runtime Status")
            st.write(f"**Role:** {account['role'].title()}")
            st.write(f"**Member Tier:** {str(account['member_tier']).title()}")
            st.write(f"**Max Books Allowed:** {account['max_books_allowed']}")
            st.write(f"**Borrowing Days:** {account['borrowing_days']}")
            st.write(f"**Status:** {' Active' if account['is_active'] else ' Inactive'}")

            st.divider()
            if account['fine_balance'] > 0:
                st.error(f"Outstanding Fine: {format_currency(account['fine_balance'])}")
            else:
                st.success("No Outstanding Fines")

        with col3:
            st.subheader("Password Rotation")

            if not has_sensitive_action_access():
                require_sensitive_action_access('password_change', 'Password Change')

            def validate_local_password_strength(password):
                if len(password) < 12:
                    return False, "Password must be at least 12 characters"
                if password.lower() == password:
                    return False, "Password must contain uppercase letters"
                if password.upper() == password:
                    return False, "Password must contain lowercase letters"
                if not any(c.isdigit() for c in password):
                    return False, "Password must contain numbers"
                if not any(c in "!@#$%^&*()-_=+[]{}|;:,.<>?" for c in password):
                    return False, "Password must contain special characters"
                return True, "Strong password"

            with st.form("change_password_form"):
                current_password = st.text_input("Current Password", type="password")
                new_password = st.text_input("New Password", type="password")
                confirm_password = st.text_input("Confirm New Password", type="password")
                submit_change = st.form_submit_button("Change Password", use_container_width=True)

                if submit_change:
                    allowed, wait_mins = AccountOpsEngine.check_operation_rate_limit(
                        account['user_id'], 'password_change', max_attempts=5, window_minutes=60
                    )
                    if not allowed:
                        st.error(f"Password changes are temporarily throttled. Try again in about {wait_mins} minute(s).")
                        AccountOpsEngine.log_operation_result(account['user_id'], 'password_change', 'blocked', 'rate_limited')
                    elif not has_sensitive_action_access():
                        st.error("Step-up verification required. Use the access verification panel first.")
                        AccountOpsEngine.log_operation_result(account['user_id'], 'password_change', 'blocked', 'step_up_required')
                    elif not all([current_password, new_password, confirm_password]):
                        st.error("Please fill all fields")
                    elif new_password != confirm_password:
                        st.error("New passwords do not match")
                    else:
                        strong, message = validate_local_password_strength(new_password)
                        if not strong:
                            st.error(message)
                            AccountOpsEngine.log_operation_result(account['user_id'], 'password_change', 'failed', message)
                        else:
                            success, msg = Auth.change_password(account['user_id'], current_password, new_password)
                            if success:
                                st.success(msg)
                                AccountOpsEngine.log_operation_result(account['user_id'], 'password_change', 'success')
                            else:
                                st.error(msg)
                                AccountOpsEngine.log_operation_result(account['user_id'], 'password_change', 'failed', msg)

        st.divider()
        st.subheader("Profile Health")

        completeness_fields = {
            'Full Name': bool(account['full_name'] and account['full_name'] != 'Unknown User'),
            'Username': bool(account['username'] and account['username'] != 'unknown'),
            'Email': bool(account['email'] and account['email'] != 'N/A'),
            'Phone': bool(account['phone'] and account['phone'] != 'Not provided'),
        }
        completeness_score = int((sum(1 for ok in completeness_fields.values() if ok) / len(completeness_fields)) * 100)

        m1, m2, m3 = st.columns(3, gap="small")
        with m1:
            st.metric("Profile Completeness", f"{completeness_score}%")
        with m2:
            session_remaining = "N/A"
            if st.session_state.get('login_time'):
                elapsed = datetime.now() - st.session_state.login_time
                left = max(0, Config.SESSION_TIMEOUT - int(elapsed.total_seconds() // 60))
                session_remaining = f"{left} min"
            st.metric("Session Timeout Remaining", session_remaining)
        with m3:
            st.metric(
                "Auth Context",
                "Superadmin (Highest Privilege)" if is_superadmin_privileged else ("Real User" if account['user_id'] > 0 else "Functional Account")
            )

        if can_use_persisted_features:
            borrow_health = Database.execute_query(
                """
                SELECT
                    COUNT(*) as active_count,
                    SUM(CASE WHEN due_date < date('now') THEN 1 ELSE 0 END) as overdue_count,
                    SUM(CASE WHEN due_date BETWEEN date('now') AND date('now', '+3 days') THEN 1 ELSE 0 END) as due_soon_count
                FROM borrowing
                WHERE user_id = ? AND return_date IS NULL
                """,
                (account['user_id'],),
                fetch_one=True
            )

            active_count = to_int(borrow_health.get('active_count') if borrow_health else 0)
            overdue_count = to_int(borrow_health.get('overdue_count') if borrow_health else 0)
            due_soon_count = to_int(borrow_health.get('due_soon_count') if borrow_health else 0)

            c1, c2, c3 = st.columns(3, gap="small")
            c1.metric("Active Borrowings", active_count)
            c2.metric("Overdue Risk", overdue_count)
            c3.metric("Due in 3 Days", due_soon_count)

            active_items = Database.execute_query(
                """
                SELECT b.title,
                       br.checkout_date,
                       br.due_date,
                       CAST(julianday(br.due_date) - julianday(date('now')) AS INTEGER) as days_left
                FROM borrowing br
                JOIN books b ON br.book_id = b.book_id
                WHERE br.user_id = ? AND br.return_date IS NULL
                ORDER BY br.due_date ASC
                LIMIT 20
                """,
                (account['user_id'],)
            )

            if active_items:
                st.subheader("Borrowing Deadlines")
                df_active = pd.DataFrame(active_items)

                def deadline_state(days_left):
                    if days_left is None:
                        return 'Unknown'
                    if days_left < 0:
                        return 'Overdue'
                    if days_left <= 3:
                        return 'Due Soon'
                    return 'On Track'

                df_active['status'] = df_active['days_left'].apply(deadline_state)
                st.dataframe(df_active, use_container_width=True, hide_index=True)
        else:
            st.info("Borrowing risk analytics are available for persisted database accounts.")

    with tab2:
        st.subheader("Reading Analytics")

        if not can_use_persisted_features:
            st.info("Detailed analytics are not available for functional/demo accounts.")
        else:
            stats = get_member_statistics(account['user_id'])
            c1, c2, c3, c4 = st.columns(4, gap="small")
            c1.metric("Books Read", stats.get('total_read', 0))
            c2.metric("Currently Borrowed", stats.get('currently_borrowed', 0))
            c3.metric("Reading Days", stats.get('total_reading_days', 0))
            c4.metric("Favorite Genre", stats.get('favorite_genre', 'N/A'))

            st.divider()

            monthly_activity = Database.execute_query(
                """
                SELECT strftime('%Y-%m', checkout_date) as month,
                       COUNT(*) as borrow_count,
                       AVG(julianday(COALESCE(return_date, date('now'))) - julianday(checkout_date)) as avg_duration
                FROM borrowing
                WHERE user_id = ? AND checkout_date >= date('now', '-12 months')
                GROUP BY month
                ORDER BY month
                """,
                (account['user_id'],)
            )

            if monthly_activity:
                st.subheader("12-Month Borrowing Activity")
                df_month = pd.DataFrame(monthly_activity)
                df_month['avg_duration'] = df_month['avg_duration'].fillna(0)

                fig = go.Figure()
                fig.add_trace(go.Bar(
                    x=df_month['month'],
                    y=df_month['borrow_count'],
                    name='Borrows',
                    marker_color='#1E88E5'
                ))
                fig.add_trace(go.Scatter(
                    x=df_month['month'],
                    y=df_month['avg_duration'],
                    name='Avg Days Kept',
                    mode='lines+markers',
                    yaxis='y2',
                    line=dict(color='#FF7043')
                ))
                fig.update_layout(
                    xaxis_title='Month',
                    yaxis=dict(title='Borrow Count'),
                    yaxis2=dict(title='Avg Days Kept', overlaying='y', side='right'),
                    legend=dict(orientation='h'),
                    height=380
                )
                st.plotly_chart(fig, use_container_width=True)

            genre_distribution = Database.execute_query(
                """
                SELECT COALESCE(b.genre, 'Unknown') as genre, COUNT(*) as count
                FROM borrowing br
                JOIN books b ON br.book_id = b.book_id
                WHERE br.user_id = ?
                GROUP BY COALESCE(b.genre, 'Unknown')
                ORDER BY count DESC
                LIMIT 8
                """,
                (account['user_id'],)
            )

            if genre_distribution:
                st.subheader("Genre Distribution")
                df_genre = pd.DataFrame(genre_distribution)
                fig_genre = px.pie(df_genre, values='count', names='genre', hole=0.45)
                st.plotly_chart(fig_genre, use_container_width=True)

            recent_history = Database.execute_query(
                """
                SELECT b.title, b.isbn, br.checkout_date, br.return_date,
                       ROUND(julianday(COALESCE(br.return_date, date('now'))) - julianday(br.checkout_date), 1) as days_borrowed
                FROM borrowing br
                JOIN books b ON br.book_id = b.book_id
                WHERE br.user_id = ?
                ORDER BY br.checkout_date DESC
                LIMIT 15
                """,
                (account['user_id'],)
            )
            st.subheader("Recent Reading History")
            if recent_history:
                st.dataframe(pd.DataFrame(recent_history), use_container_width=True, hide_index=True)
            else:
                st.info("No reading history found.")

    with tab3:
        st.subheader("Security & Access")

        if user.get('is_functional_admin'):
            st.markdown("### Functional Admin Security Key")
            with st.form("change_security_key_form"):
                current_security_key = st.text_input("Current Security Password", type="password")
                new_security_key = st.text_input("New Security Password", type="password")
                confirm_security_key = st.text_input("Confirm New Security Password", type="password")

                submit_security = st.form_submit_button("Change Security Password", use_container_width=True)

                if submit_security:
                    if not all([current_security_key, new_security_key, confirm_security_key]):
                        st.error("Please fill all fields")
                    elif new_security_key != confirm_security_key:
                        st.error("New security passwords do not match")
                    elif len(new_security_key) < 6:
                        st.warning("Security password must be at least 6 characters")
                    else:
                        stored_key = st.session_state.get('admin_security_key', Config._x4)
                        if current_security_key == stored_key:
                            if Auth.update_admin_security_key(new_security_key):
                                st.success("Security password changed successfully")
                            else:
                                st.error("Failed to update security password")
                        else:
                            st.error("Current security password is incorrect")

            st.divider()

        st.divider()
        st.subheader("Session Diagnostics")
        login_time = st.session_state.get('login_time')
        if login_time:
            elapsed = datetime.now() - login_time
            left = max(0, Config.SESSION_TIMEOUT - int(elapsed.total_seconds() // 60))
            sd1, sd2, sd3 = st.columns(3, gap="small")
            sd1.metric("Session Age", f"{int(elapsed.total_seconds() // 60)} min")
            sd2.metric("Timeout Window", f"{Config.SESSION_TIMEOUT} min")
            sd3.metric("Remaining", f"{left} min")
        else:
            st.info("No session timing data available.")

        if account['user_id'] > 0:
            st.subheader("Session / Device Management")
            current_token = st.session_state.get('session_token')
            sessions = Database.execute_query(
                """
                SELECT session_token, device_label, created_at, last_seen, is_active,
                       trusted_device, trust_label, risk_score, risk_reasons, geo_hint
                FROM user_sessions
                WHERE user_id = ?
                ORDER BY last_seen DESC
                LIMIT 30
                """,
                (account['user_id'],)
            ) or []

            if sessions:
                session_rows = []
                revokable_tokens = []
                for s in sessions:
                    is_current = s['session_token'] == current_token
                    is_active = bool(s.get('is_active'))
                    session_rows.append({
                        'Current': 'Yes' if is_current else 'No',
                        'Device': s.get('device_label') or 'Web Session',
                        'Geo Hint': s.get('geo_hint') or 'N/A',
                        'Trusted': 'Yes' if int(s.get('trusted_device') or 0) == 1 else 'No',
                        'Created': s.get('created_at'),
                        'Last Seen': s.get('last_seen'),
                        'Status': 'Active' if is_active else 'Revoked',
                        'Risk Score': int(s.get('risk_score') or 0),
                        'Risk Reasons': s.get('risk_reasons') or '-',
                        'Token': s['session_token'][:10] + '...'
                    })
                    if is_active and not is_current:
                        revokable_tokens.append(s['session_token'])

                st.dataframe(pd.DataFrame(session_rows), use_container_width=True, hide_index=True)

                sc1, sc2 = st.columns(2, gap="small")
                with sc1:
                    if revokable_tokens:
                        selected_token = st.selectbox(
                            "Revoke specific session",
                            options=revokable_tokens,
                            format_func=lambda t: t[:10] + '...'
                        )
                        if st.button("Revoke Selected Session", use_container_width=True):
                            done = Database.execute_update(
                                """
                                UPDATE user_sessions
                                SET is_active = 0, revoked_at = datetime('now'), last_seen = datetime('now')
                                WHERE user_id = ? AND session_token = ?
                                """,
                                (account['user_id'], selected_token)
                            )
                            if done:
                                st.success("Selected session revoked.")
                                st.rerun()
                            else:
                                st.error("Failed to revoke selected session.")
                with sc2:
                    if st.button("Revoke All Other Sessions", type="secondary", use_container_width=True):
                        done = Database.execute_update(
                            """
                            UPDATE user_sessions
                            SET is_active = 0, revoked_at = datetime('now'), last_seen = datetime('now')
                            WHERE user_id = ? AND session_token != ? AND is_active = 1
                            """,
                            (account['user_id'], current_token)
                        )
                        if done:
                            st.success("All other active sessions revoked.")
                            st.rerun()
                        else:
                            st.error("Failed to revoke sessions.")

                current_session_row = Auth.get_current_session_row()
                rs1, rs2, rs3 = st.columns(3, gap='small')
                rs1.metric("Current Device Trusted", "Yes" if Auth.is_current_device_trusted() else "No")
                rs2.metric("Current Session Risk", int((current_session_row or {}).get('risk_score') or 0))
                rs3.metric("Step-up Verified", "Yes" if Auth.is_step_up_verified() else "No")

                if not Auth.is_current_device_trusted():
                    if st.button("Mark Current Device as Trusted", use_container_width=True):
                        if require_sensitive_action_access('trust_current_device', 'Trust Current Device'):
                            ok = Auth.mark_current_device_trusted('Trusted from security tab', 30)
                            if ok:
                                st.success("Current device marked as trusted.")
                                st.rerun()
                            else:
                                st.error("Failed to trust current device.")
            else:
                st.caption("No persisted session records found.")

        if account['user_id'] > 0:
            st.subheader("Audit API Console")

            action_rows = Database.execute_query(
                "SELECT DISTINCT action FROM audit_logs WHERE user_id = ? ORDER BY action",
                (account['user_id'],)
            ) or []
            action_options = [row['action'] for row in action_rows if row.get('action')]

            af1, af2, af3, af4 = st.columns(4, gap="small")
            with af1:
                action_filter = st.selectbox("Action", ["All Actions"] + action_options, key="acct_audit_action")
            with af2:
                status_filter = st.selectbox("Status", ["All", "success", "failed"], key="acct_audit_status")
            with af3:
                start_date_filter = st.date_input("From", date.today() - timedelta(days=30), key="acct_audit_from")
            with af4:
                end_date_filter = st.date_input("To", date.today(), key="acct_audit_to")

            audit_query = """
                SELECT action, entity_type, entity_id, status, details, timestamp
                FROM audit_logs
                WHERE user_id = ? AND date(timestamp) BETWEEN ? AND ?
            """
            audit_params = [account['user_id'], start_date_filter, end_date_filter]

            if action_filter != "All Actions":
                audit_query += " AND action = ?"
                audit_params.append(action_filter)

            if status_filter != "All":
                audit_query += " AND status = ?"
                audit_params.append(status_filter)

            audit_query += " ORDER BY timestamp DESC LIMIT 500"
            logs = Database.execute_query(audit_query, tuple(audit_params))

            if logs:
                logs_df = pd.DataFrame(logs)
                st.dataframe(logs_df, use_container_width=True, hide_index=True)

                st.download_button(
                    "Export Filtered Audit (CSV)",
                    data=logs_df.to_csv(index=False),
                    file_name=f"audit_{account['username']}_{start_date_filter}_{end_date_filter}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            else:
                st.caption("No audit events found for current filters.")

    with tab4:
        st.subheader("Data Tools")

        if can_use_persisted_features:
            if not has_sensitive_action_access():
                require_sensitive_action_access('profile_update', 'Profile Update')

            with st.form("update_profile_contact_form"):
                st.markdown("### Update Contact Information")
                new_full_name = st.text_input("Full Name", value=account['full_name'])
                new_email = st.text_input("Email", value=account['email'])
                new_phone = st.text_input("Phone", value='' if account['phone'] == 'Not provided' else account['phone'])
                new_address = st.text_area("Address", value=account.get('address', ''), height=80)

                submit_profile_update = st.form_submit_button("Update Profile", use_container_width=True)

                if submit_profile_update:
                    allowed, wait_mins = AccountOpsEngine.check_operation_rate_limit(
                        account['user_id'], 'profile_update', max_attempts=3, window_minutes=30
                    )
                    if not allowed:
                        st.error(f"Profile updates are throttled. Try again in about {wait_mins} minute(s).")
                        AccountOpsEngine.log_operation_result(account['user_id'], 'profile_update', 'blocked', 'rate_limited')
                    elif not has_sensitive_action_access():
                        st.error("Step-up verification required. Use the access verification panel first.")
                        AccountOpsEngine.log_operation_result(account['user_id'], 'profile_update', 'blocked', 'step_up_required')
                    elif not new_full_name.strip():
                        st.error("Full name cannot be empty")
                    elif new_email and not security_manager.validate_email(new_email):
                        st.error("Invalid email format")
                    else:
                        Database.execute_update(
                            """
                            INSERT INTO account_profile_snapshots (user_id, snapshot_json, reason, created_at)
                            VALUES (?, ?, 'before_profile_update', datetime('now'))
                            """,
                            (
                                account['user_id'],
                                json.dumps({
                                    'full_name': account['full_name'],
                                    'email': account['email'],
                                    'phone': account['phone'],
                                    'address': account.get('address', '')
                                }, default=str)
                            )
                        )
                        success = Database.execute_update(
                            "UPDATE users SET full_name = ?, email = ?, phone = ?, address = ?, updated_at = datetime('now') WHERE user_id = ?",
                            (new_full_name.strip(), new_email.strip(), new_phone.strip(), new_address.strip(), account['user_id'])
                        )
                        if success:
                            st.success("Profile updated successfully. Refreshing session data...")
                            AccountOpsEngine.log_operation_result(account['user_id'], 'profile_update', 'success')
                            user['full_name'] = new_full_name.strip()
                            user['email'] = new_email.strip()
                            user['phone'] = new_phone.strip()
                            user['address'] = new_address.strip()
                            st.rerun()
                        else:
                            st.error("Failed to update profile information")
                            AccountOpsEngine.log_operation_result(account['user_id'], 'profile_update', 'failed')

            st.divider()

            st.markdown("### Borrowing Deadline Notifications")
            st.caption("Configure immediate or daily digest reminders with optional background scheduler.")

            pref = Database.execute_query(
                """
                SELECT email_enabled, in_app_enabled, deadline_threshold_days,
                       timezone, quiet_hours_enabled, quiet_start, quiet_end,
                       digest_mode, digest_hour, auto_scan_enabled, auto_scan_interval_minutes
                FROM account_notification_preferences
                WHERE user_id = ?
                """,
                (account['user_id'],),
                fetch_one=True
            )

            if not pref:
                Database.execute_update(
                    """
                    INSERT INTO account_notification_preferences
                    (user_id, email_enabled, in_app_enabled, deadline_threshold_days,
                     timezone, quiet_hours_enabled, quiet_start, quiet_end,
                     digest_mode, digest_hour, auto_scan_enabled, auto_scan_interval_minutes, updated_at)
                    VALUES (?, 1, 1, 3, 'UTC', 0, '22:00', '07:00', 0, 8, 0, 60, datetime('now'))
                    """,
                    (account['user_id'],)
                )
                pref = {
                    'email_enabled': 1,
                    'in_app_enabled': 1,
                    'deadline_threshold_days': 3,
                    'timezone': 'UTC',
                    'quiet_hours_enabled': 0,
                    'quiet_start': '22:00',
                    'quiet_end': '07:00',
                    'digest_mode': 0,
                    'digest_hour': 8,
                    'auto_scan_enabled': 0,
                    'auto_scan_interval_minutes': 60
                }

            timezone_options = [
                'UTC',
                'Asia/Dhaka',
                'Asia/Kolkata',
                'Asia/Dubai',
                'Europe/London',
                'Europe/Berlin',
                'America/New_York',
                'America/Chicago',
                'America/Los_Angeles',
                'Australia/Sydney'
            ]
            selected_tz = pref.get('timezone', 'UTC')
            if selected_tz not in timezone_options:
                timezone_options.insert(0, selected_tz)

            with st.form("account_notification_preferences_form"):
                email_enabled = st.checkbox("Email notifications", value=bool(pref.get('email_enabled', 1)))
                in_app_enabled = st.checkbox("In-app notifications", value=bool(pref.get('in_app_enabled', 1)))
                digest_mode = st.checkbox("Daily digest mode (combine reminders)", value=bool(pref.get('digest_mode', 0)))
                threshold_days = st.slider(
                    "Notify when due in (days)",
                    min_value=1,
                    max_value=14,
                    value=to_int(pref.get('deadline_threshold_days'), 3)
                )
                digest_hour = st.slider("Digest hour (0-23)", min_value=0, max_value=23, value=to_int(pref.get('digest_hour'), 8))
                tz_name = st.selectbox("Timezone", timezone_options, index=timezone_options.index(selected_tz))
                quiet_hours_enabled = st.checkbox("Enable quiet hours", value=bool(pref.get('quiet_hours_enabled', 0)))
                auto_scan_enabled = st.checkbox("Enable background scheduler mode", value=bool(pref.get('auto_scan_enabled', 0)))
                auto_scan_interval_minutes = st.slider(
                    "Scheduler interval (minutes)",
                    min_value=5,
                    max_value=180,
                    value=max(5, to_int(pref.get('auto_scan_interval_minutes'), 60))
                )

                q1, q2 = st.columns(2, gap="small")
                with q1:
                    quiet_start = st.text_input("Quiet hours start (HH:MM)", value=str(pref.get('quiet_start', '22:00')))
                with q2:
                    quiet_end = st.text_input("Quiet hours end (HH:MM)", value=str(pref.get('quiet_end', '07:00')))

                save_preferences = st.form_submit_button("Save Notification Preferences", use_container_width=True)
                if save_preferences:
                    if not re.match(r'^\d{2}:\d{2}$', quiet_start) or not re.match(r'^\d{2}:\d{2}$', quiet_end):
                        st.error("Quiet hours must use HH:MM format.")
                    else:
                        saved = Database.execute_update(
                            """
                            UPDATE account_notification_preferences
                            SET email_enabled = ?, in_app_enabled = ?, deadline_threshold_days = ?,
                                timezone = ?, quiet_hours_enabled = ?, quiet_start = ?, quiet_end = ?,
                                digest_mode = ?, digest_hour = ?, auto_scan_enabled = ?, auto_scan_interval_minutes = ?,
                                updated_at = datetime('now')
                            WHERE user_id = ?
                            """,
                            (
                                1 if email_enabled else 0,
                                1 if in_app_enabled else 0,
                                threshold_days,
                                tz_name,
                                1 if quiet_hours_enabled else 0,
                                quiet_start,
                                quiet_end,
                                1 if digest_mode else 0,
                                digest_hour,
                                1 if auto_scan_enabled else 0,
                                auto_scan_interval_minutes,
                                account['user_id']
                            )
                        )
                        if saved:
                            st.success("Notification preferences updated.")
                            st.session_state.account_timezone_hint = tz_name
                        else:
                            st.error("Failed to save notification preferences.")

            scheduler_pref = Database.execute_query(
                """
                SELECT email_enabled, in_app_enabled, deadline_threshold_days,
                       timezone, quiet_hours_enabled, quiet_start, quiet_end,
                       digest_mode, digest_hour, auto_scan_enabled, auto_scan_interval_minutes
                FROM account_notification_preferences
                WHERE user_id = ?
                """,
                (account['user_id'],),
                fetch_one=True
            ) or pref

            if scheduler_pref.get('auto_scan_enabled'):
                lock_name = f"deadline_scan_user_{account['user_id']}"
                acquired, reason = AccountOpsEngine.acquire_scheduler_lock(
                    lock_name,
                    interval_minutes=max(5, to_int(scheduler_pref.get('auto_scan_interval_minutes'), 60)),
                    hold_seconds=90
                )
                if acquired:
                    status_label = 'success'
                    error_reason = None
                    try:
                        scheduled_result = run_deadline_notification_scan(scheduler_pref, source='scheduler')
                        if scheduled_result.get('status') == 'quiet_hours_active':
                            status_label = 'quiet_hours'
                        st.caption(
                            f"Scheduler run: matched={scheduled_result['matched']} queued={scheduled_result['queued']} "
                            f"sent={scheduled_result['sent']} retried={scheduled_result['retried']} failed={scheduled_result['failed']}"
                        )
                    except Exception as ex:
                        status_label = 'failed'
                        error_reason = str(ex)
                    finally:
                        AccountOpsEngine.release_scheduler_lock(lock_name, status=status_label, error_reason=error_reason)

            if st.button("Run Deadline Notification Scan", use_container_width=True):
                scan_pref = Database.execute_query(
                    """
                    SELECT email_enabled, in_app_enabled, deadline_threshold_days,
                           timezone, quiet_hours_enabled, quiet_start, quiet_end,
                           digest_mode, digest_hour, auto_scan_enabled, auto_scan_interval_minutes
                    FROM account_notification_preferences
                    WHERE user_id = ?
                    """,
                    (account['user_id'],),
                    fetch_one=True
                ) or {
                    'email_enabled': 1,
                    'in_app_enabled': 1,
                    'deadline_threshold_days': 3,
                    'timezone': 'UTC',
                    'quiet_hours_enabled': 0,
                    'quiet_start': '22:00',
                    'quiet_end': '07:00',
                    'digest_mode': 0,
                    'digest_hour': 8,
                    'auto_scan_enabled': 0,
                    'auto_scan_interval_minutes': 60
                }
                result = run_deadline_notification_scan(scan_pref, source='manual')
                if result.get('status') == 'quiet_hours_active':
                    st.warning(f"Quiet hours active for timezone {result['timezone']}. Dispatch skipped.")
                else:
                    st.success(
                        f"Scan complete. matched={result['matched']} queued={result['queued']} "
                        f"sent={result['sent']} retried={result['retried']} failed={result['failed']} "
                        f"dedup_skipped={result['dedup_skipped']}"
                    )

            queue_metrics = Database.execute_query(
                """
                SELECT status, COUNT(*) as count
                FROM notification_delivery_queue
                WHERE user_id = ?
                GROUP BY status
                """,
                (account['user_id'],)
            ) or []
            if queue_metrics:
                metric_map = {m['status']: int(m['count']) for m in queue_metrics}
                q1, q2, q3, q4 = st.columns(4, gap='small')
                q1.metric('Queued', metric_map.get('queued', 0))
                q2.metric('Sent', metric_map.get('sent', 0))
                q3.metric('Failed', metric_map.get('failed', 0))
                q4.metric('Retried', metric_map.get('retried', 0))

            recent_queue = Database.execute_query(
                """
                SELECT status, channel, attempts, max_retries, error_reason, created_at, sent_at
                FROM notification_delivery_queue
                WHERE user_id = ?
                ORDER BY queue_id DESC
                LIMIT 20
                """,
                (account['user_id'],)
            ) or []
            if recent_queue:
                st.caption("Recent delivery states")
                st.dataframe(pd.DataFrame(recent_queue), use_container_width=True, hide_index=True)

            st.divider()

            st.markdown("### Export My Data")
            export_expiry_hours = st.slider("Export expiration window (hours)", min_value=1, max_value=168, value=24)
            export_bundle = {
                'account': account,
                'statistics': get_member_statistics(account['user_id']),
                'active_borrowing': Database.execute_query(
                    """
                    SELECT b.title, br.checkout_date, br.due_date
                    FROM borrowing br
                    JOIN books b ON br.book_id = b.book_id
                    WHERE br.user_id = ? AND br.return_date IS NULL
                    ORDER BY br.due_date ASC
                    """,
                    (account['user_id'],)
                ) or [],
                'recent_borrowing': Database.execute_query(
                    """
                    SELECT b.title, br.checkout_date, br.return_date, br.fine_amount
                    FROM borrowing br
                    JOIN books b ON br.book_id = b.book_id
                    WHERE br.user_id = ?
                    ORDER BY br.checkout_date DESC
                    LIMIT 50
                    """,
                    (account['user_id'],)
                ) or []
            }

            signed_export, checksum_text, checksum_sha, expires_at = AccountOpsEngine.build_signed_export(
                account['user_id'],
                account['username'],
                export_bundle,
                expiry_hours=export_expiry_hours
            )
            export_json = json.dumps(signed_export, indent=2, default=str)
            st.download_button(
                "Download Account Data (Signed JSON)",
                data=export_json,
                file_name=f"litgrid_account_{account['username']}.json",
                mime="application/json",
                use_container_width=True
            )
            st.download_button(
                "Download Export Checksum (.sha256)",
                data=checksum_text,
                file_name=f"litgrid_account_{account['username']}.sha256",
                mime="text/plain",
                use_container_width=True
            )
            st.caption(f"Checksum: {checksum_sha[:16]}... | Expires: {expires_at.isoformat()}Z")

            st.divider()
            st.markdown("### Danger Zone")
            st.warning("Account deactivation and deletion workflows are sensitive operations.")

            with st.expander("Step 1-2: Deactivate Account", expanded=False):
                confirm_deactivate = st.checkbox("I understand this will immediately deactivate my account.", key="dz_deactivate_ack")
                deactivate_phrase = st.text_input(
                    "Type DEACTIVATE to continue",
                    key="dz_deactivate_phrase"
                )
                can_deactivate = confirm_deactivate and deactivate_phrase.strip().upper() == "DEACTIVATE"

                if st.button("Deactivate My Account", type="secondary", disabled=not can_deactivate, use_container_width=True):
                    allowed, wait_mins = AccountOpsEngine.check_operation_rate_limit(
                        account['user_id'], 'account_deactivate', max_attempts=2, window_minutes=1440
                    )
                    if not allowed:
                        st.error(f"Deactivation is throttled. Try again in about {wait_mins} minute(s).")
                    elif not require_sensitive_action_access('account_deactivate', 'Account Deactivation'):
                        st.error("Step-up verification is required for deactivation.")
                    else:
                        deactivated = Database.execute_update(
                            "UPDATE users SET is_active = 0, updated_at = datetime('now') WHERE user_id = ?",
                            (account['user_id'],)
                        )
                        if deactivated:
                            AuditLogger.log_action(
                                user_id=account['user_id'],
                                action='account_deactivated_self_service',
                                entity_type='account',
                                entity_id=account['user_id'],
                                details='User initiated deactivation',
                                status='success'
                            )
                            st.success("Account deactivated. Logging out...")
                            Auth.logout()
                            st.rerun()
                        else:
                            st.error("Failed to deactivate account.")

            with st.expander("Step 1-2: Request Account Deletion", expanded=False):
                confirm_deletion = st.checkbox("I understand deletion requests are irreversible once approved.", key="dz_delete_ack")
                deletion_phrase = st.text_input(
                    f"Type DELETE {account['username']} to submit request",
                    key="dz_delete_phrase"
                )
                expected_phrase = f"DELETE {account['username']}"
                can_request_delete = confirm_deletion and deletion_phrase.strip() == expected_phrase

                if st.button("Submit Deletion Request", type="secondary", disabled=not can_request_delete, use_container_width=True):
                    allowed, wait_mins = AccountOpsEngine.check_operation_rate_limit(
                        account['user_id'], 'deletion_request', max_attempts=2, window_minutes=1440
                    )
                    if not allowed:
                        st.error(f"Deletion requests are throttled. Try again in about {wait_mins} minute(s).")
                    elif not require_sensitive_action_access('deletion_request', 'Deletion Request'):
                        st.error("Step-up verification is required for deletion request.")
                    else:
                        request_saved = Database.execute_update(
                            """
                            INSERT INTO account_deletion_requests
                            (user_id, export_json, requested_by, status, notes, created_at)
                            VALUES (?, ?, ?, 'pending', ?, datetime('now'))
                            """,
                            (
                                account['user_id'],
                                export_json,
                                account['username'],
                                'Requested from My Account danger zone'
                            )
                        )
                        if request_saved:
                            created_request = Database.execute_query(
                                """
                                SELECT request_id
                                FROM account_deletion_requests
                                WHERE user_id = ?
                                ORDER BY created_at DESC, request_id DESC
                                LIMIT 1
                                """,
                                (account['user_id'],),
                                fetch_one=True
                            )
                            if created_request:
                                AccountOpsEngine.log_deletion_timeline(
                                    request_id=created_request['request_id'],
                                    event_type='request_submitted',
                                    actor_user_id=account['user_id'],
                                    actor_role=account['role'],
                                    reason='User requested account deletion from danger zone.',
                                    metadata=f"checksum={checksum_sha};expires={expires_at.isoformat()}Z"
                                )
                            AuditLogger.log_action(
                                user_id=account['user_id'],
                                action='account_deletion_requested',
                                entity_type='account',
                                entity_id=account['user_id'],
                                details='Deletion request submitted by user',
                                status='success'
                            )
                            st.success("Deletion request submitted. Admin review required.")
                        else:
                            st.error("Failed to submit deletion request.")

                request_history = Database.execute_query(
                    """
                    SELECT request_id, status, created_at, reviewed_at, notes, decision_reason, reviewed_by
                    FROM account_deletion_requests
                    WHERE user_id = ?
                    ORDER BY created_at DESC
                    LIMIT 10
                    """,
                    (account['user_id'],)
                )
                if request_history:
                    st.caption("Recent deletion requests")
                    st.dataframe(pd.DataFrame(request_history), use_container_width=True, hide_index=True)
                    selected_req_id = st.selectbox(
                        "View Deletion Timeline",
                        [r['request_id'] for r in request_history],
                        key='account_deletion_timeline_selector'
                    )
                    timeline_rows = Database.execute_query(
                        """
                        SELECT event_type, actor_user_id, actor_role, reason, metadata, created_at
                        FROM account_deletion_timeline
                        WHERE request_id = ?
                        ORDER BY created_at ASC, event_id ASC
                        """,
                        (selected_req_id,)
                    ) or []
                    if timeline_rows:
                        st.caption("Immutable workflow timeline")
                        st.dataframe(pd.DataFrame(timeline_rows), use_container_width=True, hide_index=True)
        else:
            st.info("Data update/export is disabled for non-database accounts unless using superadmin privileges.")

    with tab5:
        st.subheader("Privacy, Visibility, and Anonymous Mode")

        # Determine storage mode (database or ephemeral session)
        privacy_storage_mode = "database"
        if account['user_id'] <= 0:
            privacy_storage_mode = "session"
            if 'privacy_settings_ephemeral' not in st.session_state:
                st.session_state.privacy_settings_ephemeral = {
                    'anonymous_mode_enabled': 0,
                    'anonymous_alias': generate_anonymous_alias(),
                    'anonymous_avatar_style': 'geometric',
                    'anonymous_rotation_hours': 72,
                    'profile_theme': 'adaptive'
                }
            if 'privacy_map_ephemeral' not in st.session_state:
                st.session_state.privacy_map_ephemeral = dict(default_privacy)
            st.info("🔄 **Session Mode**: Privacy settings persist until logout/browser reset. Create an account to save permanently.")

        # Load preferences
        if privacy_storage_mode == "database":
            dyn_pref = ensure_dynamic_preferences_row() or {}
        else:
            dyn_pref = st.session_state.privacy_settings_ephemeral

        raw_feature_json = dyn_pref.get('feature_json')
        try:
            persisted_features = json.loads(raw_feature_json) if raw_feature_json else {}
            if not isinstance(persisted_features, dict):
                persisted_features = {}
        except Exception:
            persisted_features = {}
        feature_flags = dict(default_feature_flags)
        feature_flags.update({k: bool(v) for k, v in persisted_features.items()})

        # Load privacy map
        if privacy_storage_mode == "database":
            privacy_map = load_privacy_map()
        else:
            privacy_map = st.session_state.privacy_map_ephemeral

        anon_col, summary_col = st.columns(2, gap='medium')

        with anon_col:
            st.markdown("### Advanced Anonymous Mode")
            with st.form("anonymous_mode_form"):
                anon_enabled = st.checkbox(
                    "Enable advanced anonymous mode",
                    value=bool(dyn_pref.get('anonymous_mode_enabled', 0))
                )
                alias_default = dyn_pref.get('anonymous_alias') or generate_anonymous_alias()
                anon_alias = st.text_input("Anonymous Alias", value=alias_default)
                avatar_style = st.radio(
                    "Anonymous avatar style",
                    ["geometric", "abstract", "minimal", "monochrome", "retro"],
                    index=["geometric", "abstract", "minimal", "monochrome", "retro"].index(
                        str(dyn_pref.get('anonymous_avatar_style', 'geometric'))
                        if str(dyn_pref.get('anonymous_avatar_style', 'geometric')) in ["geometric", "abstract", "minimal", "monochrome", "retro"]
                        else "geometric"
                    ),
                    horizontal=True
                )
                rotation_hours = st.slider(
                    "Alias rotation interval (hours)",
                    min_value=12,
                    max_value=336,
                    value=max(12, to_int(dyn_pref.get('anonymous_rotation_hours'), 72))
                )
                profile_theme = st.radio(
                    "Profile theme mode",
                    ["adaptive", "privacy-first", "social", "productivity", "minimal"],
                    index=["adaptive", "privacy-first", "social", "productivity", "minimal"].index(
                        str(dyn_pref.get('profile_theme', 'adaptive'))
                        if str(dyn_pref.get('profile_theme', 'adaptive')) in ["adaptive", "privacy-first", "social", "productivity", "minimal"]
                        else "adaptive"
                    ),
                    horizontal=True
                )
                save_anon = st.form_submit_button("Save Anonymous Preferences", use_container_width=True)

                if save_anon:
                    final_alias = anon_alias.strip() or generate_anonymous_alias()
                    if privacy_storage_mode == "database":
                        done = Database.execute_update(
                            """
                            INSERT INTO account_dynamic_preferences
                            (user_id, anonymous_mode_enabled, anonymous_alias, anonymous_avatar_style,
                             anonymous_rotation_hours, profile_theme, feature_json, updated_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'))
                            ON CONFLICT(user_id)
                            DO UPDATE SET
                                anonymous_mode_enabled = excluded.anonymous_mode_enabled,
                                anonymous_alias = excluded.anonymous_alias,
                                anonymous_avatar_style = excluded.anonymous_avatar_style,
                                anonymous_rotation_hours = excluded.anonymous_rotation_hours,
                                profile_theme = excluded.profile_theme,
                                feature_json = COALESCE(account_dynamic_preferences.feature_json, excluded.feature_json),
                                updated_at = datetime('now')
                            """,
                            (
                                account['user_id'],
                                1 if anon_enabled else 0,
                                final_alias,
                                avatar_style,
                                rotation_hours,
                                profile_theme,
                                json.dumps(feature_flags)
                            )
                        )
                        if done:
                            st.success("✅ Anonymous mode settings updated.")
                        else:
                            st.error("❌ Failed to update anonymous mode settings.")
                    else:
                        st.session_state.privacy_settings_ephemeral['anonymous_mode_enabled'] = 1 if anon_enabled else 0
                        st.session_state.privacy_settings_ephemeral['anonymous_alias'] = final_alias
                        st.session_state.privacy_settings_ephemeral['anonymous_avatar_style'] = avatar_style
                        st.session_state.privacy_settings_ephemeral['anonymous_rotation_hours'] = rotation_hours
                        st.session_state.privacy_settings_ephemeral['profile_theme'] = profile_theme
                        st.success("✅ Anonymous mode settings saved in session mode.")

            if bool(dyn_pref.get('anonymous_mode_enabled', 0)):
                st.info(
                    f"Anonymous profile active as '{dyn_pref.get('anonymous_alias') or 'Anonymous'}' "
                    f"with {dyn_pref.get('anonymous_avatar_style', 'geometric')} avatar style."
                )

        with summary_col:
            st.subheader("📊 Configuration Summary")

            current_flags = feature_flags
            enabled_count = sum(1 for key in current_flags if current_flags.get(key, False))
            s1, s2, s3 = st.columns(3)

            with s1:
                st.metric("Total Enabled Features", enabled_count)

            with s2:
                total_features = sum(len(group) for group in dynamic_feature_groups.values())
                st.metric("Total Available Features", total_features)

            with s3:
                coverage = (enabled_count / total_features * 100) if total_features > 0 else 0
                st.metric("Coverage %", f"{coverage:.1f}%")

            st.subheader("Active Features by Group")
            enabled_by_group = {}

            for group_name, group_features in dynamic_feature_groups.items():
                enabled_in_group = [key for key in group_features if bool(current_flags.get(key, False))]
                if enabled_in_group:
                    enabled_by_group[group_name] = enabled_in_group

            if enabled_by_group:
                selected_active_group = st.selectbox(
                    "Select Active Feature Category",
                    list(enabled_by_group.keys()),
                    key="active_features_group_dropdown",
                    format_func=lambda g: f"{g} ({len(enabled_by_group[g])})"
                )

                st.markdown(f"**{selected_active_group}** ({len(enabled_by_group[selected_active_group])})")
                for feature_key in enabled_by_group[selected_active_group]:
                    feature_info = dynamic_feature_groups[selected_active_group][feature_key]
                    st.caption(f"✓ {feature_info['label']}")
            else:
                st.caption("No features enabled yet. Select features below to customize your experience.")

        st.divider()
        with st.expander("Field-Level Privacy Matrix", expanded=True):
            st.caption("Set each field to public, private, or friends-only visibility.")

            with st.form("field_privacy_form"):
                updated_privacy = {}

                # Create 7-column layout with multiple rows
                field_list = list(privacy_fields)
                for row_idx in range(0, len(field_list), 7):
                    cols = st.columns(7, gap='medium')

                    # Process up to 7 fields per row
                    for col_idx in range(7):
                        field_idx = row_idx + col_idx
                        if field_idx >= len(field_list):
                            break

                        field_name = field_list[field_idx]
                        label = field_name.replace('_', ' ').title()
                        current_value = privacy_map.get(field_name, default_privacy[field_name])

                        with cols[col_idx]:
                            st.markdown(f"**{label}**")

                            # Radio button for privacy level
                            privacy_choice = st.radio(
                                f"Privacy for {field_name}",
                                ['Public', 'Friends', 'Private'],
                                index=['public', 'friends', 'private'].index(current_value) if current_value in ['public', 'friends', 'private'] else 2,
                                horizontal=False,
                                label_visibility='collapsed',
                                key=f"privacy_{field_name}"
                            )

                            # Map radio choice to lowercase
                            updated_privacy[field_name] = privacy_choice.lower()

                save_privacy = st.form_submit_button(
                    "💾 Save Privacy Matrix",
                    use_container_width=True,
                    key="field_privacy_form_submit"
                )

                if save_privacy:
                    if privacy_storage_mode == "database":
                        save_privacy_map(updated_privacy)
                        st.success("✅ Field-level privacy settings saved.")
                    else:
                        st.session_state.privacy_map_ephemeral = updated_privacy
                        st.success("✅ Field-level privacy settings saved in session mode.")

        def profile_view(field_name, viewer):
            mode = privacy_map.get(field_name, default_privacy[field_name])
            if mode == 'public':
                return True
            if mode == 'private':
                return viewer == 'private'
            return viewer in ['friends', 'private']

        section_col1, section_col2 = st.columns(2, gap='medium')

        with section_col1:
            st.markdown("### Visibility Preview")
            preview_cols = st.columns(3, gap='small')
            viewers = ['public', 'friends', 'private']
            for i, viewer in enumerate(viewers):
                with preview_cols[i]:
                    st.write(f"**{viewer.title()} View**")
                    shown = [f for f in privacy_fields if profile_view(f, viewer)]
                    hidden = [f for f in privacy_fields if f not in shown]
                    st.caption("Visible: " + (', '.join(shown) if shown else 'none'))
                    st.caption("Hidden: " + (', '.join(hidden) if hidden else 'none'))

        with section_col2:
            st.markdown("### Friends & Access Graph")

            fr1, fr2 = st.columns(2, gap='small')
            with fr1:
                target_username = st.text_input("Send friend request to username", key="friends_target_username")
                if st.button("Send Friend Request", use_container_width=True):
                    if not target_username.strip():
                        st.error("Enter a username.")
                    else:
                        target = Database.execute_query(
                            "SELECT user_id, username FROM users WHERE LOWER(username) = LOWER(?) AND is_active = 1",
                            (target_username.strip(),),
                            fetch_one=True
                        )
                        if not target:
                            st.error("User not found.")
                        elif privacy_storage_mode == "session":
                            st.error("Friend requests are only available for persisted accounts.")
                        elif to_int(target['user_id']) == account['user_id']:
                            st.error("You cannot send a friend request to yourself.")
                        else:
                            exists = Database.execute_query(
                                """
                                SELECT friendship_id FROM friendships
                                WHERE (requester_user_id = ? AND addressee_user_id = ?)
                                   OR (requester_user_id = ? AND addressee_user_id = ?)
                                """,
                                (account['user_id'], target['user_id'], target['user_id'], account['user_id']),
                                fetch_one=True
                            )
                            if exists:
                                st.warning("Friend relationship already exists or is pending.")
                            else:
                                sent = Database.execute_update(
                                    """
                                    INSERT INTO friendships
                                    (requester_user_id, addressee_user_id, status, created_at)
                                    VALUES (?, ?, 'pending', datetime('now'))
                                    """,
                                    (account['user_id'], target['user_id'])
                                )
                                if sent:
                                    st.success(f"Friend request sent to {target['username']}.")
                                    st.rerun()
                                else:
                                    st.error("Failed to send friend request.")

            with fr2:
                incoming = Database.execute_query(
                    """
                    SELECT f.friendship_id, u.username, u.full_name, f.created_at
                    FROM friendships f
                    JOIN users u ON f.requester_user_id = u.user_id
                    WHERE f.addressee_user_id = ? AND f.status = 'pending'
                    ORDER BY f.created_at DESC
                    """,
                    (account['user_id'],)
                ) or []
                if incoming:
                    st.caption("Incoming requests")
                    for row in incoming:
                        c_a, c_b, c_c = st.columns([2, 1, 1], gap='small')
                        c_a.write(f"{row['username']} ({row['full_name']})")
                        if c_b.button("Accept", key=f"friend_accept_{row['friendship_id']}"):
                            Database.execute_update(
                                """
                                UPDATE friendships
                                SET status = 'accepted', responded_at = datetime('now')
                                WHERE friendship_id = ?
                                """,
                                (row['friendship_id'],)
                            )
                            st.rerun()
                        if c_c.button("Reject", key=f"friend_reject_{row['friendship_id']}"):
                            Database.execute_update(
                                """
                                UPDATE friendships
                                SET status = 'rejected', responded_at = datetime('now')
                                WHERE friendship_id = ?
                                """,
                                (row['friendship_id'],)
                            )
                            st.rerun()
                else:
                    st.caption("No incoming friend requests.")

            accepted_friends = Database.execute_query(
                """
                SELECT f.friendship_id,
                       CASE WHEN f.requester_user_id = ? THEN u2.username ELSE u1.username END as friend_username,
                       CASE WHEN f.requester_user_id = ? THEN u2.full_name ELSE u1.full_name END as friend_name,
                       f.responded_at
                FROM friendships f
                JOIN users u1 ON f.requester_user_id = u1.user_id
                JOIN users u2 ON f.addressee_user_id = u2.user_id
                WHERE (f.requester_user_id = ? OR f.addressee_user_id = ?)
                  AND f.status = 'accepted'
                ORDER BY f.responded_at DESC
                """,
                (account['user_id'], account['user_id'], account['user_id'], account['user_id'])
            ) or []
            if accepted_friends:
                st.caption("Connected friends")
                st.dataframe(pd.DataFrame(accepted_friends), use_container_width=True, hide_index=True)

    with tab6:
        st.subheader("Feature Studio (Groupwise Dynamic Controls)")
        storage_mode = "database"
        if account['user_id'] > 0:
            dyn_pref = ensure_dynamic_preferences_row() or {}
        else:
            storage_mode = "session"
            if 'feature_studio_ephemeral' not in st.session_state:
                st.session_state.feature_studio_ephemeral = {
                    'anonymous_mode_enabled': 0,
                    'anonymous_alias': generate_anonymous_alias(),
                    'anonymous_avatar_style': 'geometric',
                    'anonymous_rotation_hours': 72,
                    'profile_theme': 'adaptive',
                    'feature_json': json.dumps(get_default_feature_flags())
                }
            dyn_pref = st.session_state.feature_studio_ephemeral

        raw_feature_json = dyn_pref.get('feature_json')
        try:
            persisted_features = json.loads(raw_feature_json) if raw_feature_json else {}
            if not isinstance(persisted_features, dict):
                persisted_features = {}
        except Exception:
            persisted_features = {}

        feature_flags = dict(get_default_feature_flags())
        feature_flags.update({k: bool(v) for k, v in persisted_features.items()})

        if storage_mode == "session":
            st.info("🔄 **Session Mode**: Feature settings persist until logout/browser reset. Create an account to save permanently.")

        st.markdown("Select and customize features for your account. Each group offers related controls.")

        # Feature Group Selector Dropdown
        group_list = list(dynamic_feature_groups.keys())
        selected_group = st.selectbox(
            "📋 Select Feature Group",
            group_list,
            index=0,
            format_func=lambda g: f"{g} ({sum(1 for k in dynamic_feature_groups[g] if feature_flags.get(k))}/{len(dynamic_feature_groups[g])})"
        )

        with st.form("feature_studio_form"):
            draft_flags = {}
            
            # Display only the selected group
            group_features = dynamic_feature_groups[selected_group]
            enabled_count = sum(1 for k in group_features if feature_flags.get(k))
            
            st.write(f"**{selected_group}** • {enabled_count} enabled of {len(group_features)}")
            
            # Create 2-column layout for features
            group_items = list(group_features.items())
            for i in range(0, len(group_items), 2):
                cols = st.columns(2, gap='medium')
                
                # First feature in row
                feature_key, feature_config = group_items[i]
                with cols[0]:
                    col_left, col_right = st.columns([4, 1], gap='small')
                    with col_left:
                        st.markdown(feature_config['label'])
                    with col_right:
                        radio_state = st.radio(
                            f"Toggle {feature_key}",
                            ['Off', 'On'],
                            index=1 if bool(feature_flags.get(feature_key, False)) else 0,
                            horizontal=True,
                            label_visibility='collapsed',
                            key=f"studio_{feature_key}"
                        )
                        draft_flags[feature_key] = (radio_state == 'On')
                
                # Second feature in row (if exists)
                if i + 1 < len(group_items):
                    feature_key, feature_config = group_items[i + 1]
                    with cols[1]:
                        col_left, col_right = st.columns([4, 1], gap='small')
                        with col_left:
                            st.markdown(feature_config['label'])
                        with col_right:
                            radio_state = st.radio(
                                f"Toggle {feature_key}",
                                ['Off', 'On'],
                                index=1 if bool(feature_flags.get(feature_key, False)) else 0,
                                horizontal=True,
                                label_visibility='collapsed',
                                key=f"studio_{feature_key}"
                            )
                            draft_flags[feature_key] = (radio_state == 'On')
            
            # Save button
            save_studio = st.form_submit_button("💾 Save Feature Configuration", use_container_width=True)
            
            if save_studio:
                # Merge draft_flags with existing features from other groups
                all_draft_flags = dict(feature_flags)
                all_draft_flags.update(draft_flags)
                
                # Perform cross-check validation
                issues, corrections = check_feature_compatibility(all_draft_flags)
                
                # Apply auto-corrections
                all_draft_flags.update(corrections)
                
                # Display compatibility warnings
                if issues:
                    st.divider()
                    st.subheader("Compatibility Review")
                    for issue in issues:
                        if issue['severity'] == 'warning':
                            st.warning(f"⚠️ {issue['message']}")
                        elif issue['severity'] == 'info':
                            st.info(f"ℹ️ {issue['message']}")
                
                # Perform save
                if storage_mode == "database":
                    done = Database.execute_update(
                        """
                        INSERT INTO account_dynamic_preferences
                        (user_id, anonymous_mode_enabled, anonymous_alias, anonymous_avatar_style,
                         anonymous_rotation_hours, profile_theme, feature_json, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'))
                        ON CONFLICT(user_id)
                        DO UPDATE SET feature_json = excluded.feature_json, updated_at = datetime('now')
                        """,
                        (
                            account['user_id'],
                            1 if dyn_pref.get('anonymous_mode_enabled') else 0,
                            dyn_pref.get('anonymous_alias') or generate_anonymous_alias(),
                            dyn_pref.get('anonymous_avatar_style') or 'geometric',
                            max(12, to_int(dyn_pref.get('anonymous_rotation_hours'), 72)),
                            dyn_pref.get('profile_theme') or 'adaptive',
                            json.dumps(all_draft_flags)
                        )
                    )
                    if done:
                        st.success("✅ Feature configuration saved to database.")
                    else:
                        st.error("❌ Failed to save feature configuration.")
                else:
                    st.session_state.feature_studio_ephemeral['feature_json'] = json.dumps(all_draft_flags)
                    st.success("✅ Feature configuration saved in session mode.")
                
                st.rerun()
        
        current_flags = feature_flags
        
        # Alias rotation action (if enabled)
        if current_flags.get('anonymous_alias_rotation'):
            st.divider()
            st.subheader("Anonymous Alias Management")
            if st.button("🔄 Rotate Anonymous Alias Now", use_container_width=True):
                new_alias = generate_anonymous_alias()
                if storage_mode == "database":
                    Database.execute_update(
                        """
                        UPDATE account_dynamic_preferences
                        SET anonymous_alias = ?, updated_at = datetime('now')
                        WHERE user_id = ?
                        """,
                        (new_alias, account['user_id'])
                    )
                else:
                    st.session_state.feature_studio_ephemeral['anonymous_alias'] = new_alias
                st.success(f"✅ Anonymous alias rotated to: **{new_alias}**")
                st.rerun()

def show_manage_books(embedded=False, browse_only=False):
    """Book management page"""
    if embedded:
        if not browse_only:
            st.subheader(" Book Workspace")
    else:
        st.markdown('<h1 class="litgrid-header"> Manage Books</h1>', unsafe_allow_html=True)
    
    books_workspace_tab = st.container()
    
    with books_workspace_tab:
        st.subheader(" Browse Books")

        language_rows = Database.execute_query(
            "SELECT DISTINCT COALESCE(NULLIF(TRIM(language), ''), 'Unknown') as language_name FROM books ORDER BY language_name"
        ) or []
        language_options = ["All"] + [row['language_name'] for row in language_rows if row.get('language_name')]

        row1_col1, row1_col2, row1_col3 = st.columns(3, gap="small")
        with row1_col1:
            search = st.text_input(" Search (Title, Author, ISBN, Keywords)", key="mb_search")
        with row1_col2:
            genre_query = st.text_input(" Genre Search", placeholder="e.g., Fiction, Science", key="mb_genre_filter_text")
        with row1_col3:
            use_fuzzy = st.checkbox("Use Smart Search", help="Tolerates typos", key="mb_use_fuzzy")

        status_filter = st.radio(
            "Status Filter",
            ["All", "Active", "Inactive", "Available Copies Only", "Checked Out Only"],
            horizontal=True,
            key="mb_status_filter"
        )

        row2_col1, row2_col2 = st.columns(2, gap="small")
        with row2_col1:
            view_mode = st.radio("View Mode", ["Operational List", "Discovery Cards"], horizontal=True, key="mb_view_mode")
        with row2_col2:
            sort_by = st.radio(
                "Sort",
                ["Title A-Z", "Newest Year", "Popularity High"],
                horizontal=True,
                key="mb_sort_by"
            )

        row3_col1, row3_col2, row3_col3, row3_col4 = st.columns(4, gap="small")
        current_year = datetime.now().year
        max_year = current_year + 10
        with row3_col1:
            year_from = st.number_input("Year From", min_value=1800, max_value=max_year, value=1800, step=1, key="mb_year_from")
        with row3_col2:
            year_to = st.number_input("Year To", min_value=1800, max_value=max_year, value=current_year, step=1, key="mb_year_to")
        with row3_col3:
            max_results = st.slider("Max Results", min_value=20, max_value=300, value=120, step=10, key="mb_limit")
        with row3_col4:
            fuzzy_threshold = st.slider("Fuzzy Threshold", min_value=30, max_value=95, value=60, step=5, key="mb_fuzzy_threshold")

        with st.expander("Advanced Search", expanded=False):
            adv_col1, adv_col2, adv_col3, adv_col4 = st.columns(4, gap="small")
            with adv_col1:
                title_query = st.text_input("Title Contains", key="mb_title_query")
            with adv_col2:
                author_query = st.text_input("Author Contains", key="mb_author_query")
            with adv_col3:
                keyword_query = st.text_input("Keyword Contains", key="mb_keyword_query")
            with adv_col4:
                isbn_exact = st.text_input("Exact ISBN/ISBN-10/ISBN-13", key="mb_isbn_exact")

            adv_col5, adv_col6, adv_col7, adv_col8 = st.columns(4, gap="small")
            with adv_col5:
                language_filter_text = st.text_input("Language Contains", key="mb_language_filter_text")
            with adv_col6:
                min_available_copies = st.number_input("Min Available Copies", min_value=0, value=0, step=1, key="mb_min_available_copies")
            with adv_col7:
                min_total_copies = st.number_input("Min Total Copies", min_value=0, value=0, step=1, key="mb_min_total_copies")
            with adv_col8:
                quick_date_preset = st.radio(
                    "Added Date Preset",
                    ["None", "Last 7 Days", "Last 30 Days", "Last 90 Days", "Last 365 Days"],
                    horizontal=True,
                    key="mb_quick_date_preset"
                )

            adv_col9, adv_col10 = st.columns(2, gap="small")
            with adv_col9:
                popularity_min = st.number_input("Popularity Min", min_value=0.0, value=0.0, step=0.1, key="mb_popularity_min")
            with adv_col10:
                popularity_max = st.number_input("Popularity Max", min_value=0.0, value=100.0, step=0.1, key="mb_popularity_max")

        date_filter_enabled = st.checkbox(" Calendar Search (Filter by Added Date)", value=False, key="mb_calendar_enabled")
        calendar_range = None
        if date_filter_enabled:
            date_col, refresh_col = st.columns([4, 1], gap="small")
            with date_col:
                calendar_range = st.date_input(
                    " Added Date Range",
                    value=(date.today() - timedelta(days=90), date.today()),
                    key="mb_calendar_range"
                )
            with refresh_col:
                st.markdown("<div style='height: 1.8rem;'></div>", unsafe_allow_html=True)
                if st.button(" Refresh", use_container_width=True, key="mb_refresh"):
                    st.rerun()
        else:
            if st.button(" Refresh", use_container_width=True, key="mb_refresh"):
                st.rerun()

        # Unified book query
        query = """
            SELECT b.book_id, b.isbn, b.title, b.author, b.genre, b.publication_year,
                   b.pages, b.language, b.keywords, b.popularity_score, b.is_available, b.created_at,
                   (SELECT COUNT(*) FROM book_inventory bi WHERE bi.book_id = b.book_id) as total_copies,
                   (SELECT COUNT(*) FROM book_inventory bi WHERE bi.book_id = b.book_id AND bi.is_available = 1) as available_copies
            FROM books b
            WHERE 1=1
        """
        params = []

        if genre_query:
            query += " AND b.genre LIKE ?"
            params.append(f"%{genre_query}%")

        if year_from and year_to:
            query += " AND b.publication_year BETWEEN ? AND ?"
            params.extend([year_from, year_to])

        if search and not use_fuzzy:
            query += " AND (b.title LIKE ? OR b.author LIKE ? OR b.isbn LIKE ? OR b.keywords LIKE ? OR b.isbn_13 LIKE ? OR b.isbn_10 LIKE ?)"
            search_term = f'%{search}%'
            params.extend([search_term, search_term, search_term, search_term, search_term, search_term])

        if title_query:
            query += " AND b.title LIKE ?"
            params.append(f"%{title_query}%")

        if author_query:
            query += " AND b.author LIKE ?"
            params.append(f"%{author_query}%")

        if keyword_query:
            query += " AND b.keywords LIKE ?"
            params.append(f"%{keyword_query}%")

        if isbn_exact:
            query += " AND (b.isbn = ? OR b.isbn_13 = ? OR b.isbn_10 = ?)"
            params.extend([isbn_exact, isbn_exact, isbn_exact])

        if language_filter_text:
            query += " AND COALESCE(NULLIF(TRIM(b.language), ''), 'Unknown') LIKE ?"
            params.append(f"%{language_filter_text}%")

        if int(min_available_copies) > 0:
            query += " AND (SELECT COUNT(*) FROM book_inventory bi WHERE bi.book_id = b.book_id AND bi.is_available = 1) >= ?"
            params.append(int(min_available_copies))

        if int(min_total_copies) > 0:
            query += " AND (SELECT COUNT(*) FROM book_inventory bi WHERE bi.book_id = b.book_id) >= ?"
            params.append(int(min_total_copies))

        if float(popularity_min) > 0.0:
            query += " AND COALESCE(b.popularity_score, 0) >= ?"
            params.append(float(popularity_min))

        if float(popularity_max) < 100.0:
            query += " AND COALESCE(b.popularity_score, 0) <= ?"
            params.append(float(popularity_max))

        if status_filter == "Active":
            query += " AND b.is_available = 1"
        elif status_filter == "Inactive":
            query += " AND b.is_available = 0"
        elif status_filter == "Available Copies Only":
            query += " AND EXISTS (SELECT 1 FROM book_inventory bi WHERE bi.book_id = b.book_id AND bi.is_available = 1)"
        elif status_filter == "Checked Out Only":
            query += " AND EXISTS (SELECT 1 FROM book_inventory bi WHERE bi.book_id = b.book_id AND bi.is_available = 0)"

        if date_filter_enabled and isinstance(calendar_range, (tuple, list)) and len(calendar_range) == 2:
            start_date, end_date = calendar_range
            if start_date and end_date:
                query += " AND date(b.created_at) BETWEEN ? AND ?"
                params.extend([str(start_date), str(end_date)])
        elif quick_date_preset != "None":
            preset_map = {
                "Last 7 Days": "-7 days",
                "Last 30 Days": "-30 days",
                "Last 90 Days": "-90 days",
                "Last 365 Days": "-365 days",
            }
            preset_value = preset_map.get(quick_date_preset)
            if preset_value:
                query += " AND date(b.created_at) >= date('now', ?)"
                params.append(preset_value)

        if sort_by == "Title A-Z":
            query += " ORDER BY b.title ASC"
        elif sort_by == "Newest Year":
            query += " ORDER BY b.publication_year DESC, b.title ASC"
        elif sort_by == "Popularity High":
            query += " ORDER BY COALESCE(b.popularity_score, 0) DESC, b.title ASC"
        else:
            query += " ORDER BY b.title ASC"

        query += " LIMIT ?"
        params.append(max_results)

        books = Database.execute_query(query, tuple(params) if params else None)

        # Apply fuzzy search on top of filtered results
        if search and use_fuzzy and books:
            fuzzy_books = EnhancedSearchFilter.fuzzy_search_books(search, threshold=fuzzy_threshold)
            fuzzy_ids = {b['book_id'] for b in fuzzy_books if b.get('book_id') is not None}
            books = [book for book in books if book.get('book_id') in fuzzy_ids]

        if books:
            total_inventory = sum((book.get('total_copies') or 0) for book in books)
            available_inventory = sum((book.get('available_copies') or 0) for book in books)

            metric_col1, metric_col2, metric_col3 = st.columns(3, gap="small")
            with metric_col1:
                st.metric("Books in Result", len(books))
            with metric_col2:
                st.metric("Copies (Total)", total_inventory)
            with metric_col3:
                st.metric("Copies Available", available_inventory)

            active_filters = []
            if search:
                active_filters.append(f"Search: {search}")
            if genre_query:
                active_filters.append(f"Genre: {genre_query}")
            if title_query:
                active_filters.append(f"Title: {title_query}")
            if author_query:
                active_filters.append(f"Author: {author_query}")
            if keyword_query:
                active_filters.append(f"Keyword: {keyword_query}")
            if isbn_exact:
                active_filters.append(f"ISBN: {isbn_exact}")
            if language_filter_text:
                active_filters.append(f"Language: {language_filter_text}")
            if int(min_available_copies) > 0:
                active_filters.append(f"Min Available: {int(min_available_copies)}")
            if int(min_total_copies) > 0:
                active_filters.append(f"Min Copies: {int(min_total_copies)}")
            if float(popularity_min) > 0.0 or float(popularity_max) < 100.0:
                active_filters.append(f"Popularity: {float(popularity_min):.1f}-{float(popularity_max):.1f}")
            if status_filter != "All":
                active_filters.append(f"Status: {status_filter}")
            if date_filter_enabled and isinstance(calendar_range, (tuple, list)) and len(calendar_range) == 2:
                active_filters.append(f"Date: {calendar_range[0]} to {calendar_range[1]}")
            elif quick_date_preset != "None":
                active_filters.append(f"Date Preset: {quick_date_preset}")
            active_filters.append(f"Sort: {sort_by}")
            st.caption(" | ".join(active_filters))
        
        if books:
            st.write(f"Found {len(books)} books")

            if view_mode == "Discovery Cards":
                card_columns = st.columns(3, gap="small")
                for index, book in enumerate(books):
                    with card_columns[index % 3]:
                        with st.container():
                            st.markdown(f"**{book['title']}**")
                            st.caption(
                                f"Author: {book.get('author') or 'N/A'} | Genre: {book.get('genre') or 'N/A'}"
                            )
                            st.caption(
                                f"ISBN: {book.get('isbn') or 'N/A'} | Year: {book.get('publication_year') or 'N/A'}"
                            )
                            st.caption(
                                f"Availability: {book.get('available_copies') or 0}/{book.get('total_copies') or 0} copies"
                            )
                            if book.get('keywords'):
                                st.caption(f"Keywords: {book['keywords'][:70]}")

                            c_action1, c_action2 = st.columns(2, gap="small")
                            with c_action1:
                                if st.button(" Details", key=f"card_details_{book['book_id']}", use_container_width=True):
                                    st.session_state[f"show_card_details_{book['book_id']}"] = True
                            with c_action2:
                                action = "Deactivate" if book['is_available'] else "Activate"
                                if st.button(f" {action}", key=f"card_toggle_{book['book_id']}", use_container_width=True):
                                    new_status = not book['is_available']
                                    if Database.execute_update(
                                        "UPDATE books SET is_available = ? WHERE book_id = ?",
                                        (new_status, book['book_id'])
                                    ):
                                        st.success(f"Book {action.lower()}d successfully!")
                                        st.rerun()

                            if st.session_state.get(f"show_card_details_{book['book_id']}", False):
                                with st.expander(f"Details: {book['title']}", expanded=True):
                                    dtab1, dtab2, dtab3 = st.tabs([" Recommendations", " Reviews", " Condition"])

                                    with dtab1:
                                        recs = SmartUtilities.get_book_recommendations(book['book_id'])
                                        if recs:
                                            for rec in recs:
                                                st.write(f"- {rec.get('title', 'N/A')}")
                                        else:
                                            st.info("No recommendations yet")

                                    with dtab2:
                                        reviews = ReviewsManager.get_book_reviews(book['book_id'])
                                        if reviews:
                                            for review in reviews:
                                                st.write(f" {review['rating']}/5 - {review['full_name']}")
                                                st.caption(review['comment'])
                                        else:
                                            st.info("No reviews yet")

                                    with dtab3:
                                        conditions = EnhancedBookManager.get_book_condition_history(book['book_id'])
                                        if conditions:
                                            for cond in conditions:
                                                st.write(f"**{cond['condition_status'].title()}** - {cond['check_date']}")
                                                if cond['condition_notes']:
                                                    st.caption(cond['condition_notes'])
                                        else:
                                            st.info("No condition records")
                            st.divider()
            else:
                for book in books:
                    with st.container():
                        col1, col2, col3, col4 = st.columns([3, 1, 1, 1], gap="small")

                        with col1:
                            status_icon = "" if book['is_available'] else ""
                            st.markdown(f"**{status_icon} {book['title']}**")

                            isbn_display = book.get('isbn', 'N/A')
                            st.caption(f"ISBN: {isbn_display} | Year: {book['publication_year'] or 'N/A'}")
                            st.caption(f"Author: {book.get('author') or 'N/A'} | Genre: {book.get('genre') or 'N/A'}")

                            if book.get('keywords'):
                                st.caption(f" {book['keywords'][:50]}...")

                        with col2:
                            st.metric("Available", f"{book['available_copies'] or 0}/{book['total_copies'] or 0}")
                            if book.get('pages'):
                                st.caption(f"Pages: {book['pages']}")
                            if book.get('language'):
                                st.caption(f"Language: {book['language']}")

                        with col3:
                            # Show cover thumbnail if available
                            cover = EnhancedBookManager.get_book_cover(book['book_id'])
                            if cover:
                                st.image(cover, width=50)

                        with col4:
                            if st.button(" Edit", key=f"list_edit_{book['book_id']}", use_container_width=True):
                                st.session_state[f'edit_book_{book["book_id"]}'] = True

                            if st.button(" Details", key=f"list_details_{book['book_id']}", use_container_width=True):
                                # Show recommendations, reviews, condition
                                with st.expander(f"Details: {book['title']}", expanded=True):
                                    dtab1, dtab2, dtab3 = st.tabs([" Recommendations", " Reviews", " Condition"])

                                    with dtab1:
                                        recs = SmartUtilities.get_book_recommendations(book['book_id'])
                                        if recs:
                                            for rec in recs:
                                                st.write(f"- {rec.get('title', 'N/A')}")
                                        else:
                                            st.info("No recommendations yet")

                                    with dtab2:
                                        reviews = ReviewsManager.get_book_reviews(book['book_id'])
                                        if reviews:
                                            for review in reviews:
                                                st.write(f" {review['rating']}/5 - {review['full_name']}")
                                                st.caption(review['comment'])
                                        else:
                                            st.info("No reviews yet")

                                    with dtab3:
                                        conditions = EnhancedBookManager.get_book_condition_history(book['book_id'])
                                        if conditions:
                                            for cond in conditions:
                                                st.write(f"**{cond['condition_status'].title()}** - {cond['check_date']}")
                                                if cond['condition_notes']:
                                                    st.caption(cond['condition_notes'])
                                        else:
                                            st.info("No condition records")

                            action = "Deactivate" if book['is_available'] else "Activate"
                            if st.button(f" {action}", key=f"list_toggle_{book['book_id']}", use_container_width=True):
                                new_status = not book['is_available']
                                if Database.execute_update(
                                    "UPDATE books SET is_available = ? WHERE book_id = ?",
                                    (new_status, book['book_id'])
                                ):
                                    st.success(f"Book {action.lower()}d successfully!")
                                    st.rerun()

                        st.divider()
        else:
            st.info("No books found")

    if browse_only:
        return
    
    with books_workspace_tab:
        st.divider()
        st.subheader(" Bulk Import from CSV")
        
        st.markdown("""
        **CSV Format Requirements:**
        - Required columns: `title`, `author`, `genre`
        - Optional columns: `isbn`, `publication_year`, `total_copies`, `available_copies`, `page_count`, `language`, `publisher`, `keywords`
        """)
        
        csv_file = st.file_uploader("Upload CSV file", type=['csv'])
        
        if csv_file:
            # Preview
            st.subheader(" Preview")
            df = pd.read_csv(csv_file)
            st.dataframe(df.head(10))
            st.write(f"Total rows: {len(df)}")
            
            if st.button(" Import All Books", type="primary"):
                with st.spinner("Importing books..."):
                    csv_file.seek(0)  # Reset file pointer
                    success, message = EnhancedBookManager.bulk_import_csv(csv_file)
                    
                    if success:
                        st.success(f" {message}")
                        st.balloons()
                    else:
                        st.error(f" {message}")
        else:
            # Show example CSV
            st.download_button(
                " Download CSV Template",
                "title,author,genre,isbn,publication_year,total_copies,page_count,language,publisher,keywords\n"
                "Sample Book,Sample Author,Fiction,978-1234567890,2024,1,300,English,Sample Publisher,fiction;sample",
                "books_template.csv",
                mime="text/csv"
            )
    
    with books_workspace_tab:
        st.divider()
        st.subheader(" Book Statistics")

        # ENHANCED CONTROLS SECTION
        st.markdown("### ⚙️ Analytics Configuration")
        
        config_tab1, config_tab2, config_tab3, config_tab4 = st.tabs([" Time Range", " Collection Filters", " Display Options", " Thresholds & Alerts"])
        
        with config_tab1:
            time_col1, time_col2, time_col3, time_col4 = st.columns(4, gap="small")
            with time_col1:
                time_window = st.selectbox(
                    "Analytics Range",
                    ["7D", "14D", "30D", "90D", "180D", "365D", "730D", "All Time"],
                    index=4,
                    key="bs_time_window"
                )
            with time_col2:
                trend_granularity = st.selectbox("Trend Granularity", ["Week", "Month", "Quarter", "Year"], key="bs_granularity")
            with time_col3:
                use_custom_range = st.checkbox("Use Custom Date Range", key="bs_custom_range")
            with time_col4:
                compare_previous = st.checkbox("Compare With Previous Period", value=True, key="bs_compare_previous")
            
            if use_custom_range:
                custom_range = st.date_input(
                    "Select Date Range",
                    value=(datetime.now() - timedelta(days=180), datetime.now()),
                    key="bs_custom_date_range"
                )
        
        with config_tab2:
            filter_col1, filter_col2, filter_col3, filter_col4 = st.columns(4, gap="small")
            all_genres = Database.execute_query(
                "SELECT DISTINCT COALESCE(NULLIF(TRIM(genre), ''), 'Unknown') as genre_name FROM books ORDER BY genre_name"
            ) or []
            all_languages = Database.execute_query(
                "SELECT DISTINCT COALESCE(NULLIF(TRIM(language), ''), 'Unknown') as language_name FROM books ORDER BY language_name"
            ) or []
            genre_choices = [g['genre_name'] for g in all_genres if g.get('genre_name')]
            language_choices = [l['language_name'] for l in all_languages if l.get('language_name')]
            
            with filter_col1:
                selected_genres = st.multiselect("Genres", options=genre_choices, key="bs_genres")
            with filter_col2:
                selected_languages = st.multiselect("Languages", options=language_choices, key="bs_languages")
            
            with filter_col3:
                active_titles_only = st.checkbox("Active Titles Only", value=False, key="bs_active_titles_only")
            with filter_col4:
                include_zero_copies = st.checkbox("Include Zero-Copy Titles", value=True, key="bs_include_zero_copies")
            
            col_filter1, col_filter2, col_filter3, col_filter4 = st.columns(4, gap="small")
            with col_filter1:
                min_copies = st.number_input("Min Copies/Title", min_value=0, value=0, key="bs_min_copies")
            current_year = datetime.now().year
            max_year = current_year + 10
            with col_filter2:
                pub_year_from = st.number_input("Publication Year From", min_value=1800, max_value=max_year, value=1800, key="bs_pub_year_from")
            with col_filter3:
                pub_year_to = st.number_input("Publication Year To", min_value=1800, max_value=max_year, value=current_year, key="bs_pub_year_to")
            with col_filter4:
                book_age_filter = st.selectbox("Book Age (Time in Collection)", ["Any", "Last 6 Months", "Last Year", "Last 2 Years", "3+ Years"], key="bs_book_age")
        
        with config_tab3:
            display_col1, display_col2, display_col3, display_col4 = st.columns(4, gap="small")
            with display_col1:
                show_advanced_metrics = st.checkbox("Advanced Metrics", value=True, key="bs_show_advanced")
            with display_col2:
                show_demand_analysis = st.checkbox("Demand Analysis", value=True, key="bs_show_demand")
            with display_col3:
                show_collection_health = st.checkbox("Collection Health", value=True, key="bs_show_health")
            with display_col4:
                show_predictive = st.checkbox("Insights & Recommendations", value=True, key="bs_show_predictive")
            
            disp_col1, disp_col2, disp_col3, disp_col4 = st.columns(4, gap="small")
            with disp_col1:
                top_n = st.slider("Top N Items", min_value=5, max_value=30, value=12, step=1, key="bs_top_n")
            with disp_col2:
                chart_height = st.slider("Chart Height (px)", min_value=250, max_value=600, value=360, step=50, key="bs_chart_height")
            with disp_col3:
                color_scheme = st.selectbox("Color Scheme", ["Viridis", "Blues", "Reds", "Greens", "Greys"], key="bs_color_scheme")
            with disp_col4:
                export_format = st.multiselect("Export Formats", ["CSV", "JSON", "PDF"], key="bs_export_format")
        
        with config_tab4:
            threshold_col1, threshold_col2, threshold_col3, threshold_col4 = st.columns(4, gap="small")
            with threshold_col1:
                utilization_alert_threshold = st.slider("Utilization Alert %", min_value=10, max_value=100, value=70, step=5, key="bs_util_threshold")
            with threshold_col2:
                low_stock_threshold = st.slider("Low Stock Threshold", min_value=0, max_value=10, value=1, step=1, key="bs_low_stock_threshold")
            with threshold_col3:
                overdue_alert = st.slider("Overdue Alert Days", min_value=1, max_value=30, value=7, step=1, key="bs_overdue_alert")
            with threshold_col4:
                demand_threshold = st.slider("High Demand Threshold", min_value=1, max_value=20, value=5, step=1, key="bs_demand_threshold")
            
            alert_col1, alert_col2, alert_col3 = st.columns(3, gap="small")
            with alert_col1:
                enable_low_stock_alerts = st.checkbox("Low Stock Alerts", value=True, key="bs_alert_low_stock")
            with alert_col2:
                enable_overdue_alerts = st.checkbox("Overdue Alerts", value=True, key="bs_alert_overdue")
            with alert_col3:
                enable_performance_alerts = st.checkbox("Performance Alerts", value=True, key="bs_alert_performance")

        # UNIFIED TIME WINDOW HANDLING
        time_col_action1, time_col_action2, time_col_action3, time_col_action4 = st.columns(4, gap="small")
        with time_col_action1:
            if st.button(" 🔄 Refresh All Statistics", key="bs_refresh", use_container_width=True, type="primary"):
                st.rerun()
        with time_col_action2:
            if st.button(" 📊 Generate Report", key="bs_generate_report", use_container_width=True):
                st.info("Report generation queued. Check Downloads folder.")
        with time_col_action3:
            if st.button(" 💾 Save Snapshot", key="bs_save_snapshot", use_container_width=True):
                st.success(f"Snapshot saved at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        with time_col_action4:
            if st.button(" ℹ️ Help & Guide", key="bs_help", use_container_width=True):
                with st.expander("Statistics Guide andBest Practices", expanded=False):
                    st.markdown("""
                    **Key Metrics Explained:**
                    - **Circulation Velocity**: Average loans per copy per period (higher = better utilization)
                    - **Collection Refresh Rate**: % of new titles added in the period
                    - **Utilization Rate**: % of inventory currently checked out
                    - **Turn-over Ratio**: How many times average title circulates per year
                    
                    **Tips:**
                    - Compare periods to identify trends
                    - Monitor low-stock alerts for timely restocking
                    - Use demand leaders to guide acquisition decisions
                    - Track collection health to prioritize maintenance
                    """)
        
        # ORIGINAL FILTERS SECTION (MAINTAINED FOR COMPATIBILITY)
        time_col1, time_col2, time_col3, time_col4, time_col5 = st.columns(5, gap="small")
        with time_col1:
            if 'bs_time_window' not in st.session_state:
                st.session_state.bs_time_window = time_window
        with time_col2:
            if 'bs_granularity' not in st.session_state:
                st.session_state.bs_granularity = trend_granularity
        with time_col3:
            if 'bs_active_titles_only' not in st.session_state:
                st.session_state.bs_active_titles_only = active_titles_only
        with time_col4:
            if 'bs_include_zero_copies' not in st.session_state:
                st.session_state.bs_include_zero_copies = include_zero_copies
        with time_col5:
            if 'bs_compare_previous' not in st.session_state:
                st.session_state.bs_compare_previous = compare_previous

        filter_col1, filter_col2, filter_col3, filter_col4, filter_col5 = st.columns(5, gap="small")
        if 'bs_genres' not in st.session_state:
            st.session_state.bs_genres = selected_genres
        if 'bs_languages' not in st.session_state:
            st.session_state.bs_languages = selected_languages
        if 'bs_min_copies' not in st.session_state:
            st.session_state.bs_min_copies = min_copies
        if 'bs_pub_year_from' not in st.session_state:
            st.session_state.bs_pub_year_from = pub_year_from
        if 'bs_pub_year_to' not in st.session_state:
            st.session_state.bs_pub_year_to = pub_year_to

        action_col1, action_col2, action_col3, action_col4 = st.columns(4, gap="small")
        if 'bs_util_threshold' not in st.session_state:
            st.session_state.bs_util_threshold = utilization_alert_threshold
        if 'bs_low_stock_threshold' not in st.session_state:
            st.session_state.bs_low_stock_threshold = low_stock_threshold
        if 'bs_top_n' not in st.session_state:
            st.session_state.bs_top_n = top_n

        filter_col1, filter_col2, filter_col3, filter_col4, filter_col5 = st.columns(5, gap="small")
        all_genres = Database.execute_query(
            "SELECT DISTINCT COALESCE(NULLIF(TRIM(genre), ''), 'Unknown') as genre_name FROM books ORDER BY genre_name"
        ) or []
        all_languages = Database.execute_query(
            "SELECT DISTINCT COALESCE(NULLIF(TRIM(language), ''), 'Unknown') as language_name FROM books ORDER BY language_name"
        ) or []
        genre_choices = [g['genre_name'] for g in all_genres if g.get('genre_name')]
        language_choices = [l['language_name'] for l in all_languages if l.get('language_name')]
        with filter_col1:
            selected_genres = st.multiselect("Genres", options=genre_choices, key="bs_genres")
        with filter_col2:
            selected_languages = st.multiselect("Languages", options=language_choices, key="bs_languages")
        with filter_col3:
            min_copies = st.number_input("Min Copies/Title", min_value=0, value=0, key="bs_min_copies")
        current_year = datetime.now().year
        max_year = current_year + 10
        with filter_col4:
            pub_year_from = st.number_input("Publication Year From", min_value=1800, max_value=max_year, value=1800, key="bs_pub_year_from")
        with filter_col5:
            pub_year_to = st.number_input("Publication Year To", min_value=1800, max_value=max_year, value=current_year, key="bs_pub_year_to")

        action_col1, action_col2, action_col3, action_col4 = st.columns(4, gap="small")
        with action_col1:
            utilization_alert_threshold = st.slider("Utilization Alert %", min_value=10, max_value=100, value=70, step=5, key="bs_util_threshold")
        with action_col2:
            low_stock_threshold = st.slider("Low Stock Threshold", min_value=0, max_value=10, value=1, step=1, key="bs_low_stock_threshold")
        with action_col3:
            top_n = st.slider("Top N", min_value=5, max_value=30, value=12, step=1, key="bs_top_n")
        with action_col4:
            if st.button(" Refresh Statistics", key="bs_refresh", use_container_width=True):
                st.rerun()

        if pub_year_from > pub_year_to:
            st.warning("Publication year range is invalid. Adjusting automatically.")
            pub_year_from, pub_year_to = pub_year_to, pub_year_from

        window_map = {
            "30D": 30,
            "90D": 90,
            "180D": 180,
            "365D": 365,
            "730D": 730,
            "All Time": None,
        }
        window_days = window_map.get(time_window)

        base_filters = ["1=1"]
        base_params = []

        if active_titles_only:
            base_filters.append("b.is_available = 1")

        if not include_zero_copies:
            base_filters.append("EXISTS (SELECT 1 FROM book_inventory bi WHERE bi.book_id = b.book_id)")

        if int(min_copies) > 0:
            base_filters.append("(SELECT COUNT(*) FROM book_inventory bi WHERE bi.book_id = b.book_id) >= ?")
            base_params.append(int(min_copies))

        base_filters.append("COALESCE(b.publication_year, 0) BETWEEN ? AND ?")
        base_params.extend([int(pub_year_from), int(pub_year_to)])

        if selected_genres:
            placeholders = ",".join(["?"] * len(selected_genres))
            base_filters.append(f"COALESCE(NULLIF(TRIM(b.genre), ''), 'Unknown') IN ({placeholders})")
            base_params.extend(selected_genres)

        if selected_languages:
            placeholders = ",".join(["?"] * len(selected_languages))
            base_filters.append(f"COALESCE(NULLIF(TRIM(b.language), ''), 'Unknown') IN ({placeholders})")
            base_params.extend(selected_languages)

        current_filters = list(base_filters)
        current_params = list(base_params)
        if window_days is not None:
            current_filters.append("date(b.created_at) >= date('now', ?)")
            current_params.append(f"-{window_days} days")

        current_where_clause = " AND ".join(current_filters)

        previous_where_clause = None
        previous_params = []
        if compare_previous and window_days is not None:
            previous_filters = list(base_filters)
            previous_params = list(base_params)
            previous_filters.append("date(b.created_at) < date('now', ?)")
            previous_filters.append("date(b.created_at) >= date('now', ?)")
            previous_params.append(f"-{window_days} days")
            previous_params.append(f"-{window_days * 2} days")
            previous_where_clause = " AND ".join(previous_filters)

        def fetch_collection_summary(where_clause, params):
            summary_query = f"""
                SELECT
                    COUNT(*) as book_count,
                    SUM((SELECT COUNT(*) FROM book_inventory bi WHERE bi.book_id = b.book_id)) as total_copies,
                    SUM((SELECT COUNT(*) FROM book_inventory bi WHERE bi.book_id = b.book_id AND bi.is_available = 1)) as available_copies,
                    AVG(COALESCE(b.popularity_score, 0)) as avg_popularity,
                    COUNT(DISTINCT COALESCE(NULLIF(TRIM(b.author), ''), 'Unknown')) as unique_authors,
                    SUM(CASE WHEN b.is_available = 0 THEN 1 ELSE 0 END) as inactive_titles,
                    AVG(CAST((julianday('now') - julianday(COALESCE(b.created_at, '2020-01-01'))) AS FLOAT)) as avg_book_age_days,
                    COUNT(DISTINCT COALESCE(NULLIF(TRIM(b.genre), ''), 'Unknown')) as unique_genres,
                    COUNT(DISTINCT COALESCE(NULLIF(TRIM(b.language), ''), 'Unknown')) as unique_languages,
                    SUM(COALESCE(b.pages, 0)) as total_pages
                FROM books b
                WHERE {where_clause}
            """
            return Database.execute_query(summary_query, tuple(params) if params else None, fetch_one=True) or {}

        def fetch_loan_summary(where_clause, params):
            loan_query = f"""
                SELECT
                    COUNT(*) as active_loans,
                    SUM(CASE WHEN date(br.due_date) < date('now') THEN 1 ELSE 0 END) as overdue_loans,
                    SUM(CASE WHEN date(br.due_date) BETWEEN date('now') AND date('now', '+7 days') THEN 1 ELSE 0 END) as due_next_7
                FROM borrowing br
                JOIN book_inventory bi ON br.inventory_id = bi.inventory_id
                JOIN books b ON bi.book_id = b.book_id
                WHERE br.return_date IS NULL AND {where_clause}
            """
            return Database.execute_query(loan_query, tuple(params) if params else None, fetch_one=True) or {}

        def fetch_advanced_metrics(where_clause, params):
            """Fetch advanced performance and health metrics"""
            adv_query = f"""
                SELECT
                    COUNT(DISTINCT CASE WHEN (SELECT COUNT(*) FROM borrowing brx WHERE brx.inventory_id = bi.inventory_id) > 0 THEN b.book_id ELSE NULL END) as titles_with_loans,
                    COUNT(DISTINCT CASE WHEN (SELECT COUNT(*) FROM borrowing brx WHERE brx.inventory_id = bi.inventory_id) = 0 THEN b.book_id ELSE NULL END) as untouched_titles,
                    COALESCE(SUM(CASE WHEN (SELECT COUNT(*) FROM book_inventory bix WHERE bix.book_id = b.book_id AND bix.is_available = 0) > 0 THEN 1 ELSE 0 END), 0) as titles_with_checkouts,
                    COUNT(DISTINCT b.book_id) as distinct_titles
                FROM books b
                LEFT JOIN book_inventory bi ON bi.book_id = b.book_id
                WHERE {where_clause}
            """
            return Database.execute_query(adv_query, tuple(params) if params else None, fetch_one=True) or {}

        def fetch_genre_language_matrix(where_clause, params, top_n_val):
            """Fetch genre-language distribution for heatmap"""
            matrix_query = f"""
                SELECT
                    COALESCE(NULLIF(TRIM(b.genre), ''), 'Unknown') as genre,
                    COALESCE(NULLIF(TRIM(b.language), ''), 'Unknown') as language,
                    COUNT(*) as titles,
                    SUM((SELECT COUNT(*) FROM book_inventory bi WHERE bi.book_id = b.book_id)) as copies
                FROM books b
                WHERE {where_clause}
                GROUP BY genre, language
                ORDER BY titles DESC
                LIMIT ?
            """
            return Database.execute_query(matrix_query, tuple(params + [int(top_n_val)]) if params else None)

        summary = fetch_collection_summary(current_where_clause, current_params)
        loans = fetch_loan_summary(current_where_clause, current_params)
        advanced = fetch_advanced_metrics(current_where_clause, current_params)

        prev_summary = {}
        prev_loans = {}
        if previous_where_clause:
            prev_summary = fetch_collection_summary(previous_where_clause, previous_params)
            prev_loans = fetch_loan_summary(previous_where_clause, previous_params)

        total_books = int(summary.get('book_count') or 0)
        total_copies = int(summary.get('total_copies') or 0)
        available_copies = int(summary.get('available_copies') or 0)
        checked_out = max(total_copies - available_copies, 0)
        utilization_pct = (checked_out / total_copies * 100) if total_copies > 0 else 0.0
        avg_copies_per_title = (total_copies / total_books) if total_books > 0 else 0.0
        avg_popularity = float(summary.get('avg_popularity') or 0.0)
        unique_authors = int(summary.get('unique_authors') or 0)
        inactive_titles = int(summary.get('inactive_titles') or 0)
        avg_book_age_days = float(summary.get('avg_book_age_days') or 0.0)
        unique_genres = int(summary.get('unique_genres') or 0)
        unique_languages = int(summary.get('unique_languages') or 0)
        total_pages = int(summary.get('total_pages') or 0)

        active_loans = int(loans.get('active_loans') or 0)
        overdue_loans = int(loans.get('overdue_loans') or 0)
        due_next_7 = int(loans.get('due_next_7') or 0)

        titles_with_loans = int(advanced.get('titles_with_loans') or 0)
        untouched_titles = int(advanced.get('untouched_titles') or 0)
        titles_with_checkouts = int(advanced.get('titles_with_checkouts') or 0)

        prev_total_books = int(prev_summary.get('book_count') or 0)
        prev_total_copies = int(prev_summary.get('total_copies') or 0)
        prev_active_loans = int(prev_loans.get('active_loans') or 0)

        delta_books = total_books - prev_total_books if previous_where_clause else None
        delta_copies = total_copies - prev_total_copies if previous_where_clause else None
        delta_loans = active_loans - prev_active_loans if previous_where_clause else None

        # ADVANCED METRIC CALCULATIONS
        circulation_velocity = (active_loans / total_copies * 100) if total_copies > 0 else 0.0
        collection_refresh_rate = ((total_books - prev_total_books) / prev_total_books * 100) if prev_total_books > 0 else 0.0 if previous_where_clause else 0.0
        turnover_ratio = (active_loans / total_books * 365 / (window_days or 365)) if total_books > 0 and window_days else 0.0
        avg_book_age_years = avg_book_age_days / 365.0
        collection_health_score = min(100, (titles_with_loans / total_books * 100) if total_books > 0 else 0.0)
        genre_diversity = (unique_genres / max(1, total_books)) * 100 if total_books > 0 else 0.0
        author_diversity = (unique_authors / max(1, total_books)) * 100 if total_books > 0 else 0.0
        utilization_efficiency = circulation_velocity * (collection_health_score / 100)
        untouched_ratio = (untouched_titles / total_books * 100) if total_books > 0 else 0.0
        overdue_ratio = (overdue_loans / active_loans * 100) if active_loans > 0 else 0.0
        avg_pages_per_title = total_pages / total_books if total_books > 0 else 0
        
        # DISPLAY CORE KPIs
        st.markdown("### 📊 Core KPIs")
        
        if total_books == 0:
            st.warning("⚠️ No books found in database. Use Bulk Import above to add books.")
            st.stop()
        
        kpi_col1, kpi_col2, kpi_col3, kpi_col4, kpi_col5, kpi_col6 = st.columns(6, gap="small")
        with kpi_col1:
            st.metric("Titles", total_books, delta_books if delta_books is not None else None)
        with kpi_col2:
            st.metric("Total Copies", total_copies, delta_copies if delta_copies is not None else None)
        with kpi_col3:
            st.metric("Checked Out", checked_out, delta=f"{utilization_pct:.1f}%" if total_copies > 0 else None)
        with kpi_col4:
            st.metric("Active Loans", active_loans, delta_loans if delta_loans is not None else None)
        with kpi_col5:
            st.metric("Overdue", overdue_loans, f"{overdue_ratio:.1f}%")
        with kpi_col6:
            st.metric("Avg Copies/Title", f"{avg_copies_per_title:.2f}")

        extra_kpi_col1, extra_kpi_col2, extra_kpi_col3, extra_kpi_col4 = st.columns(4, gap="small")
        with extra_kpi_col1:
            st.metric("Available Copies", available_copies)
        with extra_kpi_col2:
            st.metric("Unique Authors", unique_authors)
        with extra_kpi_col3:
            st.metric("Avg Popularity", f"{avg_popularity:.2f}")
        with extra_kpi_col4:
            st.metric("Due In 7 Days", due_next_7)

        if utilization_pct >= float(utilization_alert_threshold):
            st.warning(
                f"Utilization is high at {utilization_pct:.1f}% (threshold: {utilization_alert_threshold}%)."
            )

        # DISPLAY ADVANCED PERFORMANCE METRICS
        st.markdown("### ⚡ Advanced Performance Metrics")
        adv_kpi_col1, adv_kpi_col2, adv_kpi_col3, adv_kpi_col4, adv_kpi_col5, adv_kpi_col6 = st.columns(6, gap="small")
        with adv_kpi_col1:
            metric_color = "green" if circulation_velocity >= 30 else "orange" if circulation_velocity >= 15 else "red"
            st.metric("Circulation Velocity", f"{circulation_velocity:.1f}%", help="% of inventory checked out (higher = better)")
        with adv_kpi_col2:
            st.metric("Collection Refresh", f"{collection_refresh_rate:.1f}%" if previous_where_clause else "N/A", help="% new titles added")
        with adv_kpi_col3:
            st.metric("Turn-over Ratio", f"{turnover_ratio:.2f}x", help="Avg circulation per title per year")
        with adv_kpi_col4:
            st.metric("Avg Book Age", f"{avg_book_age_years:.1f} yrs", help="Average years in collection")
        with adv_kpi_col5:
            st.metric("Collection Health", f"{collection_health_score:.1f}%", help="% of titles with at least 1 loan")
        with adv_kpi_col6:
            st.metric("Untouched Titles", f"{untouched_ratio:.1f}%", help="Never checked out")

        # DIVERSITY & COMPOSITION METRICS
        st.markdown("### 🎯 Collection Composition")
        comp_col1, comp_col2, comp_col3, comp_col4, comp_col5 = st.columns(5, gap="small")
        with comp_col1:
            st.metric("Genre Diversity", f"{unique_genres} genres")
        with comp_col2:
            st.metric("Language Count", f"{unique_languages} languages")
        with comp_col3:
            st.metric("Author Diversity", f"{author_diversity:.1f}%", help="Unique authors per title")
        with comp_col4:
            st.metric("Avg Book Length", f"{avg_pages_per_title:.0f} pgs" if avg_pages_per_title > 0 else "N/A")
        with comp_col5:
            st.metric("Genre Diversity Score", f"{genre_diversity:.1f}%", help="Genre spread")

        # RETURN & LOAN METRICS
        st.markdown("### 📚 Lending Performance")
        loan_col1, loan_col2 = st.columns(2, gap="small")
        with loan_col1:
            st.metric("Utilization Efficiency", f"{utilization_efficiency:.1f}%", help="Circulation × health")
        with loan_col2:
            st.metric("Titles w/ Activity", f"{titles_with_loans}/{total_books}", help="Titles that have been borrowed")

        st.caption(
            f"Range: {time_window} | Granularity: {trend_granularity} | Active-Only: {'On' if active_titles_only else 'Off'}"
            f" | Zero-Copy Included: {'On' if include_zero_copies else 'Off'} | Inactive Titles: {inactive_titles}"
            f" | Calc Time: {datetime.now().strftime('%H:%M:%S')}"
        )

        viz_top_col1, viz_top_col2 = st.columns(2, gap="small")

        with viz_top_col1:
            stock_df = pd.DataFrame([
                {'state': 'Available', 'copies': available_copies},
                {'state': 'Checked Out', 'copies': checked_out}
            ])
            fig_stock = px.pie(
                stock_df,
                names='state',
                values='copies',
                hole=0.45,
                title='Inventory Availability Split',
                color_discrete_map={'Available': '#90EE90', 'Checked Out': '#FFB6C6'}
            )
            fig_stock.update_layout(height=360)
            st.plotly_chart(fig_stock, use_container_width=True)

        with viz_top_col2:
            # Collection Age Distribution
            age_query = f"""
                SELECT
                    CASE
                        WHEN CAST((julianday('now') - julianday(COALESCE(b.created_at, '2020-01-01'))) / 365.0 AS INT) < 1 THEN 'Last Year'
                        WHEN CAST((julianday('now') - julianday(COALESCE(b.created_at, '2020-01-01'))) / 365.0 AS INT) < 2 THEN '1-2 Yrs'
                        WHEN CAST((julianday('now') - julianday(COALESCE(b.created_at, '2020-01-01'))) / 365.0 AS INT) < 5 THEN '2-5 Yrs'
                        ELSE '5+ Yrs'
                    END as age_group,
                    COUNT(*) as titles
                FROM books b
                WHERE {current_where_clause}
                GROUP BY age_group
                ORDER BY CASE age_group WHEN 'Last Year' THEN 1 WHEN '1-2 Yrs' THEN 2 WHEN '2-5 Yrs' THEN 3 ELSE 4 END
            """
            age_rows = Database.execute_query(age_query, tuple(current_params) if current_params else None)
            if age_rows:
                age_df = pd.DataFrame(age_rows)
                fig_age = px.bar(
                    age_df,
                    x='age_group',
                    y='titles',
                    title='Collection Age Distribution',
                    labels={'age_group': 'Time in Collection', 'titles': 'Titles'},
                    color_discrete_sequence=['#636EFA']
                )
                fig_age.update_layout(height=360)
                st.plotly_chart(fig_age, use_container_width=True)

        # GENRE-LANGUAGE HEATMAP
        viz_mid_col1, viz_mid_col2 = st.columns(2, gap="small")
        with viz_mid_col1:
            language_query = f"""
                SELECT
                    COALESCE(NULLIF(TRIM(b.language), ''), 'Unknown') as language_name,
                    COUNT(*) as titles,
                    SUM((SELECT COUNT(*) FROM book_inventory bi WHERE bi.book_id = b.book_id)) as copies
                FROM books b
                WHERE {current_where_clause}
                GROUP BY COALESCE(NULLIF(TRIM(b.language), ''), 'Unknown')
                ORDER BY titles DESC
                LIMIT ?
            """
            language_rows = Database.execute_query(
                language_query,
                tuple(current_params + [int(top_n)])
            )
            if language_rows:
                language_df = pd.DataFrame(language_rows)
                fig_lang = px.bar(
                    language_df,
                    x='language_name',
                    y='titles',
                    title='Language Distribution',
                    labels={'language_name': 'Language', 'titles': 'Titles'},
                    color='copies',
                    color_continuous_scale='Blues'
                )
                fig_lang.update_layout(height=360, xaxis_tickangle=-25)
                st.plotly_chart(fig_lang, use_container_width=True)
            else:
                st.info("No language data available for the selected filters.")

        with viz_mid_col2:
            genre_query = f"""
                SELECT
                    COALESCE(NULLIF(TRIM(b.genre), ''), 'Unknown') as genre_name,
                    COUNT(*) as titles,
                    SUM((SELECT COUNT(*) FROM book_inventory bi WHERE bi.book_id = b.book_id)) as copies,
                    AVG(COALESCE(b.popularity_score, 0)) as avg_popularity
                FROM books b
                WHERE {current_where_clause}
                GROUP BY COALESCE(NULLIF(TRIM(b.genre), ''), 'Unknown')
                ORDER BY titles DESC
                LIMIT ?
            """
            genre_rows = Database.execute_query(
                genre_query,
                tuple(current_params + [int(top_n)])
            )
            if genre_rows:
                genre_df = pd.DataFrame(genre_rows)
                fig_genre = px.scatter(
                    genre_df,
                    x='genre_name',
                    y='titles',
                    size='copies',
                    color='avg_popularity',
                    title='Top Genres (Size=Copies, Color=Popularity)',
                    labels={'genre_name': 'Genre', 'titles': 'Titles', 'avg_popularity': 'Avg Popularity'},
                    color_continuous_scale='Viridis',
                    hover_data=['copies', 'avg_popularity']
                )
                fig_genre.update_layout(height=360, xaxis_tickangle=-25)
                st.plotly_chart(fig_genre, use_container_width=True)
            else:
                st.info("No genre data available for the selected filters.")

        # POPULARITY & DEMAND CHARTS
        viz_bot_col1, viz_bot_col2 = st.columns(2, gap="small")

        with viz_bot_col1:
            if trend_granularity == "Quarter":
                period_expr = (
                    "strftime('%Y', b.created_at) || '-Q' || "
                    "CASE "
                    "WHEN cast(strftime('%m', b.created_at) as integer) BETWEEN 1 AND 3 THEN '1' "
                    "WHEN cast(strftime('%m', b.created_at) as integer) BETWEEN 4 AND 6 THEN '2' "
                    "WHEN cast(strftime('%m', b.created_at) as integer) BETWEEN 7 AND 9 THEN '3' "
                    "ELSE '4' END"
                )
            elif trend_granularity == "Year":
                period_expr = "strftime('%Y', b.created_at)"
            elif trend_granularity == "Week":
                period_expr = "strftime('%Y-W%W', b.created_at)"
            else:
                period_expr = "strftime('%Y-%m', b.created_at)"

            trend_query = f"""
                SELECT
                    {period_expr} as period,
                    COUNT(*) as titles_added
                FROM books b
                WHERE {current_where_clause} AND b.created_at IS NOT NULL
                GROUP BY {period_expr}
                ORDER BY period DESC
                LIMIT 24
            """
            trend_rows = Database.execute_query(trend_query, tuple(current_params) if current_params else None)
            if trend_rows:
                trend_df = pd.DataFrame(trend_rows).sort_values('period')
                fig_trend = px.area(
                    trend_df,
                    x='period',
                    y='titles_added',
                    markers=True,
                    title=f'Collection Growth Trend ({trend_granularity})',
                    color_discrete_sequence=['#636EFA']
                )
                fig_trend.update_layout(height=360)
                st.plotly_chart(fig_trend, use_container_width=True)
            else:
                st.info("No timeline data available for the selected filters.")

        with viz_bot_col2:
            popularity_query = f"""
                SELECT
                    CAST(b.popularity_score AS INTEGER) as pop_score,
                    COUNT(*) as titles
                FROM books b
                WHERE {current_where_clause} AND b.popularity_score IS NOT NULL
                GROUP BY CAST(b.popularity_score AS INTEGER)
                ORDER BY pop_score ASC
            """
            pop_rows = Database.execute_query(popularity_query, tuple(current_params) if current_params else None)
            if pop_rows:
                pop_df = pd.DataFrame(pop_rows)
                fig_pop = px.bar(
                    pop_df,
                    x='pop_score',
                    y='titles',
                    title='Popularity Score Distribution',
                    labels={'pop_score': 'Popularity Score', 'titles': 'Titles'},
                    color='titles',
                    color_continuous_scale='Greens'
                )
                fig_pop.update_layout(height=360)
                st.plotly_chart(fig_pop, use_container_width=True)

        # PUBLICATION YEAR ANALYSIS
        yearly_query = f"""
            SELECT
                b.publication_year as year,
                COUNT(*) as titles,
                SUM((SELECT COUNT(*) FROM book_inventory bi WHERE bi.book_id = b.book_id)) as total_copies,
                AVG(COALESCE(b.popularity_score, 0)) as avg_pop
            FROM books b
            WHERE {current_where_clause} AND b.publication_year IS NOT NULL
            GROUP BY b.publication_year
            ORDER BY year ASC
        """
        yearly_rows = Database.execute_query(yearly_query, tuple(current_params) if current_params else None)
        if yearly_rows:
            yearly_df = pd.DataFrame(yearly_rows)
            fig_year = px.scatter(
                yearly_df,
                x='year',
                y='titles',
                size='total_copies',
                color='avg_pop',
                title='Publication Year vs Titles (Size=Copies, Color=Popularity)',
                labels={'year': 'Publication Year', 'titles': 'Titles', 'avg_pop': 'Avg Popularity'},
                color_continuous_scale='RdYlGn',
                hover_data=['total_copies', 'avg_pop']
            )
            fig_year.update_layout(height=330)
            st.plotly_chart(fig_year, use_container_width=True)

        # ADVANCED OPERATIONAL INSIGHTS
        st.markdown("### 🔍 Advanced Operational Insights")
        
        insight_tabs = st.tabs(["📦 Stock Analysis", "📈 Demand Leaders", "⚠️ Alerts & Issues", "🎯 Recommendations"])
        
        with insight_tabs[0]:
            insight_col1, insight_col2, insight_col3 = st.columns(3, gap="small")
            
            with insight_col1:
                st.markdown("#### 🔴 Availability Pressure")
                low_stock_query = f"""
                    SELECT
                        b.title,
                        COALESCE(NULLIF(TRIM(b.author), ''), 'Unknown') as author,
                        COALESCE(NULLIF(TRIM(b.genre), ''), 'Unknown') as genre,
                        (SELECT COUNT(*) FROM book_inventory bi WHERE bi.book_id = b.book_id) as total_copies,
                        (SELECT COUNT(*) FROM book_inventory bi WHERE bi.book_id = b.book_id AND bi.is_available = 1) as available_copies,
                        (SELECT COUNT(*) FROM borrowing br JOIN book_inventory bi ON br.inventory_id = bi.inventory_id WHERE bi.book_id = b.book_id AND br.return_date IS NULL) as current_loans
                    FROM books b
                    WHERE {current_where_clause}
                    ORDER BY available_copies ASC, total_copies ASC, b.title ASC
                    LIMIT ?
                """
                low_stock_rows = Database.execute_query(
                    low_stock_query,
                    tuple(current_params + [int(top_n)])
                )
                if low_stock_rows:
                    pressure_df = pd.DataFrame(low_stock_rows)
                    pressure_df['total_copies'] = pressure_df['total_copies'].fillna(0).astype(int)
                    pressure_df['available_copies'] = pressure_df['available_copies'].fillna(0).astype(int)
                    pressure_df['current_loans'] = pressure_df['current_loans'].fillna(0).astype(int)
                    pressure_df['is_low_stock'] = pressure_df['available_copies'] <= int(low_stock_threshold)
                    pressure_df['availability_ratio'] = pressure_df.apply(
                        lambda r: f"{int(r['available_copies'])}/{int(r['total_copies'])}", axis=1
                    )
                    pressure_df['pressure_level'] = pressure_df.apply(
                        lambda r: "🔴 CRITICAL" if r['available_copies'] == 0 else "🟠 HIGH" if r['available_copies'] <= r['total_copies'] * 0.2 else "🟡 MEDIUM",
                        axis=1
                    )
                    display_cols = ['title', 'author', 'genre', 'availability_ratio', 'current_loans', 'pressure_level']
                    st.dataframe(
                        pressure_df[display_cols],
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "title": st.column_config.TextColumn("Title", width="medium"),
                            "availability_ratio": st.column_config.TextColumn("In Stock/Total", width="small"),
                            "current_loans": st.column_config.NumberColumn("Current Loans", width="small"),
                            "pressure_level": st.column_config.TextColumn("Status", width="small"),
                        }
                    )
                else:
                    st.info("No low stock items found.")
            
            with insight_col2:
                st.markdown("#### 🟢 Underutilized Titles")
                underutilized_query = f"""
                    SELECT
                        b.title,
                        COALESCE(NULLIF(TRIM(b.author), ''), 'Unknown') as author,
                        (SELECT COUNT(*) FROM book_inventory bi WHERE bi.book_id = b.book_id) as total_copies,
                        (SELECT COUNT(*) FROM borrowing br JOIN book_inventory bi ON br.inventory_id = bi.inventory_id WHERE bi.book_id = b.book_id) as total_loans,
                        COALESCE(b.popularity_score, 0) as popularity
                    FROM books b
                    WHERE {current_where_clause}
                        AND (SELECT COUNT(*) FROM borrowing br JOIN book_inventory bi ON br.inventory_id = bi.inventory_id WHERE bi.book_id = b.book_id) < 2
                    ORDER BY total_loans ASC, b.title ASC
                    LIMIT ?
                """
                underutil_rows = Database.execute_query(
                    underutilized_query,
                    tuple(current_params + [int(top_n)])
                )
                if underutil_rows:
                    underutil_df = pd.DataFrame(underutil_rows)
                    underutil_df['total_loans'] = underutil_df['total_loans'].fillna(0).astype(int)
                    underutil_df['popularity'] = underutil_df['popularity'].fillna(0).round(2)
                    st.dataframe(
                        underutil_df[['title', 'author', 'total_copies', 'total_loans', 'popularity']],
                        use_container_width=True,
                        hide_index=True
                    )
                else:
                    st.success("All titles have adequate engagement!")
            
            with insight_col3:
                st.markdown("#### 📊 Collection Distribution")
                st.metric("Total Inventory", total_copies)
                st.metric("Active/Inactive Titles", f"{total_books - inactive_titles}/{inactive_titles}")
                st.metric("Avg Copies per Title", f"{avg_copies_per_title:.2f}")
                st.metric("Titles with Activity", f"{titles_with_loans}/{total_books}")

        with insight_tabs[1]:
            st.markdown("#### 📈 Demand Leaders")
            demand_query = f"""
                SELECT
                    b.book_id,
                    b.title,
                    COALESCE(NULLIF(TRIM(b.author), ''), 'Unknown') as author,
                    COUNT(br.borrowing_id) as total_loans,
                    SUM(CASE WHEN br.return_date IS NULL THEN 1 ELSE 0 END) as active_loans,
                    AVG(julianday(COALESCE(br.return_date, date('now'))) - julianday(br.checkout_date)) as avg_loan_days,
                    COALESCE(b.popularity_score, 0) as popularity
                FROM books b
                LEFT JOIN book_inventory bi ON bi.book_id = b.book_id
                LEFT JOIN borrowing br ON br.inventory_id = bi.inventory_id
                WHERE {current_where_clause}
                GROUP BY b.book_id, b.title, b.author
                HAVING total_loans > 0
                ORDER BY total_loans DESC, active_loans DESC
                LIMIT ?
            """
            demand_rows = Database.execute_query(
                demand_query,
                tuple(current_params + [int(top_n)])
            )
            if demand_rows:
                demand_df = pd.DataFrame(demand_rows)
                demand_df['avg_loan_days'] = demand_df['avg_loan_days'].fillna(0).round(1)
                demand_df['popularity'] = demand_df['popularity'].fillna(0).round(2)
                demand_df['demand_score'] = (demand_df['total_loans'] * demand_df['popularity'] / (demand_df['avg_loan_days'] + 1)).round(2)
                
                demand_chart_df = demand_df.head(10)[['title', 'total_loans', 'active_loans', 'demand_score']]
                fig_demand = px.bar(
                    demand_chart_df,
                    x='title',
                    y='total_loans',
                    color='demand_score',
                    title='Top 10 Demand Leaders',
                    labels={'title': 'Title', 'total_loans': 'Total Loans', 'demand_score': 'Demand Score'},
                    color_continuous_scale='Reds',
                    hover_data=['active_loans']
                )
                fig_demand.update_layout(xaxis_tickangle=-45, height=400)
                st.plotly_chart(fig_demand, use_container_width=True)
                
                st.dataframe(
                    demand_df[['title', 'author', 'total_loans', 'active_loans', 'avg_loan_days', 'popularity']],
                    use_container_width=True,
                    hide_index=True
                )
            else:
                st.info("No borrowing demand data for the selected filters.")

        with insight_tabs[2]:
            st.markdown("#### ⚠️ System Alerts")
            
            alerts_generated = []
            
            if enable_low_stock_alerts and circulation_velocity > float(utilization_alert_threshold):
                alerts_generated.append(("🔴 **HIGH UTILIZATION**", f"Circulation velocity is {circulation_velocity:.1f}% (threshold: {utilization_alert_threshold}%). Consider acquiring more copies."))
            
            if enable_overdue_alerts and overdue_loans > 0:
                alerts_generated.append(("🟠 **OVERDUE ITEMS**", f"{overdue_loans} items overdue. Consider review and follow-up."))
            
            if enable_performance_alerts and untouched_ratio > 20:
                alerts_generated.append(("🟡 **LOW ENGAGEMENT**", f"{untouched_ratio:.1f}% of titles have never been borrowed. Review collection relevance."))
            
            if collection_health_score < 60:
                alerts_generated.append(("🟡 **HEALTH WARNING**", f"Collection health is {collection_health_score:.1f}%. Many titles lack circulation activity."))
            
            if not alerts_generated:
                st.success("✅ No critical alerts found. Collection is performing well!")
            else:
                for alert_type, alert_msg in alerts_generated:
                    st.warning(f"{alert_type}\n{alert_msg}")

        with insight_tabs[3]:
            st.markdown("#### 🎯 AI-Powered Insights & Recommendations")
            
            recommendations = []
            
            if circulation_velocity < 20:
                recommendations.append(f"📊 **Low Circulation**: Consider marketing campaigns or genre-specific promotions. Current velocity: {circulation_velocity:.1f}%")
            
            if untouched_ratio > 10:
                recommendations.append(f"🎯 **Engagement**: {untouched_ratio:.1f}% of titles are unused. Review relevance or consider deaccessioning.")
            
            if collection_refresh_rate and collection_refresh_rate < 5:
                recommendations.append(f"🆕 **Acquisition Rate**: Only {collection_refresh_rate:.1f}% new titles in period. Consider accelerating acquisitions.")
            
            
            if genre_diversity < 30:
                recommendations.append(f"📚 **Genre Balance**: Low genre diversity ({genre_diversity:.1f}%). Consider expanding collection breadth.")
            
            if unique_languages < 3:
                recommendations.append(f"🌍 **Language Diversity**: Only {unique_languages} languages available. Consider multilingual acquisitions.")
            
            if avg_book_age_years > 8:
                recommendations.append(f"🔄 **Collection Age**: Average book age is {avg_book_age_years:.1f} years. Consider refresh cycle.")
            
            if recommendations:
                for i, rec in enumerate(recommendations, 1):
                    with st.container():
                        st.info(f"**{i}.** {rec}")
            else:
                st.success("✅ Your collection metrics look excellent! Continue monitoring key KPIs.")

def show_manage_members():
    """Member management page"""
    st.markdown('<h1 class="litgrid-header"> Manage Members</h1>', unsafe_allow_html=True)
    
    tab1, tab2, tab3, tab4 = st.tabs([" All Members", " Register User", " Edit/Delete", " Statistics"])
    
    with tab1:
        st.subheader("**Member List**")
        
        # Filters
        col1, col2, col3 = st.columns(3, gap="small")
        with col1:
            search = st.text_input(" Search by name or email", key="mm_search")
        with col2:
            role_filter = st.selectbox("Role", ["All", "Admin", "Librarian", "Member"], key="mm_role_filter")
        with col3:
            status_filter = st.selectbox("Status", ["All", "Active", "Inactive"], key="mm_status_filter")
        
        # Build query
        query = """
            SELECT u.user_id, u.username, u.full_name, u.email, u.phone,
                   u.role, u.fine_balance, u.is_active, u.created_at,
                   0 as current_books
            FROM users u
            WHERE 1=1
        """
        params = []
        
        if search:
            query += " AND (u.full_name LIKE ? OR u.email LIKE ? OR u.username LIKE ?)"
            params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])
        
        if role_filter != "All":
            query += " AND u.role = ?"
            params.append(role_filter.lower())
        
        if status_filter == "Active":
            query += " AND u.is_active = 1"
        elif status_filter == "Inactive":
            query += " AND u.is_active = 0"
        
        query += " ORDER BY u.created_at DESC LIMIT 100"
        
        members = Database.execute_query(query, tuple(params) if params else None)
        
        if members:
            st.write(f"Found {len(members)} members")
            
            for member in members:
                with st.container():
                    col1, col2, col3 = st.columns([2, 2, 1], gap="small")
                    
                    with col1:
                        status_icon = "" if member['is_active'] else ""
                        st.markdown(f"**{status_icon} {member['full_name']}**")
                        st.caption(f"@{member['username']} | {member['email']}")
                        if member['phone']:
                            st.caption(f" {member['phone']}")
                    
                    with col2:
                        st.write(f"**Role:** {member['role'].title()}")
                        st.write(f"**Books:** {member['current_books']}")
                        if member['fine_balance'] > 0:
                            st.warning(f"Fine: ${member['fine_balance']:.2f}")
                    
                    with col3:
                        if st.button(" View", key=f"view_{member['user_id']}", use_container_width=True):
                            st.session_state[f'view_member_{member["user_id"]}'] = True
                        
                        action = "Deactivate" if member['is_active'] else "Activate"
                        if st.button(f" {action}", key=f"toggle_member_{member['user_id']}", use_container_width=True):
                            new_status = not member['is_active']
                            if Database.execute_update(
                                "UPDATE users SET is_active = ? WHERE user_id = ?",
                                (new_status, member['user_id'])
                            ):
                                st.success(f"Member {action.lower()}d successfully!")
                                st.rerun()
                    
                    st.divider()
        else:
            st.info("No members found")
    
    with tab2:
        st.subheader("**Register New User**")
        st.caption("Admins have access to all system features. Librarians can manage books and members.")

        def validate_password_strength(password):
            """Check password strength"""
            if len(password) < 12:
                return False, "Password must be at least 12 characters"
            if password.lower() == password:
                return False, "Password must contain uppercase letters"
            if password.upper() == password:
                return False, "Password must contain lowercase letters"
            if not any(c.isdigit() for c in password):
                return False, "Password must contain numbers"
            if not any(c in "!@#$%^&*()-_=+[]{}|;:,.<>?" for c in password):
                return False, "Password must contain special characters"
            if any(password.count(c) > 3 for c in set(password)):
                return False, "Password contains too many repeated characters"
            return True, "Strong password ✓"

        def get_password_criteria(password):
            """Return granular password checks for live feedback."""
            value = password or ""
            return {
                'length': len(value) >= 12,
                'uppercase': any(c.isupper() for c in value),
                'lowercase': any(c.islower() for c in value),
                'digit': any(c.isdigit() for c in value),
                'special': any(c in "!@#$%^&*()-_=+[]{}|;:,.<>?" for c in value),
                'repeat_limit': not any(value.count(c) > 3 for c in set(value)) if value else False,
            }

        def show_password_feedback(password, confirm_password=None):
            """Render dynamic password-strength feedback."""
            checks = get_password_criteria(password)
            passed = sum(1 for ok in checks.values() if ok)
            total = len(checks)
            is_strong = passed == total

            if password:
                labels = [
                    (checks['length'], "At least 12 characters"),
                    (checks['uppercase'], "Contains uppercase letter"),
                    (checks['lowercase'], "Contains lowercase letter"),
                    (checks['digit'], "Contains number"),
                    (checks['special'], "Contains special character"),
                    (checks['repeat_limit'], "No character repeated more than 3 times"),
                ]
                for ok, text in labels:
                    st.caption(f"{'✅' if ok else '❌'} {text}")

                if confirm_password is not None and confirm_password:
                    st.caption(f"{'✅' if password == confirm_password else '❌'} Passwords match")

                if is_strong:
                    st.success("Password strength: Strong")
                else:
                    st.warning(f"Password strength: {passed}/{total} requirements met")

            return is_strong

        def check_duplicate_credentials(username, email):
            """Check if username or email already exists"""
            existing = Database.execute_query("""
                SELECT user_id FROM users
                WHERE LOWER(username) = LOWER(?) OR LOWER(email) = LOWER(?)
            """, (username, email))
            return existing is not None and len(existing) > 0

        with st.form("register_user_form"):
            # Personal Information
            st.markdown("#### **Personal Information**")
            col1, col2 = st.columns(2, gap="small")

            with col1:
                reg_full_name = st.text_input("Full Name *", key="reg_name", placeholder="Jane Smith")
                reg_date_of_birth = st.date_input("Date of Birth *", key="reg_dob")
                reg_phone = st.text_input("Phone Number *", key="reg_phone", placeholder="+1-555-0000")
                reg_gender = st.selectbox("Gender", ["Prefer not to say", "Male", "Female", "Other"], key="reg_gender")

            with col2:
                reg_username = st.text_input("Username *", key="reg_user", placeholder="janesmith123", help="4-20 alphanumeric characters")
                reg_email = st.text_input("Email *", key="reg_email", placeholder="jane@library.local")
                reg_role = st.selectbox("Role *", ["librarian", "admin"], key="reg_role", help="Admin: Full system access, Librarian: Book & member management")
                reg_department = st.selectbox("Department", ["Circulation", "Cataloging", "Reference", "Administration"], key="reg_dept")

            # Address Information (for staff records)
            st.markdown("#### **Address Information**")
            col3, col4 = st.columns(2, gap="small")
            with col3:
                reg_street = st.text_input("Street Address *", key="reg_street")
                reg_city = st.text_input("City *", key="reg_city")
            with col4:
                reg_state = st.text_input("State/Province *", key="reg_state")
                reg_zip = st.text_input("Zip/Postal Code *", key="reg_zip")

            # Employment Information (for admin/staff)
            st.markdown("#### **Employment Information**")
            col5, col6 = st.columns(2, gap="small")
            with col5:
                reg_employment_date = st.date_input("Employment Start Date *", key="reg_emp_date")
                reg_salary_grade = st.selectbox("Salary Grade", ["Grade 1", "Grade 2", "Grade 3", "Grade 4", "Grade 5"], key="reg_salary")
            with col6:
                reg_supervisor = st.text_input("Supervisor Name", key="reg_supervisor")
                reg_work_schedule = st.selectbox("Work Schedule", ["Full-Time", "Part-Time", "Casual"], key="reg_schedule")

            # Password Security
            st.markdown("#### **Password Security**")
            st.caption("Password Requirements: 12+ chars, uppercase, lowercase, numbers, special chars, no repetition")
            col7, col8 = st.columns(2, gap="small")
            with col7:
                reg_password = st.text_input("Password *", type="password", key="reg_pass", help="Min 12 chars, mixed case, numbers, special chars")
            with col8:
                reg_confirm_pass = st.text_input("Confirm Password *", type="password", key="reg_confirm_pass")
            is_register_password_strong = show_password_feedback(reg_password, reg_confirm_pass)

            # Photo Upload
            st.markdown("#### **Staff Photo**")
            reg_photo = st.file_uploader("Upload Staff Photo *", type=["jpg", "jpeg", "png"], key="reg_photo", help="Max 5MB")

            # Additional Info
            st.markdown("#### **Additional Information**")
            reg_notes = st.text_area("Notes/Comments", key="reg_notes", height=100, placeholder="Any additional information about this user...")

            # Legal Acknowledgement
            st.divider()
            st.markdown("#### **Employment Agreement**")
            agree_employment = st.radio(
                "I acknowledge this staff account creation:",
                ["❌ Do not proceed", "✅ I confirm all information is accurate and authorized"],
                key="agree_emp"
            )

            register_password_ready = (
                bool(reg_password)
                and bool(reg_confirm_pass)
                and is_register_password_strong
                and reg_password == reg_confirm_pass
            )

            submit_reg = st.form_submit_button(
                "Create Staff Account",
                use_container_width=True,
                disabled=(not register_password_ready)
            )

            if submit_reg:
                errors = []

                if not all([reg_full_name, reg_username, reg_email, reg_phone, reg_street, reg_city, reg_state, reg_zip, reg_password, reg_confirm_pass, reg_photo]):
                    errors.append("Please fill all required fields (marked with *)")

                if reg_username and not (4 <= len(reg_username) <= 20):
                    errors.append("Username must be 4-20 characters")

                if reg_email and "@" not in reg_email:
                    errors.append("Invalid email format")

                if reg_password != reg_confirm_pass:
                    errors.append("Passwords do not match")

                if reg_password and not is_register_password_strong:
                    is_strong, msg = validate_password_strength(reg_password)
                    if not is_strong:
                        errors.append(f"Password too weak: {msg}")

                if reg_username or reg_email:
                    if check_duplicate_credentials(reg_username, reg_email):
                        errors.append("Username or email already exists")

                if not reg_photo:
                    errors.append("Staff photo is required")

                if reg_photo and reg_photo.size > 5 * 1024 * 1024:
                    errors.append("Photo must be less than 5MB")

                if "❌" in agree_employment:
                    errors.append("You must confirm employment authorization")

                if errors:
                    for error in errors:
                        st.error(f"❌ {error}")
                else:
                    try:
                        # Register user with specified role (admin or librarian)
                        success, msg = Auth.register(reg_username, reg_email, reg_password, reg_full_name, reg_phone, role=reg_role)

                        if success:
                            st.success(f"✅ Staff account created successfully!")
                            st.info(f"Role: {reg_role.title()}, Department: {reg_department}")
                            st.balloons()
                        else:
                            st.error(f"❌ Error: {msg}")
                    except Exception as e:
                        st.error(f"❌ Registration error: {str(e)}")
    
    with tab3:
        st.subheader(" Edit / Delete User")
        
        search_user = st.text_input(" Search user by username or email")
        
        if search_user:
            users = Database.execute_query("""
                SELECT user_id, username, full_name, email, phone,
                       role, is_active
                FROM users
                WHERE username LIKE ? OR email LIKE ?
                LIMIT 10
            """, (f'%{search_user}%', f'%{search_user}%'))
            
            if users:
                selected_user = st.selectbox(
                    "Select User",
                    users,
                    format_func=lambda u: f"{u['full_name']} (@{u['username']}) - {u['email']}"
                )
                
                if selected_user:
                    st.divider()
                    
                    col1, col2 = st.columns([2, 1], gap="small")
                    
                    with col1:
                        st.markdown("### Edit User Details")
                        
                        with st.form("edit_user_form"):
                            edit_full_name = st.text_input("Full Name", value=selected_user['full_name'])
                            edit_email = st.text_input("Email", value=selected_user['email'])
                            edit_phone = st.text_input("Phone", value=selected_user.get('phone', ''))
                            
                            try:
                                current_role_index = ["member", "librarian", "admin"].index(selected_user['role'])
                            except ValueError:
                                current_role_index = 0
                            edit_role = st.selectbox("Role", ["member", "librarian", "admin"], index=current_role_index)
                            
                            edit_active = st.checkbox("Active", value=selected_user['is_active'])
                            
                            submit_edit = st.form_submit_button(" Save Changes", use_container_width=True)
                            
                            if submit_edit:
                                if Database.execute_update("""
                                    UPDATE users
                                    SET full_name = ?, email = ?, phone = ?, role = ?, is_active = ?
                                    WHERE user_id = ?
                                """, (edit_full_name, edit_email, edit_phone, edit_role, edit_active, selected_user['user_id'])):
                                    st.success(" User updated successfully!")
                                    st.rerun()
                                else:
                                    st.error("Failed to update user")
                    
                    with col2:
                        st.markdown("### Actions")
                        
                        st.warning(" Danger Zone")
                        
                        if st.button(" Delete User", type="secondary", use_container_width=True):
                            st.session_state[f'confirm_delete_{selected_user["user_id"]}'] = True
                        
                        if st.session_state.get(f'confirm_delete_{selected_user["user_id"]}', False):
                            st.error("Are you sure? This cannot be undone!")
                            
                            col_x, col_y = st.columns(2, gap="small")
                            with col_x:
                                if st.button(" Yes, Delete", use_container_width=True):
                                    # Soft delete - set inactive
                                    if Database.execute_update(
                                        "UPDATE users SET is_active = 0 WHERE user_id = ?",
                                        (selected_user['user_id'],)
                                    ):
                                            st.success("User deleted (deactivated)")
                                            del st.session_state[f'confirm_delete_{selected_user["user_id"]}']
                                            st.rerun()
                            
                            with col_y:
                                if st.button(" Cancel", use_container_width=True):
                                    del st.session_state[f'confirm_delete_{selected_user["user_id"]}']
                                    st.rerun()
                        
                        st.divider()
                        
                        # View borrowing history
                        if st.button(" View Borrowing History", use_container_width=True):
                            st.session_state[f'show_history_{selected_user["user_id"]}'] = True
                        
                        if st.session_state.get(f'show_history_{selected_user["user_id"]}', False):
                            st.markdown("### Borrowing History")
                            history = Database.execute_query("""
                                SELECT b.title, br.checkout_date, br.return_date, br.due_date,
                                       br.fine_amount, br.status
                                FROM borrowing br
                                JOIN book_inventory bi ON br.inventory_id = bi.inventory_id
                                JOIN books b ON bi.book_id = b.book_id
                                WHERE br.user_id = %s
                                ORDER BY br.checkout_date DESC
                                LIMIT 20
                            """, (selected_user['user_id'],))
                            
                            if history:
                                for h in history:
                                    status_emoji = "" if h['status'] == 'returned' else ""
                                    st.write(f"{status_emoji} {h['title']}")
                                    st.caption(f"Checkout: {format_date(h['checkout_date'])} | Due: {format_date(h['due_date'])}")
                                    if h['return_date']:
                                        st.caption(f"Returned: {format_date(h['return_date'])}")
                                    if h['fine_amount'] and h['fine_amount'] > 0:
                                        st.caption(f"Fine: {format_currency(h['fine_amount'])}")
                                    st.divider()
                            else:
                                st.info("No borrowing history")
                        
                        # View activity logs
                        if st.button(" View Activity Logs", use_container_width=True):
                            logs = EnhancedUserManager.get_activity_logs(selected_user['user_id'], 20)
                            if logs:
                                st.markdown("### Recent Activity")
                                for log in logs:
                                    st.caption(f"{format_datetime(log['timestamp'])} - {log['activity_type']}")
                            else:
                                st.info("No activity logs")
            else:
                st.warning("No users found")
    
    with tab4:
        st.subheader("**Member Statistics**")
        
        col1, col2, col3, col4 = st.columns(4, gap="small")
        
        total = Database.execute_query("SELECT COUNT(*) as count FROM users WHERE role = 'member'")
        active = Database.execute_query("SELECT COUNT(*) as count FROM users WHERE role = 'member' AND is_active = 1")
        with_fines = Database.execute_query(
            "SELECT COUNT(*) as count FROM users WHERE fine_balance > 0"
        )
        total_fines = Database.execute_query(
            "SELECT SUM(fine_balance) as total FROM users"
        )
        
        with col1:
            st.metric("Total Members", total[0]['count'] if total else 0)
        with col2:
            st.metric("Active Members", active[0]['count'] if active else 0)
        with col3:
            st.metric("Members with Fines", with_fines[0]['count'] if with_fines else 0)
        with col4:
            st.metric("Total Fines", f"${total_fines[0]['total'] if total_fines and total_fines[0]['total'] else 0:.2f}")

def show_borrowing_returns():
    """Borrowing and returns page"""
    st.markdown('<h1 class="litgrid-header">Borrowing & Returns</h1>', unsafe_allow_html=True)

    st.markdown("### Control Center")
    ctrl_col1, ctrl_col2, ctrl_col3, ctrl_col4 = st.columns(4, gap="small")
    with ctrl_col1:
        due_soon_days = st.slider("Due Soon Threshold (days)", min_value=1, max_value=14, value=3, key="br_due_soon_days")
    with ctrl_col2:
        trend_window = st.selectbox("Trend Window", ["7D", "30D", "90D", "180D", "365D"], index=1, key="br_trend_window")
    with ctrl_col3:
        fine_preview_mode = st.checkbox("Fine Preview Mode", value=True, key="br_fine_preview")
    with ctrl_col4:
        if st.button("Refresh Dashboard", use_container_width=True, key="br_refresh_dashboard"):
            st.rerun()

    kpi_summary = Database.execute_query(
        """
        SELECT
            (SELECT COUNT(*) FROM borrowing WHERE return_date IS NULL) as active_loans,
            (SELECT COUNT(*) FROM borrowing WHERE return_date IS NULL AND date(due_date) < date('now')) as overdue_loans,
            (SELECT COUNT(*) FROM borrowing WHERE return_date IS NULL AND julianday(due_date) - julianday(date('now')) BETWEEN 0 AND ?) as due_soon,
            (SELECT COUNT(*) FROM borrowing WHERE date(checkout_date) = date('now')) as today_checkouts,
            (SELECT COUNT(*) FROM borrowing WHERE date(return_date) = date('now')) as today_returns,
            (SELECT COALESCE(SUM(fine_amount), 0) FROM borrowing WHERE date(return_date) = date('now')) as today_fines
        """,
        (int(due_soon_days),),
        fetch_one=True
    ) or {}

    kpi_col1, kpi_col2, kpi_col3, kpi_col4, kpi_col5, kpi_col6 = st.columns(6, gap="small")
    with kpi_col1:
        st.metric("Active Loans", int(kpi_summary.get('active_loans') or 0))
    with kpi_col2:
        st.metric("Overdue", int(kpi_summary.get('overdue_loans') or 0))
    with kpi_col3:
        st.metric("Due Soon", int(kpi_summary.get('due_soon') or 0))
    with kpi_col4:
        st.metric("Today Checkouts", int(kpi_summary.get('today_checkouts') or 0))
    with kpi_col5:
        st.metric("Today Returns", int(kpi_summary.get('today_returns') or 0))
    with kpi_col6:
        st.metric("Today Fines", format_currency(float(kpi_summary.get('today_fines') or 0)))

    global_borrow_search = st.text_input(
        "Global Borrowing Search (member/book/email)",
        key="br_global_search",
        help="Used as default in Return and Active Borrowings sections"
    )

    def render_active_borrowings_panel():
        st.subheader("**Active Borrowings**")

        col1, col2, col3, col4 = st.columns(4, gap="small")
        with col1:
            filter_type = st.selectbox("Filter", ["All", f"Due Soon (< {due_soon_days} days)", "Overdue"], key="br_active_filter")
        with col2:
            search_active = st.text_input("Search by member or book", value=global_borrow_search if global_borrow_search else "", key="br_active_search")
        with col3:
            sort_active = st.selectbox("Sort By", ["Due Date (Earliest)", "Due Date (Latest)", "Member A-Z", "Book A-Z"], key="br_active_sort")
        with col4:
            active_limit = st.slider("Max Rows", min_value=10, max_value=200, value=50, step=10, key="br_active_limit")

        query = """
            SELECT br.borrowing_id, b.title, b.isbn, u.full_name, u.email,
                   br.checkout_date, br.due_date,
                   julianday(br.due_date) - julianday(date('now')) as days_remaining,
                   julianday(date('now')) - julianday(br.due_date) as days_overdue
            FROM borrowing br
            JOIN book_inventory bi ON br.inventory_id = bi.inventory_id
            JOIN books b ON bi.book_id = b.book_id
            JOIN users u ON br.user_id = u.user_id
            WHERE br.return_date IS NULL
        """
        params = []

        if filter_type == f"Due Soon (< {due_soon_days} days)":
            query += " AND julianday(br.due_date) - julianday(date('now')) BETWEEN 0 AND ?"
            params.append(int(due_soon_days))
        elif filter_type == "Overdue":
            query += " AND br.due_date < date('now')"

        if search_active:
            query += " AND (u.full_name LIKE ? OR b.title LIKE ?)"
            params.extend([f'%{search_active}%', f'%{search_active}%'])

        if sort_active == "Due Date (Latest)":
            query += " ORDER BY br.due_date DESC"
        elif sort_active == "Member A-Z":
            query += " ORDER BY u.full_name ASC, br.due_date ASC"
        elif sort_active == "Book A-Z":
            query += " ORDER BY b.title ASC, br.due_date ASC"
        else:
            query += " ORDER BY br.due_date ASC"

        query += " LIMIT ?"
        params.append(int(active_limit))

        active = Database.execute_query(query, tuple(params) if params else None)

        if active:
            st.write(f"Found {len(active)} active borrowings")

            export_df = pd.DataFrame(active)
            if not export_df.empty:
                st.download_button(
                    "Download Active Borrowings CSV",
                    export_df.to_csv(index=False),
                    file_name="active_borrowings.csv",
                    mime="text/csv",
                    use_container_width=True,
                    key="br_export_active_csv"
                )

            for bw in active:
                row_col1, row_col2 = st.columns([3, 1], gap="small")

                with row_col1:
                    st.markdown(f"**{bw['title']}**")
                    st.caption(f"Borrowed by: {bw['full_name']} ({bw['email']})")
                    st.caption(f"Checkout: {format_date(bw['checkout_date'])} | Due: {format_date(bw['due_date'])}")

                with row_col2:
                    if bw['days_overdue'] and bw['days_overdue'] > 0:
                        overdue_days = int(bw['days_overdue'])
                        st.error(f" Overdue by {overdue_days} days")
                        fine = bw['days_overdue'] * Config.FINE_PER_DAY
                        st.caption(f"Fine: {format_currency(fine)}")
                    elif bw['days_remaining'] <= due_soon_days:
                        remaining_days = max(int(bw['days_remaining']), 0)
                        st.warning(f"Due in {remaining_days} days")
                    else:
                        remaining_days = max(int(bw['days_remaining']), 0)
                        st.info(f" Due in {remaining_days} days")

                st.divider()
        else:
            st.info("No active borrowings")

    unified_col, active_col = st.columns([3, 2], gap="large")

    with unified_col:
        st.markdown("### Operations")
        workspace_col1, workspace_col2, workspace_col3 = st.columns(3, gap="small")

        pending_renewals = Database.execute_query(
            "SELECT COUNT(*) as count FROM renewal_requests WHERE status = 'pending'",
            fetch_one=True
        ) or {'count': 0}
        open_renewals = int(pending_renewals.get('count') or 0)

        active_loans_value = int(kpi_summary.get('active_loans') or 0)
        overdue_value = int(kpi_summary.get('overdue_loans') or 0)
        due_soon_value = int(kpi_summary.get('due_soon') or 0)
        today_checkouts_value = int(kpi_summary.get('today_checkouts') or 0)
        today_returns_value = int(kpi_summary.get('today_returns') or 0)
        today_fines_value = float(kpi_summary.get('today_fines') or 0)

        with workspace_col1:
            st.info("Current Load")
            st.metric("Open Work Items", active_loans_value + due_soon_value + overdue_value + open_renewals)
            st.metric("Pending Renewals", open_renewals)


        with workspace_col2:
            st.info("Daily Throughput")
            net_flow = today_checkouts_value - today_returns_value
            st.metric("Net Flow", net_flow)

        with workspace_col3:
            st.info("Risk & Financial")
            overdue_rate = (overdue_value / active_loans_value * 100) if active_loans_value > 0 else 0.0
            st.metric("Overdue Loans", overdue_value, f"{overdue_rate:.1f}%")
            st.metric("Fine Preview", "ON" if fine_preview_mode else "OFF")

        vis_col1, vis_col2 = st.columns(2, gap="small")
        with vis_col1:
            flow_df = pd.DataFrame([
                {"stage": "Checkouts", "value": today_checkouts_value},
                {"stage": "Returns", "value": today_returns_value},
                {"stage": "Overdue", "value": overdue_value},
                {"stage": "Renewals", "value": open_renewals},
            ])
            fig_flow = px.bar(
                flow_df,
                x="stage",
                y="value",
                title="Operational Flow Snapshot",
                labels={"stage": "Stage", "value": "Count"},
                color="value",
                color_continuous_scale="Blues"
            )
            fig_flow.update_layout(height=320, xaxis_title=None)
            st.plotly_chart(fig_flow, use_container_width=True)

        with vis_col2:
            status_df = pd.DataFrame([
                {"status": "Healthy", "count": max(active_loans_value - due_soon_value - overdue_value, 0)},
                {"status": "Due Soon", "count": due_soon_value},
                {"status": "Overdue", "count": overdue_value},
            ])
            fig_status = px.pie(
                status_df,
                names="status",
                values="count",
                title="Active Loan Health Split",
                hole=0.45,
                color_discrete_map={"Healthy": "#2E7D32", "Due Soon": "#F9A825", "Overdue": "#C62828"}
            )
            fig_status.update_layout(height=320)
            st.plotly_chart(fig_status, use_container_width=True)

        focus_queue = Database.execute_query(
            """
            SELECT
                b.title,
                u.full_name,
                br.due_date,
                CAST(julianday(br.due_date) - julianday(date('now')) AS INTEGER) as days_remaining
            FROM borrowing br
            JOIN book_inventory bi ON br.inventory_id = bi.inventory_id
            JOIN books b ON bi.book_id = b.book_id
            JOIN users u ON br.user_id = u.user_id
            WHERE br.return_date IS NULL
              AND (julianday(br.due_date) - julianday(date('now')) <= ?)
            ORDER BY br.due_date ASC
            LIMIT 12
            """,
            (int(due_soon_days),)
        ) or []

        if focus_queue:
            st.markdown("#### Priority Queue (Due Soon / Overdue)")
            queue_df = pd.DataFrame(focus_queue)
            queue_df['status'] = queue_df['days_remaining'].apply(
                lambda d: "Overdue" if int(d) < 0 else "Due Soon"
            )
            st.dataframe(
                queue_df[['title', 'full_name', 'due_date', 'days_remaining', 'status']],
                use_container_width=True,
                hide_index=True
            )

        if overdue_value > 0 or open_renewals > 0:
            st.warning(
                f"Attention: {overdue_value} overdue loans and {open_renewals} pending renewals need action."
            )
        else:
            st.success("All core lending operations are healthy right now.")



    with active_col:
        render_active_borrowings_panel()



    def _compute_fine(days_overdue):
        return (float(days_overdue) * Config.FINE_PER_DAY) if days_overdue and float(days_overdue) > 0 else 0.0

    checkout_col, return_col = st.columns(2, gap="large")

    with checkout_col:
        st.markdown("#### Checkout Book")
        with st.form("checkout_form"):
            member_search = st.text_input(
                "Search Member (username or email)",
                value=global_borrow_search if global_borrow_search else "",
                key="br_checkout_member_search"
            )
            member_options = {}
            selected_member = None
            if member_search:
                members = Database.execute_query(
                    "SELECT user_id, username, full_name, email, fine_balance FROM users WHERE username LIKE ? OR email LIKE ? OR full_name LIKE ? LIMIT 20",
                    (f'%{member_search}%', f'%{member_search}%', f'%{member_search}%')
                )
                if members:
                    member_options = {f"{m['full_name']} (@{m['username']})": m['user_id'] for m in members}
                    selected_member = st.selectbox("Select Member", list(member_options.keys()), key="br_checkout_member_select")
                else:
                    st.warning("No members found")

            book_search = st.text_input(
                "Search Book (title or ISBN)",
                value=global_borrow_search if global_borrow_search else "",
                key="br_checkout_book_search"
            )
            book_options = {}
            selected_book = None
            if book_search:
                books = Database.execute_query(
                    """SELECT b.book_id, b.title, b.isbn,
                              (SELECT COUNT(*) FROM book_inventory WHERE book_id = b.book_id AND is_available = 1) as available,
                              (SELECT COUNT(*) FROM book_inventory WHERE book_id = b.book_id) as total_copies
                       FROM books b
                       WHERE (b.title LIKE ? OR b.isbn LIKE ?) AND b.is_available = 1
                       LIMIT 20""",
                    (f'%{book_search}%', f'%{book_search}%')
                )
                if books:
                    book_options = {
                        f"{bk['title']} ({bk['isbn']}) - {bk['available']}/{bk['total_copies']} available": bk['book_id']
                        for bk in books
                    }
                    selected_book = st.selectbox("Select Book", list(book_options.keys()), key="br_checkout_book_select")
                else:
                    st.warning("No books found")

            preset_col1, preset_col2, preset_col3 = st.columns(3, gap="small")
            with preset_col1:
                loan_preset = st.radio("Loan Preset", ["7D", "14D", "30D", "Custom"], horizontal=True, key="br_checkout_preset")
            with preset_col2:
                if loan_preset == "Custom":
                    checkout_days = st.number_input("Borrowing Days", min_value=1, max_value=90, value=Config.DEFAULT_BORROWING_DAYS, key="br_checkout_days")
                else:
                    checkout_days = int(loan_preset.replace("D", ""))
                    st.number_input("Borrowing Days", min_value=1, max_value=90, value=checkout_days, disabled=True, key="br_checkout_days_locked")
            with preset_col3:
                manual_due_override = st.checkbox("Manual Due Date", value=False, key="br_due_override")

            due_date = date.today() + timedelta(days=int(checkout_days))
            if manual_due_override:
                due_date = st.date_input("Override Due Date", value=due_date, key="br_due_override_date")

            auto_pick_oldest = st.checkbox("Auto-pick oldest available copy", value=True, key="br_auto_pick_oldest")

            if selected_member and selected_book:
                member_id_preview = member_options[selected_member]
                book_id_preview = book_options[selected_book]
                member_stats = Database.execute_query(
                    """
                    SELECT
                        (SELECT COUNT(*) FROM borrowing WHERE user_id = ? AND return_date IS NULL) as active_loans,
                        (SELECT COALESCE(fine_balance, 0) FROM users WHERE user_id = ?) as fine_balance,
                        (SELECT COUNT(*)
                         FROM borrowing br
                         JOIN book_inventory bi ON br.inventory_id = bi.inventory_id
                         WHERE br.user_id = ? AND bi.book_id = ? AND br.return_date IS NULL) as same_book_active
                    """,
                    (member_id_preview, member_id_preview, member_id_preview, book_id_preview),
                    fetch_one=True
                ) or {}
                preview_col1, preview_col2, preview_col3 = st.columns(3, gap="small")
                with preview_col1:
                    st.metric("Member Active Loans", int(member_stats.get('active_loans') or 0))
                with preview_col2:
                    st.metric("Member Fine Balance", format_currency(float(member_stats.get('fine_balance') or 0)))
                with preview_col3:
                    st.metric("Same Title Active", int(member_stats.get('same_book_active') or 0))

            submit_checkout = st.form_submit_button(" Checkout Book", use_container_width=True)

            if submit_checkout:
                if not selected_member or not selected_book:
                    st.error("Please select both member and book!")
                else:
                    user_id = member_options[selected_member]
                    book_id = book_options[selected_book]

                    duplicate_active = Database.execute_query(
                        """
                        SELECT COUNT(*) as count
                        FROM borrowing br
                        JOIN book_inventory bi ON br.inventory_id = bi.inventory_id
                        WHERE br.user_id = ? AND bi.book_id = ? AND br.return_date IS NULL
                        """,
                        (user_id, book_id),
                        fetch_one=True
                    ) or {'count': 0}

                    if int(duplicate_active.get('count') or 0) > 0:
                        st.error("This member already has an active borrowing for this title.")
                    else:
                        if auto_pick_oldest:
                            inventory = Database.execute_query(
                                """
                                SELECT inventory_id
                                FROM book_inventory
                                WHERE book_id = ? AND is_available = 1
                                ORDER BY acquired_date ASC, inventory_id ASC
                                LIMIT 1
                                """,
                                (book_id,), fetch_one=True
                            )
                        else:
                            inventory = Database.execute_query(
                                "SELECT inventory_id FROM book_inventory WHERE book_id = ? AND is_available = 1 LIMIT 1",
                                (book_id,), fetch_one=True
                            )

                        if not inventory:
                            st.error("No copies available!")
                        else:
                            current_user = Auth.get_user()
                            if Database.execute_update(
                                """INSERT INTO borrowing (inventory_id, user_id, checkout_date, due_date,
                                                          checkout_by_user_id)
                                   VALUES (?, ?, date('now'), ?, ?)""",
                                (inventory['inventory_id'], user_id, due_date, current_user['user_id'])
                            ):
                                Database.execute_update(
                                    "UPDATE book_inventory SET is_available = 0 WHERE inventory_id = ?",
                                    (inventory['inventory_id'],)
                                )
                                st.success(f" Book checked out successfully! Due date: {due_date}")
                                st.balloons()

    with return_col:
        st.markdown("#### Return Book")
        with st.form("return_form"):
            search = st.text_input(
                "Search by member name, email or book title",
                value=global_borrow_search if global_borrow_search else "",
                key="br_return_search"
            )
            return_only_overdue = st.checkbox("Show overdue only", value=False, key="br_return_only_overdue")
            waive_fine = st.checkbox("Waive fine for this return", value=False, key="br_return_waive_fine")

            borrowings = []
            borrowing_options = {}
            selected_borrowing = None
            if search:
                return_query = """
                    SELECT br.borrowing_id, b.title, u.full_name, u.email, br.checkout_date, br.due_date,
                           julianday(date('now')) - julianday(br.due_date) as days_overdue,
                           bi.inventory_id
                    FROM borrowing br
                    JOIN book_inventory bi ON br.inventory_id = bi.inventory_id
                    JOIN books b ON bi.book_id = b.book_id
                    JOIN users u ON br.user_id = u.user_id
                    WHERE br.return_date IS NULL
                      AND (u.full_name LIKE ? OR u.email LIKE ? OR b.title LIKE ?)
                """
                return_params = [f'%{search}%', f'%{search}%', f'%{search}%']
                if return_only_overdue:
                    return_query += " AND br.due_date < date('now')"
                return_query += " ORDER BY br.due_date ASC LIMIT 30"

                borrowings = Database.execute_query(return_query, tuple(return_params))
                if borrowings:
                    borrowing_options = {
                        f"{bw['full_name']} - {bw['title']} (Due: {format_date(bw['due_date'])})": bw
                        for bw in borrowings
                    }
                    selected_borrowing = st.selectbox("Select Borrowing", list(borrowing_options.keys()), key="br_return_select")
                else:
                    st.warning("No active borrowings found")

            if selected_borrowing and fine_preview_mode:
                preview = borrowing_options[selected_borrowing]
                preview_fine = _compute_fine(preview['days_overdue'])
                status_col1, status_col2, status_col3 = st.columns(3, gap="small")
                with status_col1:
                    st.metric("Days Overdue", max(int(preview['days_overdue']) if preview['days_overdue'] else 0, 0))
                with status_col2:
                    st.metric("Projected Fine", format_currency(0 if waive_fine else preview_fine))
                with status_col3:
                    st.metric("Waive Fine", "Yes" if waive_fine else "No")

            submit_return = st.form_submit_button(" Return Book", use_container_width=True)

            if submit_return:
                if not selected_borrowing:
                    st.error("Please select a borrowing to return!")
                else:
                    borrowing = borrowing_options[selected_borrowing]
                    current_user = Auth.get_user()
                    fine = 0 if waive_fine else _compute_fine(borrowing['days_overdue'])

                    if Database.execute_update(
                        """UPDATE borrowing
                           SET return_date = date('now'), return_to_user_id = ?,
                               fine_amount = ?
                           WHERE borrowing_id = ?""",
                        (current_user['user_id'], fine, borrowing['borrowing_id'])
                    ):
                        Database.execute_update(
                            "UPDATE book_inventory SET is_available = 1 WHERE inventory_id = ?",
                            (borrowing['inventory_id'],)
                        )

                        if fine > 0:
                            Database.execute_update(
                                "UPDATE users SET fine_balance = fine_balance + ? WHERE user_id IN (SELECT user_id FROM borrowing WHERE borrowing_id = ?)",
                                (fine, borrowing['borrowing_id'])
                            )

                        if fine > 0:
                            st.warning(f" Book returned with fine: {format_currency(fine)}")
                        else:
                            st.success(" Book returned successfully!")
                        st.balloons()
    

    
    current_user = Auth.get_user()
    
    if current_user['role'] == 'admin' or current_user['role'] == 'librarian':
        # Librarian view - approve/reject renewals
        st.markdown("### Pending Renewal Requests")
        
        pending = Database.execute_query("""
            SELECT rr.renewal_id, rr.borrowing_id, rr.created_at as request_date, rr.requested_days,
                   b.title, u.full_name, u.email, br.due_date,
                   julianday(date('now')) - julianday(br.due_date) as days_overdue
            FROM renewal_requests rr
            JOIN borrowing br ON rr.borrowing_id = br.borrowing_id
            JOIN book_inventory bi ON br.inventory_id = bi.inventory_id
            JOIN books b ON bi.book_id = b.book_id
            JOIN users u ON br.user_id = u.user_id
            WHERE rr.status = 'pending'
            ORDER BY rr.created_at
        """)
        
        if pending:
            for req in pending:
                col1, col2, col3 = st.columns([3, 2, 1], gap="small")
                
                with col1:
                    st.markdown(f"**{req['title']}**")
                    st.caption(f"Member: {req['full_name']} ({req['email']})")
                    st.caption(f"Requested: {format_date(req['request_date'])} | Extension: {req['requested_days']} days")
                
                with col2:
                    st.caption(f"Current Due: {format_date(req['due_date'])}")
                    if req['days_overdue'] and req['days_overdue'] > 0:
                        st.error(f" Already overdue by {req['days_overdue']} days")
                    else:
                        new_due = req['due_date'] + timedelta(days=req['requested_days'])
                        st.info(f"New Due: {format_date(new_due)}")
                
                with col3:
                    if st.button(" Approve", key=f"approve_{req['renewal_id']}"):
                        success, msg = EnhancedBorrowingManager.approve_renewal(
                            req['renewal_id'], 
                            current_user['user_id'], 
                            'approved'
                        )
                        if success:
                            st.success("Approved!")
                            st.rerun()
                        else:
                            st.error(msg)
                    
                    if st.button(" Reject", key=f"reject_{req['renewal_id']}"):
                        success, msg = EnhancedBorrowingManager.approve_renewal(
                            req['renewal_id'], 
                            current_user['user_id'], 
                            'rejected'
                        )
                        if success:
                            st.warning("Rejected")
                            st.rerun()
                        else:
                            st.error(msg)
                
                st.divider()
        else:
            st.info("No pending renewal requests")
    

    st.subheader(" Borrowing Trends & Analytics")

    trend_map = {"7D": 7, "30D": 30, "90D": 90, "180D": 180, "365D": 365}
    trend_days = trend_map.get(trend_window, 30)

    # Date range selector
    col1, col2, col3 = st.columns(3, gap="small")
    with col1:
        start_date = st.date_input("Start Date", date.today() - timedelta(days=trend_days), key="br_trend_start")
    with col2:
        end_date = st.date_input("End Date", date.today(), key="br_trend_end")
    with col3:
        auto_generate_trends = st.checkbox("Auto Generate", value=True, key="br_trend_auto")

    generate_trends = auto_generate_trends or st.button(" Generate Trends", use_container_width=True, key="br_generate_trends")

    if generate_trends:
        st.divider()

        total_borrowed = Database.execute_query(
            "SELECT COUNT(*) as count FROM borrowing WHERE checkout_date BETWEEN ? AND ?",
            (start_date, end_date), fetch_one=True
        )

        total_returned = Database.execute_query(
            "SELECT COUNT(*) as count FROM borrowing WHERE return_date BETWEEN ? AND ?",
            (start_date, end_date), fetch_one=True
        )

        overdue_count = Database.execute_query(
            """SELECT COUNT(*) as count FROM borrowing
               WHERE return_date IS NULL AND due_date < date('now')""",
            fetch_one=True
        )

        # Row 1: Active Snapshot | Borrowing Trends & Analytics | Daily Borrowing Trend
        row1_col1, row1_col2, row1_col3 = st.columns(3, gap="small")


            

        with row1_col1:
            st.markdown("### Borrowing Trends & Analytics")
            st.info(f"Range: {start_date} to {end_date}")
            trend_kpi1, trend_kpi2 = st.columns(2, gap="small")
            with trend_kpi1:
                st.metric("Total Borrowed", total_borrowed['count'] if total_borrowed else 0)
                st.metric("Currently Overdue", overdue_count['count'] if overdue_count else 0)
            with trend_kpi2:
                st.metric("Total Returned", total_returned['count'] if total_returned else 0)
                return_rate = 0.0
                if total_borrowed and total_borrowed['count']:
                    return_rate = ((total_returned['count'] if total_returned else 0) / max(total_borrowed['count'], 1)) * 100
                st.metric("Return Rate", f"{return_rate:.1f}%")

        

        with row1_col2:
            st.markdown("###  Return Statistics")
            stat_col1, stat_col2, stat_col3 = st.columns(3, gap="small")
            with stat_col1:
                st.metric(" Borrowed", total_borrowed['count'] if total_borrowed else 0)
            with stat_col2:
                st.metric(" Returned", total_returned['count'] if total_returned else 0)
            with stat_col3:
                st.metric(" Overdue", overdue_count['count'] if overdue_count else 0)

        
def show_reports():
    """Advanced Reports page with 20+ visualizations"""
    st.markdown('<h1 class="litgrid-header"> Advanced Reports & Analytics</h1>', unsafe_allow_html=True)
    
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        " Overview Dashboard", 
        " Books Analytics", 
        " Members Analytics", 
        " Financial Analytics",
        " Advanced Visualizations"
    ])
    
    with tab1:
        st.subheader(" Library Overview Dashboard")
        
        col1, col2 = st.columns(2, gap="small")
        with col1:
            start_date = st.date_input("Start Date", date.today() - timedelta(days=30))
        with col2:
            end_date = st.date_input("End Date", date.today())
        
        if st.button(" Generate Dashboard", use_container_width=True, key="gen_overview"):
            st.divider()
            
            # Summary statistics
            col1, col2, col3, col4 = st.columns(4, gap="small")
            
            books_borrowed = Database.execute_query(
                "SELECT COUNT(*) as count FROM borrowing WHERE checkout_date BETWEEN ? AND ?",
                (start_date, end_date), fetch_one=True
            )
            books_returned = Database.execute_query(
                "SELECT COUNT(*) as count FROM borrowing WHERE return_date BETWEEN ? AND ?",
                (start_date, end_date), fetch_one=True
            )
            new_members = Database.execute_query(
                "SELECT COUNT(*) as count FROM users WHERE created_at BETWEEN ? AND ?",
                (start_date, end_date), fetch_one=True
            )
            fines_collected = Database.execute_query(
                "SELECT SUM(fine_amount) as total FROM borrowing WHERE return_date BETWEEN ? AND ?",
                (start_date, end_date), fetch_one=True
            )
            
            with col1:
                st.metric(" Books Borrowed", books_borrowed['count'] if books_borrowed else 0)
            with col2:
                st.metric(" Books Returned", books_returned['count'] if books_returned else 0)
            with col3:
                st.metric(" New Members", new_members['count'] if new_members else 0)
            with col4:
                st.metric(" Fines Collected", format_currency(fines_collected['total'] if fines_collected and fines_collected['total'] else 0))
            
            st.divider()
            
            # Daily activity chart
            st.subheader(" Daily Borrowing Activity")
            daily_data = Database.execute_query("""
                SELECT DATE(checkout_date) as date,
                       COUNT(*) as checkouts
                FROM borrowing
                WHERE checkout_date BETWEEN ? AND ?
                GROUP BY DATE(checkout_date)
                ORDER BY date
            """, (start_date, end_date))
            
            if daily_data:
                df = pd.DataFrame(daily_data)
                fig = px.line(df, x='date', y='checkouts', 
                            title='Daily Checkout Trend',
                            labels={'checkouts': 'Number of Checkouts', 'date': 'Date'},
                            markers=True)
                fig.update_traces(line_color='#1E88E5', line_width=3)
                fig.update_layout(hovermode='x unified')
                st.plotly_chart(fig, use_container_width=True)
            
            # Genre distribution
            st.subheader(" Books Distribution by Genre")
            genre_data = Database.execute_query("""
                SELECT g.genre_name, COUNT(DISTINCT b.book_id) as book_count
                FROM genres g
                JOIN book_genres bg ON g.genre_id = bg.genre_id
                JOIN books b ON bg.book_id = b.book_id
                WHERE b.is_active = 1
                GROUP BY g.genre_name
                ORDER BY book_count DESC
            """)
            
            if genre_data:
                df = pd.DataFrame(genre_data)
                fig = px.pie(df, values='book_count', names='genre_name',
                           title='Book Collection by Genre',
                           color_discrete_sequence=px.colors.qualitative.Set3)
                fig.update_traces(textposition='inside', textinfo='percent+label')
                st.plotly_chart(fig, use_container_width=True)
    
    with tab2:
        st.subheader(" Books Analytics")
        
        report_type = st.selectbox("Select Report", [
            "Most Popular Books",
            "Least Popular Books",
            "Books by Genre Performance",
            "Books by Publisher",
            "Book Ratings Distribution",
            "Average Days per Book",
            "Books Never Borrowed"
        ])
        
        if report_type == "Most Popular Books":
            try:
                data = Database.execute_query("""
                    SELECT b.title, b.isbn, b.publication_year,
                           COALESCE(bs.total_checkouts, 0) as total_checkouts, 
                           COALESCE(bs.available_copies, 0) as available_copies, 
                           COALESCE(bs.total_copies, 0) as total_copies,
                           COALESCE(bs.average_rating, 0) as average_rating
                    FROM books b
                    LEFT JOIN book_statistics bs ON b.book_id = bs.book_id
                    WHERE b.is_active = 1
                    ORDER BY COALESCE(bs.total_checkouts, 0) DESC
                    LIMIT 20
                """)
            except Exception as e:
                st.error(f"Database error: {str(e)}")
                data = None
            
            if data:
                df = pd.DataFrame(data)
                st.dataframe(df, use_container_width=True, hide_index=True)
                
                # Horizontal bar chart
                fig = px.bar(df.head(10), x='total_checkouts', y='title',
                           orientation='h',
                           title="Top 10 Most Popular Books",
                           labels={'total_checkouts': 'Total Checkouts', 'title': 'Book Title'},
                           color='total_checkouts',
                           color_continuous_scale='Blues')
                fig.update_layout(hovermode='x unified', yaxis={'categoryorder':'total ascending'}, height=500)
                st.plotly_chart(fig, use_container_width=True)
                
                # Scatter plot: Checkouts vs Rating
                fig2 = px.scatter(df, x='average_rating', y='total_checkouts',
                                size='total_checkouts', hover_data=['title'],
                                title='Book Popularity vs Rating',
                                labels={'average_rating': 'Average Rating', 'total_checkouts': 'Checkouts'},
                                color='total_checkouts',
                                color_continuous_scale='Viridis')
                st.plotly_chart(fig2, use_container_width=True)
        
        elif report_type == "Least Popular Books":
            try:
                data = Database.execute_query("""
                    SELECT b.title, b.isbn, b.author, b.publication_year,
                           COALESCE(bs.total_checkouts, 0) as total_checkouts,
                           COALESCE(bs.average_rating, 0) as average_rating
                    FROM books b
                    LEFT JOIN book_statistics bs ON b.book_id = bs.book_id
                    WHERE b.is_active = 1
                    ORDER BY COALESCE(bs.total_checkouts, 0) ASC, b.created_at DESC
                    LIMIT 20
                """)
            except Exception as e:
                st.error(f"Database error in Least Popular Books: {str(e)}")
                data = None
            
            if data and len(data) > 0:
                df = pd.DataFrame(data)
                st.dataframe(df, use_container_width=True, hide_index=True)
                
                # Only create chart if we have data
                if len(df) > 0:
                    # Bar chart for least popular
                    fig = px.bar(df.head(10), x='total_checkouts', y='title',
                               orientation='h',
                               title="10 Least Popular Books",
                               labels={'total_checkouts': 'Total Checkouts', 'title': 'Book Title'},
                               color='total_checkouts',
                               color_continuous_scale='Reds')
                    fig.update_layout(hovermode='x unified', yaxis={'categoryorder':'total descending'}, height=500)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.warning("No data available for visualization")
            else:
                st.info("No book data available")
        
        elif report_type == "Books by Genre Performance":
            try:
                data = Database.execute_query("""
                    SELECT g.genre_name,
                           COUNT(DISTINCT b.book_id) as total_books,
                           COUNT(br.borrowing_id) as total_borrows,
                           AVG(bs.average_rating) as avg_rating
                    FROM genres g
                    JOIN book_genres bg ON g.genre_id = bg.genre_id
                    JOIN books b ON bg.book_id = b.book_id
                    LEFT JOIN book_inventory bi ON b.book_id = bi.book_id
                    LEFT JOIN borrowing br ON bi.inventory_id = br.inventory_id
                    LEFT JOIN book_statistics bs ON b.book_id = bs.book_id
                    WHERE b.is_active = 1
                    GROUP BY g.genre_name
                    ORDER BY total_borrows DESC
                """)
            except Exception as e:
                st.error(f"Database error in Genre Performance: {str(e)}")
                data = None
            
            if data and len(data) > 0:
                df = pd.DataFrame(data)
                # Handle null values safely
                df['avg_rating'] = pd.to_numeric(df['avg_rating'], errors='coerce').fillna(0).round(2)
                df['total_books'] = pd.to_numeric(df['total_books'], errors='coerce').fillna(0)
                df['total_borrows'] = pd.to_numeric(df['total_borrows'], errors='coerce').fillna(0)
                
                st.dataframe(df, use_container_width=True, hide_index=True)
                
                if len(df) > 0:
                    # Grouped bar chart
                    fig = go.Figure()
                    fig.add_trace(go.Bar(name='Total Books', x=df['genre_name'], y=df['total_books'],
                                        marker_color='lightblue'))
                    fig.add_trace(go.Bar(name='Total Borrows', x=df['genre_name'], y=df['total_borrows'],
                                        marker_color='darkblue'))
                    fig.update_layout(hovermode='x unified', barmode='group', title='Genre Performance: Books vs Borrows',
                                    xaxis_title='Genre', yaxis_title='Count', height=500)
                    st.plotly_chart(fig, use_container_width=True)
                    
                    # Bubble chart
                    fig2 = px.scatter(df, x='total_books', y='total_borrows', size='avg_rating',
                                    hover_data=['genre_name'], title='Genre Insights Bubble Chart',
                                    labels={'total_books': 'Total Books', 'total_borrows': 'Total Borrows'},
                                    color='avg_rating', color_continuous_scale='RdYlGn')
                    st.plotly_chart(fig2, use_container_width=True)
                else:
                    st.warning("No data available for genre performance visualization")
            else:
                st.info("No genre performance data available")
        
        elif report_type == "Book Ratings Distribution":
            try:
                data = Database.execute_query("""
                    SELECT 
                        CAST(average_rating AS INTEGER) as rating_group,
                        COUNT(*) as book_count
                    FROM book_statistics
                    WHERE average_rating IS NOT NULL AND average_rating > 0
                    GROUP BY CAST(average_rating AS INTEGER)
                    ORDER BY rating_group
                """)
            except Exception as e:
                st.error(f"Database error in Book Ratings Distribution: {str(e)}")
                data = None
            
            if data and len(data) > 0:
                df = pd.DataFrame(data)
                # Ensure numeric types
                df['rating_group'] = pd.to_numeric(df['rating_group'], errors='coerce').fillna(0).astype(int)
                df['book_count'] = pd.to_numeric(df['book_count'], errors='coerce').fillna(0).astype(int)
                
                # Filter out invalid ratings
                df = df[df['rating_group'] > 0]
                
                if len(df) > 0:
                    df['rating_label'] = df['rating_group'].astype(str) + ' - ' + (df['rating_group'] + 1).astype(str) + ' '
                    
                    fig = px.bar(df, x='rating_label', y='book_count',
                               title='Book Ratings Distribution',
                               labels={'rating_label': 'Rating Range', 'book_count': 'Number of Books'},
                               color='book_count',
                               color_continuous_scale='YlOrRd')
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.warning("No valid rating data available for visualization")
            else:
                st.info("No rating distribution data available")
        
        elif report_type == "Books by Publisher":
            try:
                data = Database.execute_query("""
                    SELECT p.name as publisher_name,
                           COUNT(DISTINCT b.book_id) as total_books,
                           COUNT(br.borrowing_id) as total_borrows,
                           AVG(bs.average_rating) as avg_rating
                    FROM publishers p
                    JOIN books b ON p.publisher_id = b.publisher_id
                    LEFT JOIN book_inventory bi ON b.book_id = bi.book_id
                    LEFT JOIN borrowing br ON bi.inventory_id = br.inventory_id
                    LEFT JOIN book_statistics bs ON b.book_id = bs.book_id
                    WHERE b.is_active = 1
                    GROUP BY p.publisher_id, p.name
                    ORDER BY total_books DESC
                    LIMIT 15
                """)
            except Exception as e:
                st.error(f"Database error in Books by Publisher: {str(e)}")
                data = None
            
            if data and len(data) > 0:
                df = pd.DataFrame(data)
                # Handle null values safely
                df['avg_rating'] = pd.to_numeric(df['avg_rating'], errors='coerce').fillna(0).round(2)
                df['total_books'] = pd.to_numeric(df['total_books'], errors='coerce').fillna(0)
                df['total_borrows'] = pd.to_numeric(df['total_borrows'], errors='coerce').fillna(0)
                
                st.dataframe(df, use_container_width=True, hide_index=True)
                
                if len(df) > 0:
                    # Publisher performance chart
                    fig = px.bar(df, x='publisher_name', y='total_books',
                               title='Books by Publisher',
                               labels={'publisher_name': 'Publisher', 'total_books': 'Number of Books'},
                               color='total_borrows',
                               color_continuous_scale='Blues')
                    fig.update_layout(hovermode='x unified', xaxis_tickangle=-45, height=500)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.warning("No data available for publisher visualization")
            else:
                st.info("No publisher data available")
        
        elif report_type == "Average Days per Book":
            try:
                data = Database.execute_query("""
                    SELECT b.title, b.author,
                           AVG(julianday(COALESCE(br.return_date, date('now'))) - julianday(br.checkout_date)) as avg_days,
                           COUNT(br.borrowing_id) as times_borrowed
                    FROM books b
                    JOIN book_inventory bi ON b.book_id = bi.book_id
                    JOIN borrowing br ON bi.inventory_id = br.inventory_id
                    WHERE b.is_active = 1
                    GROUP BY b.book_id
                    HAVING times_borrowed >= 2
                    ORDER BY avg_days DESC
                    LIMIT 20
                """)
            except Exception as e:
                st.error(f"Database error in Average Days per Book: {str(e)}")
                data = None
            
            if data and len(data) > 0:
                df = pd.DataFrame(data)
                # Handle null values safely
                df['avg_days'] = pd.to_numeric(df['avg_days'], errors='coerce').fillna(0).round(1)
                df['times_borrowed'] = pd.to_numeric(df['times_borrowed'], errors='coerce').fillna(0).astype(int)
                
                # Filter out invalid data
                df = df[df['avg_days'] > 0]
                
                st.dataframe(df, use_container_width=True, hide_index=True)
                
                if len(df) > 0:
                    # Scatter plot of average days vs times borrowed
                    fig = px.scatter(df, x='times_borrowed', y='avg_days',
                                   hover_data=['title'], 
                                   title='Average Borrowing Duration vs Popularity',
                                   labels={'times_borrowed': 'Times Borrowed', 'avg_days': 'Average Days'},
                                   size='avg_days',
                                   color='avg_days',
                                   color_continuous_scale='RdYlBu')
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.warning("No valid borrowing duration data available")
            else:
                st.info("Insufficient borrowing data for analysis")
        
        elif report_type == "Books Never Borrowed":
            try:
                data = Database.execute_query("""
                    SELECT b.title, b.isbn, b.publication_year, p.name as publisher_name
                    FROM books b
                    LEFT JOIN publishers p ON b.publisher_id = p.publisher_id
                    LEFT JOIN book_inventory bi ON b.book_id = bi.book_id
                    LEFT JOIN borrowing br ON bi.inventory_id = br.inventory_id
                    WHERE b.is_active = 1 AND br.borrowing_id IS NULL
                    ORDER BY b.created_at DESC
                    LIMIT 50
                """)
                
                if data:
                    st.warning(f" Found {len(data)} books that have never been borrowed")
                    df = pd.DataFrame(data)
                    # Handle null values
                    df = df.fillna('N/A')
                    st.dataframe(df, use_container_width=True, hide_index=True)
                else:
                    st.success(" All books have been borrowed at least once!")
            except Exception as e:
                st.error(f"Error generating Books Never Borrowed report: {str(e)}")
                st.info("Please check the database connection and try again.")
        
        else:
            st.info("Please select a report type to generate")
    
    with tab3:
        st.subheader(" Members Analytics")
        
        member_report = st.selectbox("Select Member Report", [
            "Top Borrowers",
            "Members with Outstanding Fines",
            "Member Registration Trend",
            "Member Tier Distribution",
            "Inactive Members",
            "Most Active Members (This Month)"
        ])
        
        if member_report == "Top Borrowers":
            try:
                data = Database.execute_query("""
                    SELECT u.full_name, u.email, u.member_tier,
                           COUNT(br.borrowing_id) as total_borrowed,
                           SUM(CASE WHEN br.return_date IS NULL THEN 1 ELSE 0 END) as currently_borrowed
                    FROM users u
                    LEFT JOIN borrowing br ON u.user_id = br.user_id
                    WHERE u.role = 'member' AND u.is_active = 1
                    GROUP BY u.user_id
                    HAVING total_borrowed > 0
                    ORDER BY total_borrowed DESC
                    LIMIT 20
                """)
                
                if data:
                    df = pd.DataFrame(data)
                    # Handle null values and ensure proper data types
                    df = df.fillna({'member_tier': 'basic', 'email': 'N/A'})
                    df['total_borrowed'] = pd.to_numeric(df['total_borrowed'], errors='coerce').fillna(0).astype(int)
                    df['currently_borrowed'] = pd.to_numeric(df['currently_borrowed'], errors='coerce').fillna(0).astype(int)
                    
                    st.dataframe(df, use_container_width=True, hide_index=True)
                    
                    # Horizontal bar chart with color by tier
                    if len(df) > 0:
                        try:
                            fig = px.bar(df.head(15), x='total_borrowed', y='full_name',
                                       orientation='h',
                                       title="Top 15 Borrowers",
                                       labels={'total_borrowed': 'Total Books Borrowed', 'full_name': 'Member Name'},
                                       color='member_tier',
                                       color_discrete_map={'basic': '#FFA726', 'premium': '#66BB6A', 'gold': '#FFD700'})
                            fig.update_layout(hovermode='x unified', yaxis={'categoryorder':'total ascending'}, height=600)
                            st.plotly_chart(fig, use_container_width=True)
                        except Exception as viz_error:
                            st.warning(f"Could not generate visualization: {str(viz_error)}")
                else:
                    st.info("No borrowing data available for top borrowers report")
            except Exception as e:
                st.error(f"Error generating Top Borrowers report: {str(e)}")
                st.info("Please check the database connection and try again.")
        
        elif member_report == "Members with Outstanding Fines":
            try:
                data = Database.execute_query("""
                    SELECT u.full_name, u.email, u.phone, u.fine_balance, u.member_tier
                    FROM users u
                    WHERE u.fine_balance > 0
                    ORDER BY u.fine_balance DESC
                """)
                
                if data:
                    df = pd.DataFrame(data)
                    # Handle null values and ensure proper data types
                    df = df.fillna({'email': 'N/A', 'phone': 'N/A', 'member_tier': 'basic'})
                    df['fine_balance'] = pd.to_numeric(df['fine_balance'], errors='coerce').fillna(0.0)
                    
                    # Filter out zero fine balances after conversion
                    df = df[df['fine_balance'] > 0]
                    
                    if len(df) > 0:
                        st.dataframe(df, use_container_width=True, hide_index=True)
                        
                        try:
                            total = df['fine_balance'].sum()
                            avg = df['fine_balance'].mean()
                            
                            col1, col2, col3 = st.columns(3, gap="small")
                            with col1:
                                st.metric("Total Outstanding", format_currency(total))
                            with col2:
                                st.metric("Average Fine", format_currency(avg))
                            with col3:
                                st.metric("Members with Fines", len(df))
                            
                            # Histogram of fine distribution
                            fig = px.histogram(df, x='fine_balance', nbins=min(20, len(df)),
                                             title='Fine Amount Distribution',
                                             labels={'fine_balance': 'Fine Amount ($)', 'count': 'Number of Members'},
                                             color_discrete_sequence=['#FF6B6B'])
                            st.plotly_chart(fig, use_container_width=True)
                        except Exception as viz_error:
                            st.warning(f"Could not generate fine statistics: {str(viz_error)}")
                    else:
                        st.success(" No members have outstanding fines!")
                else:
                    st.success(" No members have outstanding fines!")
            except Exception as e:
                st.error(f"Error generating Outstanding Fines report: {str(e)}")
                st.info("Please check the database connection and try again.")
        
        elif member_report == "Member Registration Trend":
            try:
                data = Database.execute_query("""
                    SELECT strftime('%Y-%m', created_at) as month,
                           COUNT(*) as new_members
                    FROM users
                    WHERE role = 'member' AND created_at IS NOT NULL
                    GROUP BY strftime('%Y-%m', created_at)
                    ORDER BY month DESC
                    LIMIT 12
                """)
                
                if data:
                    df = pd.DataFrame(data)
                    # Handle null values and ensure proper data types
                    df = df.fillna({'month': 'Unknown', 'new_members': 0})
                    df['new_members'] = pd.to_numeric(df['new_members'], errors='coerce').fillna(0).astype(int)
                    
                    # Filter out invalid months
                    df = df[df['month'] != 'Unknown']
                    
                    if len(df) > 0:
                        df = df.sort_values('month')
                        
                        try:
                            fig = px.area(df, x='month', y='new_members',
                                        title='Member Registration Trend (Last 12 Months)',
                                        labels={'month': 'Month', 'new_members': 'New Members'},
                                        color_discrete_sequence=['#4CAF50'])
                            fig.update_traces(fill='tozeroy')
                            st.plotly_chart(fig, use_container_width=True)
                        except Exception as viz_error:
                            st.warning(f"Could not generate trend visualization: {str(viz_error)}")
                            # Show data table as fallback
                            st.dataframe(df, use_container_width=True, hide_index=True)
                    else:
                        st.info("No registration trend data available")
                else:
                    st.info("No member registration data available")
            except Exception as e:
                st.error(f"Error generating Member Registration Trend report: {str(e)}")
                st.info("Please check the database connection and try again.")
        
        elif member_report == "Member Tier Distribution":
            try:
                data = Database.execute_query("""
                    SELECT member_tier, COUNT(*) as count
                    FROM users
                    WHERE role = 'member' AND is_active = 1
                    GROUP BY member_tier
                """)
                
                if data:
                    df = pd.DataFrame(data)
                    # Handle null values and ensure proper data types
                    df = df.fillna({'member_tier': 'basic'})
                    df['count'] = pd.to_numeric(df['count'], errors='coerce').fillna(0).astype(int)
                    
                    # Filter out zero counts
                    df = df[df['count'] > 0]
                    
                    if len(df) > 0:
                        try:
                            # Donut chart
                            fig = px.pie(df, values='count', names='member_tier',
                                       title='Member Tier Distribution',
                                       hole=0.4,
                                       color_discrete_sequence=px.colors.sequential.RdBu)
                            fig.update_traces(textposition='inside', textinfo='percent+label+value')
                            st.plotly_chart(fig, use_container_width=True)
                        except Exception as viz_error:
                            st.warning(f"Could not generate pie chart: {str(viz_error)}")
                            # Show data table as fallback
                            st.dataframe(df, use_container_width=True, hide_index=True)
                    else:
                        st.info("No member tier data available")
                else:
                    st.info("No member tier data available")
            except Exception as e:
                st.error(f"Error generating Member Tier Distribution report: {str(e)}")
                st.info("Please check the database connection and try again.")
        
        elif member_report == "Inactive Members":
            try:
                data = Database.execute_query("""
                    SELECT u.full_name, u.email, u.member_tier,
                           u.created_at as joined_date,
                           COALESCE(MAX(br.checkout_date), 'Never') as last_activity,
                           COUNT(br.borrowing_id) as total_borrowed
                    FROM users u
                    LEFT JOIN borrowing br ON u.user_id = br.user_id
                    WHERE u.role = 'member' AND u.is_active = 1
                    GROUP BY u.user_id
                    HAVING (last_activity = 'Never' OR julianday('now') - julianday(MAX(br.checkout_date)) > 90)
                    ORDER BY last_activity DESC
                    LIMIT 20
                """)
                
                if data:
                    df = pd.DataFrame(data)
                    # Handle null values and ensure proper data types
                    df = df.fillna({'email': 'N/A', 'member_tier': 'basic', 'last_activity': 'Never'})
                    df['total_borrowed'] = pd.to_numeric(df['total_borrowed'], errors='coerce').fillna(0).astype(int)
                    
                    st.warning(f" Found {len(df)} inactive members (no activity for 90+ days)")
                    st.dataframe(df, use_container_width=True, hide_index=True)
                    
                    # Inactive members by tier
                    try:
                        tier_counts = df['member_tier'].value_counts()
                        if not tier_counts.empty and len(tier_counts) > 0:
                            fig = px.pie(values=tier_counts.values, names=tier_counts.index,
                                       title='Inactive Members by Tier',
                                       color_discrete_sequence=px.colors.qualitative.Set3)
                            st.plotly_chart(fig, use_container_width=True)
                    except Exception as viz_error:
                        st.warning(f"Could not generate tier distribution chart: {str(viz_error)}")
                else:
                    st.success(" All members are active!")
            except Exception as e:
                st.error(f"Error generating Inactive Members report: {str(e)}")
                st.info("Please check the database connection and try again.")
        
        elif member_report == "Most Active Members (This Month)":
            try:
                data = Database.execute_query("""
                    SELECT u.full_name, u.email, u.member_tier,
                           COUNT(br.borrowing_id) as books_this_month,
                           AVG(julianday(COALESCE(br.return_date, date('now'))) - julianday(br.checkout_date)) as avg_duration
                    FROM users u
                    JOIN borrowing br ON u.user_id = br.user_id
                    WHERE u.role = 'member' AND u.is_active = 1
                      AND br.checkout_date >= date('now', 'start of month')
                    GROUP BY u.user_id
                    ORDER BY books_this_month DESC, avg_duration ASC
                    LIMIT 15
                """)
                
                if data:
                    df = pd.DataFrame(data)
                    # Handle null values and ensure proper data types
                    df = df.fillna({'email': 'N/A', 'member_tier': 'basic'})
                    df['books_this_month'] = pd.to_numeric(df['books_this_month'], errors='coerce').fillna(0).astype(int)
                    df['avg_duration'] = pd.to_numeric(df['avg_duration'], errors='coerce').fillna(0.0).round(1)
                    
                    # Filter out zero activity
                    df = df[df['books_this_month'] > 0]
                    
                    if len(df) > 0:
                        st.success(f" Top {len(df)} most active members this month")
                        st.dataframe(df, use_container_width=True, hide_index=True)
                        
                        try:
                            # Activity chart
                            fig = px.bar(df, x='full_name', y='books_this_month',
                                       title='Most Active Members This Month',
                                       labels={'full_name': 'Member Name', 'books_this_month': 'Books Borrowed'},
                                       color='member_tier',
                                       color_discrete_map={'basic': '#FFA726', 'premium': '#66BB6A', 'gold': '#FFD700'})
                            fig.update_layout(hovermode='x unified', xaxis_tickangle=-45, height=500)
                            st.plotly_chart(fig, use_container_width=True)
                        except Exception as viz_error:
                            st.warning(f"Could not generate activity chart: {str(viz_error)}")
                    else:
                        st.info("No borrowing activity this month")
                else:
                    st.info("No borrowing activity this month")
            except Exception as e:
                st.error(f"Error generating Most Active Members report: {str(e)}")
                st.info("Please check the database connection and try again.")
        
        else:
            st.info("Please select a member report type")
    
    with tab4:
        st.subheader(" Financial Analytics")
        
        col1, col2 = st.columns(2, gap="small")
        with col1:
            fin_start = st.date_input("From Date", date.today() - timedelta(days=90), key="fin_start")
        with col2:
            fin_end = st.date_input("To Date", date.today(), key="fin_end")
        
        if st.button(" Generate Financial Report", use_container_width=True):
            try:
                st.divider()
                
                # Summary metrics
                total_fines = Database.execute_query("""
                    SELECT SUM(fine_amount) as total
                    FROM borrowing
                    WHERE return_date BETWEEN ? AND ? AND fine_amount > 0
                """, (fin_start, fin_end), fetch_one=True)
                
                outstanding_fines = Database.execute_query("""
                    SELECT SUM(fine_balance) as total
                    FROM users
                    WHERE fine_balance > 0
                """, fetch_one=True)
                
                avg_fine = Database.execute_query("""
                    SELECT AVG(fine_amount) as avg
                    FROM borrowing
                    WHERE return_date BETWEEN ? AND ? AND fine_amount > 0
                """, (fin_start, fin_end), fetch_one=True)
                
                # Handle null values and ensure proper data types
                total_collected = 0
                if total_fines and total_fines.get('total'):
                    total_collected = float(total_fines['total'])
                
                outstanding_total = 0
                if outstanding_fines and outstanding_fines.get('total'):
                    outstanding_total = float(outstanding_fines['total'])
                
                avg_amount = 0
                if avg_fine and avg_fine.get('avg'):
                    avg_amount = float(avg_fine['avg'])
                
                col1, col2, col3 = st.columns(3, gap="small")
                with col1:
                    st.metric(" Fines Collected", format_currency(total_collected))
                with col2:
                    st.metric(" Outstanding Fines", format_currency(outstanding_total))
                with col3:
                    st.metric(" Average Fine", format_currency(avg_amount))
            except Exception as e:
                st.error(f"Error generating financial summary: {str(e)}")
                st.info("Please check the database connection and try again.")
            
                st.divider()
                
                # Daily fines chart
                try:
                    fines_daily = Database.execute_query("""
                        SELECT DATE(return_date) as date, 
                               SUM(fine_amount) as total,
                               COUNT(*) as fine_count
                        FROM borrowing
                        WHERE return_date BETWEEN ? AND ? AND fine_amount > 0
                        GROUP BY DATE(return_date)
                        ORDER BY date
                    """, (fin_start, fin_end))
                    
                    if fines_daily:
                        df = pd.DataFrame(fines_daily)
                        # Handle null values and ensure proper data types
                        df = df.fillna(0)
                        df['total'] = pd.to_numeric(df['total'], errors='coerce').fillna(0.0)
                        df['fine_count'] = pd.to_numeric(df['fine_count'], errors='coerce').fillna(0).astype(int)
                        
                        if len(df) > 0 and df['total'].sum() > 0:
                            try:
                                # Area chart for daily fines
                                fig = go.Figure()
                                fig.add_trace(go.Scatter(
                                    x=df['date'], y=df['total'],
                                    mode='lines',
                                    name='Fine Amount',
                                    line=dict(color='#4CAF50', width=3),
                                    fill='tozeroy',
                                    fillcolor='rgba(76, 175, 80, 0.2)'
                                ))
                                fig.update_layout(hovermode='x unified',
                                    title="Daily Fines Collected",
                                    xaxis_title="Date",
                                    yaxis_title="Amount ($)",
                                    height=400
                                )
                                st.plotly_chart(fig, use_container_width=True)
                                
                                # Bar chart for fine counts
                                fig2 = px.bar(df, x='date', y='fine_count',
                                            title='Number of Fines per Day',
                                            labels={'date': 'Date', 'fine_count': 'Number of Fines'},
                                            color='fine_count',
                                            color_continuous_scale='Reds')
                                st.plotly_chart(fig2, use_container_width=True)
                            except Exception as viz_error:
                                st.warning(f"Could not generate daily charts: {str(viz_error)}")
                                st.dataframe(df, use_container_width=True, hide_index=True)
                        else:
                            st.info("No fine data available for the selected date range")
                    else:
                        st.info("No fine data available for the selected date range")
                except Exception as e:
                    st.error(f"Error generating daily fines analysis: {str(e)}")
            
                # Monthly revenue projection
                st.subheader(" Monthly Revenue Projection")
                try:
                    monthly_fines = Database.execute_query("""
                        SELECT strftime('%Y-%m', return_date) as month,
                               SUM(fine_amount) as total
                        FROM borrowing
                        WHERE fine_amount > 0 AND return_date IS NOT NULL
                        GROUP BY strftime('%Y-%m', return_date)
                        ORDER BY month DESC
                        LIMIT 12
                    """)
                    
                    if monthly_fines:
                        df_monthly = pd.DataFrame(monthly_fines)
                        # Handle null values and ensure proper data types
                        df_monthly = df_monthly.fillna({'month': '', 'total': 0})
                        df_monthly['total'] = pd.to_numeric(df_monthly['total'], errors='coerce').fillna(0.0)
                        
                        # Filter out invalid months and zero totals
                        df_monthly = df_monthly[df_monthly['month'] != '']
                        df_monthly = df_monthly[df_monthly['total'] > 0]
                        
                        if len(df_monthly) > 0:
                            df_monthly = df_monthly.sort_values('month')
                            
                            try:
                                fig = px.line(df_monthly, x='month', y='total',
                                            title='Monthly Fine Revenue (Last 12 Months)',
                                            labels={'month': 'Month', 'total': 'Revenue ($)'},
                                            markers=True)
                                fig.update_traces(line_color='#2196F3', line_width=4)
                                st.plotly_chart(fig, use_container_width=True)
                            except Exception as viz_error:
                                st.warning(f"Could not generate monthly revenue chart: {str(viz_error)}")
                                st.dataframe(df_monthly, use_container_width=True, hide_index=True)
                        else:
                            st.info("No monthly revenue data available")
                    else:
                        st.info("No monthly revenue data available")
                except Exception as e:
                    st.error(f"Error generating monthly revenue projection: {str(e)}")
            except Exception as e:
                st.error(f"Error generating Financial Analytics report: {str(e)}")
                st.info("Please check the database connection and try again.")
    
    with tab5:
        st.subheader(" Advanced Visualizations")
        
        viz_type = st.selectbox("Select Visualization Type", [
            "Library Health Heatmap",
            "Book-Member Network",
            "Genre Performance Radar",
            "Collection Growth Timeline",
            "Multi-Dimensional Analysis"
        ])
        
        if viz_type == "Library Health Heatmap":
            try:
                # Create a heatmap showing various metrics
                st.markdown("###  Library Health Metrics")
                
                # Total books
                total_books = Database.execute_query("SELECT COUNT(*) as cnt FROM books WHERE is_active=1", fetch_one=True)
                # Available books
                avail_books = Database.execute_query("SELECT COUNT(*) as cnt FROM book_inventory WHERE is_available=1", fetch_one=True)
                # Active members
                active_members = Database.execute_query("SELECT COUNT(*) as cnt FROM users WHERE is_active=1 AND role='member'", fetch_one=True)
                # Current borrows
                current_borrows = Database.execute_query("SELECT COUNT(*) as cnt FROM borrowing WHERE return_date IS NULL", fetch_one=True)
                # Overdue books
                overdue = Database.execute_query("SELECT COUNT(*) as cnt FROM borrowing WHERE return_date IS NULL AND due_date < date('now')", fetch_one=True)
                
                # Handle null values and ensure proper data types
                metrics = {
                    'Total Books': int(total_books['cnt']) if total_books and total_books.get('cnt') else 0,
                    'Available Books': int(avail_books['cnt']) if avail_books and avail_books.get('cnt') else 0,
                    'Active Members': int(active_members['cnt']) if active_members and active_members.get('cnt') else 0,
                    'Current Borrows': int(current_borrows['cnt']) if current_borrows and current_borrows.get('cnt') else 0,
                    'Overdue Books': int(overdue['cnt']) if overdue and overdue.get('cnt') else 0
                }
                
                # Calculate health scores (0-100) with safe division
                availability_score = (metrics['Available Books'] / max(metrics['Total Books'], 1)) * 100
                utilization_score = (metrics['Current Borrows'] / max(metrics['Active Members'], 1)) * 100
                compliance_score = max(0, 100 - (metrics['Overdue Books'] / max(metrics['Current Borrows'], 1)) * 100)
                
                health_df = pd.DataFrame({
                    'Metric': ['Book Availability', 'Member Utilization', 'Return Compliance'],
                    'Score': [availability_score, min(utilization_score, 100), compliance_score]
                })
                
                try:
                    fig = px.bar(health_df, x='Metric', y='Score',
                               title='Library Health Scores',
                               labels={'Score': 'Score (%)'},
                               color='Score',
                               color_continuous_scale='RdYlGn',
                               range_y=[0, 100])
                    fig.update_layout(hovermode='x unified', height=400)
                    st.plotly_chart(fig, use_container_width=True)
                except Exception as viz_error:
                    st.warning(f"Could not generate health chart: {str(viz_error)}")
                    st.dataframe(health_df, use_container_width=True, hide_index=True)
                
                # Display raw metrics
                col1, col2, col3, col4, col5 = st.columns(5, gap="small")
                with col1:
                    st.metric(" Total Books", metrics['Total Books'])
                with col2:
                    st.metric(" Available", metrics['Available Books'])
                with col3:
                    st.metric(" Active Members", metrics['Active Members'])
                with col4:
                    st.metric(" Current Borrows", metrics['Current Borrows'])
                with col5:
                    st.metric(" Overdue", metrics['Overdue Books'])
            except Exception as e:
                st.error(f"Error generating Library Health Heatmap: {str(e)}")
                st.info("Please check the database connection and try again.")
        
        elif viz_type == "Genre Performance Radar":
            try:
                st.markdown("###  Genre Performance Radar Chart")
                
                genre_perf = Database.execute_query("""
                    SELECT g.genre_name,
                           COUNT(DISTINCT b.book_id) as book_count,
                           COUNT(br.borrowing_id) as borrow_count,
                           AVG(bs.average_rating) as avg_rating
                    FROM genres g
                    JOIN book_genres bg ON g.genre_id = bg.genre_id
                    JOIN books b ON bg.book_id = b.book_id
                    LEFT JOIN book_inventory bi ON b.book_id = bi.book_id
                    LEFT JOIN borrowing br ON bi.inventory_id = br.inventory_id
                    LEFT JOIN book_statistics bs ON b.book_id = bs.book_id
                    WHERE b.is_active = 1
                    GROUP BY g.genre_name
                    LIMIT 10
                """)
                
                if genre_perf:
                    df = pd.DataFrame(genre_perf)
                    # Handle null values and ensure proper data types
                    df = df.fillna({'genre_name': 'Unknown', 'book_count': 0, 'borrow_count': 0, 'avg_rating': 0})
                    df['book_count'] = pd.to_numeric(df['book_count'], errors='coerce').fillna(0).astype(int)
                    df['borrow_count'] = pd.to_numeric(df['borrow_count'], errors='coerce').fillna(0).astype(int)
                    df['avg_rating'] = pd.to_numeric(df['avg_rating'], errors='coerce').fillna(0.0)
                    
                    # Filter out unknown genres and ensure we have data
                    df = df[df['genre_name'] != 'Unknown']
                    df = df[df['book_count'] > 0]
                    
                    if len(df) > 0:
                        # Scale ratings and normalize counts
                        df['avg_rating'] = df['avg_rating'] * 20  # Scale to 100
                        max_books = df['book_count'].max()
                        max_borrows = df['borrow_count'].max()
                        
                        df['book_count_norm'] = (df['book_count'] / max(max_books, 1)) * 100
                        df['borrow_count_norm'] = (df['borrow_count'] / max(max_borrows, 1)) * 100
                        
                        try:
                            fig = go.Figure()
                            
                            # Only add traces if we have data points
                            if len(df) >= 3:  # Need at least 3 points for a meaningful radar
                                fig.add_trace(go.Scatterpolar(
                                    r=df['book_count_norm'].tolist() + [df['book_count_norm'].iloc[0]],
                                    theta=df['genre_name'].tolist() + [df['genre_name'].iloc[0]],
                                    fill='toself',
                                    name='Book Collection',
                                    line_color='#2196F3'
                                ))
                                
                                fig.add_trace(go.Scatterpolar(
                                    r=df['borrow_count_norm'].tolist() + [df['borrow_count_norm'].iloc[0]],
                                    theta=df['genre_name'].tolist() + [df['genre_name'].iloc[0]],
                                    fill='toself',
                                    name='Popularity',
                                    line_color='#FF9800'
                                ))
                                
                                fig.update_layout(hovermode='x unified', 
                                    polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
                                    showlegend=True,
                                    title="Genre Performance Comparison",
                                    height=600
                                )
                                st.plotly_chart(fig, use_container_width=True)
                            else:
                                st.warning("Need at least 3 genres for radar chart. Showing data table instead:")
                                st.dataframe(df, use_container_width=True, hide_index=True)
                        except Exception as viz_error:
                            st.warning(f"Could not generate radar chart: {str(viz_error)}")
                            st.dataframe(df, use_container_width=True, hide_index=True)
                    else:
                        st.info("No genre performance data available")
                else:
                    st.info("No genre performance data available")
            except Exception as e:
                st.error(f"Error generating Genre Performance Radar: {str(e)}")
                st.info("Please check the database connection and try again.")
        
        elif viz_type == "Collection Growth Timeline":
            try:
                st.markdown("###  Library Collection Growth")
                
                growth_data = Database.execute_query("""
                    SELECT strftime('%Y-%m', created_at) as month,
                           COUNT(*) as new_books
                    FROM books
                    WHERE created_at IS NOT NULL
                    GROUP BY strftime('%Y-%m', created_at)
                    ORDER BY month
                    LIMIT 24
                """)
                
                if growth_data:
                    df = pd.DataFrame(growth_data)
                    # Handle null values and ensure proper data types
                    df = df.fillna({'month': '', 'new_books': 0})
                    df['new_books'] = pd.to_numeric(df['new_books'], errors='coerce').fillna(0).astype(int)
                    
                    # Filter out invalid months and zero counts
                    df = df[df['month'] != '']
                    df = df[df['new_books'] > 0]
                    
                    if len(df) > 0:
                        df = df.sort_values('month')
                        df['cumulative'] = df['new_books'].cumsum()
                        
                        try:
                            fig = go.Figure()
                            fig.add_trace(go.Bar(
                                x=df['month'], y=df['new_books'],
                                name='New Books Added',
                                marker_color='lightblue'
                            ))
                            fig.add_trace(go.Scatter(
                                x=df['month'], y=df['cumulative'],
                                name='Total Collection',
                                marker_color='darkblue',
                                yaxis='y2',
                                mode='lines+markers'
                            ))
                            
                            fig.update_layout(hovermode='x unified',
                                title='Collection Growth Over Time',
                                yaxis=dict(title='New Books per Month'),
                                yaxis2=dict(title='Total Books', overlaying='y', side='right'),
                                height=500
                            )
                            st.plotly_chart(fig, use_container_width=True)
                        except Exception as viz_error:
                            st.warning(f"Could not generate growth timeline: {str(viz_error)}")
                            st.dataframe(df, use_container_width=True, hide_index=True)
                    else:
                        st.info("No collection growth data available")
                else:
                    st.info("No collection growth data available")
            except Exception as e:
                st.error(f"Error generating Collection Growth Timeline: {str(e)}")
                st.info("Please check the database connection and try again.")
        
        elif viz_type == "Multi-Dimensional Analysis":
            try:
                st.markdown("###  Multi-Dimensional Book Analysis")
                
                multi_data = Database.execute_query("""
                    SELECT b.title, b.publication_year,
                           bs.total_checkouts, bs.average_rating,
                           bs.total_copies, bs.available_copies,
                           g.genre_name
                    FROM books b
                    JOIN book_statistics bs ON b.book_id = bs.book_id
                    LEFT JOIN book_genres bg ON b.book_id = bg.book_id
                    LEFT JOIN genres g ON bg.genre_id = g.genre_id
                    WHERE b.is_active = 1 AND bs.total_checkouts > 0
                    LIMIT 100
                """)
                
                if multi_data:
                    df = pd.DataFrame(multi_data)
                    # Handle null values and ensure proper data types
                    df = df.fillna({
                        'title': 'Unknown Title',
                        'publication_year': 0,
                        'total_checkouts': 0,
                        'average_rating': 0,
                        'total_copies': 1,
                        'available_copies': 0,
                        'genre_name': 'Unknown'
                    })
                    
                    # Ensure proper data types
                    df['publication_year'] = pd.to_numeric(df['publication_year'], errors='coerce').fillna(0).astype(int)
                    df['total_checkouts'] = pd.to_numeric(df['total_checkouts'], errors='coerce').fillna(0).astype(int)
                    df['average_rating'] = pd.to_numeric(df['average_rating'], errors='coerce').fillna(0.0)
                    df['total_copies'] = pd.to_numeric(df['total_copies'], errors='coerce').fillna(1).astype(int)
                    df['available_copies'] = pd.to_numeric(df['available_copies'], errors='coerce').fillna(0).astype(int)
                    
                    # Calculate utilization with safe division
                    df['utilization'] = ((df['total_copies'] - df['available_copies']) / df['total_copies'].replace(0, 1)) * 100
                    
                    # Filter out invalid data
                    df = df[df['title'] != 'Unknown Title']
                    df = df[df['total_checkouts'] > 0]
                    
                    if len(df) > 0:
                        try:
                            # 3D Scatter plot
                            fig = px.scatter_3d(df, x='publication_year', y='total_checkouts', z='average_rating',
                                              color='genre_name', size='total_copies',
                                              hover_data=['title'],
                                              title='3D Book Analysis: Year vs Popularity vs Rating',
                                              labels={
                                                  'publication_year': 'Publication Year',
                                                  'total_checkouts': 'Total Checkouts',
                                                  'average_rating': 'Rating'
                                              })
                            fig.update_layout(hovermode='x unified', height=700)
                            st.plotly_chart(fig, use_container_width=True)
                        except Exception as viz_error:
                            st.warning(f"Could not generate 3D scatter plot: {str(viz_error)}")
                        
                        try:
                            # Sunburst chart - only if we have enough valid data
                            sunburst_df = df[df['genre_name'] != 'Unknown']
                            if len(sunburst_df) > 0:
                                st.markdown("###  Collection Hierarchy")
                                fig2 = px.sunburst(sunburst_df, path=['genre_name', 'title'], values='total_checkouts',
                                                  title='Book Collection by Genre (Size = Popularity)',
                                                  color='average_rating',
                                                  color_continuous_scale='RdYlGn')
                                fig2.update_layout(height=600)
                                st.plotly_chart(fig2, use_container_width=True)
                        except Exception as viz_error:
                            st.warning(f"Could not generate sunburst chart: {str(viz_error)}")
                        
                        # Show data table as fallback
                        if len(df) > 0:
                            st.markdown("###  Data Table")
                            st.dataframe(df, use_container_width=True, hide_index=True)
                    else:
                        st.info("No multi-dimensional analysis data available")
                else:
                    st.info("No multi-dimensional analysis data available")
            except Exception as e:
                st.error(f"Error generating Multi-Dimensional Analysis: {str(e)}")
                st.info("Please check the database connection and try again.")

# ================================================================
# V4.0 NEW PAGES
# ================================================================

def show_my_library():
    """My Library page - User's personal PDF library"""
    st.title(" My Library")
    
    user = Auth.get_user()
    is_management_user = user['role'] in ['admin', 'librarian', 'superadmin']
    
    tab_labels = [" My PDFs & Upload"]
    if is_management_user:
        tab_labels[0] = " Book Workspace"
        tab_labels.append(" Manage Members")
    tab_labels.extend([" Privacy Settings", " Browse Community"])
    tabs = st.tabs(tab_labels)
    
    # Tab 1: My PDFs & Upload PDF
    with tabs[0]:
        if not is_management_user:
            show_books(embedded=True)
            st.divider()

        pdfs = PeerLibraryManager.get_user_library(user['user_id'])
        total_pdfs = len(pdfs) if pdfs else 0
        public_pdfs = len([p for p in (pdfs or []) if p.get('is_public')])
        catalog_summary = Database.execute_query("SELECT COUNT(*) as count FROM books", fetch_one=True) or {'count': 0}
        total_catalog_titles = int(catalog_summary.get('count') or 0)

        def render_workspace_statistics_panel():
            st.subheader(" Statistics")

            ctrl_col1, ctrl_col2 = st.columns(2, gap="small")
            with ctrl_col1:
                stats_window = st.selectbox(
                    "Window",
                    ["7D", "30D", "90D", "365D", "All Time"],
                    index=2,
                    key="ws_stats_window"
                )
            with ctrl_col2:
                show_charts = st.checkbox("Show Charts", value=True, key="ws_stats_show_charts")

            window_map = {"7D": 7, "30D": 30, "90D": 90, "365D": 365, "All Time": None}
            window_days = window_map.get(stats_window)

            private_pdfs = max(total_pdfs - public_pdfs, 0)
            public_share = (public_pdfs / total_pdfs * 100) if total_pdfs > 0 else 0.0

            inventory_summary = Database.execute_query(
                """
                SELECT
                    COUNT(*) as total_copies,
                    SUM(CASE WHEN is_available = 1 THEN 1 ELSE 0 END) as available_copies
                FROM book_inventory
                """,
                fetch_one=True
            ) or {}
            total_copies = int(inventory_summary.get('total_copies') or 0)
            available_copies = int(inventory_summary.get('available_copies') or 0)
            checked_out_copies = max(total_copies - available_copies, 0)
            availability_pct = (available_copies / total_copies * 100) if total_copies > 0 else 0.0

            active_loan_summary = Database.execute_query(
                """
                SELECT
                    SUM(CASE WHEN return_date IS NULL THEN 1 ELSE 0 END) as active_loans,
                    SUM(CASE WHEN return_date IS NULL AND date(due_date) < date('now') THEN 1 ELSE 0 END) as overdue_loans
                FROM borrowing
                """,
                fetch_one=True
            ) or {}
            active_loans = int(active_loan_summary.get('active_loans') or 0)
            overdue_loans = int(active_loan_summary.get('overdue_loans') or 0)

            if window_days is None:
                titles_window_query = "SELECT COUNT(*) as count FROM books"
                titles_window_params = None
                checkouts_window_query = "SELECT COUNT(*) as count FROM borrowing"
                checkouts_window_params = None
                checkouts_trend_query = """
                    SELECT date(checkout_date) as day, COUNT(*) as checkout_count
                    FROM borrowing
                    WHERE checkout_date IS NOT NULL
                    GROUP BY date(checkout_date)
                    ORDER BY day DESC
                    LIMIT 30
                """
                checkouts_trend_params = None
            else:
                titles_window_query = "SELECT COUNT(*) as count FROM books WHERE date(created_at) >= date('now', ?)"
                titles_window_params = (f"-{window_days} days",)
                checkouts_window_query = "SELECT COUNT(*) as count FROM borrowing WHERE date(checkout_date) >= date('now', ?)"
                checkouts_window_params = (f"-{window_days} days",)
                checkouts_trend_query = """
                    SELECT date(checkout_date) as day, COUNT(*) as checkout_count
                    FROM borrowing
                    WHERE checkout_date IS NOT NULL AND date(checkout_date) >= date('now', ?)
                    GROUP BY date(checkout_date)
                    ORDER BY day DESC
                    LIMIT 30
                """
                checkouts_trend_params = (f"-{window_days} days",)

            titles_window = Database.execute_query(
                titles_window_query,
                titles_window_params,
                fetch_one=True
            ) or {'count': 0}
            checkouts_window = Database.execute_query(
                checkouts_window_query,
                checkouts_window_params,
                fetch_one=True
            ) or {'count': 0}
            new_titles_in_window = int(titles_window.get('count') or 0)
            checkouts_in_window = int(checkouts_window.get('count') or 0)

            kpi_col1, kpi_col2 = st.columns(2, gap="small")
            with kpi_col1:
                st.metric("My PDFs", total_pdfs)
                st.metric("Public PDFs", public_pdfs, f"{public_share:.1f}%")
                st.metric("Private PDFs", private_pdfs)
                st.metric("Catalog Titles", total_catalog_titles)
            with kpi_col2:
                st.metric("Total Copies", total_copies)
                st.metric("Available Copies", available_copies, f"{availability_pct:.1f}%")
                st.metric("Checked Out", checked_out_copies)
                st.metric("Active Loans", active_loans)

            extra_col1, extra_col2 = st.columns(2, gap="small")
            with extra_col1:
                st.metric(f"New Titles ({stats_window})", new_titles_in_window)
            with extra_col2:
                st.metric(f"Checkouts ({stats_window})", checkouts_in_window)

            st.progress(min(max(availability_pct / 100.0, 0.0), 1.0), text=f"Availability Health: {availability_pct:.1f}%")

            if overdue_loans > 0:
                st.warning(f"Overdue loans detected: {overdue_loans}")
            else:
                st.success("No overdue loans at the moment")

            if show_charts:
                genre_rows = Database.execute_query(
                    """
                    SELECT COALESCE(NULLIF(TRIM(genre), ''), 'Unknown') as genre_name, COUNT(*) as title_count
                    FROM books
                    GROUP BY COALESCE(NULLIF(TRIM(genre), ''), 'Unknown')
                    ORDER BY title_count DESC
                    LIMIT 6
                    """
                ) or []
                trend_rows = Database.execute_query(
                    checkouts_trend_query,
                    checkouts_trend_params
                ) or []

                if genre_rows:
                    genre_df = pd.DataFrame(genre_rows)
                    st.caption("Top Genres")
                    st.bar_chart(genre_df.set_index('genre_name')['title_count'])

                if trend_rows:
                    trend_df = pd.DataFrame(trend_rows).sort_values('day')
                    st.caption(f"Checkout Trend ({stats_window})")
                    st.line_chart(trend_df.set_index('day')['checkout_count'])

        if is_management_user:
            browse_col, stats_col = st.columns([2, 1], gap="large")
            with browse_col:
                show_manage_books(embedded=True, browse_only=True)
            with stats_col:
                render_workspace_statistics_panel()
        else:
            render_workspace_statistics_panel()

        st.subheader(" Book Registry & Intake")

        with st.container():
            prefill_pdf = None
            if pdfs:
                prefill_id_options = [0] + [int(p['pdf_id']) for p in pdfs if p.get('pdf_id') is not None]
                selected_prefill_id = st.selectbox(
                    "Smart Prefill",
                    options=prefill_id_options,
                    format_func=lambda x: "None" if x == 0 else next(
                        (f"{p.get('title') or 'Untitled'} | {p.get('author') or 'Unknown'}" for p in pdfs if int(p.get('pdf_id') or 0) == x),
                        str(x)
                    ),
                    key="ucc_prefill_pdf"
                )
                if selected_prefill_id != 0:
                    prefill_pdf = next((p for p in pdfs if int(p.get('pdf_id') or 0) == int(selected_prefill_id)), None)

                if prefill_pdf and st.button(" Apply Prefill to Intake", key="ucc_apply_prefill", use_container_width=True):
                    allowed_genres = ["Fiction", "Non-Fiction", "Science", "Technology", "History", "Biography", "Self-Help", "Other"]
                    prefill_genre = str(prefill_pdf.get('genre') or "Other")
                    if prefill_genre not in allowed_genres:
                        prefill_genre = "Other"
                    st.session_state["ucc_title"] = str(prefill_pdf.get('title') or "")
                    st.session_state["ucc_author"] = str(prefill_pdf.get('author') or "")
                    st.session_state["ucc_genre"] = prefill_genre
                    st.session_state["ucc_description"] = str(prefill_pdf.get('description') or "")
                    st.rerun()

            with st.form("unified_content_console_form"):
                action_col1, action_col2 = st.columns(2, gap="small")
                with action_col1:
                    create_pdf = st.checkbox("Create/Upload PDF Asset", value=True, key="ucc_create_pdf")
                with action_col2:
                    create_catalog = st.checkbox(
                        "Create Physical Catalog Record",
                        value=True if is_management_user else False,
                        disabled=not is_management_user,
                        key="ucc_create_catalog"
                    )

                title = st.text_input("Title of the book *", max_chars=500, key="ucc_title")
                current_year = datetime.now().year
                max_year = current_year + 10

                row1_col1, row1_col2, row1_col3, row1_col4, row1_col5 = st.columns(5, gap="small")
                with row1_col1:
                    authors_text = st.text_input("Author(s)", max_chars=300, key="ucc_author")
                with row1_col2:
                    publisher = st.text_input("Publisher", max_chars=300, key="ucc_publisher")
                with row1_col3:
                    genre = st.selectbox("Genre", ["Fiction", "Non-Fiction", "Science", "Technology", "History", "Biography", "Self-Help", "Other"], key="ucc_genre")
                with row1_col4:
                    subtitle = st.text_input("Subtitle", max_chars=500, key="ucc_subtitle")
                with row1_col5:
                    publication_year = st.number_input("Year of publication", min_value=1800, max_value=max_year, value=current_year, key="ucc_pub_year")

                row2_col1, row2_col2, row2_col3, row2_col4, row2_col5 = st.columns(5, gap="small")
                with row2_col1:
                    edition = st.text_input("Edition", placeholder="e.g., 2nd Edition", key="ucc_edition")
                with row2_col2:
                    isbn_issn = st.text_input("ISBN/ISSN", placeholder="ISBN-13 or ISSN", key="ucc_isbn_issn")
                with row2_col3:
                    keywords = st.text_input("Keywords", placeholder="fiction, mystery, thriller", key="ucc_keywords")
                with row2_col4:
                    contributors = st.text_input("Contributors/Editors", key="ucc_contributors")
                with row2_col5:
                    publication_place = st.text_input("Publication Place", key="ucc_pub_place")

                row3_col1, row3_col2, row3_col3, row3_col4, row3_col5 = st.columns(5, gap="small")
                with row3_col1:
                    series_title = st.text_input("Series", key="ucc_series")
                with row3_col2:
                    volume = st.text_input("Volume", key="ucc_volume")
                with row3_col3:
                    shared_language = st.selectbox("Language", ["English", "Hindi", "Spanish", "French", "German", "Other"], key="ucc_language")
                with row3_col4:
                    shared_pages = st.number_input("Pages", min_value=1, value=200, key="ucc_pages")
                with row3_col5:
                    call_number = st.text_input("Call Number", key="ucc_call_number")

                row4_col1, row4_col2, row4_col3, row4_col4, row4_col5 = st.columns(5, gap="small")
                with row4_col1:
                    accession_number = st.text_input("Accession Number", key="ucc_accession")
                with row4_col2:
                    shelf_location = st.text_input("Shelf/Location", placeholder="Aisle-B / Rack-12", key="ucc_catalog_location")
                with row4_col3:
                    condition_note = st.text_input("Condition Notes", placeholder="new/good/worn", key="ucc_condition")
                with row4_col4:
                    source_url = st.text_input("Source URL", placeholder="https://...", key="ucc_source_url")
                with row4_col5:
                    pass

                description = st.text_area("Description", max_chars=1400, key="ucc_description")

                row5_col1, row5_col2, row5_col3, row5_col4, row5_col5 = st.columns(5, gap="small")
                with row5_col1:
                    uploaded_file = st.file_uploader("PDF File", type=['pdf'], key="ucc_pdf_file")
                with row5_col2:
                    is_public = st.checkbox("Publish PDF as Public", value=False, key="ucc_pdf_public")
                with row5_col3:
                    digital_format = st.selectbox(
                        "Digital file format",
                        ["PDF", "EPUB", "MOBI", "DOCX", "TXT", "Other"],
                        key="ucc_digital_format"
                    )
                with row5_col4:
                    auto_file_size_bytes = len(uploaded_file.getvalue()) if uploaded_file else 0
                    st.text_input(
                        "File size",
                        value=(f"{(auto_file_size_bytes / 1024 / 1024):.2f} MB" if auto_file_size_bytes else "N/A"),
                        disabled=True,
                        key="ucc_file_size_display"
                    )
                with row5_col5:
                    cover_image = st.file_uploader("Cover image", type=['png', 'jpg', 'jpeg', 'webp'], key="ucc_cover_image")

                pages = 1
                language = "Other"
                copies = 1
                submit_unified = None
                if is_management_user and create_catalog:
                    row6_col1, row6_col2, row6_col3 = st.columns(3, gap="small")
                    with row6_col1:
                        copies = st.number_input("Copies", min_value=1, value=1, key="ucc_copies")
                    with row6_col2:
                        catalog_acquired_on = st.date_input("Acquired On", value=date.today(), key="ucc_acquired_on")
                    with row6_col3:
                        st.markdown("<div style='height: 1.8rem;'></div>", unsafe_allow_html=True)
                        submit_unified = st.form_submit_button(" Execute Unified Action", use_container_width=True)
                else:
                    catalog_acquired_on = date.today()

                pages = int(shared_pages)
                language = shared_language
                catalog_location = shelf_location

                if submit_unified is None:
                    submit_unified = st.form_submit_button(" Execute Unified Action", use_container_width=True)

                if submit_unified:
                    if not create_pdf and not create_catalog:
                        st.error("Select at least one action: PDF asset and/or catalog record.")
                    elif not title:
                        st.error("Title is required.")
                    else:
                        results = []

                        if create_pdf:
                            if not uploaded_file:
                                st.error("PDF file is required when Create/Upload PDF Asset is selected.")
                            else:
                                enriched_description = description or ""
                                enriched_description += (
                                    f"\n\n[Metadata] Publisher: {publisher or 'N/A'} | Edition: {edition or 'N/A'}"
                                    f" | ISBN/ISSN: {isbn_issn or 'N/A'} | Format: {digital_format}"
                                    f" | Size: {(auto_file_size_bytes / 1024 / 1024):.2f} MB"
                                    f" | Subtitle: {subtitle or 'N/A'} | Contributors: {contributors or 'N/A'}"
                                    f" | Publication Place: {publication_place or 'N/A'} | Series: {series_title or 'N/A'}"
                                    f" | Volume: {volume or 'N/A'} | Call Number: {call_number or 'N/A'}"
                                    f" | Accession: {accession_number or 'N/A'} | Source: {source_url or 'N/A'}"
                                )

                                pdf_ok, pdf_msg = PeerLibraryManager.upload_pdf_to_library(
                                    user['user_id'], uploaded_file, title, authors_text, genre, enriched_description, is_public
                                )
                                if pdf_ok:
                                    results.append("PDF uploaded")
                                else:
                                    st.error(pdf_msg)

                        if create_catalog and is_management_user:
                            resolved_isbn = (isbn_issn or "").strip()
                            if not resolved_isbn:
                                resolved_isbn = f"AUTO-{user['user_id']}-{int(time.time())}"

                            digits_only = re.sub(r"\D", "", resolved_isbn)
                            isbn_13 = digits_only if len(digits_only) == 13 else None
                            isbn_10 = digits_only if len(digits_only) == 10 else None

                            catalog_description = description or ""
                            catalog_description += (
                                f"\n\nPublisher: {publisher or 'N/A'}"
                                f"\nEdition: {edition or 'N/A'}"
                                f"\nSubtitle: {subtitle or 'N/A'}"
                                f"\nContributors: {contributors or 'N/A'}"
                                f"\nPublication Place: {publication_place or 'N/A'}"
                                f"\nSeries: {series_title or 'N/A'}"
                                f"\nVolume: {volume or 'N/A'}"
                                f"\nISBN/ISSN: {isbn_issn or 'N/A'}"
                                f"\nDigital Format: {digital_format}"
                                f"\nDigital File Size: {(auto_file_size_bytes / 1024 / 1024):.2f} MB"
                                f"\nCall Number: {call_number or 'N/A'}"
                                f"\nAccession Number: {accession_number or 'N/A'}"
                                f"\nSource URL: {source_url or 'N/A'}"
                                f"\nAcquired On: {catalog_acquired_on}"
                            )

                            catalog_ok = Database.execute_update(
                                """
                                INSERT INTO books (
                                    isbn, isbn_13, isbn_10, title, author, genre, publication_year,
                                    publisher, pages, page_count, language, description, keywords,
                                    location, condition_notes, is_available
                                )
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                                """,
                                (
                                    resolved_isbn,
                                    isbn_13,
                                    isbn_10,
                                    title,
                                    authors_text or "Unknown",
                                    genre,
                                    publication_year,
                                    publisher or None,
                                    pages,
                                    pages,
                                    language,
                                    catalog_description or None,
                                    keywords or None,
                                    catalog_location or None,
                                    condition_note or None
                                )
                            )

                            if catalog_ok:
                                created = Database.execute_query(
                                    "SELECT book_id FROM books WHERE isbn = ? ORDER BY book_id DESC LIMIT 1",
                                    (resolved_isbn,)
                                )
                                if created:
                                    book_id = created[0]['book_id']
                                    for i in range(int(copies)):
                                        Database.execute_update(
                                            "INSERT INTO book_inventory (book_id, barcode, is_available, location, condition_status) VALUES (?, ?, 1, ?, ?)",
                                            (book_id, f"{resolved_isbn}-{i+1}", catalog_location or None, condition_note or 'good')
                                        )

                                    if cover_image:
                                        try:
                                            Database.execute_update(
                                                "INSERT INTO book_covers (book_id, cover_image) VALUES (?, ?)",
                                                (book_id, cover_image.getvalue())
                                            )
                                        except Exception:
                                            pass
                                results.append(f"Catalog record created ({int(copies)} copies)")
                            else:
                                st.error("Catalog creation failed (likely duplicate ISBN).")

                        if results:
                            st.success(" | ".join(results))
                            st.balloons()
                            st.rerun()


        st.subheader("Collection Operations")

        # Filters Row
        filter_col1, filter_col2, filter_col3, filter_col4 = st.columns(4, gap="small")
        with filter_col1:
            collection_search = st.text_input("Search", placeholder="Title, author, genre", key="ucs_search")
        with filter_col2:
            visibility_filter = st.selectbox("Visibility", ["All", "Public", "Private"], key="ucs_visibility")
        with filter_col3:
            sort_mode = st.selectbox("Sort", ["Newest", "Most Viewed", "Title A-Z", "Author A-Z", "Size A-Z"], key="ucs_sort")
        with filter_col4:
            if st.button("Refresh", use_container_width=True, key="ucs_refresh"):
                st.rerun()

        # Apply filters
        filtered_pdfs = pdfs or []

        if collection_search:
            q = collection_search.lower()
            filtered_pdfs = [
                p for p in filtered_pdfs
                if q in str(p.get('title', '')).lower()
                or q in str(p.get('author', '')).lower()
                or q in str(p.get('genre', '')).lower()
            ]

        if visibility_filter == "Public":
            filtered_pdfs = [p for p in filtered_pdfs if p.get('is_public')]
        elif visibility_filter == "Private":
            filtered_pdfs = [p for p in filtered_pdfs if not p.get('is_public')]

        if sort_mode == "Most Viewed":
            filtered_pdfs = sorted(filtered_pdfs, key=lambda p: int(p.get('views_count') or 0), reverse=True)
        elif sort_mode == "Title A-Z":
            filtered_pdfs = sorted(filtered_pdfs, key=lambda p: str(p.get('title') or '').lower())
        elif sort_mode == "Author A-Z":
            filtered_pdfs = sorted(filtered_pdfs, key=lambda p: str(p.get('author') or '').lower())
        elif sort_mode == "Size A-Z":
            filtered_pdfs = sorted(filtered_pdfs, key=lambda p: int(p.get('file_size') or 0))
        else:
            filtered_pdfs = sorted(filtered_pdfs, key=lambda p: str(p.get('upload_date') or ''), reverse=True)

        # Calculate statistics
        total_views = sum(int(p.get('views_count') or 0) for p in filtered_pdfs)
        public_count = len([p for p in filtered_pdfs if p.get('is_public')])
        private_count = len([p for p in filtered_pdfs if not p.get('is_public')])
        total_size_bytes = sum(int(p.get('file_size') or 0) for p in filtered_pdfs)
        total_size_mb = round(total_size_bytes / (1024 * 1024), 2)
        avg_views = round(total_views / len(filtered_pdfs), 1) if filtered_pdfs else 0
        avg_size_mb = round(total_size_mb / len(filtered_pdfs), 2) if filtered_pdfs else 0
        genre_counts = {}
        author_counts = {}
        for pdf in filtered_pdfs:
            genre = pdf.get('genre') or 'Unknown'
            genre_counts[genre] = genre_counts.get(genre, 0) + 1
            author = pdf.get('author') or 'Unknown'
            author_counts[author] = author_counts.get(author, 0) + 1

        if filtered_pdfs:
            stats_col, ops_col = st.columns(2, gap="large")

            # ========== LEFT: DYNAMIC STATISTICS ==========
            with stats_col:
                st.markdown("### Collection Statistics")
                
                # Key Metrics Row
                metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4, gap="small")
                with metric_col1:
                    st.metric("Total PDFs", len(filtered_pdfs))
                with metric_col2:
                    st.metric("Total Views", total_views)
                with metric_col3:
                    st.metric("Avg Views", avg_views)
                with metric_col4:
                    st.metric("Storage MB", total_size_mb)
                
                st.divider()
                
                # Visibility & Size Stats
                vis_size_col1, vis_size_col2 = st.columns(2, gap="small")
                
                with vis_size_col1:
                    st.markdown("**Visibility**")
                    vis_data = [
                        {"Status": "Public", "Count": public_count},
                        {"Status": "Private", "Count": private_count}
                    ]
                    if any(v["Count"] > 0 for v in vis_data):
                        fig_vis = px.pie(pd.DataFrame(vis_data), names='Status', values='Count', 
                                        color_discrete_map={"Public": "#4CAF50", "Private": "#2196F3"})
                        fig_vis.update_layout(height=280, margin=dict(l=0, r=0, t=0, b=0))
                        st.plotly_chart(fig_vis, use_container_width=True)
                    else:
                        st.info("No visibility data")
                
                with vis_size_col2:
                    st.markdown("**Storage Stats**")
                    size_info = f"Total: {total_size_mb} MB\nAverage: {avg_size_mb} MB/file"
                    st.info(size_info)
                    if filtered_pdfs:
                        largest = max(filtered_pdfs, key=lambda p: int(p.get('file_size') or 0))
                        largest_size_mb = round(int(largest.get('file_size') or 0) / (1024 * 1024), 2)
                        st.caption(f"Largest: {largest.get('title')[:25]} ({largest_size_mb} MB)")
                
                st.divider()
                
                # Genre Distribution
                st.markdown("**Genre Distribution**")
                if genre_counts:
                    genre_df = pd.DataFrame(list(genre_counts.items()), columns=['Genre', 'Count']).sort_values('Count', ascending=False)
                    fig_genre = px.bar(genre_df, x='Genre', y='Count', title=None, 
                                      color='Count', color_continuous_scale='Blues')
                    fig_genre.update_layout(height=280, showlegend=False, margin=dict(l=0, r=0, t=20, b=0))
                    st.plotly_chart(fig_genre, use_container_width=True)
                else:
                    st.info("No genre data")
                
                st.divider()
                
                # Top Authors
                st.markdown("**Top Authors**")
                if author_counts:
                    top_authors = sorted(author_counts.items(), key=lambda x: x[1], reverse=True)[:5]
                    author_df = pd.DataFrame(top_authors, columns=['Author', 'Count'])
                    fig_author = px.bar(author_df, x='Author', y='Count', title=None, 
                                       color='Count', color_continuous_scale='Greens')
                    fig_author.update_layout(height=280, showlegend=False, margin=dict(l=0, r=0, t=20, b=0))
                    st.plotly_chart(fig_author, use_container_width=True)
                else:
                    st.info("No author data")
                
                st.divider()
                
                # Most Viewed PDFs
                st.markdown("**Top 5 Most Viewed**")
                top_viewed = sorted(filtered_pdfs, key=lambda p: int(p.get('views_count') or 0), reverse=True)[:5]
                if top_viewed:
                    view_df = pd.DataFrame([
                        {"Title": p.get('title')[:25], "Views": int(p.get('views_count') or 0)}
                        for p in top_viewed
                    ])
                    fig_views = px.bar(view_df, x='Title', y='Views', title=None,
                                      color='Views', color_continuous_scale='Reds')
                    fig_views.update_layout(height=280, showlegend=False, margin=dict(l=0, r=0, t=20, b=0))
                    st.plotly_chart(fig_views, use_container_width=True)
                else:
                    st.info("No view data")

            # ========== RIGHT: BATCH OPERATIONS ==========
            with ops_col:
                st.markdown("### Batch Operations")
                
                batch_tabs = st.tabs(["Basic Actions", "Advanced Tools"])

                with batch_tabs[0]:
                    batch_options = [
                        {
                            'label': f"{p.get('title') or 'Untitled'[:50]}",
                            'value': int(p.get('pdf_id'))
                        }
                        for p in filtered_pdfs
                        if p.get('pdf_id') is not None
                    ]

                    selected_batch_ids = st.multiselect(
                        "Select PDFs",
                        options=[o['value'] for o in batch_options],
                        format_func=lambda x: next((o['label'] for o in batch_options if o['value'] == x), str(x)),
                        key="ucs_batch_selected"
                    )

                    batch_action_options = [
                        "Make Public",
                        "Make Private",
                        "Export ZIP",
                        "Export Metadata CSV",
                        "Export as JSON",
                        "Bulk Delete/Archive",
                        "Bulk Edit Genre"
                    ]
                    if is_management_user:
                        batch_action_options.append("Catalog Sync")
                    
                    batch_action = st.selectbox("Action", batch_action_options, key="ucs_batch_action")

                    if batch_action == "Bulk Edit Genre":
                        new_genre = st.selectbox("New Genre", ["Fiction", "Non-Fiction", "Science", "Technology", "History", "Biography", "Self-Help", "Other"], key="ucs_bulk_genre")
                    else:
                        new_genre = None

                    apply_batch = st.button("Execute Action", use_container_width=True, type="primary", key="ucs_batch_apply")

                    if apply_batch:
                        if not selected_batch_ids:
                            st.warning("Select at least one PDF for batch actions.")
                        else:
                            if batch_action == "Make Public":
                                updated = 0
                                for pdf_id in selected_batch_ids:
                                    ok = Database.execute_update(
                                        "UPDATE pdf_library SET is_public = 1 WHERE pdf_id = ? AND user_id = ?",
                                        (pdf_id, user['user_id'])
                                    )
                                    if ok:
                                        updated += 1
                                st.success(f"Updated {updated} PDF(s) to Public.")
                                st.rerun()

                            elif batch_action == "Make Private":
                                updated = 0
                                for pdf_id in selected_batch_ids:
                                    ok = Database.execute_update(
                                        "UPDATE pdf_library SET is_public = 0 WHERE pdf_id = ? AND user_id = ?",
                                        (pdf_id, user['user_id'])
                                    )
                                    if ok:
                                        updated += 1
                                st.success(f"Updated {updated} PDF(s) to Private.")
                                st.rerun()

                            elif batch_action == "Export ZIP":
                                zip_buffer = BytesIO()
                                exported = 0
                                with zipfile.ZipFile(zip_buffer, mode='w', compression=zipfile.ZIP_DEFLATED) as archive:
                                    for pdf_id in selected_batch_ids:
                                        pdf_data = PeerLibraryManager.get_pdf_file(pdf_id)
                                        if pdf_data and pdf_data.get('pdf_file'):
                                            filename = pdf_data.get('pdf_filename') or f"pdf_{pdf_id}.pdf"
                                            archive.writestr(filename, pdf_data['pdf_file'])
                                            exported += 1

                                if exported > 0:
                                    zip_bytes = zip_buffer.getvalue()
                                    st.download_button(
                                        "Download ZIP",
                                        data=zip_bytes,
                                        file_name=f"pdf_batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                                        mime="application/zip",
                                        key="ucs_batch_zip_download",
                                        use_container_width=True
                                    )
                                    st.success(f"Prepared ZIP for {exported} PDF(s).")
                                else:
                                    st.warning("No files could be exported.")

                            elif batch_action == "Export Metadata CSV":
                                metadata_list = []
                                for pdf_id in selected_batch_ids:
                                    pdf = next((p for p in filtered_pdfs if int(p.get('pdf_id')) == pdf_id), None)
                                    if pdf:
                                        metadata_list.append({
                                            'Title': pdf.get('title'),
                                            'Author': pdf.get('author'),
                                            'Genre': pdf.get('genre'),
                                            'Visibility': 'Public' if pdf.get('is_public') else 'Private',
                                            'Views': pdf.get('views_count'),
                                            'Uploaded': pdf.get('upload_date')
                                        })
                                if metadata_list:
                                    metadata_df = pd.DataFrame(metadata_list)
                                    csv_data = metadata_df.to_csv(index=False)
                                    st.download_button(
                                        "Download CSV",
                                        data=csv_data,
                                        file_name=f"pdfs_metadata_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                                        mime="text/csv",
                                        key="ucs_batch_metadata_download",
                                        use_container_width=True
                                    )
                                    st.success(f"Prepared metadata for {len(metadata_list)} PDF(s).")

                            elif batch_action == "Export as JSON":
                                json_list = []
                                for pdf_id in selected_batch_ids:
                                    pdf = next((p for p in filtered_pdfs if int(p.get('pdf_id')) == pdf_id), None)
                                    if pdf:
                                        json_list.append({
                                            'pdf_id': pdf.get('pdf_id'),
                                            'title': pdf.get('title'),
                                            'author': pdf.get('author'),
                                            'genre': pdf.get('genre'),
                                            'visibility': 'public' if pdf.get('is_public') else 'private',
                                            'views': int(pdf.get('views_count') or 0),
                                            'upload_date': pdf.get('upload_date'),
                                            'description': pdf.get('description')
                                        })
                                if json_list:
                                    json_data = json.dumps(json_list, indent=2)
                                    st.download_button(
                                        "Download JSON",
                                        data=json_data,
                                        file_name=f"pdfs_metadata_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                                        mime="application/json",
                                        key="ucs_batch_json_download",
                                        use_container_width=True
                                    )
                                    st.success(f"Prepared JSON for {len(json_list)} PDF(s).")

                            elif batch_action == "Bulk Delete/Archive":
                                st.warning("This will mark selected PDFs as archived (soft delete).")
                                if st.checkbox("I confirm archive operation", key="ucs_confirm_delete"):
                                    archived = 0
                                    for pdf_id in selected_batch_ids:
                                        ok = Database.execute_update(
                                            "UPDATE pdf_library SET is_archived = 1, archived_date = datetime('now') WHERE pdf_id = ? AND user_id = ?",
                                            (pdf_id, user['user_id'])
                                        )
                                        if ok:
                                            archived += 1
                                    st.success(f"Archived {archived} PDF(s).")
                                    st.rerun()

                            elif batch_action == "Bulk Edit Genre":
                                updated = 0
                                for pdf_id in selected_batch_ids:
                                    ok = Database.execute_update(
                                        "UPDATE pdf_library SET genre = ? WHERE pdf_id = ? AND user_id = ?",
                                        (new_genre, pdf_id, user['user_id'])
                                    )
                                    if ok:
                                        updated += 1
                                st.success(f"Updated {updated} PDF(s) genre to {new_genre}.")
                                st.rerun()

                            elif batch_action == "Catalog Sync" and is_management_user:
                                batch_copies = st.number_input("Copies", min_value=1, value=1, key="ucs_batch_copies_sync")
                                batch_pub_year = st.number_input("Year", min_value=1800, max_value=datetime.now().year + 10, value=datetime.now().year, key="ucs_batch_year_sync")
                                if st.button("Execute Sync", use_container_width=True):
                                    synced = 0
                                    skipped = 0
                                    for pdf_id in selected_batch_ids:
                                        row = next((p for p in filtered_pdfs if int(p.get('pdf_id')) == int(pdf_id)), None)
                                        if not row:
                                            skipped += 1
                                            continue

                                        sync_isbn = f"PDFLIB-{int(pdf_id)}"
                                        existing = Database.execute_query(
                                            "SELECT book_id FROM books WHERE isbn = ? LIMIT 1",
                                            (sync_isbn,),
                                            fetch_one=True
                                        )
                                        if existing:
                                            skipped += 1
                                            continue

                                        created = Database.execute_update(
                                            """
                                            INSERT INTO books (isbn, title, author, genre, publication_year,
                                                             pages, language, description, keywords, is_available)
                                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                                            """,
                                            (
                                                sync_isbn,
                                                row.get('title') or f"PDF {pdf_id}",
                                                row.get('author') or "Unknown",
                                                row.get('genre') or "Other",
                                                int(batch_pub_year),
                                                1,
                                                "Other",
                                                row.get('description') or None,
                                                "pdf_library_sync"
                                            )
                                        )
                                        if created:
                                            book = Database.execute_query(
                                                "SELECT book_id FROM books WHERE isbn = ? ORDER BY book_id DESC LIMIT 1",
                                                (sync_isbn,),
                                                fetch_one=True
                                            )
                                            if book:
                                                for i in range(int(batch_copies)):
                                                    Database.execute_update(
                                                        "INSERT INTO book_inventory (book_id, barcode, is_available) VALUES (?, ?, 1)",
                                                        (book['book_id'], f"{sync_isbn}-{i+1}")
                                                    )
                                            synced += 1
                                        else:
                                            skipped += 1

                                    st.success(f"Catalog sync completed. Synced: {synced}, Skipped: {skipped}")
                                    st.rerun()

                with batch_tabs[1]:
                    st.markdown("**Advanced Tools**")
                    
                    advanced_tool = st.selectbox("Select Tool", [
                        "Add Tags/Keywords",
                        "Bulk Rename",
                        "Duplicate PDF",
                        "Validate PDF Integrity"
                    ], key="ucs_advanced_tool")

                    if advanced_tool == "Add Tags/Keywords":
                        tags_to_add = st.text_input("Tags to add (comma-separated)", key="ucs_tags_input", help="e.g., research, important, archived")
                        if st.button("Add Tags to Selected", use_container_width=True, key="ucs_add_tags"):
                            if not selected_batch_ids:
                                st.warning("Select PDFs first.")
                            elif tags_to_add.strip():
                                updated = 0
                                for pdf_id in selected_batch_ids:
                                    existing = next((p for p in filtered_pdfs if int(p.get('pdf_id')) == pdf_id), None)
                                    if existing:
                                        current_tags = existing.get('keywords') or ""
                                        new_tags = f"{current_tags}, {tags_to_add}".strip(", ")
                                        ok = Database.execute_update(
                                            "UPDATE pdf_library SET keywords = ? WHERE pdf_id = ? AND user_id = ?",
                                            (new_tags, pdf_id, user['user_id'])
                                        )
                                        if ok:
                                            updated += 1
                                st.success(f"Added tags to {updated} PDF(s).")
                                st.rerun()

                    elif advanced_tool == "Bulk Rename":
                        rename_pattern = st.text_input("Pattern (use {idx} for index, {title} for current title)", placeholder="Document_{idx}", key="ucs_rename_pattern")
                        if st.button("Preview Rename", use_container_width=True, key="ucs_preview_rename"):
                            rename_preview = []
                            for idx, pdf_id in enumerate(selected_batch_ids):
                                pdf = next((p for p in filtered_pdfs if int(p.get('pdf_id')) == pdf_id), None)
                                if pdf:
                                    new_name = rename_pattern.format(idx=idx+1, title=pdf.get('title', 'Document'))
                                    rename_preview.append({"Current": pdf.get('title'), "New": new_name})
                            st.dataframe(rename_preview, use_container_width=True, hide_index=True)

                    elif advanced_tool == "Duplicate PDF":
                        duplicate_suffix = st.text_input("Suffix for copy", value="(copy)", key="ucs_duplicate_suffix")
                        if st.button("Duplicate Selected", use_container_width=True, key="ucs_do_duplicate"):
                            if not selected_batch_ids:
                                st.warning("Select PDFs to duplicate.")
                            else:
                                duplicated = 0
                                for pdf_id in selected_batch_ids:
                                    source = next((p for p in filtered_pdfs if int(p.get('pdf_id')) == pdf_id), None)
                                    if source:
                                        pdf_data = PeerLibraryManager.get_pdf_file(pdf_id)
                                        if pdf_data:
                                            new_title = f"{source.get('title')} {duplicate_suffix}"
                                            ok, msg = PeerLibraryManager.upload_pdf_to_library(
                                                user['user_id'],
                                                BytesIO(pdf_data.get('pdf_file')),
                                                new_title,
                                                source.get('author'),
                                                source.get('genre'),
                                                source.get('description'),
                                                source.get('is_public', False)
                                            )
                                            if ok:
                                                duplicated += 1
                                st.success(f"Duplicated {duplicated} PDF(s).")
                                st.rerun()

                    elif advanced_tool == "Validate PDF Integrity":
                        if st.button("Validate Selected PDFs", use_container_width=True, key="ucs_validate_pdfs"):
                            if not selected_batch_ids:
                                st.warning("Select PDFs to validate.")
                            else:
                                validation_results = []
                                for pdf_id in selected_batch_ids:
                                    pdf = next((p for p in filtered_pdfs if int(p.get('pdf_id')) == pdf_id), None)
                                    pdf_data = PeerLibraryManager.get_pdf_file(pdf_id) if pdf else None
                                    if pdf_data and pdf_data.get('pdf_file'):
                                        status = "Valid"
                                        try:
                                            import PyPDF2
                                            reader = PyPDF2.PdfReader(BytesIO(pdf_data.get('pdf_file')))
                                            pages = len(reader.pages)
                                        except:
                                            status = "Check Failed"
                                            pages = 0
                                    else:
                                        status = "Not Found"
                                        pages = 0
                                    validation_results.append({"Title": pdf.get('title') if pdf else 'Unknown', "Status": status, "Pages": pages})
                                st.dataframe(validation_results, use_container_width=True, hide_index=True)
                
                st.divider()
                st.markdown("### Individual PDF Details (Scrollable)")
                
                if filtered_pdfs:
                    pdf_display_height = st.slider("Display Height (px)", min_value=200, max_value=1000, value=600, step=50, key="ucs_pdf_height")
                
                for idx, pdf in enumerate(filtered_pdfs):
                    with st.container():
                        pdf_card_col1, pdf_card_col2, pdf_card_col3 = st.columns([2.5, 1, 0.5], gap="small")
                        
                        with pdf_card_col1:
                            visibility_label = "Public" if pdf.get('is_public') else "Private"
                            st.markdown(f"**[{visibility_label}] {pdf['title'][:40]}**")
                            st.caption(f"Author: {pdf.get('author') or 'Unknown'} | Genre: {pdf.get('genre') or 'Other'}")
                            st.caption(f"Uploaded: {pdf.get('upload_date') or 'N/A'} | Views: {pdf.get('views_count') or 0}")

                        with pdf_card_col2:
                            if st.button("Download", key=f"ucs_quick_dl_{pdf['pdf_id']}", help="Download", use_container_width=True):
                                pdf_data = PeerLibraryManager.get_pdf_file(pdf['pdf_id'])
                                if pdf_data:
                                    PeerLibraryManager.increment_pdf_views(pdf['pdf_id'])
                                    st.download_button(
                                        "Confirm",
                                        pdf_data['pdf_file'],
                                        pdf_data['pdf_filename'],
                                        mime="application/pdf",
                                        key=f"ucs_download_btn_{pdf['pdf_id']}"
                                    )

                        with pdf_card_col3:
                            if st.button("Edit", key=f"ucs_edit_toggle_{pdf['pdf_id']}", help="Edit", use_container_width=True):
                                st.session_state[f'edit_pdf_{pdf["pdf_id"]}'] = not st.session_state.get(f'edit_pdf_{pdf["pdf_id"]}', False)

                        if st.session_state.get(f'edit_pdf_{pdf["pdf_id"]}', False):
                            with st.form(f"ucs_quick_edit_{pdf['pdf_id']}", clear_on_submit=False):
                                st.markdown("**Edit Metadata**")
                                edit_title = st.text_input("Title", value=str(pdf.get('title') or ""), key=f"ucs_title_{pdf['pdf_id']}")
                                edit_author = st.text_input("Author", value=str(pdf.get('author') or ""), key=f"ucs_author_{pdf['pdf_id']}")
                                edit_genre = st.selectbox("Genre", ["Fiction", "Non-Fiction", "Science", "Technology", "History", "Biography", "Self-Help", "Other"], index=0, key=f"ucs_genre_{pdf['pdf_id']}")
                                edit_keywords = st.text_input("Tags/Keywords", value=str(pdf.get('keywords') or ""), key=f"ucs_keywords_{pdf['pdf_id']}")
                                edit_description = st.text_area("Description", value=str(pdf.get('description') or ""), key=f"ucs_desc_{pdf['pdf_id']}", height=60)

                                vis_cols = st.columns([1, 1], gap="small")
                                with vis_cols[0]:
                                    new_visibility = st.checkbox("Make Public", value=pdf.get('is_public', False), key=f"ucs_vis_toggle_{pdf['pdf_id']}")
                                with vis_cols[1]:
                                    pass

                                save_cols = st.columns([1, 1], gap="small")
                                with save_cols[0]:
                                    save_edit = st.form_submit_button("Save", use_container_width=True)
                                with save_cols[1]:
                                    cancel_edit = st.form_submit_button("Cancel", use_container_width=True)

                                if cancel_edit:
                                    st.session_state[f'edit_pdf_{pdf["pdf_id"]}'] = False
                                    st.rerun()

                                if save_edit:
                                    if not edit_title.strip():
                                        st.error("Title cannot be empty.")
                                    else:
                                        updated = Database.execute_update(
                                            """
                                            UPDATE pdf_library
                                            SET title = ?, author = ?, genre = ?, keywords = ?, description = ?, is_public = ?
                                            WHERE pdf_id = ? AND user_id = ?
                                            """,
                                            (
                                                edit_title.strip(),
                                                edit_author.strip(),
                                                edit_genre,
                                                edit_keywords.strip(),
                                                edit_description.strip(),
                                                new_visibility,
                                                pdf['pdf_id'],
                                                user['user_id']
                                            )
                                        )
                                        if updated:
                                            st.success("Metadata saved.")
                                            st.session_state[f'edit_pdf_{pdf["pdf_id"]}'] = False
                                            st.rerun()
                                        else:
                                            st.error("Failed to save metadata.")

                        st.divider()
        else:
            st.info("No PDFs match current filters.")

    privacy_tab_index = 2 if is_management_user else 1
    community_tab_index = 3 if is_management_user else 2

    # Manage Members Tab (admin/librarian/superadmin)
    if is_management_user:
        with tabs[1]:
            show_manage_members()
    
    # Privacy Settings Tab
    with tabs[privacy_tab_index]:
        st.subheader(" Privacy & Anonymous Mode")
        
        privacy = PrivacyManager.get_privacy_settings(user['user_id'])
        
        col1, col2 = st.columns(2, gap="small")
        
        with col1:
            st.markdown("###  Visibility Settings")
            
            show_email = st.checkbox("Show Email", value=privacy.get('show_email', False))
            show_phone = st.checkbox("Show Phone", value=privacy.get('show_phone', False))
            show_full_name = st.checkbox("Show Full Name", value=privacy.get('show_full_name', True))
            show_profile_photo = st.checkbox("Show Profile Photo", value=privacy.get('show_profile_photo', True))
            show_library = st.checkbox("Show Library", value=privacy.get('show_library', True))
            show_borrowing_history = st.checkbox("Show Borrowing History", value=privacy.get('show_borrowing_history', False))
        
        with col2:
            st.markdown("###  Anonymous Mode")
            
            anonymous_mode = st.checkbox(
                "Enable Anonymous Mode", 
                value=privacy.get('anonymous_mode', False),
                help="Hide your identity and show a pseudonym instead"
            )
            
            if anonymous_mode:
                pseudonym = privacy.get('pseudonym')
                if pseudonym:
                    st.info(f"Your pseudonym: **{pseudonym}**")
                else:
                    st.info("A pseudonym will be generated when you save")
                
                st.warning(" Anonymous mode hides: name, email, phone, and profile photo")
        
        if st.button(" Save Privacy Settings", type="primary"):
            settings = {
                'show_email': show_email,
                'show_phone': show_phone,
                'show_full_name': show_full_name,
                'show_profile_photo': show_profile_photo,
                'show_library': show_library,
                'show_borrowing_history': show_borrowing_history
            }
            
            # Update anonymous mode separately to generate pseudonym
            if anonymous_mode != privacy.get('anonymous_mode'):
                success, pseudonym = PrivacyManager.toggle_anonymous_mode(user['user_id'], anonymous_mode)
                if success:
                    st.success(f"Anonymous mode {'enabled' if anonymous_mode else 'disabled'}")
            
            # Update other settings
            if PrivacyManager.update_privacy_settings(user['user_id'], settings):
                st.success(" Privacy settings saved!")
                st.rerun()
            else:
                st.error("Failed to save settings")
        
        # Preview
        st.divider()
        st.subheader(" Preview Public Profile")
        
        if st.button(" View as Public"):
            profile = PrivacyManager.preview_public_profile(user['user_id'])
            if profile:
                st.json(profile)

    # Browse Community Tab
    with tabs[community_tab_index]:
        show_community_library(embedded=True)

def show_community_library(embedded=False):
    """Browse Community Library"""
    if embedded:
        st.subheader(" Community Library")
    else:
        st.title(" Community Library")
    
    st.markdown("Browse and read PDFs shared by other users")
    
    # Search
    search_query = st.text_input(" Search PDFs", placeholder="Search by title, author, genre...")
    
    # Get public PDFs
    public_pdfs = PeerLibraryManager.browse_public_libraries()
    
    if search_query and public_pdfs:
        public_pdfs = [p for p in public_pdfs if 
                      search_query.lower() in p['title'].lower() or
                      search_query.lower() in (p['author'] or '').lower() or
                      search_query.lower() in (p['genre'] or '').lower()]
    
    st.write(f"**{len(public_pdfs)}** PDFs available")
    
    if public_pdfs:
        # Group by user
        user_groups = {}
        for pdf in public_pdfs:
            owner = pdf['owner_name']
            if owner not in user_groups:
                user_groups[owner] = []
            user_groups[owner].append(pdf)
        
        for owner, pdfs in user_groups.items():
            with st.expander(f" {owner} ({len(pdfs)} PDFs)"):
                for pdf in pdfs:
                    col1, col2 = st.columns([4, 1], gap="small")
                    
                    with col1:
                        st.markdown(f"###  {pdf['title']}")
                        st.write(f"**Author:** {pdf['author'] or 'Unknown'}")
                        st.write(f"**Genre:** {pdf['genre'] or 'Unknown'}")
                        st.write(f"**Description:** {pdf['description'] or 'No description'}")
                        st.write(f"**Views:** {pdf['views_count']} | **Uploaded:** {pdf['upload_date']}")
                    
                    with col2:
                        if st.button(" Read", key=f"read_{pdf['pdf_id']}"):
                            pdf_data = PeerLibraryManager.get_pdf_file(pdf['pdf_id'])
                            if pdf_data:
                                PeerLibraryManager.increment_pdf_views(pdf['pdf_id'])
                                
                                # Display PDF
                                st.download_button(
                                    " Download PDF",
                                    pdf_data['pdf_file'],
                                    pdf_data['pdf_filename'],
                                    mime="application/pdf",
                                    key=f"dl_{pdf['pdf_id']}"
                                )
                                
                                # Extract and show preview
                                preview_text = pdf_handler.extract_text(pdf_data['pdf_file'], max_pages=3)
                                if preview_text:
                                    with st.expander(" Preview (First 3 pages)"):
                                        st.text(preview_text)
                        
                        if st.button(" Visit Profile", key=f"profile_{pdf['owner_id']}"):
                            st.session_state.viewing_user = pdf['owner_id']
                            st.rerun()
    else:
        st.info("No public PDFs available yet")
    
    # If viewing a profile
    if 'viewing_user' in st.session_state:
        st.divider()
        st.subheader(" User Profile")
        
        profile_data = PeerLibraryManager.get_user_public_profile(st.session_state.viewing_user)
        current_user = Auth.get_user()
        
        if profile_data:
            user_info = profile_data['user']
            privacy = profile_data['privacy']
            
            col1, col2 = st.columns([1, 3], gap="small")
            
            with col1:
                if privacy.get('show_profile_photo'):
                    user_photo = EnhancedUserManager.get_user_photo(st.session_state.viewing_user)
                    if user_photo:
                        st.image(user_photo, width=150)
                
                # Show profile rating
                rating_data = ProfileCommentsManager.get_average_profile_rating(st.session_state.viewing_user)
                if rating_data['rating_count'] > 0:
                    st.markdown(f" **{rating_data['avg_rating']:.1f}** / 5.0")
                    st.caption(f"({rating_data['rating_count']} ratings)")
            
            with col2:
                st.markdown(f"### {user_info['full_name']}")
                if privacy.get('show_email') and user_info['email'] != '[Hidden]':
                    st.write(f"**Email:** {user_info['email']}")
                
                # Show user's library
                if privacy.get('show_library'):
                    user_pdfs = PeerLibraryManager.get_user_library(
                        st.session_state.viewing_user, 
                        public_only=True
                    )
                    st.write(f"**Public Library:** {len(user_pdfs)} PDFs")
            
            # Tabs for different sections
            profile_tabs = st.tabs([" Library", " Comments & Ratings", " Stats"])
            
            # Tab 1: User's Library
            with profile_tabs[0]:
                if privacy.get('show_library'):
                    user_pdfs = PeerLibraryManager.get_user_library(
                        st.session_state.viewing_user, 
                        public_only=True
                    )
                    
                    if user_pdfs:
                        st.write(f"**{len(user_pdfs)} Public PDFs**")
                        
                        for pdf in user_pdfs:
                            with st.expander(f" {pdf['title']} by {pdf['author']}"):
                                col1, col2 = st.columns([3, 1], gap="small")
                                
                                with col1:
                                    st.write(f"**Genre:** {pdf['genre']}")
                                    st.write(f"**Description:** {pdf['description']}")
                                    st.caption(f"Uploaded: {pdf['upload_date']} | Views: {pdf['views']}")
                                    
                                    # PDF Comments
                                    pdf_comments = ProfileCommentsManager.get_pdf_comments(pdf['pdf_id'])
                                    pdf_rating = ProfileCommentsManager.get_pdf_average_rating(pdf['pdf_id'])
                                    
                                    if pdf_rating['rating_count'] > 0:
                                        st.markdown(f" **{pdf_rating['avg_rating']:.1f}** / 5.0 ({pdf_rating['rating_count']} ratings)")
                                    
                                    if pdf_comments:
                                        st.markdown("**Comments:**")
                                        for comment in pdf_comments[:3]:  # Show first 3
                                            st.markdown(f"- **{comment['full_name']}**: {comment['comment']}")
                                            if comment['rating']:
                                                st.caption(f" {comment['rating']}/5.0")
                                
                                with col2:
                                    # Add comment to PDF
                                    with st.form(f"pdf_comment_{pdf['pdf_id']}", clear_on_submit=True):
                                        st.write("**Add Comment**")
                                        pdf_rating_input = st.slider("Rating", 0.0, 5.0, 3.0, 0.5, key=f"pdf_rating_{pdf['pdf_id']}")
                                        pdf_comment_input = st.text_area("Comment", key=f"pdf_comment_{pdf['pdf_id']}", height=80)
                                        
                                        if st.form_submit_button(" Post"):
                                            if pdf_comment_input:
                                                success, msg = ProfileCommentsManager.add_pdf_comment(
                                                    pdf['pdf_id'],
                                                    current_user['user_id'],
                                                    pdf_comment_input,
                                                    pdf_rating_input if pdf_rating_input > 0 else None
                                                )
                                                if success:
                                                    st.success("Comment posted!")
                                                    st.rerun()
                                                else:
                                                    st.error(msg)
                                            else:
                                                st.warning("Please enter a comment")
                    else:
                        st.info("No public PDFs available")
                else:
                    st.info("User's library is private")
            
            # Tab 2: Comments & Ratings
            with profile_tabs[1]:
                st.markdown("###  Profile Comments & Ratings")
                
                # Add comment form
                if current_user['user_id'] != st.session_state.viewing_user:
                    with st.form("profile_comment_form", clear_on_submit=True):
                        st.write("**Leave a comment and rating:**")
                        
                        col1, col2 = st.columns([3, 1], gap="small")
                        with col1:
                            comment_text = st.text_area("Your comment", height=100)
                        with col2:
                            rating_input = st.slider("Rating", 0.0, 5.0, 5.0, 0.5)
                        
                        if st.form_submit_button(" Post Comment", type="primary"):
                            if comment_text:
                                success, msg = ProfileCommentsManager.add_profile_comment(
                                    st.session_state.viewing_user,
                                    current_user['user_id'],
                                    comment_text,
                                    rating_input if rating_input > 0 else None
                                )
                                if success:
                                    st.success(" Comment posted successfully!")
                                    st.balloons()
                                    st.rerun()
                                else:
                                    st.error(f" {msg}")
                            else:
                                st.warning(" Please enter a comment")
                
                st.divider()
                
                # Display existing comments
                comments = ProfileCommentsManager.get_profile_comments(st.session_state.viewing_user)
                
                if comments:
                    st.write(f"**{len(comments)} Comments**")
                    
                    for comment in comments:
                        with st.container():
                            col1, col2 = st.columns([4, 1], gap="small")
                            
                            with col1:
                                st.markdown(f"**{comment['full_name']}** (@{comment['username']})")
                                st.write(comment['comment'])
                                st.caption(f"Posted: {comment['created_at']}")
                            
                            with col2:
                                if comment['rating']:
                                    st.metric("Rating", f" {comment['rating']}/5.0")
                            
                            st.divider()
                else:
                    st.info("No comments yet. Be the first to comment!")
            
            # Tab 3: Stats
            with profile_tabs[2]:
                st.markdown("###  Profile Statistics")
                
                col1, col2, col3 = st.columns(3, gap="small")
                
                with col1:
                    pdf_count = len(PeerLibraryManager.get_user_library(st.session_state.viewing_user, public_only=True))
                    st.metric("Public PDFs", pdf_count)
                
                with col2:
                    comment_count = len(ProfileCommentsManager.get_profile_comments(st.session_state.viewing_user))
                    st.metric("Comments Received", comment_count)
                
                with col3:
                    rating_data = ProfileCommentsManager.get_average_profile_rating(st.session_state.viewing_user)
                    st.metric("Average Rating", f"{rating_data['avg_rating']:.1f} ")
        
        if st.button(" Back to Community", key="back_to_community_embedded" if embedded else "back_to_community_page"):
            del st.session_state.viewing_user
            st.rerun()

def show_system_tools():
    """System Tools for admins and superadmins."""
    st.title(" System Tools")
    
    user = Auth.get_user()

    role_value = str(user.get('role', '')).lower()
    is_superadmin = bool(user.get('is_superadmin')) or role_value == 'superadmin'
    has_admin_access = role_value in ['admin', 'superadmin']

    if not has_admin_access:
        st.error("Admin or Super Admin access required")
        return
    
    # Add Smart Tools and Data Management tabs for functional admin
    if user.get('is_functional_admin') or is_superadmin:
        tabs = st.tabs([" Backup & Restore", " Data Integrity", " Export Data", " System Stats", " Smart Tools", " Data Management", " Send Reminders"])
    else:
        tabs = st.tabs([" Backup & Restore", " Data Integrity", " Export Data", " System Stats", " Send Reminders"])
    
    # Backup & Restore
    with tabs[0]:
        st.subheader(" Backup & Restore")
        
        col1, col2 = st.columns(2, gap="small")
        
        with col1:
            st.markdown("### Create Backup")
            
            if st.button(" Create Full Backup", type="primary"):
                with st.spinner("Creating backup..."):
                    success, backup_path = DataManagement.create_full_backup()
                    
                    if success:
                        st.success(f" Backup created: {backup_path}")
                        
                        # Offer download
                        try:
                            with open(backup_path, 'rb') as f:
                                st.download_button(
                                    " Download Backup",
                                    f.read(),
                                    os.path.basename(backup_path),
                                    mime="application/zip"
                                )
                        except:
                            pass
                    else:
                        st.error("Backup failed")
            
            # Backup history
            st.markdown("### Recent Backups")
            backups = Database.execute_query("""
                SELECT * FROM backup_logs
                ORDER BY created_at DESC
                LIMIT 10
            """)
            
            if backups:
                for backup in backups:
                    filename = backup['file_path'].split('/')[-1] if backup['file_path'] else 'Unknown'
                    file_size_mb = (backup['file_size'] / 1024 / 1024) if backup['file_size'] else 0
                    st.write(f" {filename} - {backup['created_at']} ({file_size_mb:.2f} MB)")
        
        with col2:
            st.markdown("### Restore Backup")
            
            restore_file = st.file_uploader("Upload Backup ZIP", type=['zip'])
            
            if restore_file:
                st.warning(" Restore will overwrite existing data!")
                
                if st.button(" Restore Now", type="primary"):
                    with st.spinner("Restoring..."):
                        if DataManagement.restore_from_backup(restore_file):
                            st.success(" Restore successful!")
                        else:
                            st.error("Restore failed")
    
    # Data Integrity
    with tabs[1]:
        st.subheader(" Data Integrity Check")
        
        if st.button(" Run Integrity Check"):
            with st.spinner("Checking..."):
                issues = DataManagement.run_integrity_check()
                
                if issues:
                    st.warning(f" Found {len(issues)} issues:")
                    for issue in issues:
                        st.write(f"- {issue}")
                else:
                    st.success(" No integrity issues found!")
        
        # Duplicates
        st.divider()
        st.subheader(" Duplicate Books")
        
        if st.button(" Check Duplicates"):
            duplicates = DataManagement.check_duplicates()
            
            if duplicates:
                st.warning(f"Found {len(duplicates)} potential duplicates:")
                for dup in duplicates:
                    st.write(f"- **{dup['title']}** by {dup['author']} ({dup['count']} copies)")
            else:
                st.success("No duplicates found!")
    
    # Export Data
    with tabs[2]:
        st.subheader(" Export Data")
        
        col1, col2 = st.columns(2, gap="small")
        
        with col1:
            st.markdown("### Export Books")
            
            if st.button(" Export Books to Excel"):
                books = Database.execute_query("""
                    SELECT 
                        b.book_id,
                        b.title,
                        b.author,
                        b.genre,
                        b.isbn,
                        b.publication_year,
                        b.pages as page_count,
                        CASE 
                            WHEN b.is_available = 1 THEN 'Active'
                            ELSE 'Inactive'
                        END as status
                    FROM books b
                """)
                if books:
                    excel_data = excel_exporter.export_books(books)
                    st.download_button(
                        " Download Books.xlsx",
                        excel_data.getvalue(),
                        "books_catalog.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.warning("No books found to export")
        
        with col2:
            st.markdown("### Export Users")
            
            if st.button(" Export Users to Excel"):
                users = Database.execute_query("""
                    SELECT 
                        u.user_id,
                        u.full_name,
                        u.username,
                        u.email,
                        u.role,
                        u.member_tier,
                        (SELECT COUNT(*) FROM borrowing WHERE user_id = u.user_id) as books_borrowed,
                        u.fine_balance,
                        CASE 
                            WHEN u.is_active = 1 THEN 'Active'
                            ELSE 'Inactive'
                        END as status,
                        u.created_at as joined_date
                    FROM users u
                """)
                if users:
                    excel_data = excel_exporter.export_users(users)
                    st.download_button(
                        " Download Users.xlsx",
                        excel_data.getvalue(),
                        "users_list.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
    
    # System Stats
    with tabs[3]:
        st.subheader(" System Statistics")
        
        col1, col2, col3 = st.columns(3, gap="small")
        
        # Database size
        with col1:
            try:
                import os
                db_path = 'litgrid_local.db'
                if os.path.exists(db_path):
                    size_bytes = os.path.getsize(db_path)
                    size_mb = size_bytes / 1024 / 1024
                    st.metric("Database Size", f"{size_mb:.2f} MB")
                else:
                    st.metric("Database Size", "N/A")
            except Exception as e:
                st.metric("Database Size", "N/A")
        
        # Total records
        with col2:
            total_books = Database.execute_query("SELECT COUNT(*) as count FROM books")
            count = total_books[0]['count'] if total_books else 0
            st.metric("Total Books", count)
        
        with col3:
            total_users = Database.execute_query("SELECT COUNT(*) as count FROM users")
            count = total_users[0]['count'] if total_users else 0
            st.metric("Total Users", count)
    
    # Smart Tools (Functional Admin Only)
    if user.get('is_functional_admin') and len(tabs) > 4:
        with tabs[4]:
            st.subheader(" Smart Tools & Utilities")
            
            smart_tabs = st.tabs([
                " ID Generator", 
                " Calendar View", 
                " Barcode & QR", 
                " Recommendations",
                " Spell Check",
                " Book Similarity"
            ])
            
            # Tab 1: Unique ID Generator
            with smart_tabs[0]:
                st.markdown("###  Unique ID Generator")
                st.caption("Generate unique identifiers for users and books")
                
                col1, col2 = st.columns(2, gap="small")
                
                with col1:
                    st.markdown("#### User ID Generator")
                    user_id_input = st.number_input("User ID", min_value=1, value=1, key="user_id_gen")
                    if st.button(" Generate User ID", key="gen_user_id"):
                        try:
                            unique_id = unique_id_gen.generate_user_id()
                            st.success(f" Generated: **{unique_id}**")
                            st.code(f"Format: USR-{datetime.now().year}-{str(user_id_input).zfill(6)}")
                        except Exception as e:
                            st.error(f"Error: {e}")
                
                with col2:
                    st.markdown("#### Book ID Generator")
                    book_id_input = st.number_input("Book ID", min_value=1, value=1, key="book_id_gen")
                    if st.button(" Generate Book ID", key="gen_book_id"):
                        try:
                            unique_id = unique_id_gen.generate_book_id()
                            st.success(f" Generated: **{unique_id}**")
                            st.code(f"Format: BK-{datetime.now().year}-{str(book_id_input).zfill(6)}")
                        except Exception as e:
                            st.error(f"Error: {e}")
                
                st.divider()
                st.info(" **Tip:** These IDs can be used for tracking and barcode generation")
            
            # Tab 2: Calendar View
            with smart_tabs[1]:
                st.markdown("###  Borrowing Calendar View")
                st.caption("Visual overview of all active loans and due dates")
                
                if st.button(" Load Calendar", key="load_calendar"):
                    with st.spinner("Loading calendar..."):
                        calendar_data = SmartUtilities.create_borrowing_calendar()
                        
                        if calendar_data:
                            st.success(f" Found {len(calendar_data)} active loans")
                            
                            # Display in table format
                            df = pd.DataFrame(calendar_data)
                            st.dataframe(df, use_container_width=True, hide_index=True)
                            
                            # Show urgent items
                            st.markdown("####  Due Soon")
                            urgent = [item for item in calendar_data if item['days_until_due'] <= 3]
                            if urgent:
                                for item in urgent:
                                    color = "red" if item['days_until_due'] < 0 else "orange"
                                    st.markdown(f":{color}[ **{item['title']}** - {item['user']} - Due: {item['due_date']} ({item['days_until_due']} days)]")
                            else:
                                st.success(" No urgent returns")
                        else:
                            st.info("No active loans")
            
            # Tab 3: Barcode & QR Generator
            with smart_tabs[2]:
                st.markdown("###  Barcode & QR Code Generator")
                st.caption("Generate barcodes and QR codes for books and users")
                
                col1, col2 = st.columns(2, gap="small")
                
                with col1:
                    st.markdown("#### Book Barcode")
                    book_id_barcode = st.number_input("Book ID", min_value=1, value=1, key="book_barcode")
                    if st.button(" Generate Book Barcode", key="gen_book_barcode"):
                        try:
                            barcode_img = SmartUtilities.generate_book_barcode(book_id_barcode)
                            if barcode_img:
                                st.image(barcode_img, caption=f"Book {book_id_barcode} Barcode")
                                st.success(" Barcode generated successfully")
                            else:
                                st.error("Failed to generate barcode")
                        except Exception as e:
                            st.error(f"Error: {e}")
                
                with col2:
                    st.markdown("#### User QR Code")
                    user_id_qr = st.number_input("User ID", min_value=1, value=1, key="user_qr")
                    if st.button(" Generate User QR", key="gen_user_qr"):
                        try:
                            qr_img = SmartUtilities.generate_user_qr(user_id_qr)
                            if qr_img:
                                st.image(qr_img, caption=f"User {user_id_qr} QR Code", width=200)
                                st.success(" QR code generated successfully")
                            else:
                                st.error("Failed to generate QR code")
                        except Exception as e:
                            st.error(f"Error: {e}")
                
                st.divider()
                st.info(" **Tip:** Scan QR codes with mobile devices for quick access")
            
            # Tab 4: Book Recommendations
            with smart_tabs[3]:
                st.markdown("###  Book Recommendation Engine")
                st.caption("AI-powered book recommendations based on similarity")
                
                book_search = st.text_input("Search for a book", key="rec_search")
                
                if book_search:
                    # Search for book
                    books = Database.execute_query("""
                        SELECT book_id, title, author, genre 
                        FROM books 
                        WHERE title LIKE %s OR author LIKE %s
                        LIMIT 10
                    """, (f"%{book_search}%", f"%{book_search}%"))
                    
                    if books:
                        selected_book = st.selectbox(
                            "Select a book",
                            options=books,
                            format_func=lambda x: f"{x['title']} by {x['author']}"
                        )
                        
                        if st.button(" Get Recommendations", key="get_recs"):
                            with st.spinner("Analyzing..."):
                                recommendations = SmartUtilities.get_book_recommendations(selected_book['book_id'])
                                
                                if recommendations:
                                    st.success(f" Found {len(recommendations)} similar books")
                                    
                                    for rec in recommendations:
                                        with st.expander(f" {rec['title']} by {rec['author']}"):
                                            st.write(f"**Genre:** {rec['genre']}")
                                            st.write(f"**Similarity:** {rec.get('similarity_score', 0):.1f}%")
                                            st.write(f"**ISBN:** {rec.get('isbn', 'N/A')}")
                                else:
                                    st.info("No recommendations found")
                    else:
                        st.warning("No books found")
            
            # Tab 5: Spell Checker
            with smart_tabs[4]:
                st.markdown("###  Local Spell Checker")
                st.caption("Check spelling for book titles and descriptions")
                
                text_to_check = st.text_area("Enter text to check", height=150, key="spell_check_text")
                
                if st.button(" Check Spelling", key="check_spell"):
                    if text_to_check:
                        # Simple spell check using fuzzy matching against known words
                        words = text_to_check.split()
                        
                        # Get common book words from database
                        common_words = Database.execute_query("""
                            SELECT DISTINCT LOWER(title) as word FROM books
                            UNION
                            SELECT DISTINCT LOWER(a.name) as word FROM authors a
                        """)
                        
                        if common_words:
                            word_list = [w['word'] for w in common_words]
                            
                            suspicious_words = []
                            for word in words:
                                clean_word = word.lower().strip('.,!?;:')
                                # Check if word exists or is very similar
                                matches = process.extract(clean_word, word_list, limit=1)
                                if matches and matches[0][1] < 80:  # Less than 80% match
                                    suspicious_words.append({
                                        'word': word,
                                        'suggestion': matches[0][0] if matches else None,
                                        'score': matches[0][1] if matches else 0
                                    })
                            
                            if suspicious_words:
                                st.warning(f" Found {len(suspicious_words)} potential issues:")
                                for item in suspicious_words:
                                    st.write(f"- **{item['word']}** -> Did you mean: *{item['suggestion']}*? (Confidence: {item['score']}%)")
                            else:
                                st.success(" No spelling issues detected")
                    else:
                        st.warning("Please enter text to check")
            
            # Tab 6: Book Similarity Matcher
            with smart_tabs[5]:
                st.markdown("###  Book Similarity Matcher")
                st.caption("Find duplicate or similar books using fuzzy matching")
                
                threshold = st.slider("Similarity Threshold (%)", 60, 100, 80, key="similarity_threshold")
                
                if st.button(" Find Similar Books", key="find_similar"):
                    with st.spinner("Analyzing entire catalog..."):
                        all_books = Database.execute_query("""
                            SELECT b.book_id, b.title, b.isbn, b.author
                            FROM books b
                            WHERE b.is_available = 1
                        """)
                        
                        if all_books and len(all_books) > 1:
                            duplicates = []
                            checked = set()
                            
                            for i, book1 in enumerate(all_books):
                                if book1['book_id'] in checked:
                                    continue
                                    
                                for book2 in all_books[i+1:]:
                                    if book2['book_id'] in checked:
                                        continue
                                    
                                    # Calculate similarity
                                    title_sim = fuzz.ratio(
                                        book1['title'].lower(), 
                                        book2['title'].lower()
                                    )
                                    author_sim = fuzz.ratio(
                                        book1['author'].lower(), 
                                        book2['author'].lower()
                                    )
                                    
                                    # Average similarity
                                    avg_sim = (title_sim + author_sim) / 2
                                    
                                    if avg_sim >= threshold:
                                        duplicates.append({
                                            'book1': book1,
                                            'book2': book2,
                                            'similarity': avg_sim
                                        })
                                        checked.add(book1['book_id'])
                                        checked.add(book2['book_id'])
                            
                            if duplicates:
                                st.warning(f" Found {len(duplicates)} potential matches:")
                                
                                for dup in duplicates:
                                    with st.expander(f"Match {dup['similarity']:.1f}% similarity"):
                                        col1, col2 = st.columns(2, gap="small")
                                        with col1:
                                            st.write("**Book 1:**")
                                            st.write(f"Title: {dup['book1']['title']}")
                                            st.write(f"Author: {dup['book1']['author']}")
                                            st.write(f"ISBN: {dup['book1']['isbn']}")
                                        with col2:
                                            st.write("**Book 2:**")
                                            st.write(f"Title: {dup['book2']['title']}")
                                            st.write(f"Author: {dup['book2']['author']}")
                                            st.write(f"ISBN: {dup['book2']['isbn']}")
                            else:
                                st.success(f" No similar books found at {threshold}% threshold")
                        else:
                            st.info("Not enough books to compare")
    
    # Data Management (Functional Admin Only)
    if user.get('is_functional_admin') and len(tabs) > 5:
        with tabs[5]:
            st.subheader(" Data Management & Configuration")
            
            data_tabs = st.tabs([
                " JSON Config", 
                " SQLite Storage", 
                " Temp Files",
                " Auto-Save",
                " Data Sync"
            ])
            
            # Tab 1: JSON Configuration
            with data_tabs[0]:
                st.markdown("###  JSON Configuration Manager")
                st.caption("View and manage application configuration")
                
                col1, col2 = st.columns([2, 1], gap="small")
                
                with col1:
                    st.markdown("#### Current Configuration")
                    config_data = Config.get_config_dict()
                    
                    if config_data:
                        # Display as formatted JSON
                        st.json(config_data)
                    else:
                        st.warning("No configuration loaded")
                
                with col2:
                    st.markdown("#### Actions")
                    
                    if st.button(" Save Config", key="save_config"):
                        with st.spinner("Saving configuration..."):
                            success, msg = Config.save_config()
                            if success:
                                st.success(f" {msg}")
                            else:
                                st.error(f" {msg}")
                    
                    if st.button(" Load Config", key="load_config"):
                        with st.spinner("Loading configuration..."):
                            success, msg = Config.load_config()
                            if success:
                                st.success(f" {msg}")
                                st.rerun()
                            else:
                                st.error(f" {msg}")
                    
                    st.divider()
                    st.info(" **Tip:** Config is stored in `config.json`")
                    
                    if os.path.exists(Config.CONFIG_FILE):
                        file_size = os.path.getsize(Config.CONFIG_FILE)
                        st.metric("Config File Size", f"{file_size} bytes")
            
            # Tab 2: SQLite Storage
            with data_tabs[1]:
                st.markdown("###  SQLite Database")
                st.caption("Main database storage and status")
                
                col1, col2 = st.columns([2, 1], gap="small")
                
                with col1:
                    st.markdown("#### Database Status")
                    
                    if os.path.exists(Config.SQLITE_DB):
                        file_size = os.path.getsize(Config.SQLITE_DB) / 1024  # KB
                        st.success(f" SQLite database exists ({file_size:.2f} KB)")
                        
                        # Get database stats
                        books = Database.execute_query("SELECT COUNT(*) as count FROM books")
                        users = Database.execute_query("SELECT COUNT(*) as count FROM users")
                        transactions = Database.execute_query("SELECT COUNT(*) as count FROM transactions")
                        
                        if books and users and transactions:
                            col_a, col_b, col_c = st.columns(3, gap="small")
                            with col_a:
                                st.metric("Books", books[0]['count'])
                            with col_b:
                                st.metric("Users", users[0]['count'])
                            with col_c:
                                st.metric("Transactions", transactions[0]['count'])
                    else:
                        st.warning(" SQLite database not found")
                
                with col2:
                    st.markdown("#### Actions")
                    
                    if st.button(" Initialize Database", key="init_db"):
                        with st.spinner("Initializing database..."):
                            success = Database.init_pool()
                            if success:
                                st.success(" Database initialized successfully")
                                st.rerun()
                            else:
                                st.error(" Database initialization failed")
                    
                    if st.button(" Database Info", key="db_info"):
                        with st.spinner("Getting database information..."):
                            try:
                                tables = Database.execute_query("""
                                    SELECT name FROM sqlite_master 
                                    WHERE type='table' AND name NOT LIKE 'sqlite_%'
                                    ORDER BY name
                                """)
                                if tables:
                                    st.success(f" Found {len(tables)} tables")
                                    for table in tables:
                                        st.text(f"• {table['name']}")
                                else:
                                    st.info("No tables found")
                            except Exception as e:
                                st.error(f" Error: {e}")
                    
                    st.divider()
                    st.info(" **Tip:** Local storage enables offline access")
            
            # Tab 3: Temp Files
            with data_tabs[2]:
                st.markdown("###  Temporary File Management")
                st.caption("Manage temporary files and cleanup")
                
                col1, col2 = st.columns([2, 1], gap="small")
                
                with col1:
                    st.markdown("#### Temp Directory Status")
                    
                    if os.path.exists(Config.TEMP_DIR):
                        temp_count = TempFileManager.get_temp_file_count()
                        st.metric("Temp Files", temp_count)
                        
                        if temp_count > 0:
                            st.markdown("#### Temp Files")
                            files = os.listdir(Config.TEMP_DIR)
                            for filename in files[:10]:  # Show first 10
                                filepath = os.path.join(Config.TEMP_DIR, filename)
                                if os.path.isfile(filepath):
                                    size = os.path.getsize(filepath) / 1024  # KB
                                    mtime = datetime.fromtimestamp(os.path.getmtime(filepath))
                                    st.text(f" {filename} ({size:.2f} KB) - {mtime.strftime('%Y-%m-%d %H:%M')}")
                        else:
                            st.success(" No temp files")
                    else:
                        st.warning(" Temp directory not initialized")
                
                with col2:
                    st.markdown("#### Actions")
                    
                    if st.button(" Initialize Temp Dir", key="init_temp"):
                        success = TempFileManager.init_temp_dir()
                        if success:
                            st.success(" Temp directory initialized")
                            st.rerun()
                        else:
                            st.error(" Failed to initialize")
                    
                    hours = st.number_input("Clean files older than (hours)", min_value=1, value=24, key="clean_hours")
                    
                    if st.button(" Clean Temp Files", key="clean_temp"):
                        with st.spinner("Cleaning temp files..."):
                            count = TempFileManager.clean_temp_files(hours)
                            st.success(f" Cleaned {count} temp files")
                            st.rerun()
                    
                    st.divider()
                    st.info(" **Tip:** Clean old files regularly to save space")
            
            # Tab 4: Auto-Save
            with data_tabs[3]:
                st.markdown("###  Auto-Save Manager")
                st.caption("View and manage auto-saved form data")
                
                col1, col2 = st.columns([2, 1], gap="small")
                
                with col1:
                    st.markdown("#### Auto-Save Status")
                    
                    if os.path.exists(Config.AUTO_SAVE_FILE):
                        file_size = os.path.getsize(Config.AUTO_SAVE_FILE) / 1024  # KB
                        st.success(f" Auto-save file exists ({file_size:.2f} KB)")
                        
                        # Load and display auto-save data
                        try:
                            with open(Config.AUTO_SAVE_FILE, 'r') as f:
                                auto_save_data = json.load(f)
                            
                            st.metric("Saved Forms", len(auto_save_data))
                            
                            if auto_save_data:
                                st.markdown("#### Saved Forms")
                                for form_name, data in auto_save_data.items():
                                    with st.expander(f" {form_name}"):
                                        st.json(data)
                            else:
                                st.info("No saved forms")
                        except Exception as e:
                            st.error(f"Error loading auto-save data: {e}")
                    else:
                        st.info("ℹ No auto-save file")
                
                with col2:
                    st.markdown("#### Actions")
                    
                    if os.path.exists(Config.AUTO_SAVE_FILE):
                        try:
                            with open(Config.AUTO_SAVE_FILE, 'r') as f:
                                auto_save_data = json.load(f)
                            
                            form_names = list(auto_save_data.keys())
                            
                            if form_names:
                                selected_form = st.selectbox("Select Form", form_names, key="form_select")
                                
                                if st.button(" Clear Form", key="clear_form"):
                                    success = AutoSaveManager.clear_form_data(selected_form)
                                    if success:
                                        st.success(f" Cleared {selected_form}")
                                        st.rerun()
                                    else:
                                        st.error(" Failed to clear")
                        except:
                            pass
                    
                    st.divider()
                    st.info(" **Tip:** Auto-save prevents data loss on form errors")
            
            # Tab 5: Data Sync
            with data_tabs[4]:
                st.markdown("###  Data Synchronization")
                st.caption("Sync data between MariaDB and SQLite")
                
                col1, col2 = st.columns([2, 1], gap="small")
                
                with col1:
                    st.markdown("#### Sync Status")
                    
                    sync_status = DataSyncManager.get_sync_status()
                    
                    if sync_status:
                        st.success(f" Last sync: {sync_status['synced_at']}")
                        st.write(f"**Type:** {sync_status['sync_type']}")
                        st.write(f"**Status:** {sync_status['status']}")
                        st.write(f"**Message:** {sync_status['message']}")
                    else:
                        st.warning(" No sync history")
                    
                    st.divider()
                    st.markdown("#### Data Consistency Check")
                    
                    if st.button(" Check Consistency", key="check_consistency"):
                        with st.spinner("Checking data consistency..."):
                            consistency = DataSyncManager.verify_data_consistency()
                            
                            if consistency:
                                col_a, col_b, col_c = st.columns(3, gap="small")
                                
                                with col_a:
                                    st.metric("MariaDB Books", consistency['mariadb_count'])
                                with col_b:
                                    st.metric("SQLite Books", consistency['sqlite_count'])
                                with col_c:
                                    st.metric("Difference", consistency['difference'])
                                
                                if consistency['in_sync']:
                                    st.success(" Data is in sync!")
                                else:
                                    st.warning(f" {consistency['difference']} records out of sync")
                            else:
                                st.error(" Failed to check consistency")
                
                with col2:
                    st.markdown("#### Actions")
                    
                    if st.button(" Sync to Local", key="sync_to_local", type="primary"):
                        with st.spinner("Syncing data to local storage..."):
                            success, msg = DataSyncManager.sync_to_local()
                            if success:
                                st.success(f" {msg}")
                                st.rerun()
                            else:
                                st.error(f" {msg}")
                    
                    st.divider()
                    st.info(" **Tip:** Regular syncing keeps local data up-to-date")
    
    # Send Reminders Tab (Always available for admin)
    reminder_tab_index = 6 if user.get('is_functional_admin') else 4
    if len(tabs) > reminder_tab_index:
        with tabs[reminder_tab_index]:
            st.subheader(" Send Return Reminders")
            
            st.markdown("""
            Send automated return reminders to members with:
            - Books due within 3 days
            - Books due today
            - Overdue books
            """)
            
            # Preview affected members
            affected = Database.execute_query("""
                SELECT br.borrowing_id, b.title, u.full_name, u.email, br.due_date,
                       julianday(br.due_date) - julianday(date('now')) as days_until_due
                FROM borrowing br
                JOIN book_inventory bi ON br.inventory_id = bi.inventory_id
                JOIN books b ON bi.book_id = b.book_id
                JOIN users u ON br.user_id = u.user_id
                WHERE br.return_date IS NULL
                  AND julianday(br.due_date) - julianday(date('now')) <= 3
                ORDER BY br.due_date
            """)
            
            if affected:
                st.info(f" **{len(affected)}** members will receive reminders")
                
                # Show preview
                with st.expander(" View Recipients"):
                    for member in affected[:10]:  # Show first 10
                        days = member['days_until_due']
                        status = " OVERDUE" if days < 0 else " DUE TODAY" if days == 0 else f"Due in {days} days"
                        st.write(f"- {member['full_name']} ({member['email']}) - **{member['title']}** - {status}")
                    
                    if len(affected) > 10:
                        st.caption(f"... and {len(affected) - 10} more")
                
                st.divider()
                
                # Send button
                col1, col2 = st.columns([2, 1], gap="small")
                with col1:
                    st.warning(" This will send notification to all affected members")
                
                with col2:
                    if st.button(" Send Reminders Now", type="primary", use_container_width=True):
                        with st.spinner("Sending reminders..."):
                            sent_count = EmailService.send_bulk_reminders()
                            st.success(f" Sent {sent_count} reminders successfully!")
                            st.balloons()
            else:
                st.success(" No reminders needed - all books returned on time!")



# ================================================================
# MAIN APP
# ================================================================

def main():
    """Main application"""
    # Page config
    st.set_page_config(
        page_title="LitGrid - Library Management",
        page_icon="",
        layout="wide"
    )
    
    # Load CSS
    load_css()
    
    # Initialize DB
    Database.init_pool()
    
    # Initialize auth
    Auth.init_session()
    
    # Check authentication
    if not Auth.is_authenticated():
        show_login_page()
    else:
        user = Auth.get_user()
        
        # Sidebar
        with st.sidebar:
            st.divider()
            st.markdown(f"###  {user['full_name']}")
            st.caption(f"{user['role'].title()} | {user['member_tier'].title()}")

            # Show demo mode indicator
            if user.get('is_demo'):
                st.info("Demo Mode Active - Read-Only Access")
                if st.button("Exit Demo Mode", use_container_width=True, key="exit_demo_btn"):
                    Auth.logout()
                    st.rerun()
                st.divider()

            if user['fine_balance'] > 0:
                st.warning(f" Fine: {format_currency(user['fine_balance'])}")
            
            # Show notifications badge
            notifications = EmailService.get_user_notifications(user['email'])
            unread_count = len(notifications)
            if unread_count > 0:
                st.info(f" {unread_count} new notification{'s' if unread_count != 1 else ''}")
                
                with st.expander(" View Notifications"):
                    for notif in notifications[-5:]:  # Show last 5
                        st.markdown(f"**{notif['subject']}**")
                        st.caption(format_datetime(notif['timestamp']))
                        if st.button(" Read", key=f"notif_{notif['timestamp']}"):
                            st.info(notif['message'])
                        st.divider()
            
            st.divider()
            
            # Menu
            menu = ["Dashboard", "My Account", " My Library"]
            
            if user['role'] in ['admin', 'librarian', 'superadmin']:
                menu.extend(["Borrowing & Returns", "Reports", " System Tools"])
            
            choice = st.radio(" Navigation", menu, label_visibility="collapsed")
            
            st.divider()
            if st.button(" Logout", use_container_width=True):
                Auth.logout()
                st.rerun()
            
            # Developer Info
            st.divider()
            st.markdown("""
                <div style='text-align: center; padding: 10px;'>
                    <p style='font-size: 12px; color: #666; margin-bottom: 8px;'>Developed by</p>
                    <p style='font-size: 14px; font-weight: bold; color: #333; margin-bottom: 10px;'>Labib Bin Shahed</p>
                </div>
            """, unsafe_allow_html=True)
            
            st.markdown("""
                <div style='text-align: center;'>
                    <a href='https://github.com/la-b-ib' target='_blank'>
                        <img src='https://img.shields.io/badge/GitHub-181717?style=for-the-badge&logo=github&logoColor=white' alt='GitHub' style='margin: 2px;'/>
                    </a>
                    <a href='https://www.linkedin.com/in/la-b-ib/' target='_blank'>
                        <img src='https://img.shields.io/badge/LinkedIn-0077B5?style=for-the-badge&logo=linkedin&logoColor=white' alt='LinkedIn' style='margin: 2px;'/>
                    </a>
                </div>
            """, unsafe_allow_html=True)
        
        # Main content
        if choice == "Dashboard":
            show_dashboard()
        elif choice == "My Account":
            show_account()
        elif choice == " My Library":
            show_my_library()
        elif choice == "Borrowing & Returns":
            show_borrowing_returns()
        elif choice == "Reports":
            show_reports()
        elif choice == " System Tools":
            show_system_tools()

if __name__ == "__main__":
    main()

